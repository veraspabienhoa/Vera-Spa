"""Authenticated official-rules routes for VERA SPA Web V2.

PostgreSQL ``official_policy/leave_rules`` is canonical.  The complete dynamic
grid is kept as one versioned JSON document and mirrored to the legacy
``LoaiNghi`` worksheet so Streamlit and Web V2 continue to enforce the same
rules while the migration is in progress.
"""
from __future__ import annotations

import hashlib
from io import BytesIO
import json
import re
from typing import Any, Callable
from urllib.parse import quote

from fastapi import Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import text


CATEGORY = "official_policy"
SETTING_KEY = "leave_rules"
SNAPSHOT_CATEGORY = "leave_rules"
SNAPSHOT_KEY = "loai_nghi_snapshot_v2"
REQUIRED_COLUMNS = (
    "STT",
    "Lý do nghỉ",
    "Loại nghỉ",
    "Số ngày tính phép",
    "Phạt vi phạm",
)
DEFAULT_COLUMNS = [
    "STT", "Lý do nghỉ", "Loại nghỉ", "Chi tiết", "Số ngày tính phép",
    "Phạt vi phạm", "Chỉ nhập được cuối tuần", "User có quyền được nhập",
    "Kiều đăng ký", "Giá trị", "Ngoại trừ đăng ký", "Kiểu hủy",
    "Số ngày hủy trước", "Ngoại trừ hủy", "Ghi chú",
]
MAX_COLUMNS = 100
MAX_ROWS = 2000
MAX_CELL_LENGTH = 5000
MAX_XLSX_BYTES = 5 * 1024 * 1024
RULES_FEATURES = (
    "official_rules_view", "official_rules_edit",
    "official_rules_export", "official_rules_import",
)
DAILY_QUOTA_CATEGORY = "leave_rules"
DAILY_QUOTA_KEY = "daily_quota"
AUTO_CHECK_CATEGORY = "auto_check"
AUTO_CHECK_KEY = "config"
DEFAULT_LATE_THRESHOLD_MINUTES = 5
AUTO_CHECK_STATUS_DEFAULT = "RUNNING"
DAILY_QUOTA_WEEKDAYS = (
    (1, "Thứ 2", 5, 2),
    (2, "Thứ 3", 5, 2),
    (3, "Thứ 4", 5, 2),
    (4, "Thứ 5", 5, 2),
    (5, "Thứ 6", 5, 2),
    (6, "Thứ 7", 3, 0),
    (7, "Chủ nhật", 3, 0),
)


class RulesUpdate(BaseModel):
    columns: list[str] = Field(min_length=1, max_length=MAX_COLUMNS)
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=MAX_ROWS)
    expected_revision: int = Field(ge=0)


class DailyQuotaDay(BaseModel):
    weekday: int = Field(ge=1, le=7)
    paid_limit: int = Field(ge=0, le=100)
    generated_limit: int = Field(ge=0, le=100)


class DailyQuotaUpdate(BaseModel):
    days: list[DailyQuotaDay] = Field(min_length=7, max_length=7)
    expected_revision: int = Field(ge=0)


class LateThresholdUpdate(BaseModel):
    threshold_minutes: int = Field(ge=5, le=180)
    expected_revision: int = Field(ge=0)


def _load_late_threshold(conn) -> dict[str, Any]:
    row = conn.execute(text("""
        SELECT value_json, revision, updated_at, updated_by
        FROM vera_app_setting
        WHERE category=:category AND setting_key=:setting_key
        LIMIT 1
    """), {"category": AUTO_CHECK_CATEGORY, "setting_key": AUTO_CHECK_KEY}).mappings().first()
    value = dict(row["value_json"] or {}) if row else {}
    try:
        threshold = max(5, min(180, int(value.get("threshold_minutes", DEFAULT_LATE_THRESHOLD_MINUTES))))
    except (TypeError, ValueError):
        threshold = DEFAULT_LATE_THRESHOLD_MINUTES
    return {
        "threshold_minutes": threshold,
        "revision": int(row["revision"] or 0) if row else 0,
        "updated_at": row["updated_at"].isoformat() if row and row["updated_at"] else "",
        "updated_by": str(row["updated_by"] or "") if row else "",
    }


def _default_daily_quota_days() -> list[dict[str, Any]]:
    return [
        {
            "weekday": weekday,
            "weekday_label": label,
            "paid_limit": paid,
            "generated_limit": generated,
        }
        for weekday, label, paid, generated in DAILY_QUOTA_WEEKDAYS
    ]


def _normalize_daily_quota(value: Any) -> list[dict[str, Any]]:
    defaults = _default_daily_quota_days()
    default_by_day = {item["weekday"]: item for item in defaults}
    raw = value if isinstance(value, dict) else {}
    raw_days = raw.get("days") if isinstance(raw.get("days"), list) else []
    parsed: dict[int, dict[str, Any]] = {}
    for item in raw_days:
        if not isinstance(item, dict):
            continue
        try:
            weekday = int(item.get("weekday") or 0)
            if weekday not in default_by_day or weekday in parsed:
                continue
            parsed[weekday] = {
                **default_by_day[weekday],
                "paid_limit": max(0, min(100, int(float(item.get("paid_limit", default_by_day[weekday]["paid_limit"]))))),
                "generated_limit": max(0, min(100, int(float(item.get("generated_limit", default_by_day[weekday]["generated_limit"]))))),
            }
        except (TypeError, ValueError):
            continue

    if len(parsed) == 7:
        return [parsed[index] for index in range(1, 8)]

    # Read the three-key configuration used by the current Streamlit app and
    # earlier Web V2 builds. This makes the first UI load non-destructive.
    try:
        weekday_limit = max(0, min(100, int(float(raw.get("weekday_limit", 5)))))
    except (TypeError, ValueError):
        weekday_limit = 5
    try:
        weekend_limit = max(0, min(100, int(float(raw.get("weekend_limit", 3)))))
    except (TypeError, ValueError):
        weekend_limit = 3
    try:
        generated_limit = max(0, min(100, int(float(raw.get("phat_sinh_limit", 2)))))
    except (TypeError, ValueError):
        generated_limit = 2
    return [
        {
            **item,
            "paid_limit": weekend_limit if item["weekday"] >= 6 else weekday_limit,
            "generated_limit": 0 if item["weekday"] >= 6 else generated_limit,
        }
        for item in defaults
    ]


def _load_daily_quota(conn) -> dict[str, Any]:
    row = conn.execute(text("""
        SELECT value_json, revision, updated_at, updated_by
        FROM vera_app_setting
        WHERE category=:category AND setting_key=:setting_key
        LIMIT 1
    """), {"category": DAILY_QUOTA_CATEGORY, "setting_key": DAILY_QUOTA_KEY}).mappings().first()
    value = row["value_json"] if row else {}
    return {
        "days": _normalize_daily_quota(value),
        "revision": int(row["revision"] or 0) if row else 0,
        "updated_at": row["updated_at"].isoformat() if row and row["updated_at"] else "",
        "updated_by": str(row["updated_by"] or "") if row else "",
    }


def _daily_quota_payload(days: list[DailyQuotaDay]) -> dict[str, Any]:
    items = [item.model_dump() for item in days]
    if {item["weekday"] for item in items} != set(range(1, 8)):
        raise HTTPException(400, "Bảng hạn mức phải có đủ và đúng một dòng từ Thứ 2 đến Chủ nhật.")
    ordered = sorted(items, key=lambda item: item["weekday"])
    labels = {weekday: label for weekday, label, _paid, _generated in DAILY_QUOTA_WEEKDAYS}
    normalized = [{**item, "weekday_label": labels[item["weekday"]]} for item in ordered]
    weekday_paid = [item["paid_limit"] for item in normalized if item["weekday"] <= 5]
    weekend_paid = [item["paid_limit"] for item in normalized if item["weekday"] >= 6]
    weekday_generated = [item["generated_limit"] for item in normalized if item["weekday"] <= 5]
    return {
        "days": normalized,
        # Keep the old keys so the current Streamlit version remains safe.
        # When individual days differ, the legacy screen uses the strictest
        # value while Web V2 uses the exact value for each weekday.
        "weekday_limit": min(weekday_paid),
        "weekend_limit": min(weekend_paid),
        "phat_sinh_limit": min(weekday_generated),
    }


def _write_daily_quota_sheet(ws, payload: dict[str, Any]) -> None:
    old_values = ws.get_all_values()
    rows = [list(row) for row in old_values]
    if not rows:
        rows = [["Key", "Value"]]
    if len(rows[0]) < 2:
        rows[0] += [""] * (2 - len(rows[0]))
    rows[0][0:2] = ["Key", "Value"]
    updates = {
        "weekday_limit": payload["weekday_limit"],
        "weekend_limit": payload["weekend_limit"],
        "phat_sinh_limit": payload["phat_sinh_limit"],
    }
    for item in payload["days"]:
        updates[f"weekday_{item['weekday']}_paid_limit"] = item["paid_limit"]
        updates[f"weekday_{item['weekday']}_generated_limit"] = item["generated_limit"]
    index_by_key = {
        str(row[0] if row else "").strip(): index
        for index, row in enumerate(rows[1:], start=1)
        if row and str(row[0]).strip()
    }
    for key, value in updates.items():
        if key in index_by_key:
            index = index_by_key[key]
            rows[index] += [""] * max(0, 2 - len(rows[index]))
            rows[index][0:2] = [key, value]
        else:
            rows.append([key, value])
    width = max(2, max(len(row) for row in rows))
    padded = [row + [""] * (width - len(row)) for row in rows]
    ws.clear()
    ws.update(
        range_name=f"A1:{get_column_letter(width)}{len(padded)}",
        values=padded,
        value_input_option="USER_ENTERED",
    )


def _clean_scalar(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value != value or value in (float("inf"), float("-inf")):
            return ""
        return value
    if isinstance(value, (list, tuple, dict, set)):
        raise HTTPException(400, "Ô Nội quy chỉ được chứa chữ hoặc số.")
    result = str(value).strip()
    if len(result) > MAX_CELL_LENGTH:
        raise HTTPException(400, f"Một ô Nội quy vượt quá {MAX_CELL_LENGTH} ký tự.")
    return result


def _normalize_document(columns: list[Any], rows: list[Any], *, validate_apply: bool = True) -> dict:
    clean_columns = [str(column or "").strip() for column in columns]
    if not clean_columns or any(not column for column in clean_columns):
        raise HTTPException(400, "Tên cột Nội quy không được để trống.")
    if len(clean_columns) > MAX_COLUMNS:
        raise HTTPException(400, f"Bảng Nội quy tối đa {MAX_COLUMNS} cột.")
    duplicate_columns = sorted({column for column in clean_columns if clean_columns.count(column) > 1})
    if duplicate_columns:
        raise HTTPException(400, "Tên cột bị trùng: " + ", ".join(duplicate_columns[:20]))
    if validate_apply:
        missing = [column for column in REQUIRED_COLUMNS if column not in clean_columns]
        if missing:
            raise HTTPException(400, "Thiếu cột nghiệp vụ bắt buộc: " + ", ".join(missing))
    if len(rows) > MAX_ROWS:
        raise HTTPException(400, f"Bảng Nội quy tối đa {MAX_ROWS} dòng.")

    clean_rows: list[dict[str, Any]] = []
    for row in rows:
        if isinstance(row, dict):
            values = row
        elif isinstance(row, (list, tuple)):
            values = dict(zip(clean_columns, list(row)))
        else:
            values = {}
        clean_rows.append({column: _clean_scalar(values.get(column, "")) for column in clean_columns})

    if validate_apply:
        if not clean_rows:
            raise HTTPException(400, "Nội quy không được để trống.")
        reasons = [str(row.get("Lý do nghỉ", "") or "").strip() for row in clean_rows]
        blank_rows = [str(index + 2) for index, reason in enumerate(reasons) if not reason]
        if blank_rows:
            raise HTTPException(400, "Cột Lý do nghỉ đang trống tại dòng: " + ", ".join(blank_rows[:20]))
        seen: dict[str, str] = {}
        duplicates: list[str] = []
        for reason in reasons:
            key = reason.casefold()
            if key in seen and reason not in duplicates:
                duplicates.append(reason)
            seen[key] = reason
        if duplicates:
            raise HTTPException(400, "Lý do nghỉ bị trùng: " + ", ".join(duplicates[:20]))

    body = {"columns": clean_columns, "rows": clean_rows}
    raw = json.dumps(body, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    body["checksum"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return body


def _snapshot_document(value: Any) -> dict:
    value = value if isinstance(value, dict) else {}
    columns = value.get("columns") or value.get("headers") or DEFAULT_COLUMNS
    rows = value.get("rows") or []
    return _normalize_document(list(columns), list(rows), validate_apply=False)


def _permissions(conn, ident, feature_allowed: Callable[[Any, Any, str], bool]) -> dict[str, bool]:
    return {feature: feature_allowed(conn, ident, feature) for feature in RULES_FEATURES}


def _load_document(conn) -> dict:
    row = conn.execute(text("""
        SELECT value_json, source, updated_by, revision, updated_at
        FROM vera_app_setting
        WHERE category=:category AND setting_key=:setting_key
        LIMIT 1
    """), {"category": CATEGORY, "setting_key": SETTING_KEY}).mappings().first()
    if row:
        document = _snapshot_document(row["value_json"])
        return {
            **document,
            "revision": int(row["revision"] or 0),
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else "",
            "updated_by": str(row["updated_by"] or ""),
            "source": str(row["source"] or "web_v2_rules"),
            "canonical": True,
        }

    fallback = conn.execute(text("""
        SELECT value_json, source, updated_by, revision, updated_at
        FROM vera_app_setting
        WHERE category=:category AND setting_key=:setting_key
        LIMIT 1
    """), {"category": SNAPSHOT_CATEGORY, "setting_key": SNAPSHOT_KEY}).mappings().first()
    if not fallback:
        raise HTTPException(503, "Chưa có dữ liệu Nội quy/LoaiNghi trong PostgreSQL.")
    value = fallback["value_json"] if isinstance(fallback["value_json"], dict) else {}
    document = _snapshot_document(value)
    return {
        **document,
        # Revision zero signals that the first Web V2 save will bootstrap the
        # official canonical document without pretending the snapshot is it.
        "revision": 0,
        "updated_at": str(value.get("updated_at") or fallback["updated_at"] or ""),
        "updated_by": str(value.get("updated_by") or fallback["updated_by"] or ""),
        "source": "loai_nghi_snapshot_v2",
        "canonical": False,
    }


def _sheet_values(document: dict) -> list[list[Any]]:
    columns = document["columns"]
    return [columns] + [
        [row.get(column, "") for column in columns]
        for row in document["rows"]
    ]


def _write_legacy_sheet(ws, document: dict) -> None:
    values = _sheet_values(document)
    wanted_rows = max(100, len(values) + 20)
    wanted_cols = max(15, len(document["columns"]) + 3)
    try:
        ws.resize(
            rows=max(int(getattr(ws, "row_count", 0) or 0), wanted_rows),
            cols=max(int(getattr(ws, "col_count", 0) or 0), wanted_cols),
        )
    except Exception:
        pass
    ws.clear()
    ws.update(
        range_name=f"A1:{get_column_letter(len(document['columns']))}{len(values)}",
        values=values,
        value_input_option="USER_ENTERED",
    )


def _restore_legacy_sheet(ws, values: list[list[Any]]) -> None:
    try:
        ws.clear()
        if values:
            width = max(len(row) for row in values)
            padded = [list(row) + [""] * (width - len(row)) for row in values]
            ws.update(
                range_name=f"A1:{get_column_letter(width)}{len(padded)}",
                values=padded,
                value_input_option="USER_ENTERED",
            )
    except Exception:
        pass


def _safe_excel_value(value: Any) -> Any:
    value = _clean_scalar(value)
    if isinstance(value, str) and re.match(r"^[=+@]", value):
        return "'" + value
    return value


def _excel_bytes(document: dict, revision: int, updated_at: str, updated_by: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NoiQuy"
    header_fill = PatternFill("solid", fgColor="214639")
    header_font = Font(color="FFFFFF", bold=True)
    thin_fill = PatternFill("solid", fgColor="F2F6F3")

    columns = document["columns"]
    sheet.append(columns)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    for row_index, row in enumerate(document["rows"], start=2):
        sheet.append([_safe_excel_value(row.get(column, "")) for column in columns])
        if row_index % 2 == 0:
            for cell in sheet[row_index]:
                cell.fill = thin_fill
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(document['rows']) + 1)}"
    for index, column in enumerate(columns, start=1):
        longest = max([len(str(column))] + [len(str(row.get(column, "") or "")) for row in document["rows"][:300]])
        sheet.column_dimensions[get_column_letter(index)].width = min(42, max(10, longest + 2))
        if column in {"Số ngày tính phép", "Phạt vi phạm", "Giá trị", "Số ngày hủy trước"}:
            for cell in sheet[get_column_letter(index)][1:]:
                cell.number_format = "#,##0.##"

    guide = workbook.create_sheet("HuongDan")
    guide.append(["VERA SPA · BẢNG NỘI QUY"])
    guide.append(["Phiên bản", revision])
    guide.append(["Cập nhật", updated_at])
    guide.append(["Người cập nhật", updated_by])
    guide.append(["Lưu ý", "Giữ nguyên tên các cột bắt buộc. Import chỉ nạp vào vùng chỉnh sửa; cần bấm Ghi thay đổi & áp dụng."])
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90
    guide["A1"].font = Font(bold=True, color="214639", size=14)
    guide.sheet_state = "hidden"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _read_excel(raw: bytes) -> dict:
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=False)
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được file Excel: {exc}") from exc
    if not workbook.sheetnames:
        raise HTTPException(400, "File Excel không có sheet dữ liệu.")
    sheet_name = "LoaiNghi" if "LoaiNghi" in workbook.sheetnames else (
        "NoiQuy" if "NoiQuy" in workbook.sheetnames else workbook.sheetnames[0]
    )
    sheet = workbook[sheet_name]
    if sheet.max_column > MAX_COLUMNS or sheet.max_row > MAX_ROWS + 1:
        raise HTTPException(400, f"File Excel tối đa {MAX_COLUMNS} cột và {MAX_ROWS} dòng dữ liệu.")
    cell_rows = list(sheet.iter_rows(values_only=False))
    if not cell_rows:
        raise HTTPException(400, "File Excel không có dữ liệu.")
    columns = [str(cell.value or "").strip() for cell in cell_rows[0]]
    while columns and not columns[-1]:
        columns.pop()
    if not columns:
        raise HTTPException(400, "Dòng đầu tiên của Excel phải là tên cột.")
    rows = []
    for excel_row in cell_rows[1:]:
        source_cells = list(excel_row[:len(columns)])
        if any(cell.data_type == "f" for cell in source_cells):
            raise HTTPException(400, "Nội quy không nhận ô công thức Excel.")
        cells = [cell.value for cell in source_cells] + [""] * max(0, len(columns) - len(source_cells))
        if not any(str(value or "").strip() for value in cells):
            continue
        rows.append(dict(zip(columns, cells)))
    return _normalize_document(columns, rows, validate_apply=True)


def install_rules_routes(
    app,
    *,
    engine_instance: Callable[[], Any],
    current_identity: Callable[..., Any],
    require_feature: Callable[[Any, Any, str], None],
    feature_allowed: Callable[[Any, Any, str], bool],
    google_client: Callable[[], Any],
    leave_sheet_id: str,
    identity_type: type,
    vn_tz,
) -> None:
    def response_payload(conn, ident) -> dict:
        document = _load_document(conn)
        daily_quota = _load_daily_quota(conn)
        late_threshold = _load_late_threshold(conn)
        is_admin = str(getattr(ident, "role", "") or "").strip().lower() == "admin"
        return {
            **document,
            "required_columns": list(REQUIRED_COLUMNS),
            "permissions": _permissions(conn, ident, feature_allowed),
            "daily_quota": daily_quota,
            "late_threshold": late_threshold,
            "can_edit_daily_quota": is_admin,
            "can_edit_late_threshold": is_admin,
        }

    @app.get("/v2/rules")
    def get_rules(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "official_rules_view")
            return response_payload(conn, ident)

    @app.put("/v2/rules")
    def save_rules(body: RulesUpdate, ident: identity_type = Depends(current_identity)):
        document = _normalize_document(body.columns, body.rows, validate_apply=True)
        conn = engine_instance().connect()
        tx = conn.begin()
        worksheet = None
        old_sheet_values: list[list[Any]] = []
        sheet_changed = False
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:web-v2:official-rules'))"))
            require_feature(conn, ident, "official_rules_edit")
            current = conn.execute(text("""
                SELECT revision
                FROM vera_app_setting
                WHERE category=:category AND setting_key=:setting_key
                FOR UPDATE
            """), {"category": CATEGORY, "setting_key": SETTING_KEY}).scalar_one_or_none()
            current_revision = int(current or 0)
            if current_revision != body.expected_revision:
                raise HTTPException(
                    409,
                    "Bảng Nội quy đã được người khác cập nhật. Hãy bấm Làm mới rồi kiểm tra lại trước khi ghi.",
                )

            if current is None:
                saved = conn.execute(text("""
                    INSERT INTO vera_app_setting(
                        category, setting_key, value_json, source, updated_by,
                        revision, created_at, updated_at
                    ) VALUES (
                        :category, :setting_key, CAST(:value_json AS jsonb),
                        'web_v2_rules', :updated_by, 1, now(), now()
                    )
                    RETURNING revision, updated_at
                """), {
                    "category": CATEGORY,
                    "setting_key": SETTING_KEY,
                    "value_json": json.dumps(document, ensure_ascii=False),
                    "updated_by": ident.employee_username,
                }).mappings().one()
            else:
                saved = conn.execute(text("""
                    UPDATE vera_app_setting
                    SET value_json=CAST(:value_json AS jsonb),
                        source='web_v2_rules', updated_by=:updated_by,
                        revision=revision + 1, updated_at=now()
                    WHERE category=:category AND setting_key=:setting_key
                    RETURNING revision, updated_at
                """), {
                    "category": CATEGORY,
                    "setting_key": SETTING_KEY,
                    "value_json": json.dumps(document, ensure_ascii=False),
                    "updated_by": ident.employee_username,
                }).mappings().one()

            worksheet = google_client().open_by_key(leave_sheet_id).worksheet("LoaiNghi")
            old_sheet_values = worksheet.get_all_values()
            # Mark before clear/update so a partially failed Google request is
            # also restored from the captured worksheet snapshot.
            sheet_changed = True
            _write_legacy_sheet(worksheet, document)
            tx.commit()
            return {
                "ok": True,
                "message": "Đã ghi Bảng nội quy và áp dụng THÀNH CÔNG.",
                "revision": int(saved["revision"]),
                "updated_at": saved["updated_at"].isoformat(),
                "updated_by": ident.employee_username,
                "checksum": document["checksum"],
            }
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            if sheet_changed and worksheet is not None:
                _restore_legacy_sheet(worksheet, old_sheet_values)
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            if sheet_changed and worksheet is not None:
                _restore_legacy_sheet(worksheet, old_sheet_values)
            raise HTTPException(500, f"Không thể áp dụng Bảng nội quy an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()

    @app.put("/v2/rules/late-threshold")
    def save_late_threshold(body: LateThresholdUpdate, ident: identity_type = Depends(current_identity)):
        conn = engine_instance().connect()
        tx = conn.begin()
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:web-v2:late-threshold'))"))
            require_feature(conn, ident, "official_rules_edit")
            if str(getattr(ident, "role", "") or "").strip().lower() != "admin":
                raise HTTPException(403, "Chỉ tài khoản admin được thay đổi ngưỡng tự động phạt đi trễ.")
            row = conn.execute(text("""
                SELECT value_json, revision
                FROM vera_app_setting
                WHERE category=:category AND setting_key=:setting_key
                FOR UPDATE
            """), {"category": AUTO_CHECK_CATEGORY, "setting_key": AUTO_CHECK_KEY}).mappings().first()
            current_revision = int(row["revision"] or 0) if row else 0
            if current_revision != body.expected_revision:
                raise HTTPException(
                    409,
                    "Ngưỡng đi trễ đã được người khác cập nhật. Hãy bấm Làm mới trước khi áp dụng lại.",
                )
            value = dict(row["value_json"] or {}) if row else {}
            value.setdefault("status", AUTO_CHECK_STATUS_DEFAULT)
            value.setdefault("manual_run_requested", False)
            value["threshold_minutes"] = int(body.threshold_minutes)
            if row is None:
                saved = conn.execute(text("""
                    INSERT INTO vera_app_setting(
                        category, setting_key, value_json, source, updated_by,
                        revision, created_at, updated_at
                    ) VALUES (
                        :category, :setting_key, CAST(:value_json AS jsonb),
                        'web_v2_rules', :updated_by, 1, now(), now()
                    )
                    RETURNING revision, updated_at
                """), {
                    "category": AUTO_CHECK_CATEGORY,
                    "setting_key": AUTO_CHECK_KEY,
                    "value_json": json.dumps(value, ensure_ascii=False),
                    "updated_by": ident.employee_username,
                }).mappings().one()
            else:
                saved = conn.execute(text("""
                    UPDATE vera_app_setting
                    SET value_json=CAST(:value_json AS jsonb),
                        source='web_v2_rules', updated_by=:updated_by,
                        revision=revision + 1, updated_at=now()
                    WHERE category=:category AND setting_key=:setting_key
                    RETURNING revision, updated_at
                """), {
                    "category": AUTO_CHECK_CATEGORY,
                    "setting_key": AUTO_CHECK_KEY,
                    "value_json": json.dumps(value, ensure_ascii=False),
                    "updated_by": ident.employee_username,
                }).mappings().one()
            tx.commit()
            return {
                "ok": True,
                "message": f"Đã áp dụng ngưỡng tự động phạt đi trễ {body.threshold_minutes} phút.",
                "threshold_minutes": int(body.threshold_minutes),
                "revision": int(saved["revision"]),
                "updated_at": saved["updated_at"].isoformat(),
                "updated_by": ident.employee_username,
            }
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            raise HTTPException(500, f"Không thể áp dụng ngưỡng đi trễ an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()

    @app.put("/v2/rules/daily-quota")
    def save_daily_quota(body: DailyQuotaUpdate, ident: identity_type = Depends(current_identity)):
        payload = _daily_quota_payload(body.days)
        conn = engine_instance().connect()
        tx = conn.begin()
        config_ws = None
        old_sheet_values: list[list[Any]] = []
        sheet_changed = False
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:web-v2:daily-quota'))"))
            require_feature(conn, ident, "official_rules_edit")
            if str(getattr(ident, "role", "") or "").strip().lower() != "admin":
                raise HTTPException(403, "Chỉ tài khoản admin được thay đổi hạn mức nghỉ theo ngày.")
            current = conn.execute(text("""
                SELECT revision
                FROM vera_app_setting
                WHERE category=:category AND setting_key=:setting_key
                FOR UPDATE
            """), {"category": DAILY_QUOTA_CATEGORY, "setting_key": DAILY_QUOTA_KEY}).scalar_one_or_none()
            current_revision = int(current or 0)
            if current_revision != body.expected_revision:
                raise HTTPException(
                    409,
                    "Hạn mức nghỉ đã được người khác cập nhật. Hãy bấm Làm mới trước khi áp dụng lại.",
                )

            if current is None:
                saved = conn.execute(text("""
                    INSERT INTO vera_app_setting(
                        category, setting_key, value_json, source, updated_by,
                        revision, created_at, updated_at
                    ) VALUES (
                        :category, :setting_key, CAST(:value_json AS jsonb),
                        'web_v2_daily_quota', :updated_by, 1, now(), now()
                    )
                    RETURNING revision, updated_at
                """), {
                    "category": DAILY_QUOTA_CATEGORY,
                    "setting_key": DAILY_QUOTA_KEY,
                    "value_json": json.dumps(payload, ensure_ascii=False),
                    "updated_by": ident.employee_username,
                }).mappings().one()
            else:
                saved = conn.execute(text("""
                    UPDATE vera_app_setting
                    SET value_json=CAST(:value_json AS jsonb),
                        source='web_v2_daily_quota', updated_by=:updated_by,
                        revision=revision + 1, updated_at=now()
                    WHERE category=:category AND setting_key=:setting_key
                    RETURNING revision, updated_at
                """), {
                    "category": DAILY_QUOTA_CATEGORY,
                    "setting_key": DAILY_QUOTA_KEY,
                    "value_json": json.dumps(payload, ensure_ascii=False),
                    "updated_by": ident.employee_username,
                }).mappings().one()

            spreadsheet = google_client().open_by_key(leave_sheet_id)
            try:
                config_ws = spreadsheet.worksheet("Config")
            except Exception:
                config_ws = spreadsheet.add_worksheet(title="Config", rows=100, cols=10)
            old_sheet_values = config_ws.get_all_values()
            sheet_changed = True
            _write_daily_quota_sheet(config_ws, payload)
            tx.commit()
            return {
                "ok": True,
                "message": "Đã áp dụng hạn mức nghỉ theo ngày THÀNH CÔNG.",
                "days": payload["days"],
                "revision": int(saved["revision"]),
                "updated_at": saved["updated_at"].isoformat(),
                "updated_by": ident.employee_username,
            }
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            if sheet_changed and config_ws is not None:
                _restore_legacy_sheet(config_ws, old_sheet_values)
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            if sheet_changed and config_ws is not None:
                _restore_legacy_sheet(config_ws, old_sheet_values)
            raise HTTPException(500, f"Không thể áp dụng hạn mức nghỉ an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()

    @app.get("/v2/rules/export.xlsx")
    def export_rules(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "official_rules_export")
            document = _load_document(conn)
        binary = _excel_bytes(
            document,
            int(document.get("revision") or 0),
            str(document.get("updated_at") or ""),
            str(document.get("updated_by") or ""),
        )
        filename = "NoiQuy_VeraSpa.xlsx"
        return StreamingResponse(
            BytesIO(binary),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    @app.post("/v2/rules/import.xlsx")
    async def import_rules(request: Request, ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "official_rules_import")
        raw = await request.body()
        if not raw:
            raise HTTPException(400, "Chưa chọn file Excel.")
        if len(raw) > MAX_XLSX_BYTES:
            raise HTTPException(400, "File Excel vượt quá 5 MB.")
        document = _read_excel(raw)
        return {
            **document,
            "message": f"Đã nạp {len(document['rows'])} dòng từ Excel vào vùng chỉnh sửa. Chưa áp dụng cho hệ thống.",
        }
