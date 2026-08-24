"""Admin-only archive export and guarded retention controls for Web V2."""
from __future__ import annotations

from datetime import date, datetime, timedelta
import hashlib
from io import BytesIO
import json
from typing import Any, Callable
from urllib.parse import quote

from fastapi import Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import text

from vera_web_v2_payroll import ADMIN_FIELDS as PAYROLL_FIELDS
from vera_web_v2_payroll import MONEY_FIELDS as PAYROLL_MONEY_FIELDS
from vera_web_v2_payroll import _number as payroll_number
from vera_web_v2_snapshot import _records as attendance_records


DATASETS = {"leave", "payroll", "attendance"}
DELETE_CONFIRMATION = "XÓA DỮ LIỆU"


class StorageDelete(BaseModel):
    dataset: str = Field(min_length=1, max_length=30)
    start: date
    end: date
    expected_count: int = Field(ge=0)
    confirmation: str = Field(min_length=1, max_length=100)


def _validate_dates(start: date, end: date) -> None:
    if end < start:
        raise HTTPException(400, "Đến ngày phải bằng hoặc sau Từ ngày.")
    if end - start > timedelta(days=1826):
        raise HTTPException(400, "Mỗi thao tác chỉ xử lý tối đa 5 năm dữ liệu.")


def _parse_date(value: Any) -> date | None:
    raw = str(value or "").strip().split(" ", 1)[0]
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    return None


def _payroll_date(item: dict[str, Any]) -> date | None:
    for key in ("Đến ngày", "Ngày lưu", "Từ ngày"):
        parsed = _parse_date(item.get(key))
        if parsed:
            return parsed
    return None


def _payroll_payload(conn) -> list[dict[str, Any]]:
    payload = conn.execute(text("""
        SELECT payload FROM vera_dataset_cache
        WHERE dataset_key='payroll_history' LIMIT 1
    """)).scalar_one_or_none()
    return [dict(item) for item in (payload or []) if isinstance(item, dict)]


def _payroll_rows(conn, start: date, end: date) -> list[dict[str, Any]]:
    return [item for item in _payroll_payload(conn) if (day := _payroll_date(item)) and start <= day <= end]


def _leave_rows(conn, start: date, end: date) -> list[dict[str, Any]]:
    rows = conn.execute(text("""
        SELECT record_uid, leave_date, weekday_label, employee_name, leave_reason,
               leave_type, detail, calculated_days, accumulated_leave, penalty,
               update_date, update_time, updated_by
        FROM leave_records
        WHERE leave_date BETWEEN :start AND :end
        ORDER BY leave_date, source_row, employee_name
    """), {"start": start, "end": end}).mappings().all()
    return [dict(item) for item in rows]


def _rows(conn, dataset: str, start: date, end: date) -> list[dict[str, Any]]:
    if dataset == "leave":
        return _leave_rows(conn, start, end)
    if dataset == "payroll":
        return _payroll_rows(conn, start, end)
    if dataset == "attendance":
        return attendance_records(conn, start, end)
    raise HTTPException(400, "Nhóm dữ liệu không hợp lệ.")


def _checksum(payload: list[dict[str, Any]]) -> str:
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F513F")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _append_sheet(workbook, title: str, fields: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append([label for _key, label in fields])
    for item in rows:
        sheet.append([item.get(key, "") for key, _label in fields])
    _style_header(sheet)
    for column in sheet.columns:
        width = min(42, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width


def _delete_payroll(conn, start: date, end: date) -> int:
    payload = _payroll_payload(conn)
    retained = [item for item in payload if not ((day := _payroll_date(item)) and start <= day <= end)]
    deleted = len(payload) - len(retained)
    conn.execute(text("""
        UPDATE vera_dataset_cache
        SET payload=CAST(:payload AS jsonb), row_count=:row_count, checksum=:checksum,
            source_version='web-v2-retention', updated_at=NOW()
        WHERE dataset_key='payroll_history'
    """), {
        "payload": json.dumps(retained, ensure_ascii=False, default=str),
        "row_count": len(retained), "checksum": _checksum(retained),
    })
    return deleted


def _delete_attendance(conn, start: date, end: date) -> int:
    rows = conn.execute(text("""
        SELECT dataset_key, payload FROM vera_dataset_cache
        WHERE dataset_key='timesoft_employee_checkin_today'
           OR dataset_key LIKE 'timesoft_employee_checkin_20%'
        FOR UPDATE
    """)).mappings().all()
    deleted = 0
    for dataset in rows:
        payload = [dict(item) for item in (dataset.get("payload") or []) if isinstance(item, dict)]
        retained = []
        for item in payload:
            item_date = _parse_date(item.get("WorkDateStr"))
            if item_date and start <= item_date <= end:
                deleted += 1
            else:
                retained.append(item)
        conn.execute(text("""
            UPDATE vera_dataset_cache
            SET payload=CAST(:payload AS jsonb), row_count=:row_count, checksum=:checksum,
                source_version='web-v2-retention', updated_at=NOW()
            WHERE dataset_key=:dataset_key
        """), {
            "dataset_key": dataset["dataset_key"],
            "payload": json.dumps(retained, ensure_ascii=False, default=str),
            "row_count": len(retained), "checksum": _checksum(retained),
        })
    return deleted


def install_storage_routes(
    app, *, engine_instance: Callable[[], Any], current_identity,
    require_feature, delete_leave_uids: Callable[[list[str], Any], dict[str, Any]], identity_type,
):
    def require_admin(conn, ident, feature: str) -> None:
        if str(ident.role or "").lower() != "admin":
            raise HTTPException(403, "Chỉ Admin được quản lý bộ nhớ hệ thống.")
        require_feature(conn, ident, feature)

    @app.get("/v2/storage/preview")
    def storage_preview(
        start: date = Query(...), end: date = Query(...),
        ident: identity_type = Depends(current_identity),
    ):
        _validate_dates(start, end)
        with engine_instance().connect() as conn:
            require_admin(conn, ident, "storage_admin_view")
            counts = {dataset: len(_rows(conn, dataset, start, end)) for dataset in sorted(DATASETS)}
        return {"start": start.isoformat(), "end": end.isoformat(), "counts": counts, "total": sum(counts.values())}

    @app.get("/v2/storage/export.xlsx")
    def storage_export(
        start: date = Query(...), end: date = Query(...), dataset: str = Query(default="all"),
        ident: identity_type = Depends(current_identity),
    ):
        _validate_dates(start, end)
        wanted = sorted(DATASETS) if dataset == "all" else [dataset]
        if any(item not in DATASETS for item in wanted):
            raise HTTPException(400, "Nhóm dữ liệu export không hợp lệ.")
        with engine_instance().connect() as conn:
            require_admin(conn, ident, "storage_export")
            data = {item: _rows(conn, item, start, end) for item in wanted}

        workbook = Workbook()
        workbook.remove(workbook.active)
        if "leave" in data:
            _append_sheet(workbook, "Lịch nghỉ", [
                ("record_uid", "Mã bản ghi"), ("leave_date", "Ngày"), ("weekday_label", "Thứ ngày"),
                ("employee_name", "Nhân viên"), ("leave_reason", "Lý do"), ("leave_type", "Loại nghỉ"),
                ("detail", "Chi tiết"), ("calculated_days", "Số ngày tính"),
                ("accumulated_leave", "Phép cộng dồn"), ("penalty", "Phạt"),
                ("update_date", "Ngày cập nhật"), ("update_time", "Giờ cập nhật"), ("updated_by", "Người cập nhật"),
            ], data["leave"])
        if "payroll" in data:
            _append_sheet(workbook, "Bảng lương", [(field, field) for field in PAYROLL_FIELDS], [
                {key: (payroll_number(item.get(key)) if key in PAYROLL_MONEY_FIELDS else item.get(key, "")) for key in PAYROLL_FIELDS}
                for item in data["payroll"]
            ])
        if "attendance" in data:
            _append_sheet(workbook, "Chấm công", [
                ("date", "Ngày"), ("employee_code", "Mã nhân viên"), ("employee_name", "Nhân viên"),
                ("shift", "Ca"), ("shift_start", "Bắt đầu ca"), ("shift_end", "Kết thúc ca"),
                ("check_in", "Giờ vào"), ("check_out", "Giờ ra"), ("arrival_status", "Trạng thái vào"),
                ("departure_status", "Trạng thái ra"), ("late_minutes", "Phút trễ"),
                ("early_minutes", "Phút về sớm"), ("total_minutes", "Tổng phút"), ("punch_count", "Số lần chấm"),
            ], data["attendance"])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"VERA_LuuTru_{start.isoformat()}_{end.isoformat()}.xlsx"
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
        })

    @app.delete("/v2/storage")
    def storage_delete(body: StorageDelete, ident: identity_type = Depends(current_identity)):
        dataset = body.dataset.strip().lower()
        if dataset not in DATASETS:
            raise HTTPException(400, "Nhóm dữ liệu cần xóa không hợp lệ.")
        _validate_dates(body.start, body.end)
        if body.confirmation.strip() != DELETE_CONFIRMATION:
            raise HTTPException(400, f"Vui lòng nhập đúng: {DELETE_CONFIRMATION}")

        with engine_instance().connect() as conn:
            require_admin(conn, ident, "storage_delete")
            preview_rows = _rows(conn, dataset, body.start, body.end)
        if len(preview_rows) != body.expected_count:
            raise HTTPException(409, "Số lượng dữ liệu đã thay đổi. Hãy xem trước lại trước khi xóa.")
        if not preview_rows:
            return {"ok": True, "deleted": 0, "message": "Không có dữ liệu trong khoảng đã chọn."}

        if dataset == "leave":
            return delete_leave_uids([str(item["record_uid"]) for item in preview_rows], ident)

        engine = engine_instance()
        conn = engine.connect()
        tx = conn.begin()
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:v2:storage_retention'))"))
            current_count = len(_rows(conn, dataset, body.start, body.end))
            if current_count != body.expected_count:
                raise HTTPException(409, "Số lượng dữ liệu đã thay đổi. Hãy xem trước lại trước khi xóa.")
            raw_deleted = _delete_payroll(conn, body.start, body.end) if dataset == "payroll" else _delete_attendance(conn, body.start, body.end)
            conn.execute(text("""
                INSERT INTO vera_sync_event(dataset_key, event_type, detail, created_at)
                VALUES (:dataset_key, 'admin_retention_delete', :detail, NOW())
            """), {
                "dataset_key": dataset,
                "detail": f"{ident.employee_username}: {body.start.isoformat()}..{body.end.isoformat()}, logical={current_count}, raw={raw_deleted}",
            })
            tx.commit()
            return {
                "ok": True, "deleted": current_count,
                "message": f"Đã xóa {current_count} bản ghi {dataset} trong khoảng đã chọn.",
            }
        except Exception:
            if tx.is_active:
                tx.rollback()
            raise
        finally:
            conn.close()
