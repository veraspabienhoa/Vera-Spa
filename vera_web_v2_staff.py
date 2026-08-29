"""Authenticated employee-management routes for VERA SPA Web V2.

PostgreSQL is the canonical store.  Mutations are mirrored to the legacy
credential/status worksheets so Streamlit and Web V2 remain convergent while
the migration is in progress.  The browser never receives passwords or token
hashes and never writes the employees table directly.
"""
from datetime import date, datetime
from io import BytesIO
import json
import os
import re
import threading
from typing import Any, Callable
from urllib.parse import quote

from fastapi import Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, Field
from sqlalchemy import text

from vera_web_v2_security import password_policy_error


CREDENTIAL_SHEET_ID = os.getenv(
    "VERA_CREDENTIAL_SHEET_ID", "1DGXy3kPyMPwtz-3CnG8i6BiQbXFDApasoXVFzSmUe24"
)
STATUS_WORKSHEET = "TrangThaiNhanSu"
STATUS_HEADERS = ["STT", "Tên nhân viên", "Trạng thái", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"]
CREDENTIAL_HEADERS = [
    "STT", "Tên nhân viên", "Mật khẩu", "Phân quyền", "Họ và tên đầy đủ", "Ngày sinh",
    "Điện thoại", "Email", "Địa chỉ", "Số tài khoản ngân hàng", "Tên ngân hàng",
    "Phát sinh tháng", "Có phép tháng", "Phép năm", "Ca làm việc", "Ngày bắt đầu ca",
    "Chu kỳ", "Khóa đăng nhập", "Remember Token Hash", "Remember Token Expiry",
    "Ngày bắt đầu làm",
]
STAFF_EXPORT_COLUMNS = [
    "Tên nhân viên", "Họ và tên đầy đủ", "Ngày bắt đầu làm", "Ngày sinh",
    "Phân quyền", "Trạng thái làm việc", "Điện thoại", "Email", "Địa chỉ",
    "Số tài khoản ngân hàng", "Tên ngân hàng", "Phát sinh tháng", "Có phép tháng",
    "Phép năm", "Ca làm việc", "Ngày bắt đầu ca", "Chu kỳ", "Khóa đăng nhập",
]
ALL_ROLES = ["nhanvien", "leader", "quanly", "letan", "locker", "tapvu", "admin"]
ROLE_ORDER = ["leader", "nhanvien", "quanly", "letan", "locker", "tapvu", "admin"]
FRONTDESK_ROLES = {"nhanvien", "locker", "tapvu"}
STATUS_OPTIONS = ["Đang làm việc", "Tạm thời nghỉ việc", "Đã nghỉ việc"]
STATUS_ALIASES = {
    "dang lam viec": "Đang làm việc",
    "nghi viec tam thoi": "Tạm thời nghỉ việc",
    "tam thoi nghi viec": "Tạm thời nghỉ việc",
    "da nghi viec han": "Đã nghỉ việc",
    "da nghi viec": "Đã nghỉ việc",
}
CYCLE_OPTIONS = ["Luân phiên (14 ngày)", "Theo chu kỳ Tháng", "Cố định (Không đổi)"]
DEPARTMENT_ORDER = ["Nhân viên + Leader", "Lễ tân", "Quản lý", "Locker", "Tạp vụ"]
TICHLUY_ELIGIBLE_ROLES = {"nhanvien", "leader"}
TICHLUY_WORKSHEET = "TichLuy"
TICHLUY_HEADERS = [
    "STT", "Tên nhân viên", "Ngày bắt đầu làm", "Mục tiêu tích lũy", "Đã tích lũy",
    "Còn lại", "Kỳ gần nhất", "Số tiền kỳ gần nhất", "Chi tiết các kỳ",
]


class StaffCreate(BaseModel):
    username: str = Field(min_length=1, max_length=200)
    password: str = Field(min_length=8, max_length=300)
    role: str = Field(default="nhanvien", min_length=1, max_length=50)
    full_name: str = Field(min_length=1, max_length=300)
    birth_date: str = Field(default="", max_length=30)
    phone: str = Field(default="", max_length=80)
    email: str = Field(default="", max_length=300)
    address: str = Field(default="", max_length=1000)
    bank_account: str = Field(default="", max_length=100)
    bank_name: str = Field(default="", max_length=300)
    employment_start_date: str = Field(default="", max_length=30)


class StaffUpdate(BaseModel):
    role: str | None = Field(default=None, max_length=50)
    full_name: str | None = Field(default=None, max_length=300)
    birth_date: str | None = Field(default=None, max_length=30)
    phone: str | None = Field(default=None, max_length=80)
    email: str | None = Field(default=None, max_length=300)
    address: str | None = Field(default=None, max_length=1000)
    bank_account: str | None = Field(default=None, max_length=100)
    bank_name: str | None = Field(default=None, max_length=300)
    monthly_generated: float | None = Field(default=None, ge=0)
    monthly_leave: float | None = Field(default=None, ge=0)
    annual_leave: float | None = Field(default=None, ge=0)
    work_shift: str | None = Field(default=None, max_length=300)
    shift_start_date: str | None = Field(default=None, max_length=30)
    rotation_cycle: str | None = Field(default=None, max_length=100)
    login_locked: bool | None = None
    employment_start_date: str | None = Field(default=None, max_length=30)
    employment_status: str | None = Field(default=None, max_length=100)


class StaffDelete(BaseModel):
    usernames: list[str] = Field(min_length=1, max_length=100)


def _date_text(value: Any, *, field_name: str, allow_blank: bool = True) -> str:
    if value is None or str(value).strip() == "":
        if allow_blank:
            return ""
        raise HTTPException(400, f"{field_name} không được để trống.")
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    raise HTTPException(400, f"{field_name} không hợp lệ. Dùng định dạng DD/MM/YYYY.")


def _excel_date(value: Any, *, field_name: str) -> str:
    if value is None or str(value).strip() == "":
        return ""
    return _date_text(value, field_name=field_name)


def _number(value: Any, *, field_name: str) -> float:
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        result = float(str(value).strip().replace(",", "."))
    except ValueError as exc:
        raise HTTPException(400, f"{field_name} phải là số không âm.") from exc
    if result < 0:
        raise HTTPException(400, f"{field_name} phải là số không âm.")
    return result


def _department(role: str) -> str:
    value = str(role or "").strip().lower()
    if value in {"nhanvien", "leader"}:
        return "Nhân viên + Leader"
    return {
        "letan": "Lễ tân", "quanly": "Quản lý", "locker": "Locker", "tapvu": "Tạp vụ",
    }.get(value, value or "Khác")


def _status_value(value: Any, norm: Callable[[Any], str]) -> str:
    key = norm(value)
    if not key:
        return STATUS_OPTIONS[0]
    result = STATUS_ALIASES.get(key)
    if not result:
        raise HTTPException(400, f"Trạng thái làm việc không hợp lệ: {value}")
    return result


def _shift_label(item: dict[str, Any]) -> str:
    name = str(item.get("Tên ca") or "").strip()
    start = str(item.get("Giờ bắt đầu") or "").strip()
    end = str(item.get("Giờ kết thúc") or "").strip()
    label = f"{name} ({start} - {end})" if name and start and end else name
    if str(item.get("Ghi chú") or "").strip().casefold() == "không đổi" and "không đổi" not in label.casefold():
        label = f"{label} (Không đổi)"
    return label


def _shift_catalog(conn, employee_rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    output = {department: [] for department in DEPARTMENT_ORDER}
    payload = conn.execute(text("""
        SELECT value_json FROM vera_app_setting
        WHERE category='shift' AND setting_key='shift_definitions'
        LIMIT 1
    """)).scalar_one_or_none()
    if isinstance(payload, list):
        ordered = sorted(
            (item for item in payload if isinstance(item, dict)),
            key=lambda item: (str(item.get("Bộ phận") or ""), int(item.get("Thứ tự") or 999), str(item.get("Tên ca") or "")),
        )
        for item in ordered:
            if str(item.get("Trạng thái") or "").strip().casefold() == "đã xóa":
                continue
            department = str(item.get("Bộ phận") or "Nhân viên + Leader").strip()
            if department not in output:
                output[department] = []
            label = _shift_label(item)
            if label and label not in output[department]:
                output[department].append(label)
    for row in employee_rows:
        department = _department(str(row.get("role") or ""))
        shift = str(row.get("work_shift") or "").strip()
        if shift and shift not in output.setdefault(department, []):
            output[department].append(shift)
    return output


def _employee_payload(row: dict[str, Any], status: str) -> dict[str, Any]:
    payload = dict(row.get("payload") or {}) if isinstance(row.get("payload"), dict) else {}
    payload.update({
        "STT": int(row.get("stt") or 0),
        "Tên nhân viên": str(row.get("username") or ""),
        "Mật khẩu": str(row.get("password_value") or ""),
        "Phân quyền": str(row.get("role") or ""),
        "Họ và tên đầy đủ": str(row.get("full_name") or ""),
        "Ngày sinh": str(row.get("birth_date") or ""),
        "Điện thoại": str(row.get("phone") or ""),
        "Email": str(row.get("email") or ""),
        "Địa chỉ": str(row.get("address") or ""),
        "Số tài khoản ngân hàng": str(row.get("bank_account") or ""),
        "Tên ngân hàng": str(row.get("bank_name") or ""),
        "Phát sinh tháng": float(row.get("monthly_generated") or 0),
        "Có phép tháng": float(row.get("monthly_leave") or 0),
        "Phép năm": float(row.get("annual_leave") or 0),
        "Ca làm việc": str(row.get("work_shift") or ""),
        "Ngày bắt đầu ca": str(row.get("shift_start_date") or ""),
        "Chu kỳ": str(row.get("rotation_cycle") or ""),
        "Khóa đăng nhập": "KHÓA" if bool(row.get("login_locked")) else "",
        "Remember Token Hash": str(row.get("remember_token_hash") or ""),
        "Remember Token Expiry": str(row.get("remember_token_expiry") or ""),
        "Ngày bắt đầu làm": str(row.get("employment_start_date") or ""),
        "Trạng thái làm việc": status,
    })
    return payload


def _credential_values(row: dict[str, Any], status: str) -> list[Any]:
    payload = _employee_payload(row, status)
    return [payload.get(header, "") for header in CREDENTIAL_HEADERS]


def _public_employee(row: dict[str, Any], status: str) -> dict[str, Any]:
    return {
        "username": str(row.get("username") or ""),
        "full_name": str(row.get("full_name") or ""),
        "birth_date": str(row.get("birth_date") or ""),
        "phone": str(row.get("phone") or ""),
        "email": str(row.get("email") or ""),
        "address": str(row.get("address") or ""),
        "bank_account": str(row.get("bank_account") or ""),
        "bank_name": str(row.get("bank_name") or ""),
        "monthly_generated": float(row.get("monthly_generated") or 0),
        "monthly_leave": float(row.get("monthly_leave") or 0),
        "annual_leave": float(row.get("annual_leave") or 0),
        "role": str(row.get("role") or "").lower(),
        "department": _department(str(row.get("role") or "")),
        "employment_status": status,
        "work_shift": str(row.get("work_shift") or ""),
        "shift_start_date": str(row.get("shift_start_date") or ""),
        "rotation_cycle": str(row.get("rotation_cycle") or ""),
        "login_locked": bool(row.get("login_locked")),
        "employment_start_date": str(row.get("employment_start_date") or ""),
    }


def _select_staff_rows(conn, *, for_update: bool = False) -> list[dict[str, Any]]:
    suffix = " FOR UPDATE" if for_update else ""
    rows = conn.execute(text("""
        SELECT username, stt, password_value, role, full_name, birth_date, phone, email,
               address, bank_account, bank_name, monthly_generated, monthly_leave,
               annual_leave, work_shift, shift_start_date, rotation_cycle, login_locked,
               remember_token_hash, remember_token_expiry, employment_start_date,
               source_sheet_id, source_row, payload
        FROM employees
        WHERE COALESCE(source_sheet_id,'credentials')='credentials'
        ORDER BY COALESCE(source_row, 2147483647), COALESCE(stt, 2147483647), username
    """ + suffix)).mappings().all()
    return [dict(row) for row in rows]


def _effective_status(row: dict[str, Any], google_status: dict[str, str], norm: Callable[[Any], str]) -> str:
    key = norm(row.get("username"))
    if key in google_status:
        return google_status[key]
    payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
    raw = payload.get("Trạng thái làm việc") or payload.get("employment_status") or STATUS_OPTIONS[0]
    return STATUS_ALIASES.get(norm(raw), STATUS_OPTIONS[0])


def install_staff_routes(
    app,
    *,
    engine_instance: Callable[[], Any],
    current_identity: Callable[..., Any],
    require_feature: Callable[[Any, Any, str], None],
    feature_allowed: Callable[[Any, Any, str], bool],
    norm: Callable[[Any], str],
    google_client: Callable[[], Any],
    leave_sheet_id: str,
    progressive_key: Callable[[Any], str],
    reindex_after_delete: Callable[[Any, list[int]], None],
    rebalance_progressive_rows: Callable[[Any, date, set[str]], list[tuple[int, dict]]],
    sheet_values_for_record: Callable[[list[Any], dict, int], list[Any]],
    restore_sheet_updates: Callable[[Any, dict[int, list[Any]]], None],
    identity_type: type,
    vn_tz,
) -> None:
    def credential_spreadsheet():
        return google_client().open_by_key(CREDENTIAL_SHEET_ID)

    def credential_ws():
        return credential_spreadsheet().get_worksheet(0)

    def worksheet(title: str, rows: int, cols: int, headers: list[str]):
        spreadsheet = credential_spreadsheet()
        try:
            ws = spreadsheet.worksheet(title)
        except Exception:
            ws = spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)
        current = ws.row_values(1)
        if current[:len(headers)] != headers:
            ws.update(
                range_name=f"A1:{get_column_letter(len(headers))}1",
                values=[headers], value_input_option="USER_ENTERED",
            )
        return ws

    def google_status_map() -> dict[str, str]:
        result: dict[str, str] = {}
        try:
            ws = worksheet(STATUS_WORKSHEET, 1000, len(STATUS_HEADERS), STATUS_HEADERS)
            for row in ws.get_all_values()[1:]:
                if len(row) < 2 or not norm(row[1]):
                    continue
                result[norm(row[1])] = STATUS_ALIASES.get(norm(row[2] if len(row) > 2 else ""), STATUS_OPTIONS[0])
        except Exception:
            pass
        return result

    def reconcile_employment_status_login_gates() -> None:
        """Seed PostgreSQL login gates from the legacy status worksheet at rollout."""
        statuses = google_status_map()
        if not statuses:
            return
        try:
            with engine_instance().begin() as conn:
                rows = conn.execute(text("SELECT username FROM employees")).mappings().all()
                for row in rows:
                    status = statuses.get(norm(row["username"]))
                    if not status:
                        continue
                    conn.execute(text("""
                        UPDATE employees
                        SET payload=jsonb_set(
                              COALESCE(payload, '{}'::jsonb),
                              '{Trạng thái làm việc}',
                              to_jsonb(CAST(:status AS text)),
                              true
                            ),
                            remember_token_hash=CASE WHEN :is_active THEN remember_token_hash ELSE '' END,
                            remember_token_expiry=CASE WHEN :is_active THEN remember_token_expiry ELSE '' END,
                            updated_at=NOW()
                        WHERE username=:username
                          AND (
                            COALESCE(payload->>'Trạng thái làm việc','') IS DISTINCT FROM :status
                            OR (NOT :is_active AND (
                              COALESCE(remember_token_hash,'') <> ''
                              OR COALESCE(remember_token_expiry,'') <> ''
                            ))
                          )
                    """), {
                        "username": row["username"], "status": status,
                        "is_active": status == STATUS_OPTIONS[0],
                    })
                    conn.execute(text("""
                        UPDATE vera_v2_user_profile
                        SET is_active=:is_active, updated_at=NOW()
                        WHERE lower(btrim(employee_username))=lower(btrim(:username))
                          AND is_active IS DISTINCT FROM :is_active
                    """), {
                        "username": row["username"],
                        "is_active": status == STATUS_OPTIONS[0],
                    })
        except Exception:
            # The API must remain available when Google is temporarily unreachable;
            # every later Web V2 status change still writes PostgreSQL synchronously.
            return

    @app.on_event("startup")
    def schedule_employment_status_reconciliation() -> None:
        # Google availability must never delay Cloud Run readiness. The worker
        # only seeds legacy rows; every Web V2 status write remains synchronous.
        threading.Thread(
            target=reconcile_employment_status_login_gates,
            name="vera-employment-status-reconcile",
            daemon=True,
        ).start()

    def write_status(username: str, status: str, actor: str) -> None:
        ws = worksheet(STATUS_WORKSHEET, 1000, len(STATUS_HEADERS), STATUS_HEADERS)
        values = ws.get_all_values()
        row_number = None
        for index, row in enumerate(values[1:], start=2):
            if len(row) > 1 and norm(row[1]) == norm(username):
                row_number = index
                break
        now = datetime.now(vn_tz)
        data = [username, status, now.strftime("%d/%m/%Y"), now.strftime("%H:%M:%S"), actor]
        if row_number:
            ws.update(range_name=f"B{row_number}:F{row_number}", values=[data], value_input_option="USER_ENTERED")
        else:
            ws.append_row([max(1, len(values))] + data, value_input_option="USER_ENTERED")

    def restore_pruned_leave(change: dict[str, Any] | None) -> None:
        if not change or change.get("restored"):
            return
        ws = change.get("worksheet")
        if ws is None:
            return
        restore_sheet_updates(ws, change.get("changed_backups") or {})
        for source_row in sorted(change.get("deleted_rows") or []):
            try:
                ws.insert_row(
                    change["deleted_values"][source_row],
                    index=source_row,
                    value_input_option="USER_ENTERED",
                )
            except Exception:
                pass
        change["restored"] = True

    def prune_registered_leave(conn, username: str, effective_date: date) -> dict[str, Any]:
        """Remove mirrored leave rows from the employment-status date onward.

        Rows before ``effective_date`` are deliberately untouched so every
        historical list, statistic and Excel export keeps the employee's full
        record before temporary/permanent departure.
        """
        conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:leave_primary'))"))
        rows = [dict(row) for row in conn.execute(text("""
            SELECT record_uid, source_sheet_id, source_row, leave_date, employee_name,
                   leave_reason, leave_type, detail, calculated_days, accumulated_leave,
                   penalty, update_date, update_time, updated_by, weekday_label
            FROM leave_records
            WHERE lower(btrim(employee_name))=lower(btrim(:employee))
              AND leave_date >= :effective_date
            ORDER BY source_row, record_uid
            FOR UPDATE
        """), {"employee": username, "effective_date": effective_date}).mappings().all()]
        if not rows:
            return {"count": 0, "effective_date": effective_date, "employee": username}

        for row in rows:
            if str(row.get("source_sheet_id") or "") != leave_sheet_id or int(row.get("source_row") or 0) < 2:
                raise HTTPException(
                    409,
                    "Có lịch nghỉ tương lai không có vị trí MainData hợp lệ; từ chối đổi trạng thái để tránh lệch dữ liệu.",
                )

        ws = google_client().open_by_key(leave_sheet_id).get_worksheet(0)
        values = ws.get_all_values()
        headers = values[0][:13] if values else []
        if not headers:
            raise RuntimeError("MainData chưa có header A:M")
        source_rows = sorted({int(row["source_row"]) for row in rows})
        deleted_values: dict[int, list[Any]] = {}
        for source_row in source_rows:
            if source_row > len(values):
                raise RuntimeError(f"MainData thiếu dòng nguồn {source_row}")
            deleted_values[source_row] = list(values[source_row - 1][:13])

        change: dict[str, Any] = {
            "count": len(rows),
            "effective_date": effective_date,
            "employee": username,
            "worksheet": ws,
            "deleted_rows": [],
            "deleted_values": deleted_values,
            "changed_backups": {},
            "restored": False,
        }
        try:
            affected: dict[date, set[str]] = {}
            for row in rows:
                key = progressive_key(row.get("leave_reason"))
                if key:
                    affected.setdefault(row["leave_date"], set()).add(key)
                result = conn.execute(text("DELETE FROM leave_records WHERE record_uid=:uid"), {"uid": row["record_uid"]})
                if int(result.rowcount or 0) != 1:
                    raise RuntimeError("Xóa PostgreSQL không đúng một lịch nghỉ.")
            reindex_after_delete(conn, source_rows)
            rebalanced: list[tuple[int, dict]] = []
            for target_date, keys in affected.items():
                rebalanced.extend(rebalance_progressive_rows(conn, target_date, keys))

            for source_row in sorted(source_rows, reverse=True):
                ws.delete_rows(source_row)
                change["deleted_rows"].append(source_row)

            after_values = ws.get_all_values() if rebalanced else []
            for row_number, record in rebalanced:
                change["changed_backups"][row_number] = (
                    list(after_values[row_number - 1][:13]) if row_number <= len(after_values) else []
                )
                ws.update(
                    range_name=f"A{row_number}:M{row_number}",
                    values=[sheet_values_for_record(headers, record, row_number)],
                    value_input_option="USER_ENTERED",
                )
            return change
        except Exception:
            restore_pruned_leave(change)
            raise

    def delete_rows_by_name(ws, names: set[str], name_col: int) -> list[tuple[int, list[Any]]]:
        values = ws.get_all_values()
        targets = []
        for index, row in enumerate(values[1:], start=2):
            value = row[name_col - 1] if len(row) >= name_col else ""
            if norm(value) in names:
                targets.append((index, list(row)))
        for index, _ in sorted(targets, reverse=True):
            ws.delete_rows(index)
        return targets

    def renumber_sheet(ws, stt_col: int = 1) -> None:
        values = ws.get_all_values()
        count = max(0, len(values) - 1)
        if count:
            column = get_column_letter(stt_col)
            ws.update(
                range_name=f"{column}2:{column}{count + 1}",
                values=[[index] for index in range(1, count + 1)],
                value_input_option="USER_ENTERED",
            )

    def sync_tichluy_members(conn) -> None:
        try:
            ws = worksheet(TICHLUY_WORKSHEET, 2000, len(TICHLUY_HEADERS), TICHLUY_HEADERS)
            old_values = ws.get_all_values()
            old_by_name = {
                norm(row[1]): list(row[:len(TICHLUY_HEADERS)])
                for row in old_values[1:]
                if len(row) > 1 and norm(row[1])
            }
            eligible = [
                row for row in _select_staff_rows(conn)
                if str(row.get("role") or "").lower() in TICHLUY_ELIGIBLE_ROLES
            ]
            output = []
            for index, row in enumerate(eligible, start=1):
                existing = old_by_name.get(norm(row["username"]), [])
                existing += [""] * max(0, len(TICHLUY_HEADERS) - len(existing))
                start_date = str(row.get("employment_start_date") or "") or str(existing[2] or "")
                output.append([
                    index, row["username"], start_date,
                    existing[3] or 5000000, existing[4] or 0, existing[5] or 5000000,
                    existing[6], existing[7] or 0, existing[8],
                ])
            if len(old_values) > 1:
                ws.batch_clear([f"A2:I{max(len(old_values), len(output) + 1)}"])
            if output:
                ws.update(range_name=f"A2:I{len(output) + 1}", values=output, value_input_option="USER_ENTERED")
        except Exception:
            # Tích lũy is an auxiliary legacy mirror; employee CRUD must not be lost
            # when that worksheet is temporarily unavailable.
            pass

    def permissions(conn, ident) -> dict[str, bool]:
        keys = (
            "staff_list", "staff_export", "staff_import", "employee_add", "employee_add_save",
            "employee_edit", "employee_edit_save", "employment_status", "employment_status_edit",
            "employee_delete", "employee_delete_confirm", "shift_assignment_edit", "account_lock_edit",
        )
        output = {key: feature_allowed(conn, ident, key) for key in keys}
        # Xóa nhân viên là thao tác quản trị tài khoản chỉ dựa trên vai trò
        # Admin. Không để một override phân quyền cũ làm giao diện hiện nút Xóa
        # nhưng API lại từ chối ở quyền xác nhận ẩn.
        is_admin = str(ident.role).lower() == "admin"
        output["employee_delete"] = is_admin
        output["employee_delete_confirm"] = is_admin
        return output

    def allowed_roles(ident) -> list[str]:
        return ALL_ROLES if str(ident.role).lower() == "admin" else ["nhanvien", "locker", "tapvu"]

    def ensure_manageable(ident, target_role: str) -> None:
        if str(ident.role).lower() == "admin":
            return
        if str(target_role or "").lower() not in FRONTDESK_ROLES:
            raise HTTPException(403, "Quản lý/Lễ tân chỉ được thao tác tài khoản nhanvien, locker hoặc tapvu.")

    def validate_role(ident, value: str) -> str:
        role = str(value or "").strip().lower()
        if role not in ALL_ROLES:
            raise HTTPException(400, f"Phân quyền không hợp lệ: {value}")
        if role not in allowed_roles(ident):
            raise HTTPException(403, f"Tài khoản hiện tại không được gán phân quyền '{role}'.")
        return role

    def validate_shift(conn, role: str, shift: str, current_shift: str, rows: list[dict[str, Any]]) -> str:
        value = str(shift or "").strip()
        if not value or value == str(current_shift or "").strip():
            return value
        catalog = _shift_catalog(conn, rows)
        if value not in catalog.get(_department(role), []):
            raise HTTPException(400, f"Ca làm việc '{value}' không thuộc bộ phận {_department(role)}.")
        return value

    def staff_result(conn, ident) -> dict[str, Any]:
        rows = _select_staff_rows(conn)
        statuses = google_status_map()
        public_rows = [
            _public_employee(row, _effective_status(row, statuses, norm))
            for row in rows
            if str(row.get("role") or "").lower() != "admin"
        ]
        all_public = list(public_rows)
        if str(ident.role).lower() != "admin":
            public_rows = [row for row in public_rows if row["role"] != "quanly"]
        role_rank = {role: index for index, role in enumerate(ROLE_ORDER)}
        status_rank = {status: index for index, status in enumerate(STATUS_OPTIONS)}
        public_rows.sort(key=lambda row: (
            role_rank.get(row["role"], 99),
            status_rank.get(row["employment_status"], 99),
            norm(row["username"]),
        ))
        summary = {
            "total": len(all_public),
            "active": sum(row["employment_status"] == STATUS_OPTIONS[0] for row in all_public),
            "temporary": sum(row["employment_status"] == STATUS_OPTIONS[1] for row in all_public),
            "left": sum(row["employment_status"] == STATUS_OPTIONS[2] for row in all_public),
        }
        return {
            "employees": public_rows,
            "summary": summary,
            "permissions": permissions(conn, ident),
            "role_options": allowed_roles(ident),
            "status_options": STATUS_OPTIONS,
            "cycle_options": CYCLE_OPTIONS,
            "shifts_by_department": _shift_catalog(conn, rows),
        }

    def find_row(rows: list[dict[str, Any]], username: str) -> dict[str, Any]:
        target = norm(username)
        for row in rows:
            if norm(row.get("username")) == target:
                return row
        raise HTTPException(404, "Không tìm thấy nhân viên.")

    def update_database_row(
        conn,
        ident,
        row: dict[str, Any],
        values: dict[str, Any],
        *,
        status_map: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        ensure_manageable(ident, str(row.get("role") or ""))
        profile_fields = {
            "role", "full_name", "birth_date", "phone", "email", "address", "bank_account",
            "bank_name", "monthly_generated", "monthly_leave", "annual_leave", "employment_start_date",
        }
        if profile_fields.intersection(values):
            require_feature(conn, ident, "employee_edit_save")
        if "employment_status" in values:
            require_feature(conn, ident, "employment_status_edit")
        if {"work_shift", "shift_start_date", "rotation_cycle"}.intersection(values):
            require_feature(conn, ident, "shift_assignment_edit")
        if "login_locked" in values:
            require_feature(conn, ident, "account_lock_edit")

        merged = dict(row)
        if "role" in values:
            merged["role"] = validate_role(ident, values["role"])
        if "full_name" in values:
            merged["full_name"] = str(values["full_name"] or "").strip()
        for key in ("phone", "email", "address", "bank_account", "bank_name"):
            if key in values:
                merged[key] = str(values[key] or "").strip()
        for key, label in (
            ("birth_date", "Ngày sinh"),
            ("shift_start_date", "Ngày bắt đầu ca"),
            ("employment_start_date", "Ngày bắt đầu làm"),
        ):
            if key in values:
                merged[key] = _date_text(values[key], field_name=label)
        for key, label in (
            ("monthly_generated", "Phát sinh tháng"),
            ("monthly_leave", "Có phép tháng"),
            ("annual_leave", "Phép năm"),
        ):
            if key in values:
                merged[key] = _number(values[key], field_name=label)
        if "rotation_cycle" in values:
            cycle = str(values["rotation_cycle"] or "").strip()
            if cycle and cycle not in CYCLE_OPTIONS:
                raise HTTPException(400, f"Chu kỳ không hợp lệ: {cycle}")
            merged["rotation_cycle"] = cycle
        if "work_shift" in values:
            merged["work_shift"] = validate_shift(
                conn, str(merged.get("role") or ""), str(values["work_shift"] or ""),
                str(row.get("work_shift") or ""), _select_staff_rows(conn),
            )
        if "login_locked" in values:
            merged["login_locked"] = bool(values["login_locked"])
            if merged["login_locked"]:
                merged["remember_token_hash"] = ""
                merged["remember_token_expiry"] = ""
        old_status = _effective_status(row, status_map if status_map is not None else google_status_map(), norm)
        status = _status_value(values.get("employment_status", old_status), norm)
        if str(merged.get("role") or "").lower() == "admin" and status != STATUS_OPTIONS[0]:
            raise HTTPException(400, "Không áp dụng trạng thái nghỉ việc cho tài khoản admin.")
        if status != STATUS_OPTIONS[0]:
            # Employment status is an independent login gate.  Clear legacy
            # remembered-login material immediately when employment pauses or ends.
            merged["remember_token_hash"] = ""
            merged["remember_token_expiry"] = ""

        payload = _employee_payload(merged, status)
        updated = conn.execute(text("""
            UPDATE employees SET
                role=:role, full_name=:full_name, birth_date=:birth_date, phone=:phone,
                email=:email, address=:address, bank_account=:bank_account, bank_name=:bank_name,
                monthly_generated=:monthly_generated, monthly_leave=:monthly_leave,
                annual_leave=:annual_leave, work_shift=:work_shift,
                shift_start_date=:shift_start_date, rotation_cycle=:rotation_cycle,
                login_locked=:login_locked, remember_token_hash=:remember_token_hash,
                remember_token_expiry=:remember_token_expiry,
                employment_start_date=:employment_start_date,
                payload=CAST(:payload AS jsonb), updated_at=NOW()
            WHERE username=:username
            RETURNING username, stt, password_value, role, full_name, birth_date, phone, email,
                      address, bank_account, bank_name, monthly_generated, monthly_leave,
                      annual_leave, work_shift, shift_start_date, rotation_cycle, login_locked,
                      remember_token_hash, remember_token_expiry, employment_start_date,
                      source_sheet_id, source_row, payload
        """), {
            **merged,
            "username": row["username"],
            "payload": json.dumps(payload, ensure_ascii=False),
        }).mappings().first()
        if not updated:
            raise HTTPException(404, "Nhân viên không còn tồn tại.")
        if (
            str(merged.get("role") or "").lower() != str(row.get("role") or "").lower()
            or status != old_status
        ):
            conn.execute(text("""
                UPDATE vera_v2_user_profile
                SET role=:role, is_active=:is_active, updated_at=NOW()
                WHERE lower(btrim(employee_username))=lower(btrim(:username))
            """), {
                "role": merged["role"], "is_active": status == STATUS_OPTIONS[0],
                "username": row["username"],
            })
        return {
            **dict(updated),
            "_status": status,
            "_old_status": old_status,
            "_status_changed": "employment_status" in values and status != old_status,
        }

    def mirror_database_row(row: dict[str, Any], status: str) -> tuple[Any, int, list[Any]]:
        ws = credential_ws()
        all_values = ws.get_all_values()
        if not all_values:
            ws.update(range_name="A1:U1", values=[CREDENTIAL_HEADERS], value_input_option="USER_ENTERED")
            all_values = [CREDENTIAL_HEADERS]
        source_row = int(row.get("source_row") or 0)
        if source_row < 2 or source_row > len(all_values) or norm(all_values[source_row - 1][1] if len(all_values[source_row - 1]) > 1 else "") != norm(row["username"]):
            source_row = 0
            for index, values in enumerate(all_values[1:], start=2):
                if len(values) > 1 and norm(values[1]) == norm(row["username"]):
                    source_row = index
                    break
        if source_row < 2:
            raise RuntimeError(f"Không tìm thấy {row['username']} trong Sheet1 để đồng bộ.")
        backup = list(all_values[source_row - 1][:len(CREDENTIAL_HEADERS)])
        ws.update(
            range_name=f"A{source_row}:U{source_row}",
            values=[_credential_values(row, status)], value_input_option="USER_ENTERED",
        )
        return ws, source_row, backup

    @app.get("/v2/staff")
    def get_staff(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "staff_list")
            return staff_result(conn, ident)

    @app.post("/v2/staff")
    def create_staff(body: StaffCreate, ident: identity_type = Depends(current_identity)):
        engine = engine_instance()
        conn = engine.connect()
        tx = conn.begin()
        inserted_sheet_row = 0
        status_created = False
        ws = None
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:employees'))"))
            require_feature(conn, ident, "employee_add")
            require_feature(conn, ident, "employee_add_save")
            username = body.username.strip()
            full_name = body.full_name.strip()
            if not full_name:
                raise HTTPException(400, "Họ và tên không được để trống.")
            password_error = password_policy_error(body.password, username=username, full_name=full_name)
            if password_error:
                raise HTTPException(400, password_error)
            if norm(username) in {"quan tri vien", "admin"}:
                raise HTTPException(400, "Tên nhân viên này được dành cho tài khoản hệ thống.")
            rows = _select_staff_rows(conn, for_update=True)
            if any(norm(row["username"]) == norm(username) for row in rows):
                raise HTTPException(409, "Tên nhân viên đã tồn tại; hệ thống không phân biệt dấu hoặc HOA/thường.")
            role = validate_role(ident, body.role)
            start_work = _date_text(
                body.employment_start_date or datetime.now(vn_tz).strftime("%d/%m/%Y"),
                field_name="Ngày bắt đầu làm", allow_blank=False,
            )
            stt = max([int(row.get("stt") or 0) for row in rows] + [0]) + 1
            ws = credential_ws()
            sheet_values = ws.get_all_values()
            if not sheet_values:
                ws.update(range_name="A1:U1", values=[CREDENTIAL_HEADERS], value_input_option="USER_ENTERED")
                sheet_values = [CREDENTIAL_HEADERS]
            if any(len(row) > 1 and norm(row[1]) == norm(username) for row in sheet_values[1:]):
                raise HTTPException(409, "Tên nhân viên đã tồn tại trong Sheet1.")
            source_row = len(sheet_values) + 1
            record = {
                "username": username, "stt": stt, "password_value": body.password,
                "role": role, "full_name": full_name,
                "birth_date": _date_text(body.birth_date, field_name="Ngày sinh"),
                "phone": body.phone.strip(), "email": body.email.strip(),
                "address": body.address.strip(), "bank_account": body.bank_account.strip(),
                "bank_name": body.bank_name.strip(), "monthly_generated": 0,
                "monthly_leave": 0, "annual_leave": 0, "work_shift": "",
                "shift_start_date": "", "rotation_cycle": "", "login_locked": False,
                "remember_token_hash": "", "remember_token_expiry": "",
                "employment_start_date": start_work, "source_sheet_id": "credentials",
                "source_row": source_row, "payload": {"must_change_password": True},
            }
            status = STATUS_OPTIONS[0]
            payload = _employee_payload(record, status)
            conn.execute(text("""
                INSERT INTO employees(
                    username, stt, password_value, role, full_name, birth_date, phone, email,
                    address, bank_account, bank_name, monthly_generated, monthly_leave,
                    annual_leave, work_shift, shift_start_date, rotation_cycle, login_locked,
                    remember_token_hash, remember_token_expiry, employment_start_date,
                    source_sheet_id, source_row, payload, updated_at
                ) VALUES (
                    :username, :stt, :password_value, :role, :full_name, :birth_date, :phone, :email,
                    :address, :bank_account, :bank_name, :monthly_generated, :monthly_leave,
                    :annual_leave, :work_shift, :shift_start_date, :rotation_cycle, :login_locked,
                    :remember_token_hash, :remember_token_expiry, :employment_start_date,
                    :source_sheet_id, :source_row, CAST(:payload AS jsonb), NOW()
                )
            """), {**record, "payload": json.dumps(payload, ensure_ascii=False)})
            ws.append_row(_credential_values(record, status), value_input_option="USER_ENTERED")
            inserted_sheet_row = source_row
            write_status(username, status, ident.employee_username)
            status_created = True
            tx.commit()
            sync_tichluy_members(conn)
            return {"ok": True, "message": f"Đã thêm nhân viên {username} THÀNH CÔNG."}
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            if inserted_sheet_row and ws is not None:
                try:
                    ws.delete_rows(inserted_sheet_row)
                except Exception:
                    pass
            if status_created:
                try:
                    delete_rows_by_name(worksheet(STATUS_WORKSHEET, 1000, len(STATUS_HEADERS), STATUS_HEADERS), {norm(body.username)}, 2)
                except Exception:
                    pass
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            if inserted_sheet_row and ws is not None:
                try:
                    ws.delete_rows(inserted_sheet_row)
                except Exception:
                    pass
            if status_created:
                try:
                    delete_rows_by_name(
                        worksheet(STATUS_WORKSHEET, 1000, len(STATUS_HEADERS), STATUS_HEADERS),
                        {norm(body.username)},
                        2,
                    )
                except Exception:
                    pass
            raise HTTPException(500, f"Không thêm được nhân viên an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()

    @app.patch("/v2/staff/{username}")
    def update_staff(username: str, body: StaffUpdate, ident: identity_type = Depends(current_identity)):
        values = body.model_dump(exclude_unset=True)
        if not values:
            raise HTTPException(400, "Chưa có thay đổi cần lưu.")
        engine = engine_instance()
        conn = engine.connect()
        tx = conn.begin()
        sheet_backup = None
        sheet_ref = None
        status_before = None
        leave_prune = None
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:employees'))"))
            require_feature(conn, ident, "staff_list")
            rows = _select_staff_rows(conn, for_update=True)
            row = find_row(rows, username)
            status_before = _effective_status(row, google_status_map(), norm)
            updated = update_database_row(conn, ident, row, values)
            sheet_ref, sheet_row, sheet_values = mirror_database_row(updated, updated["_status"])
            sheet_backup = (sheet_row, sheet_values)
            if updated["_status_changed"]:
                write_status(updated["username"], updated["_status"], ident.employee_username)
                if updated["_status"] in {STATUS_OPTIONS[1], STATUS_OPTIONS[2]}:
                    leave_prune = prune_registered_leave(
                        conn,
                        updated["username"],
                        datetime.now(vn_tz).date(),
                    )
            tx.commit()
            sync_tichluy_members(conn)
            deleted_leave = int((leave_prune or {}).get("count") or 0)
            suffix = (
                f" Đã xóa {deleted_leave} lịch nghỉ từ ngày "
                f"{leave_prune['effective_date'].strftime('%d/%m/%Y')} trở về sau; lịch sử trước ngày này được giữ nguyên."
                if deleted_leave else ""
            )
            return {
                "ok": True,
                "message": f"Đã cập nhật {updated['username']} THÀNH CÔNG.{suffix}",
                "employee": _public_employee(updated, updated["_status"]),
                "deleted_leave_records": deleted_leave,
            }
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            restore_pruned_leave(leave_prune)
            if sheet_backup and sheet_ref is not None:
                try:
                    row_number, old_values = sheet_backup
                    sheet_ref.update(range_name=f"A{row_number}:U{row_number}", values=[old_values], value_input_option="USER_ENTERED")
                except Exception:
                    pass
            if status_before is not None and "employment_status" in values:
                try:
                    write_status(username, status_before, ident.employee_username)
                except Exception:
                    pass
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            restore_pruned_leave(leave_prune)
            if sheet_backup and sheet_ref is not None:
                try:
                    row_number, old_values = sheet_backup
                    sheet_ref.update(range_name=f"A{row_number}:U{row_number}", values=[old_values], value_input_option="USER_ENTERED")
                except Exception:
                    pass
            if status_before is not None and "employment_status" in values:
                try:
                    write_status(username, status_before, ident.employee_username)
                except Exception:
                    pass
            raise HTTPException(500, f"Không cập nhật được nhân viên an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()

    @app.delete("/v2/staff")
    def delete_staff(body: StaffDelete, ident: identity_type = Depends(current_identity)):
        names = list(dict.fromkeys(str(name or "").strip() for name in body.usernames if str(name or "").strip()))
        if not names:
            raise HTTPException(400, "Chưa chọn nhân viên cần xóa.")
        if any(norm(name) == norm(ident.employee_username) for name in names):
            raise HTTPException(400, "Không thể xóa tài khoản đang đăng nhập.")
        engine = engine_instance()
        conn = engine.connect()
        tx = conn.begin()
        deleted_sheet_rows: list[tuple[int, list[Any]]] = []
        deleted_status_rows: list[tuple[int, list[Any]]] = []
        staff_ws = None
        status_ws = None
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:employees'))"))
            if str(ident.role).lower() != "admin":
                raise HTTPException(403, "Chỉ Admin được xóa nhân viên.")
            rows = _select_staff_rows(conn, for_update=True)
            targets = [find_row(rows, name) for name in names]
            for row in targets:
                ensure_manageable(ident, str(row.get("role") or ""))
                if str(row.get("role") or "").lower() == "admin" or norm(row.get("username")) == "quan tri vien":
                    raise HTTPException(400, "Không thể xóa tài khoản hệ thống/admin.")
            target_keys = {norm(row["username"]) for row in targets}
            for row in targets:
                conn.execute(text("DELETE FROM employees WHERE username=:username"), {"username": row["username"]})
                conn.execute(text("""
                    UPDATE vera_v2_user_profile SET is_active=false, updated_at=NOW()
                    WHERE lower(btrim(employee_username))=lower(btrim(:username))
                """), {"username": row["username"]})
            remaining = _select_staff_rows(conn)
            for index, row in enumerate(remaining, start=1):
                payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
                payload["STT"] = index
                conn.execute(text("""
                    UPDATE employees SET stt=:stt, source_row=:source_row,
                        payload=CAST(:payload AS jsonb), updated_at=NOW()
                    WHERE username=:username
                """), {
                    "stt": index, "source_row": index + 1, "username": row["username"],
                    "payload": json.dumps(payload, ensure_ascii=False),
                })
            staff_ws = credential_ws()
            deleted_sheet_rows = delete_rows_by_name(staff_ws, target_keys, 2)
            renumber_sheet(staff_ws)
            status_ws = worksheet(STATUS_WORKSHEET, 1000, len(STATUS_HEADERS), STATUS_HEADERS)
            deleted_status_rows = delete_rows_by_name(status_ws, target_keys, 2)
            renumber_sheet(status_ws)
            tx.commit()
            sync_tichluy_members(conn)
            return {"ok": True, "deleted": len(targets), "message": f"Đã xóa {len(targets)} nhân viên THÀNH CÔNG. Lịch sử lịch nghỉ được giữ nguyên."}
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            if staff_ws is not None:
                for index, values in sorted(deleted_sheet_rows):
                    try:
                        staff_ws.insert_row(values, index=index, value_input_option="USER_ENTERED")
                    except Exception:
                        pass
            if status_ws is not None:
                for index, values in sorted(deleted_status_rows):
                    try:
                        status_ws.insert_row(values, index=index, value_input_option="USER_ENTERED")
                    except Exception:
                        pass
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            if staff_ws is not None:
                for index, values in sorted(deleted_sheet_rows):
                    try:
                        staff_ws.insert_row(values, index=index, value_input_option="USER_ENTERED")
                    except Exception:
                        pass
            if status_ws is not None:
                for index, values in sorted(deleted_status_rows):
                    try:
                        status_ws.insert_row(values, index=index, value_input_option="USER_ENTERED")
                    except Exception:
                        pass
            raise HTTPException(500, f"Không xóa được nhân viên an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()

    def filtered_staff(conn, ident, search: str, role: str, status: str, shift: str) -> list[dict[str, Any]]:
        data = staff_result(conn, ident)["employees"]
        if search.strip():
            needle = norm(search)
            exact = [row for row in data if needle in {norm(row["username"]), norm(row["full_name"])}]
            data = exact or [row for row in data if needle in norm(f"{row['username']} {row['full_name']}")]
        if role.strip():
            data = [row for row in data if row["role"] == role.strip().lower()]
        if status.strip():
            wanted = _status_value(status, norm)
            data = [row for row in data if row["employment_status"] == wanted]
        if shift.strip():
            data = [row for row in data if row["work_shift"] == shift.strip()]
        return data

    def build_staff_workbook(rows: list[dict[str, Any]], shifts: dict[str, list[str]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhSachNhanSu"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.append(STAFF_EXPORT_COLUMNS)
        header_fill = PatternFill("solid", fgColor="214639")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="DDE5E0")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        ws.row_dimensions[1].height = 34

        date_columns = {"Ngày bắt đầu làm", "Ngày sinh", "Ngày bắt đầu ca"}
        number_columns = {"Phát sinh tháng", "Có phép tháng", "Phép năm"}
        for row in rows:
            values = {
                "Tên nhân viên": row["username"], "Họ và tên đầy đủ": row["full_name"],
                "Ngày bắt đầu làm": row["employment_start_date"], "Ngày sinh": row["birth_date"],
                "Phân quyền": row["role"], "Trạng thái làm việc": row["employment_status"],
                "Điện thoại": row["phone"], "Email": row["email"], "Địa chỉ": row["address"],
                "Số tài khoản ngân hàng": row["bank_account"], "Tên ngân hàng": row["bank_name"],
                "Phát sinh tháng": row["monthly_generated"], "Có phép tháng": row["monthly_leave"],
                "Phép năm": row["annual_leave"], "Ca làm việc": row["work_shift"],
                "Ngày bắt đầu ca": row["shift_start_date"], "Chu kỳ": row["rotation_cycle"],
                "Khóa đăng nhập": "KHÓA" if row["login_locked"] else "",
            }
            excel_values = []
            for column in STAFF_EXPORT_COLUMNS:
                value = values.get(column, "")
                if column in date_columns and value:
                    try:
                        value = datetime.strptime(str(value), "%d/%m/%Y").date()
                    except ValueError:
                        pass
                excel_values.append(value)
            ws.append(excel_values)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row[0].row].height = 22
        for column in date_columns:
            index = STAFF_EXPORT_COLUMNS.index(column) + 1
            for cell in ws.iter_cols(min_col=index, max_col=index, min_row=2, max_row=max(2, ws.max_row)):
                for item in cell:
                    item.number_format = "dd/mm/yyyy"
        for column in number_columns:
            index = STAFF_EXPORT_COLUMNS.index(column) + 1
            for cell in ws.iter_cols(min_col=index, max_col=index, min_row=2, max_row=max(2, ws.max_row)):
                for item in cell:
                    item.number_format = "0.0"
                    item.alignment = Alignment(horizontal="right", vertical="center")
        for column in ("Điện thoại", "Số tài khoản ngân hàng"):
            index = STAFF_EXPORT_COLUMNS.index(column) + 1
            for cell in ws.iter_cols(min_col=index, max_col=index, min_row=2, max_row=max(2, ws.max_row)):
                for item in cell:
                    item.number_format = "@"

        widths = {
            "Tên nhân viên": 24, "Họ và tên đầy đủ": 28, "Ngày bắt đầu làm": 18, "Ngày sinh": 16,
            "Phân quyền": 14, "Trạng thái làm việc": 22, "Điện thoại": 16, "Email": 30,
            "Địa chỉ": 42, "Số tài khoản ngân hàng": 22, "Tên ngân hàng": 24,
            "Phát sinh tháng": 16, "Có phép tháng": 16, "Phép năm": 14, "Ca làm việc": 28,
            "Ngày bắt đầu ca": 18, "Chu kỳ": 23, "Khóa đăng nhập": 18,
        }
        for index, column in enumerate(STAFF_EXPORT_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(index)].width = widths[column]
        ws.auto_filter.ref = f"A1:R{max(1, ws.max_row)}"

        catalog = wb.create_sheet("DanhMuc")
        catalog.sheet_state = "hidden"
        lists = {
            "A": ALL_ROLES, "B": STATUS_OPTIONS, "C": CYCLE_OPTIONS, "D": ["", "KHÓA"],
            "E": list(dict.fromkeys(shift for values in shifts.values() for shift in values)),
        }
        for column, values in lists.items():
            for index, value in enumerate(values, start=1):
                catalog[f"{column}{index}"] = value
        validation_map = {
            "Phân quyền": ("A", len(lists["A"])),
            "Trạng thái làm việc": ("B", len(lists["B"])),
            "Chu kỳ": ("C", len(lists["C"])),
            "Khóa đăng nhập": ("D", len(lists["D"])),
            "Ca làm việc": ("E", max(1, len(lists["E"]))),
        }
        last_row = max(2, ws.max_row)
        for column, (source_col, count) in validation_map.items():
            target_col = get_column_letter(STAFF_EXPORT_COLUMNS.index(column) + 1)
            validation = DataValidation(type="list", formula1=f"=DanhMuc!${source_col}$1:${source_col}${count}", allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f"{target_col}2:{target_col}{last_row}")
        status_col = get_column_letter(STAFF_EXPORT_COLUMNS.index("Trạng thái làm việc") + 1)
        ws.conditional_formatting.add(
            f"A2:R{last_row}",
            FormulaRule(formula=[f'${status_col}2="Đã nghỉ việc"'], fill=PatternFill("solid", fgColor="FDECEC")),
        )

        guide = wb.create_sheet("HuongDan")
        guide.sheet_view.showGridLines = False
        guide["A1"] = "HƯỚNG DẪN IMPORT DANH SÁCH NHÂN SỰ"
        guide["A1"].font = Font(bold=True, size=15, color="214639")
        notes = [
            "1. Không đổi Tên nhân viên vì đây là khóa đối chiếu tài khoản.",
            "2. Import chỉ cập nhật tài khoản đã tồn tại; không tạo và không xóa tài khoản.",
            "3. Mật khẩu và Remember Token không được xuất hoặc ghi đè từ file này.",
            "4. Ngày dùng định dạng DD/MM/YYYY. Hãy dùng các danh sách chọn cho vai trò, trạng thái, ca và chu kỳ.",
            "5. Toàn bộ file được kiểm tra trước khi ghi; nếu có một dòng sai, hệ thống không áp dụng file.",
        ]
        for index, note in enumerate(notes, start=3):
            guide[f"A{index}"] = note
        guide.column_dimensions["A"].width = 110
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @app.get("/v2/staff/export.xlsx")
    def export_staff_excel(
        search: str = Query(default="", max_length=200),
        role: str = Query(default="", max_length=50),
        status: str = Query(default="", max_length=100),
        shift: str = Query(default="", max_length=300),
        ident: identity_type = Depends(current_identity),
    ):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "staff_export")
            rows = filtered_staff(conn, ident, search, role, status, shift)
            shifts = staff_result(conn, ident)["shifts_by_department"]
        content = build_staff_workbook(rows, shifts)
        filename = f"VeraSpa_DanhSachNhanSu_{datetime.now(vn_tz).strftime('%d%m%Y')}.xlsx"
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    def parse_import(content: bytes) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(400, f"Không đọc được file Excel: {exc}") from exc
        ws = workbook["DanhSachNhanSu"] if "DanhSachNhanSu" in workbook.sheetnames else workbook.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "File Excel không có dữ liệu.")
        headers = [str(value or "").strip() for value in rows[0]]
        if "Tên nhân viên" not in headers:
            raise HTTPException(400, "File import phải có cột 'Tên nhân viên'.")
        output = []
        seen = set()
        for row_number, values in enumerate(rows[1:], start=2):
            item = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers)) if headers[index]}
            username = str(item.get("Tên nhân viên") or "").strip()
            if not username:
                continue
            key = norm(username)
            if key in seen:
                raise HTTPException(400, f"Tên nhân viên bị trùng trong file import: {username}")
            seen.add(key)
            item["_row"] = row_number
            output.append(item)
        if not output:
            raise HTTPException(400, "File import không có nhân viên.")
        return output

    def import_values(item: dict[str, Any]) -> dict[str, Any]:
        mapping = {
            "Họ và tên đầy đủ": "full_name", "Phân quyền": "role", "Điện thoại": "phone",
            "Email": "email", "Địa chỉ": "address", "Số tài khoản ngân hàng": "bank_account",
            "Tên ngân hàng": "bank_name", "Ca làm việc": "work_shift", "Chu kỳ": "rotation_cycle",
        }
        output = {}
        for source, target in mapping.items():
            if source not in item:
                continue
            value = str(item.get(source) or "").strip()
            # A blank role means "keep current". Blank profile, shift and
            # cycle cells intentionally clear their current values.
            if target == "role" and not value:
                continue
            output[target] = value
        for source, target in (
            ("Ngày bắt đầu làm", "employment_start_date"),
            ("Ngày sinh", "birth_date"),
            ("Ngày bắt đầu ca", "shift_start_date"),
        ):
            if source in item:
                output[target] = _excel_date(item.get(source), field_name=source)
        for source, target in (
            ("Phát sinh tháng", "monthly_generated"),
            ("Có phép tháng", "monthly_leave"),
            ("Phép năm", "annual_leave"),
        ):
            if source in item:
                output[target] = _number(item.get(source), field_name=source)
        if "Trạng thái làm việc" in item and str(item.get("Trạng thái làm việc") or "").strip():
            output["employment_status"] = _status_value(item.get("Trạng thái làm việc"), norm)
        if "Khóa đăng nhập" in item:
            output["login_locked"] = norm(item.get("Khóa đăng nhập")) in {"khoa", "1", "true", "yes", "x"}
        return output

    @app.post("/v2/staff/import.xlsx")
    async def import_staff_excel(request: Request, ident: identity_type = Depends(current_identity)):
        length = int(request.headers.get("content-length") or 0)
        if length > 5 * 1024 * 1024:
            raise HTTPException(413, "File Excel vượt quá 5 MB.")
        content = await request.body()
        if not content:
            raise HTTPException(400, "Chưa chọn file Excel.")
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(413, "File Excel vượt quá 5 MB.")
        imported = parse_import(content)
        engine = engine_instance()
        conn = engine.connect()
        tx = conn.begin()
        staff_ws = None
        old_sheet_values = None
        old_status_values = None
        status_ws = None
        leave_prunes: list[dict[str, Any]] = []
        try:
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:employees'))"))
            require_feature(conn, ident, "staff_import")
            rows = _select_staff_rows(conn, for_update=True)
            known = {norm(row["username"]): row for row in rows}
            unknown = [str(item["Tên nhân viên"]) for item in imported if norm(item["Tên nhân viên"]) not in known]
            if unknown:
                raise HTTPException(400, "Không tìm thấy trong hệ thống: " + ", ".join(unknown))
            status_snapshot = google_status_map()
            updates = []
            for item in imported:
                row = known[norm(item["Tên nhân viên"])]
                ensure_manageable(ident, str(row.get("role") or ""))
                values = import_values(item)
                updates.append(update_database_row(conn, ident, row, values, status_map=status_snapshot))

            statuses = dict(status_snapshot)
            for updated in updates:
                statuses[norm(updated["username"])] = updated["_status"]
            refreshed = _select_staff_rows(conn)
            staff_ws = credential_ws()
            old_sheet_values = staff_ws.get_all_values()
            output = [
                _credential_values(row, statuses.get(norm(row["username"]), _effective_status(row, statuses, norm)))
                for row in refreshed
            ]
            if output:
                staff_ws.update(range_name=f"A2:U{len(output) + 1}", values=output, value_input_option="USER_ENTERED")

            status_ws = worksheet(STATUS_WORKSHEET, 1000, len(STATUS_HEADERS), STATUS_HEADERS)
            old_status_values = status_ws.get_all_values()
            now = datetime.now(vn_tz)
            status_rows = []
            for index, row in enumerate(refreshed, start=1):
                if str(row.get("role") or "").lower() == "admin":
                    continue
                status_rows.append([
                    index, row["username"], statuses.get(norm(row["username"]), STATUS_OPTIONS[0]),
                    now.strftime("%d/%m/%Y"), now.strftime("%H:%M:%S"), ident.employee_username,
                ])
            if len(old_status_values) > 1:
                status_ws.batch_clear([f"A2:F{max(len(old_status_values), len(status_rows) + 1)}"])
            if status_rows:
                status_ws.update(range_name=f"A2:F{len(status_rows) + 1}", values=status_rows, value_input_option="USER_ENTERED")
            effective_date = datetime.now(vn_tz).date()
            for updated in updates:
                if updated["_status_changed"] and updated["_status"] in {STATUS_OPTIONS[1], STATUS_OPTIONS[2]}:
                    change = prune_registered_leave(conn, updated["username"], effective_date)
                    if int(change.get("count") or 0):
                        leave_prunes.append(change)
            tx.commit()
            sync_tichluy_members(conn)
            deleted_leave = sum(int(change.get("count") or 0) for change in leave_prunes)
            suffix = f" Đã xóa {deleted_leave} lịch nghỉ từ ngày hiệu lực trở về sau." if deleted_leave else ""
            return {
                "ok": True,
                "updated": len(updates),
                "deleted_leave_records": deleted_leave,
                "message": f"Đã Import và cập nhật {len(updates)} nhân viên THÀNH CÔNG.{suffix}",
            }
        except HTTPException:
            if tx.is_active:
                tx.rollback()
            for change in reversed(leave_prunes):
                restore_pruned_leave(change)
            if staff_ws is not None and old_sheet_values:
                try:
                    staff_ws.batch_clear([f"A1:U{max(len(old_sheet_values), 2)}"])
                    staff_ws.update(range_name=f"A1:U{len(old_sheet_values)}", values=old_sheet_values, value_input_option="USER_ENTERED")
                except Exception:
                    pass
            if status_ws is not None and old_status_values:
                try:
                    status_ws.batch_clear([f"A1:F{max(len(old_status_values), 2)}"])
                    status_ws.update(range_name=f"A1:F{len(old_status_values)}", values=old_status_values, value_input_option="USER_ENTERED")
                except Exception:
                    pass
            raise
        except Exception as exc:
            if tx.is_active:
                tx.rollback()
            for change in reversed(leave_prunes):
                restore_pruned_leave(change)
            if staff_ws is not None and old_sheet_values:
                try:
                    staff_ws.batch_clear([f"A1:U{max(len(old_sheet_values), 2)}"])
                    staff_ws.update(range_name=f"A1:U{len(old_sheet_values)}", values=old_sheet_values, value_input_option="USER_ENTERED")
                except Exception:
                    pass
            if status_ws is not None and old_status_values:
                try:
                    status_ws.batch_clear([f"A1:F{max(len(old_status_values), 2)}"])
                    status_ws.update(range_name=f"A1:F{len(old_status_values)}", values=old_status_values, value_input_option="USER_ENTERED")
                except Exception:
                    pass
            raise HTTPException(500, f"Không Import được danh sách nhân viên an toàn: {type(exc).__name__}: {exc}") from exc
        finally:
            conn.close()
