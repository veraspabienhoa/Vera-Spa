"""VERA SPA Web V2 Python API.

This service is the server-side write boundary for the React/Web V2 pilot.
The browser never writes leave_records directly.  The API validates the
Supabase session, resolves the mapped VERA employee, derives leave days and
penalty from the canonical LoaiNghi policy, applies the same safety invariants
used by the Streamlit leave registration flow, writes PostgreSQL with a
record_uid, and mirrors the row to the legacy MainData Google Sheet.

The API deliberately fails closed when policy data is missing or a rule cannot
be interpreted.  That is safer than allowing Web V2 to bypass a legacy rule.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
import hashlib
import hmac
from io import BytesIO
import json
import math
import os
import re
import threading
import time
import unicodedata
import uuid
from typing import Any
from urllib.parse import quote

import google.auth
import gspread
import requests
from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import bindparam, create_engine, text
from sqlalchemy.engine import URL

from vera_leave_registration_shared import summarize_leave_day
from vera_json import json_safe, json_text

VN_TZ = timezone(timedelta(hours=7))
LEAVE_SHEET_ID = os.getenv(
    "VERA_LEAVE_SHEET_ID", "1Kz0aw-JatptAN9G7YSwZ6rJO09urOPaD-rS-18eZSY0"
)
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://nunxfjhrszmlyyrvphuq.supabase.co").rstrip("/")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "").strip()
CORS_ORIGINS = [
    x.strip()
    for x in os.getenv(
        "VERA_V2_CORS_ORIGINS",
        "https://app.veraspa.vn,https://veraspabienhoa.github.io,http://localhost:5173,http://127.0.0.1:5173",
    ).split(",")
    if x.strip()
]

app = FastAPI(title="VERA SPA API", version="3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
    expose_headers=["Content-Disposition"],
)

_engine = None
_gspread = None


def _engine_instance():
    global _engine
    if _engine is None:
        host = os.getenv("DB_HOST", "").strip()
        user = os.getenv("DB_USER", "").strip()
        password = os.getenv("DB_PASS", "")
        database = os.getenv("DB_NAME", "postgres")
        port = int(os.getenv("DB_PORT", "5432"))
        if not host or not user or not password:
            raise RuntimeError("PostgreSQL environment is incomplete")
        url = URL.create(
            "postgresql+psycopg", username=user, password=password,
            host=host, port=port, database=database,
        )
        _engine = create_engine(
            url,
            pool_pre_ping=True,
            pool_size=int(os.getenv("DB_POOL_SIZE", "2")),
            max_overflow=int(os.getenv("DB_MAX_OVERFLOW", "0")),
            # Keep healthy pooler connections longer.  Recreating a connection
            # repeatedly triggers PostgreSQL timezone-catalog discovery.
            pool_recycle=max(3600, int(os.getenv("DB_POOL_RECYCLE", "3600"))),
        )
    return _engine


def _google_client():
    global _gspread
    if _gspread is None:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds, _ = google.auth.default(scopes=scopes)
        _gspread = gspread.authorize(creds)
    return _gspread


def _norm(value: Any) -> str:
    s = str(value or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.replace("đ", "d")
    return re.sub(r"\s+", " ", s).strip()


def _employee_name_matches(value: Any, query: Any) -> bool:
    needle = _norm(query)
    if not needle:
        return True
    short_name = re.split(r"\s*[-–—]\s*", str(value or ""), maxsplit=1)[0]
    return needle in {_norm(value), _norm(short_name)}


_WATCHED_PAID_REASON_KEYS = {
    _norm("Nghỉ CÓ phép"),
    _norm("Nghỉ CUỐI TUẦN CÓ phép"),
    _norm("Đi trễ CÓ phép"),
    _norm("Đi trễ CUỐI TUẦN CÓ phép"),
    _norm("Về sớm CÓ phép"),
    _norm("Về sớm CUỐI TUẦN CÓ phép"),
}


def _num(value: Any, default=0.0, money=False) -> float:
    try:
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value or "").strip()
        if not s or s.lower() in {"nan", "none", "-"}:
            return float(default)
        if money:
            s = re.sub(r"[^0-9-]", "", s)
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return float(default)


def _policy_payload(conn) -> dict:
    # Preferred official policy introduced by the Nội quy page.
    row = conn.execute(text("""
        SELECT value_json FROM vera_app_setting
        WHERE category='official_policy' AND setting_key='leave_rules'
        LIMIT 1
    """)).scalar_one_or_none()
    if isinstance(row, dict) and (row.get("rows") or row.get("columns")):
        return row

    # Production-safe fallback to the PostgreSQL LoaiNghi snapshot.  It is still
    # PostgreSQL data, not a browser-provided rule set.
    row = conn.execute(text("""
        SELECT value_json FROM vera_app_setting
        WHERE category='leave_rules' AND setting_key='loai_nghi_snapshot_v2'
        LIMIT 1
    """)).scalar_one_or_none()
    if isinstance(row, dict) and row.get("rows"):
        return row
    raise HTTPException(503, "Chưa có Nội quy/LoaiNghi trong PostgreSQL; Web V2 không được phép ghi.")


def _policy_rows(conn) -> list[dict]:
    payload = _policy_payload(conn)
    headers = payload.get("columns") or payload.get("headers") or []
    rows = payload.get("rows") or []
    if rows and isinstance(rows[0], dict):
        return [dict(r) for r in rows]
    if not headers:
        headers = [
            "STT", "Lý do nghỉ", "Loại nghỉ", "Chi tiết", "Số ngày tính phép",
            "Phạt vi phạm", "Chỉ nhập được cuối tuần", "User có quyền được nhập",
            "Kiều đăng ký", "Giá trị", "Ngoại trừ đăng ký", "Kiểu hủy",
            "Số ngày hủy trước", "Ngoại trừ hủy", "Ghi chú",
        ]
    return [dict(zip(headers, list(r) + [""] * max(0, len(headers) - len(r)))) for r in rows]


def _field(row: dict, *names: str, default=""):
    by_norm = {_norm(k): v for k, v in row.items()}
    for name in names:
        key = _norm(name)
        if key in by_norm:
            return by_norm[key]
    return default


def _reason_item(conn, reason: str) -> dict:
    wanted = _norm(reason)
    for row in _policy_rows(conn):
        name = str(_field(row, "Lý do nghỉ", default="") or "").strip()
        if _norm(name) == wanted:
            detail = str(_field(row, "Chi tiết", default="") or "").strip()
            return {
                "name": name,
                "leave_type": str(_field(row, "Loại nghỉ", default="") or "").strip(),
                "detail_config": detail,
                "days": _num(_field(row, "Số ngày tính phép", "Số ngày tính", default=0)),
                "penalty": _num(_field(row, "Phạt vi phạm", default=0), money=True),
                "allowed_days": str(_field(row, "Chỉ nhập được cuối tuần", "Ngày được phép nhập", default="") or "").strip(),
                "allowed_roles": str(_field(row, "User có quyền được nhập", default="") or "").strip(),
                "register_type": str(_field(row, "Kiểu đăng ký", "Kiều đăng ký", default="") or "").strip(),
                "register_value": str(_field(row, "Giá trị", "Giá trị đăng ký", default="") or "").strip(),
                "register_exceptions": str(_field(row, "Ngoại trừ đăng ký", default="") or "").strip(),
                "cancel_type": str(_field(row, "Kiểu hủy", default="") or "").strip(),
                "cancel_value": str(_field(row, "Số ngày hủy trước", "Giá trị hủy", default="") or "").strip(),
                "cancel_exceptions": str(_field(row, "Ngoại trừ hủy", default="") or "").strip(),
                "requires_manual_penalty": "can nhap so tien" in _norm(detail),
            }
    raise HTTPException(400, f"Không tìm thấy Lý do nghỉ '{reason}' trong Nội quy/LoaiNghi.")


def _role_tokens(value: str) -> set[str]:
    n = _norm(value)
    roles = {"admin", "quanly", "letan", "leader", "nhanvien", "locker", "tapvu", "auto update"}
    return {r for r in roles if r in n}


def _weekday_label(d: date) -> str:
    return ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"][d.weekday()]


def _weekday_short_label(d: date) -> str:
    return ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"][d.weekday()]


def _stats_group(leave_type: str, reason: str) -> str:
    """Match the canonical daily-stat grouping used by the current app."""
    type_key = _norm(leave_type)
    reason_key = _norm(reason)
    if "khong phep" in type_key:
        return "khong_phep"
    if "phat sinh" in type_key:
        return "phat_sinh"
    if "co phep" in type_key:
        return "co_phep"
    if "khong phep" in reason_key:
        return "khong_phep"
    if "phat sinh" in reason_key:
        return "phat_sinh"
    if (
        "co phep" in reason_key
        or "nghi phep" in reason_key
        or "nghi dam hieu" in reason_key
        or re.search(r"(^|\s)cp($|\s)", reason_key)
    ):
        return "co_phep"
    return ""


def _daily_quota_config(conn) -> dict[str, Any]:
    default_days = [
        {"weekday": weekday, "paid_limit": 3 if weekday >= 6 else 5, "generated_limit": 0 if weekday >= 6 else 2}
        for weekday in range(1, 8)
    ]
    defaults: dict[str, Any] = {
        "weekday_limit": 5,
        "weekend_limit": 3,
        "phat_sinh_limit": 2,
        "days": default_days,
    }
    payload = conn.execute(text("""
        SELECT value_json
        FROM vera_app_setting
        WHERE category='leave_rules' AND setting_key='daily_quota'
        LIMIT 1
    """)).scalar_one_or_none()
    if not isinstance(payload, dict):
        return defaults
    output = dict(defaults)
    for key in ("weekday_limit", "weekend_limit", "phat_sinh_limit"):
        try:
            output[key] = max(0, int(float(payload.get(key, output[key]))))
        except (TypeError, ValueError):
            pass
    raw_days = payload.get("days") if isinstance(payload.get("days"), list) else []
    parsed_days: dict[int, dict[str, int]] = {}
    for item in raw_days:
        if not isinstance(item, dict):
            continue
        try:
            weekday = int(item.get("weekday") or 0)
            if weekday not in range(1, 8) or weekday in parsed_days:
                continue
            parsed_days[weekday] = {
                "weekday": weekday,
                "paid_limit": max(0, int(float(item.get("paid_limit", 0)))),
                "generated_limit": max(0, int(float(item.get("generated_limit", 0)))),
            }
        except (TypeError, ValueError):
            continue
    if len(parsed_days) == 7:
        output["days"] = [parsed_days[index] for index in range(1, 8)]
    else:
        output["days"] = [
            {
                "weekday": weekday,
                "paid_limit": output["weekend_limit"] if weekday >= 6 else output["weekday_limit"],
                "generated_limit": 0 if weekday >= 6 else output["phat_sinh_limit"],
            }
            for weekday in range(1, 8)
        ]
    return output


def _day_allowed(rule: str, d: date) -> bool:
    n = _norm(rule)
    if not n or n in {"tat ca", "all", "none", "nan"}:
        return True
    tokens = {
        0: ("thu hai", "t2"), 1: ("thu ba", "t3"), 2: ("thu tu", "t4"),
        3: ("thu nam", "t5"), 4: ("thu sau", "t6"),
        5: ("thu bay", "thu bay", "t7", "cuoi tuan"), 6: ("chu nhat", "cn", "cuoi tuan"),
    }[d.weekday()]
    return any(t in n for t in tokens)


def _parse_first_number(value: str):
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(value or ""))
    return float(m.group(0).replace(",", ".")) if m else None


def _registration_rule(item: dict, role: str, target: date):
    if role == "admin":
        return
    allowed = _role_tokens(item["allowed_roles"])
    if allowed and role not in allowed:
        raise HTTPException(403, f"Tài khoản {role} không được dùng lý do '{item['name']}'.")
    if not _day_allowed(item["allowed_days"], target):
        raise HTTPException(400, f"'{item['name']}' không được nhập vào {_weekday_label(target)} {target.strftime('%d/%m/%Y')}.")

    now = datetime.now(VN_TZ)
    today = now.date()
    if target < today:
        raise HTTPException(400, "Không được đăng ký lịch ở ngày quá khứ.")
    if role in _role_tokens(item["register_exceptions"]):
        return

    typ = _norm(item["register_type"])
    val = item["register_value"]
    if "khong gioi han" in typ:
        return
    if "truoc n ngay" in typ or typ in {"truoc ngay", "before days"}:
        n = _parse_first_number(val)
        if n is None:
            raise HTTPException(400, f"Giá trị đăng ký của '{item['name']}' không hợp lệ.")
        earliest = today + timedelta(days=max(0, int(n)))
        if target < earliest:
            raise HTTPException(400, f"'{item['name']}' phải đăng ký trước ít nhất {int(n)} ngày. Ngày sớm nhất: {earliest.strftime('%d/%m/%Y')}.")
        return
    if "ngay hien tai tu gio" in typ:
        m = re.search(r"(\d{1,2})\s*(?::|h)?\s*(\d{1,2})?", str(val or ""))
        if not m:
            raise HTTPException(400, f"Giờ đăng ký của '{item['name']}' không hợp lệ.")
        hh, mm = int(m.group(1)), int(m.group(2) or 0)
        if target != today or now.time() < now.replace(hour=hh, minute=mm, second=0, microsecond=0).time():
            raise HTTPException(400, f"'{item['name']}' chỉ được đăng ký cho ngày hiện tại từ {hh:02d}:{mm:02d}.")
        return
    if "khong cho phep" in typ or "khong duoc dang ky" in typ:
        raise HTTPException(400, f"'{item['name']}' đang được cấu hình không cho phép đăng ký.")
    raise HTTPException(400, f"Không nhận diện được Kiểu đăng ký '{item['register_type']}' của '{item['name']}'.")


def _group(reason: str) -> str:
    n = _norm(reason)
    if "khong phep" in n:
        return "khong_phep"
    if "phat sinh" in n:
        return "phat_sinh"
    excluded = ("di tre", "ve som", "ra som", "qua tour", "loi vi pham", "khong don", "xuong phong", "ho tro")
    if "co phep" in n and not any(x in n for x in excluded):
        return "co_phep"
    return ""


def _is_video(reason: str) -> bool:
    return _norm(reason) == "nghi phep quay video"


def _is_long_sick(reason: str) -> bool:
    return _norm(reason) == "nghi benh co giay kham hoac duoc quan ly duyet"


def _is_annual(reason: str) -> bool:
    return "phep nam" in _norm(reason)


def _progressive_key(reason: str) -> str:
    n = _norm(reason)
    if "nghi" in n and "khong phep" in n:
        return "nghi_khong_phep"
    if "di tre" in n and "khong phep" in n:
        return "di_tre_khong_phep"
    if ("ve som" in n or "ra som" in n) and "khong phep" in n:
        return "ve_som_khong_phep"
    return ""


class Identity(BaseModel):
    auth_user_id: str
    employee_username: str
    role: str
    full_name: str = ""
    email: str = ""
    must_change_password: bool = False


from vera_web_v2_permissions import DEFAULT_ROLE_FEATURES as WEB_V2_DEFAULT_FEATURES, FEATURES as WEB_V2_FEATURES

_PERMISSION_CACHE_SECONDS = max(
    1.0, float(os.getenv("VERA_PERMISSION_CACHE_SECONDS", "60") or 60),
)
_permission_cache_lock = threading.Lock()
_permission_cache: dict[str, Any] = {"loaded_at": 0.0, "payload": {}}


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _norm(value) in {"1", "true", "yes", "y", "co", "có", "x"}


def _clear_permission_cache() -> None:
    with _permission_cache_lock:
        _permission_cache.update({"loaded_at": 0.0, "payload": {}})


def _permission_payload(conn) -> dict[str, Any]:
    now = time.monotonic()
    with _permission_cache_lock:
        if now - float(_permission_cache["loaded_at"] or 0) <= _PERMISSION_CACHE_SECONDS:
            return dict(_permission_cache["payload"])
        payload = conn.execute(text("""
            SELECT value_json
            FROM vera_app_setting
            WHERE category='authorization' AND setting_key='feature_permissions'
            LIMIT 1
        """)).scalar_one_or_none()
        payload = payload if isinstance(payload, dict) else {}
        _permission_cache.update({"loaded_at": time.monotonic(), "payload": dict(payload)})
        return dict(payload)


def _feature_allowed(
    conn, ident: Identity, feature: str, permission_payload: dict[str, Any] | None = None,
) -> bool:
    """Mirror Streamlit precedence: admin -> account -> role -> defaults."""
    feature = str(feature or "").strip()
    role = str(ident.role or "").strip().lower()
    if role == "admin":
        return True

    payload = permission_payload if isinstance(permission_payload, dict) else _permission_payload(conn)

    username_key = _norm(ident.employee_username)
    for item in payload.get("accounts", []) or []:
        if not isinstance(item, dict):
            continue
        if _norm(item.get("target")) == username_key and str(item.get("feature") or "").strip() == feature:
            return _as_bool(item.get("allowed"))

    for item in payload.get("roles", []) or []:
        if not isinstance(item, dict):
            continue
        if str(item.get("target") or "").strip().lower() == role and str(item.get("feature") or "").strip() == feature:
            return _as_bool(item.get("allowed"))

    return feature in WEB_V2_DEFAULT_FEATURES.get(role, set())


def _registration_role_locked(conn, role: str) -> bool:
    if str(role or "").strip().lower() == "admin":
        return False
    payload = conn.execute(text("""
        SELECT value_json
        FROM vera_app_setting
        WHERE category='control' AND setting_key='registration_role_locks'
        LIMIT 1
    """)).scalar_one_or_none()
    payload = payload if isinstance(payload, dict) else {}
    return _as_bool(payload.get(str(role or "").strip().lower(), False))


def _require_password_changed(ident: Identity):
    if ident.must_change_password:
        raise HTTPException(428, "Bạn phải đổi mật khẩu lần đầu trước khi sử dụng chức năng này.")


def _require_feature(conn, ident: Identity, feature: str):
    if feature not in {"profile", "profile_edit"}:
        _require_password_changed(ident)
    if not _feature_allowed(conn, ident, feature):
        raise HTTPException(403, "Tài khoản hiện tại chưa được cấp quyền dùng chức năng này.")


async def current_identity(authorization: str | None = Header(default=None)) -> Identity:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "Thiếu phiên đăng nhập Supabase.")
    token = authorization.split(" ", 1)[1].strip()
    if not token or not SUPABASE_ANON_KEY:
        raise HTTPException(503, "API Auth chưa được cấu hình.")
    try:
        resp = requests.get(
            f"{SUPABASE_URL}/auth/v1/user",
            headers={"Authorization": f"Bearer {token}", "apikey": SUPABASE_ANON_KEY},
            timeout=10,
        )
    except requests.RequestException as exc:
        raise HTTPException(503, f"Không xác minh được phiên đăng nhập: {exc}") from exc
    if resp.status_code != 200:
        raise HTTPException(401, "Phiên đăng nhập không hợp lệ hoặc đã hết hạn.")
    user = resp.json()
    auth_uid = str(user.get("id") or "")
    if not auth_uid:
        raise HTTPException(401, "Không xác định được tài khoản Supabase.")

    with _engine_instance().connect() as conn:
        row = conn.execute(text("""
            SELECT p.auth_user_id::text, p.employee_username, p.role,
                   COALESCE(e.full_name,''), COALESCE(e.email,''),
                   lower(COALESCE(e.payload->>'must_change_password','')) IN ('1','true','yes','y')
            FROM vera_v2_user_profile p
            JOIN employees e ON e.username=p.employee_username
            WHERE p.auth_user_id=CAST(:uid AS uuid)
              AND p.is_active=true
              AND COALESCE(e.login_locked,false)=false
              AND COALESCE(
                    e.payload->>'Trạng thái làm việc',
                    e.payload->>'employment_status',
                    'Đang làm việc'
                  ) = 'Đang làm việc'
            LIMIT 1
        """), {"uid": auth_uid}).first()
    if not row:
        raise HTTPException(403, "Tài khoản chưa được liên kết với nhân viên VERA đang hoạt động.")
    return Identity(
        auth_user_id=row[0], employee_username=row[1], role=str(row[2] or "").lower(),
        full_name=row[3], email=row[4], must_change_password=bool(row[5]),
    )


class LeaveCreate(BaseModel):
    leave_date: date
    employee_name: str = Field(min_length=1, max_length=200)
    leave_reason: str = Field(min_length=1, max_length=300)
    detail: str = Field(default="", max_length=3000)
    manual_penalty: float | None = Field(default=None, ge=0)


class LeaveUpdate(BaseModel):
    leave_reason: str = Field(min_length=1, max_length=300)
    manual_penalty: float | None = Field(default=None, ge=0)


class LeaveDelete(BaseModel):
    record_uids: list[str] = Field(min_length=1, max_length=100)


class LeaveWatchUpdate(BaseModel):
    watched_date: date
    watching: bool = True


class LeaveWatchAcknowledge(BaseModel):
    watched_dates: list[date] = Field(min_length=1, max_length=100)


class PushSubscriptionCreate(BaseModel):
    subscription: dict[str, Any]


class PushSubscriptionDelete(BaseModel):
    endpoint: str = Field(min_length=10, max_length=4000)


class PushDispatch(BaseModel):
    dates: list[date] = Field(default_factory=list, max_length=31)


_LEAVE_EDIT_FEATURES = ("leave_manage_edit", "leave_detail_edit")
_LEAVE_DELETE_FEATURES = ("leave_manage_delete", "leave_detail_delete")
_EMPLOYEE_LIKE_ROLES = {"nhanvien", "leader", "locker", "tapvu"}


def _has_any_feature(conn, ident: Identity, features: tuple[str, ...]) -> bool:
    return any(_feature_allowed(conn, ident, feature) for feature in features)


def _paid_interest_counts(conn, watched_dates: list[date]) -> dict[date, int]:
    dates = sorted(set(watched_dates))
    counts = {value: 0 for value in dates}
    if not dates:
        return counts
    rows = conn.execute(text("""
        SELECT leave_date, leave_reason
        FROM leave_records
        WHERE leave_date BETWEEN :start_date AND :end_date
        ORDER BY leave_date
    """), {"start_date": dates[0], "end_date": dates[-1]}).mappings().all()
    date_set = set(dates)
    for row in rows:
        target = row.get("leave_date")
        if target in date_set and _norm(row.get("leave_reason")) in _WATCHED_PAID_REASON_KEYS:
            counts[target] += 1
    return counts


def _refresh_leave_watches(conn, ident: Identity) -> list[dict[str, Any]]:
    rows = conn.execute(text("""
        SELECT watched_date, last_seen_paid_count, current_paid_count, has_unread,
               created_at, updated_at
        FROM vera_v2_leave_watch
        WHERE auth_user_id=CAST(:auth_user_id AS uuid)
        ORDER BY watched_date
        FOR UPDATE
    """), {"auth_user_id": ident.auth_user_id}).mappings().all()
    counts = _paid_interest_counts(conn, [row["watched_date"] for row in rows])
    output = []
    for raw in rows:
        row = dict(raw)
        target = row["watched_date"]
        current = int(counts.get(target, 0))
        stored = int(row.get("current_paid_count") or 0)
        unread = bool(row.get("has_unread"))
        if current != stored:
            unread = True
            conn.execute(text("""
                UPDATE vera_v2_leave_watch
                SET employee_username=:employee_username,
                    current_paid_count=:current_paid_count,
                    has_unread=true,
                    updated_at=NOW()
                WHERE auth_user_id=CAST(:auth_user_id AS uuid)
                  AND watched_date=:watched_date
            """), {
                "auth_user_id": ident.auth_user_id,
                "employee_username": ident.employee_username,
                "watched_date": target,
                "current_paid_count": current,
            })
        output.append({
            "date": target.isoformat(),
            "last_seen_paid_count": int(row.get("last_seen_paid_count") or 0),
            "current_paid_count": current,
            "has_unread": unread,
        })
    return output


def _vault_secret(conn, name: str) -> str:
    value = conn.execute(text("""
        SELECT decrypted_secret
        FROM vault.decrypted_secrets
        WHERE name=:name
        LIMIT 1
    """), {"name": name}).scalar_one_or_none()
    return str(value or "").strip()


def _push_subscription_values(body: PushSubscriptionCreate) -> tuple[str, str, str]:
    subscription = body.subscription if isinstance(body.subscription, dict) else {}
    endpoint = str(subscription.get("endpoint") or "").strip()
    keys = subscription.get("keys") if isinstance(subscription.get("keys"), dict) else {}
    p256dh = str(keys.get("p256dh") or "").strip()
    auth_secret = str(keys.get("auth") or "").strip()
    if not endpoint.startswith("https://") or len(endpoint) > 4000:
        raise HTTPException(400, "Đăng ký Web Push không có endpoint HTTPS hợp lệ.")
    if not p256dh or len(p256dh) > 512 or not auth_secret or len(auth_secret) > 256:
        raise HTTPException(400, "Đăng ký Web Push thiếu khóa thiết bị hợp lệ.")
    return endpoint, p256dh, auth_secret


def _send_web_push(delivery: dict[str, Any], private_key: str, subject: str) -> tuple[bool, int | None, str]:
    from pywebpush import WebPushException, webpush

    payload = delivery.get("payload") if isinstance(delivery.get("payload"), dict) else {
        "title": "VERA SPA · Lịch nghỉ thay đổi",
        "body": (
            f"Ngày {delivery['watched_date'].strftime('%d/%m/%Y')}: số lịch nghỉ CÓ phép "
            f"thay đổi từ {delivery['previous_count']} thành {delivery['current_count']}."
        ),
        "url": "https://veraspabienhoa.github.io/Vera-Spa/",
        "tag": f"vera-leave-{delivery['watched_date'].isoformat()}",
        "watched_date": delivery["watched_date"].isoformat(),
    }
    try:
        response = webpush(
            subscription_info={
                "endpoint": delivery["endpoint"],
                "keys": {"p256dh": delivery["p256dh"], "auth": delivery["auth_secret"]},
            },
            data=json.dumps(payload, ensure_ascii=False),
            vapid_private_key=private_key,
            vapid_claims={"sub": subject},
            timeout=15,
        )
        return True, getattr(response, "status_code", None), ""
    except WebPushException as exc:
        status = getattr(getattr(exc, "response", None), "status_code", None)
        return False, status, str(exc)[:1000]
    except Exception as exc:
        return False, None, f"{type(exc).__name__}: {exc}"[:1000]


def _dispatch_paid_watch_pushes(target_dates: list[date]) -> dict[str, int]:
    dates = sorted(set(target_dates))
    deliveries: list[dict[str, Any]] = []
    with _engine_instance().begin() as conn:
        private_key = _vault_secret(conn, "vera_v2_vapid_private_key")
        subject = _vault_secret(conn, "vera_v2_vapid_subject") or "https://veraspabienhoa.github.io/Vera-Spa/"
        if not private_key:
            raise HTTPException(503, "Máy chủ chưa cấu hình khóa riêng Web Push.")
        counts = _paid_interest_counts(conn, dates)
        for target in dates:
            watch_rows = conn.execute(text("""
                SELECT auth_user_id::text AS auth_user_id, current_paid_count
                FROM vera_v2_leave_watch
                WHERE watched_date=:watched_date
                FOR UPDATE
            """), {"watched_date": target}).mappings().all()
            current = int(counts.get(target, 0))
            for watch_row in watch_rows:
                previous = int(watch_row.get("current_paid_count") or 0)
                if previous == current:
                    continue
                auth_user_id = str(watch_row["auth_user_id"])
                conn.execute(text("""
                    UPDATE vera_v2_leave_watch
                    SET current_paid_count=:current_paid_count,
                        has_unread=true,
                        updated_at=NOW()
                    WHERE auth_user_id=CAST(:auth_user_id AS uuid)
                      AND watched_date=:watched_date
                """), {
                    "auth_user_id": auth_user_id,
                    "watched_date": target,
                    "current_paid_count": current,
                })
                subscriptions = conn.execute(text("""
                    SELECT subscription_id::text AS subscription_id, endpoint, p256dh, auth_secret
                    FROM vera_v2_push_subscription
                    WHERE auth_user_id=CAST(:auth_user_id AS uuid)
                      AND is_active=true
                    ORDER BY updated_at DESC
                """), {"auth_user_id": auth_user_id}).mappings().all()
                for subscription in subscriptions:
                    deliveries.append({
                        **dict(subscription),
                        "watched_date": target,
                        "previous_count": previous,
                        "current_count": current,
                    })

    successes = failures = deactivated = 0
    results = []
    for delivery in deliveries:
        ok, status, error_text = _send_web_push(delivery, private_key, subject)
        successes += int(ok)
        failures += int(not ok)
        inactive = not ok and status in {404, 410}
        deactivated += int(inactive)
        results.append({
            "subscription_id": delivery["subscription_id"],
            "ok": ok,
            "inactive": inactive,
            "error_text": error_text,
        })

    if results:
        with _engine_instance().begin() as conn:
            for result in results:
                conn.execute(text("""
                    UPDATE vera_v2_push_subscription
                    SET is_active=CASE WHEN :inactive THEN false ELSE is_active END,
                        last_success_at=CASE WHEN :ok THEN NOW() ELSE last_success_at END,
                        failure_count=CASE WHEN :ok THEN 0 ELSE failure_count + 1 END,
                        last_error=CASE WHEN :ok THEN NULL ELSE :last_error END,
                        updated_at=NOW()
                    WHERE subscription_id=CAST(:subscription_id AS uuid)
                """), result)
    return {
        "dates": len(dates),
        "deliveries": len(deliveries),
        "sent": successes,
        "failed": failures,
        "deactivated": deactivated,
    }


def _dispatch_admin_daily_pushes() -> dict[str, int]:
    with _engine_instance().connect() as conn:
        private_key = _vault_secret(conn, "vera_v2_vapid_private_key")
        subject = _vault_secret(conn, "vera_v2_vapid_subject") or "https://veraspabienhoa.github.io/Vera-Spa/"
        if not private_key:
            raise HTTPException(503, "Máy chủ chưa cấu hình khóa riêng Web Push.")
        changes = int(conn.execute(text("""
            SELECT COUNT(*) FROM vera_sync_event
            WHERE created_at >= NOW() - INTERVAL '24 hours'
              AND dataset_key='leave_records'
              AND event_type IN ('insert','update','delete')
        """)).scalar() or 0)
        groups = conn.execute(text("""
            SELECT dataset_key, COUNT(*) count FROM vera_sync_event
            WHERE created_at >= NOW() - INTERVAL '24 hours'
              AND dataset_key='leave_records'
              AND event_type IN ('insert','update','delete')
            GROUP BY dataset_key ORDER BY count DESC, dataset_key LIMIT 4
        """)).mappings().all()
        subscriptions = conn.execute(text("""
            SELECT s.subscription_id::text subscription_id, s.endpoint, s.p256dh, s.auth_secret
            FROM vera_v2_push_subscription s JOIN vera_v2_user_profile p ON p.auth_user_id=s.auth_user_id
            WHERE s.is_active=true AND p.is_active=true AND lower(COALESCE(p.role,''))='admin'
        """)).mappings().all()
    summary = ", ".join(f"{row['dataset_key']}: {row['count']}" for row in groups) or "Không có thay đổi mới"
    payload = {
        "title": "VERA SPA · Báo cáo thay đổi hằng ngày",
        "body": f"24 giờ qua có {changes} thay đổi. {summary}.",
        "url": "https://veraspabienhoa.github.io/Vera-Spa/",
        "tag": f"vera-admin-daily-{datetime.now(VN_TZ).date().isoformat()}",
    }
    successes = failures = deactivated = 0
    for row in subscriptions:
        delivery = {**dict(row), "payload": payload}
        ok, status, error_text = _send_web_push(delivery, private_key, subject)
        successes += int(ok); failures += int(not ok)
        inactive = not ok and status in {404, 410}; deactivated += int(inactive)
        with _engine_instance().begin() as conn:
            conn.execute(text("""
                UPDATE vera_v2_push_subscription SET
                  is_active=CASE WHEN :inactive THEN false ELSE is_active END,
                  last_success_at=CASE WHEN :ok THEN NOW() ELSE last_success_at END,
                  failure_count=CASE WHEN :ok THEN 0 ELSE failure_count+1 END,
                  last_error=CASE WHEN :ok THEN NULL ELSE :error END, updated_at=NOW()
                WHERE subscription_id=CAST(:subscription_id AS uuid)
            """), {"inactive": inactive, "ok": ok, "error": error_text, "subscription_id": row["subscription_id"]})
    with _engine_instance().begin() as conn:
        conn.execute(text("INSERT INTO vera_sync_event(dataset_key,event_type,detail,created_at) VALUES ('system_audit','admin_daily_push',:detail,NOW())"), {"detail": f"sent={successes}; failed={failures}; deactivated={deactivated}"})
    return {"changes": changes, "deliveries": len(subscriptions), "sent": successes, "failed": failures, "deactivated": deactivated}


def _is_employee_co_phep(reason: str) -> bool:
    key = _norm(reason)
    if not key or "khong phep" in key or "nghi phat sinh" in key:
        return False
    excluded = (
        "di tre", "khong don ve sinh", "loi vi pham", "qua tour", "xuong phong",
        "ra som", "vao muon", "di tua", "ngung nhan", "ho tro ca",
    )
    return not any(token in key for token in excluded)


def _cancel_notice(conn, reason: str, target: date, role: str) -> None:
    if role == "admin":
        return
    today = datetime.now(VN_TZ).date()
    if target < today:
        raise HTTPException(403, "Không được hủy/thay đổi lịch trong quá khứ.")
    item = _reason_item(conn, reason)
    if role in _role_tokens(item.get("cancel_exceptions", "")):
        return
    typ = _norm(item.get("cancel_type", ""))
    value = item.get("cancel_value", "")
    if not typ:
        if role in {"letan", "quanly"}:
            return
        raise HTTPException(403, f"'{item['name']}' chưa cấu hình Kiểu hủy; tài khoản hiện tại không được tự hủy/thay đổi.")
    if "khong gioi han" in typ:
        return
    if "truoc n ngay" in typ or typ in {"truoc ngay", "before days"}:
        days = _parse_first_number(value)
        if days is None:
            raise HTTPException(400, f"Giá trị hủy của '{item['name']}' không hợp lệ.")
        earliest = today + timedelta(days=max(0, int(days)))
        if target < earliest:
            raise HTTPException(403, f"'{item['name']}' phải hủy/thay đổi trước ít nhất {int(days)} ngày; chỉ xử lý từ {earliest.strftime('%d/%m/%Y')}.")
        return
    if "khong duoc huy ngay hien tai" in typ:
        if target == today:
            raise HTTPException(403, f"'{item['name']}' không được hủy/thay đổi trong ngày hiện tại.")
        return
    if "khong cho phep" in typ or "khong duoc huy" in typ:
        raise HTTPException(403, f"'{item['name']}' đang được cấu hình không cho phép hủy.")
    raise HTTPException(400, f"Không nhận diện được Kiểu hủy '{item.get('cancel_type', '')}' của '{item['name']}'.")


def _catalog_rule_for_edit(conn, reason: str, target: date, role: str) -> dict:
    item = _reason_item(conn, reason)
    allowed = _role_tokens(item.get("allowed_roles", ""))
    if allowed and role not in allowed:
        raise HTTPException(403, f"Tài khoản {role} không được dùng lý do '{item['name']}' theo cột H của LoaiNghi.")
    if not _day_allowed(item.get("allowed_days", ""), target):
        raise HTTPException(400, f"'{item['name']}' không được áp dụng cho {_weekday_label(target)} {target.strftime('%d/%m/%Y')} theo cột G của LoaiNghi.")
    return item


def _validate_delete_permission(conn, row: dict, ident: Identity) -> None:
    role = ident.role
    if role == "admin":
        return
    target = row["leave_date"]
    reason = row["leave_reason"]
    today = datetime.now(VN_TZ).date()
    if target < today:
        raise HTTPException(403, "Không được xóa lịch nghỉ của ngày trong quá khứ.")
    if role in _EMPLOYEE_LIKE_ROLES:
        if not _has_any_feature(conn, ident, _LEAVE_DELETE_FEATURES):
            raise HTTPException(403, "Tài khoản chưa được cấp quyền xóa lịch nghỉ.")
        if _norm(row["employee_name"]) != _norm(ident.employee_username):
            raise HTTPException(403, "Nhân viên chỉ được xóa lịch nghỉ của chính mình.")
        if not (_is_video(reason) or "khong phep" in _norm(reason) or _is_employee_co_phep(reason) or (role == "leader" and "leader" in _norm(reason))):
            raise HTTPException(403, "Lý do hiện tại không thuộc nhóm Nhân viên/Leader được phép hủy.")
        _cancel_notice(conn, reason, target, role)
        return
    if role in {"letan", "quanly"}:
        special = (
            _feature_allowed(conn, ident, "leave_today_khong_phep_edit_delete")
            and target == today and _group(reason) == "khong_phep"
        )
        if not special and not _has_any_feature(conn, ident, _LEAVE_DELETE_FEATURES):
            raise HTTPException(403, "Tài khoản chưa được cấp quyền xóa lịch nghỉ.")
        _cancel_notice(conn, reason, target, role)
        return
    raise HTTPException(403, "Vai trò hiện tại không được phép xóa lịch nghỉ.")


def _validate_edit_permission(conn, row: dict, new_reason: str, ident: Identity) -> tuple[dict, bool]:
    role = ident.role
    target = row["leave_date"]
    old_reason = row["leave_reason"]
    item = _catalog_rule_for_edit(conn, new_reason, target, role)
    if role == "admin":
        return item, False

    today = datetime.now(VN_TZ).date()
    next_month_end = (date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    if target < today:
        raise HTTPException(403, "Không được sửa lịch nghỉ của ngày trong quá khứ.")
    if target > next_month_end:
        raise HTTPException(403, f"Chỉ được sửa lịch nghỉ đến hết {next_month_end.strftime('%d/%m/%Y')}.")

    future_conversion = bool(
        target >= today + timedelta(days=1)
        and _is_employee_co_phep(old_reason)
        and abs(float(row.get("calculated_days") or 0) - 1.0) < 1e-9
        and ("khong phep" in _norm(new_reason) or (_is_employee_co_phep(new_reason) and abs(float(item.get("days") or 0) - 0.5) < 1e-9))
    )
    if role in _EMPLOYEE_LIKE_ROLES:
        if not _has_any_feature(conn, ident, _LEAVE_EDIT_FEATURES):
            raise HTTPException(403, "Tài khoản chưa được cấp quyền sửa lịch nghỉ.")
        if _norm(row["employee_name"]) != _norm(ident.employee_username):
            raise HTTPException(403, "Nhân viên chỉ được sửa lịch nghỉ của chính mình.")
        if not future_conversion:
            _cancel_notice(conn, old_reason, target, role)
        old_video, new_video = _is_video(old_reason), _is_video(new_reason)
        old_unpaid, new_unpaid = "khong phep" in _norm(old_reason), "khong phep" in _norm(new_reason)
        old_paid, new_paid = _is_employee_co_phep(old_reason), _is_employee_co_phep(new_reason)
        leader_policy = role == "leader" and "leader" in _norm(old_reason)
        valid = (
            (leader_policy and ("leader" in _norm(new_reason) or new_video))
            or (old_unpaid and (new_unpaid or new_video or new_paid))
            or (old_paid and (new_paid or new_video or new_unpaid))
            or old_video
        )
        if not valid:
            raise HTTPException(403, "Lý do hiện tại không thuộc nhóm Nhân viên/Leader được phép sửa sang lý do đã chọn.")
        if not future_conversion:
            _registration_rule(item, role, target)
        return item, future_conversion
    if role in {"letan", "quanly"}:
        special = (
            _feature_allowed(conn, ident, "leave_today_khong_phep_edit_delete")
            and target == today and _group(old_reason) == "khong_phep" and _group(new_reason) == "khong_phep"
        )
        if not special and not _has_any_feature(conn, ident, _LEAVE_EDIT_FEATURES):
            raise HTTPException(403, "Tài khoản chưa được cấp quyền sửa lịch nghỉ.")
        _cancel_notice(conn, old_reason, target, role)
        _registration_rule(item, role, target)
        return item, False
    raise HTTPException(403, "Vai trò hiện tại không được phép sửa lịch nghỉ.")


def _validate_and_prepare(conn, body: LeaveCreate, ident: Identity) -> tuple[dict, list[str]]:
    employee = body.employee_name.strip()
    role = ident.role
    if role not in {"admin", "quanly", "letan"} and _norm(employee) != _norm(ident.employee_username):
        raise HTTPException(403, "Tài khoản hiện tại chỉ được đăng ký lịch nghỉ của chính mình.")

    emp = conn.execute(text("""
        SELECT username, monthly_generated, monthly_leave, annual_leave
        FROM employees WHERE lower(btrim(username))=lower(btrim(:u))
          AND COALESCE(login_locked,false)=false
          AND COALESCE(payload->>'Trạng thái làm việc', payload->>'employment_status', 'Đang làm việc') = 'Đang làm việc'
        LIMIT 1
    """), {"u": employee}).mappings().first()
    if not emp:
        raise HTTPException(400, "Không tìm thấy nhân viên đang hoạt động.")
    employee = emp["username"]

    item = _reason_item(conn, body.leave_reason)
    _registration_rule(item, role, body.leave_date)
    days = float(item["days"])
    if days not in {0.0, 0.5, 1.0}:
        raise HTTPException(400, "Số ngày tính trong 1 ngày chỉ được 0, 0.5 hoặc 1.")
    if item["requires_manual_penalty"]:
        if body.manual_penalty is None:
            raise HTTPException(400, "Lý do này bắt buộc nhập Mức phạt vi phạm.")
        base_penalty = float(body.manual_penalty)
    else:
        base_penalty = float(item["penalty"])

    if "loi vi pham khac" in _norm(item["name"]) and not body.detail.strip():
        raise HTTPException(400, "Chưa có Chi tiết vi phạm cho 'Lỗi vi phạm khác'.")
    if "nghi ly do khac" in _norm(item["name"]) and not body.detail.strip():
        raise HTTPException(400, "Bắt buộc nhập Chi tiết/Ghi chú đối với 'Nghỉ lý do khác'.")

    today = datetime.now(VN_TZ).date()
    if role != "admin":
        last_next_month = (date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        if body.leave_date < today:
            raise HTTPException(400, "Không được đăng ký lịch nghỉ cho ngày trong quá khứ.")
        if not _is_long_sick(item["name"]) and body.leave_date > last_next_month:
            raise HTTPException(400, f"Chỉ được đăng ký lịch nghỉ đến hết {last_next_month.strftime('%d/%m/%Y')}.")

    existing = conn.execute(text("""
        SELECT record_uid, leave_reason, calculated_days, penalty
        FROM leave_records
        WHERE leave_date=:d AND lower(btrim(employee_name))=lower(btrim(:e))
        FOR SHARE
    """), {"d": body.leave_date, "e": employee}).mappings().all()
    if any(_norm(r["leave_reason"]) == _norm(item["name"]) for r in existing):
        raise HTTPException(409, f"{employee} đã có đúng lý do '{item['name']}' ngày {body.leave_date.strftime('%d/%m/%Y')}.")

    if role != "admin":
        if days > 0 and any(float(r["calculated_days"] or 0) > 0 for r in existing):
            raise HTTPException(400, "Trong cùng 1 ngày, mỗi nhân viên chỉ được có 1 dòng có Số ngày tính > 0; không cho phép 0.5 + 0.5.")
        new_group = _group(item["name"])
        if new_group and any(_group(r["leave_reason"]) == new_group for r in existing):
            raise HTTPException(400, "Trong cùng 1 ngày, một nhân viên không được đăng ký 2 lần cùng nhóm Có phép/Không phép/Phát sinh.")

        if body.leave_date.weekday() >= 5 and not (_is_annual(item["name"]) or _is_long_sick(item["name"]) or "khong phep" in _norm(item["name"])):
            weekend_count = conn.execute(text("""
                SELECT count(DISTINCT leave_date) FROM leave_records
                WHERE lower(btrim(employee_name))=lower(btrim(:e))
                  AND extract(year from leave_date)=:y AND extract(month from leave_date)=:m
                  AND extract(isodow from leave_date) IN (6,7)
                  AND lower(leave_reason) NOT LIKE '%không phép%'
                  AND lower(leave_reason) NOT LIKE '%quay video%'
                  AND lower(leave_reason) NOT LIKE '%giấy khám%'
            """), {"e": employee, "y": body.leave_date.year, "m": body.leave_date.month}).scalar() or 0
            if int(weekend_count) >= 2:
                raise HTTPException(400, f"{employee} chỉ được đăng ký tối đa 2 lần cuối tuần trong tháng đối với lý do chịu giới hạn.")

        if not _is_video(item["name"]):
            month_rows = conn.execute(text("""
                SELECT leave_reason, calculated_days FROM leave_records
                WHERE lower(btrim(employee_name))=lower(btrim(:e))
                  AND extract(year from leave_date)=:y AND extract(month from leave_date)=:m
            """), {"e": employee, "y": body.leave_date.year, "m": body.leave_date.month}).mappings().all()
            if _is_annual(item["name"]):
                year_used = conn.execute(text("""
                    SELECT COALESCE(sum(calculated_days),0) FROM leave_records
                    WHERE lower(btrim(employee_name))=lower(btrim(:e))
                      AND extract(year from leave_date)=:y AND lower(leave_reason) LIKE '%phép năm%'
                """), {"e": employee, "y": body.leave_date.year}).scalar() or 0
                limit = float(emp["annual_leave"] or 0)
                if limit > 0 and float(year_used) + days > limit:
                    raise HTTPException(400, f"Vượt quá số ngày Phép năm; quỹ còn {max(0, limit-float(year_used)):g} ngày.")
            elif "phat sinh" in _norm(item["name"]):
                used = sum(1 for r in month_rows if "phat sinh" in _norm(r["leave_reason"]))
                limit = float(emp["monthly_generated"] or 0)
                if limit > 0 and used >= limit:
                    raise HTTPException(400, f"Vượt giới hạn Phát sinh {limit:g} lần/tháng.")
            elif _group(item["name"]) == "co_phep" and days > 0 and not _is_long_sick(item["name"]):
                used = sum(float(r["calculated_days"] or 0) for r in month_rows if _group(r["leave_reason"]) == "co_phep" and not _is_long_sick(r["leave_reason"]))
                limit = float(emp["monthly_leave"] or 0)
                if limit > 0 and used + days > limit:
                    raise HTTPException(400, f"Vượt số ngày Có phép trong tháng; tối đa {limit:g} ngày/tháng.")

        # Current legacy group quota: 4 Có phép and 1 Phát sinh/day. KHÔNG phép is not capped.
        group = _group(item["name"])
        if group in {"co_phep", "phat_sinh"} and not (_is_video(item["name"]) or _is_long_sick(item["name"])):
            day_reasons = conn.execute(text("SELECT leave_reason FROM leave_records WHERE leave_date=:d"), {"d": body.leave_date}).scalars().all()
            count = sum(1 for r in day_reasons if _group(r) == group)
            limit = 4 if group == "co_phep" else 1
            if count >= limit:
                label = "CÓ phép" if group == "co_phep" else "PHÁT SINH"
                raise HTTPException(400, f"Ngày {body.leave_date.strftime('%d/%m/%Y')} đã đủ {limit} người {label}.")

    progressive = _progressive_key(item["name"])
    warnings = []
    ordinal = None
    extra = 0.0
    if progressive:
        day_reasons = conn.execute(text("SELECT leave_reason FROM leave_records WHERE leave_date=:d"), {"d": body.leave_date}).scalars().all()
        ordinal = sum(1 for r in day_reasons if _progressive_key(r) == progressive) + 1
        extra = max(0, ordinal - 2) * 100000.0
        warnings.append(f"Người Thứ {ordinal} · phạt lũy tiến cộng thêm {extra:,.0f} VNĐ.")

    # Monthly accumulated value mirrors the existing main-data convention.
    accumulated = conn.execute(text("""
        SELECT COALESCE(sum(calculated_days),0) FROM leave_records
        WHERE lower(btrim(employee_name))=lower(btrim(:e))
          AND extract(year from leave_date)=:y AND extract(month from leave_date)=:m
          AND lower(leave_reason) NOT LIKE '%quay video%'
          AND lower(leave_reason) NOT LIKE '%giấy khám%'
    """), {"e": employee, "y": body.leave_date.year, "m": body.leave_date.month}).scalar() or 0
    if not (_is_video(item["name"]) or _is_long_sick(item["name"])):
        accumulated = float(accumulated) + days

    detail = body.detail.strip()
    if ordinal:
        prefix = f"Người Thứ {ordinal} {item['name'].lower()}"
        detail = f"{prefix} | {detail}" if detail else prefix
    now = datetime.now(VN_TZ)
    record_uid = str(uuid.uuid4())
    record = {
        "record_uid": record_uid,
        "leave_date": body.leave_date,
        "employee_name": employee,
        "leave_reason": item["name"],
        "leave_type": item["leave_type"],
        "detail": detail,
        "calculated_days": days,
        "accumulated_leave": float(accumulated),
        "penalty": base_penalty + extra,
        "update_date": now.strftime("%d/%m/%Y"),
        "update_time": now.strftime("%H:%M:%S"),
        "updated_by": ident.employee_username,
        "weekday_label": _weekday_label(body.leave_date),
    }
    return record, warnings


def _sheet_row_for_record(ws, record: dict) -> tuple[int, list[Any]]:
    all_values = ws.get_all_values()
    headers = all_values[0][:13] if all_values else []
    if not headers:
        raise RuntimeError("MainData chưa có header A:M")
    target_row = 2
    for idx, row in enumerate(all_values[1:], start=2):
        if any(str(x).strip() for x in row[:13]):
            target_row = idx + 1
    mapping = {
        "stt": target_row - 1,
        "ngay": record["leave_date"].strftime("%d/%m/%Y"),
        "thu ngay": record["weekday_label"],
        "ten nhan vien": record["employee_name"],
        "ly do nghi": record["leave_reason"],
        "loai nghi": record["leave_type"],
        "chi tiet": record["detail"],
        "so ngay tinh": record["calculated_days"],
        "so ngay tinh phep": record["calculated_days"],
        "so ngay phep cong don": record["accumulated_leave"],
        "phat vi pham": record["penalty"],
        "ngay cap nhat": record["update_date"],
        "gio cap nhat": record["update_time"],
        "nguoi cap nhat": record["updated_by"],
    }
    return target_row, [mapping.get(_norm(h), "") for h in headers]


def _insert_record(conn, record: dict, source_row: int):
    payload = {
        "Ngày": record["leave_date"].strftime("%d/%m/%Y"),
        "Thứ ngày": record["weekday_label"], "Tên nhân viên": record["employee_name"],
        "Lý do nghỉ": record["leave_reason"], "Loại nghỉ": record["leave_type"],
        "Chi tiết": record["detail"], "Số ngày tính": record["calculated_days"],
        "Số ngày phép cộng dồn": record["accumulated_leave"], "Phạt vi phạm": record["penalty"],
        "Ngày cập nhật": record["update_date"], "Giờ cập nhật": record["update_time"],
        "Người cập nhật": record["updated_by"], "record_uid": record["record_uid"],
    }
    conn.execute(text("""
        INSERT INTO leave_records(
            source_sheet_id, source_row, leave_date, employee_name, leave_reason,
            leave_type, detail, calculated_days, accumulated_leave, penalty,
            update_date, update_time, updated_by, weekday_label, payload, record_uid,
            created_at, updated_at
        ) VALUES (
            :sid, :srow, :leave_date, :employee_name, :leave_reason,
            :leave_type, :detail, :calculated_days, :accumulated_leave, :penalty,
            :update_date, :update_time, :updated_by, :weekday_label, CAST(:payload AS jsonb), :record_uid,
            NOW(), NOW()
        )
    """), {**record, "sid": LEAVE_SHEET_ID, "srow": source_row, "payload": json_text(payload)})
    # Operational trace; this is additive and does not replace the legacy
    # activity log.  A database error must roll back only this optional write,
    # not poison the surrounding leave-registration transaction.
    try:
        with conn.begin_nested():
            conn.execute(text("""
                INSERT INTO vera_sync_event(dataset_key,event_type,detail,created_at)
                VALUES ('leave_primary','web_v2_leave_create',:detail,NOW())
            """), {"detail": f"record_uid={record['record_uid']}; employee={record['employee_name']}; actor={record['updated_by']}"})
    except Exception:
        # Do not make leave registration depend on an optional telemetry table shape.
        pass


def _record_payload(record: dict, source_row: int) -> dict:
    return {
        "Ngày": record["leave_date"].strftime("%d/%m/%Y"),
        "Thứ ngày": record["weekday_label"],
        "Tên nhân viên": record["employee_name"],
        "Lý do nghỉ": record["leave_reason"],
        "Loại nghỉ": record["leave_type"],
        "Chi tiết": record["detail"],
        "Số ngày tính": record["calculated_days"],
        "Số ngày phép cộng dồn": record["accumulated_leave"],
        "Phạt vi phạm": record["penalty"],
        "Ngày cập nhật": record["update_date"],
        "Giờ cập nhật": record["update_time"],
        "Người cập nhật": record["updated_by"],
        "record_uid": record["record_uid"],
        "__record_uid": record["record_uid"],
        "__source_sheet_id": LEAVE_SHEET_ID,
        "__source_row": int(source_row),
    }


def _sheet_values_for_record(headers: list[str], record: dict, source_row: int) -> list[Any]:
    mapping = {
        "stt": int(source_row) - 1,
        "ngay": record["leave_date"].strftime("%d/%m/%Y"),
        "thu ngay": record["weekday_label"],
        "ten nhan vien": record["employee_name"],
        "ly do nghi": record["leave_reason"],
        "loai nghi": record["leave_type"],
        "chi tiet": record["detail"],
        "so ngay tinh": json_safe(record["calculated_days"]),
        "so ngay tinh phep": json_safe(record["calculated_days"]),
        "so ngay phep cong don": json_safe(record["accumulated_leave"]),
        "phat vi pham": json_safe(record["penalty"]),
        "ngay cap nhat": record["update_date"],
        "gio cap nhat": record["update_time"],
        "nguoi cap nhat": record["updated_by"],
    }
    return [mapping.get(_norm(header), "") for header in headers[:13]]


def _update_record(conn, record: dict, source_row: int) -> None:
    payload = _record_payload(record, source_row)
    result = conn.execute(text("""
        UPDATE leave_records SET
            leave_date=:leave_date, employee_name=:employee_name, leave_reason=:leave_reason,
            leave_type=:leave_type, detail=:detail, calculated_days=:calculated_days,
            accumulated_leave=:accumulated_leave, penalty=:penalty,
            update_date=:update_date, update_time=:update_time, updated_by=:updated_by,
            weekday_label=:weekday_label, payload=CAST(:payload AS jsonb), updated_at=NOW()
        WHERE record_uid=:record_uid
    """), {**record, "payload": json_text(payload)})
    if int(result.rowcount or 0) != 1:
        raise RuntimeError(f"record_uid update affected {result.rowcount} rows")


def _strip_progressive_prefix(detail: str) -> str:
    value = str(detail or "").strip()
    if not _norm(value).startswith("nguoi thu"):
        return value
    return value.split("|", 1)[1].strip() if "|" in value else ""


def _existing_progressive_ordinal(detail: str) -> int | None:
    match = re.match(r"^\s*Người\s+Thứ\s+(\d+)", str(detail or ""), flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def _row_record(row: dict) -> dict:
    return {
        key: row.get(key)
        for key in (
            "record_uid", "leave_date", "employee_name", "leave_reason", "leave_type",
            "detail", "calculated_days", "accumulated_leave", "penalty", "update_date",
            "update_time", "updated_by", "weekday_label",
        )
    }


def _rebalance_progressive_rows(conn, target: date, keys: set[str]) -> list[tuple[int, dict]]:
    keys = {key for key in keys if key}
    if not keys:
        return []
    rows = conn.execute(text("""
        SELECT record_uid, source_row, leave_date, employee_name, leave_reason, leave_type,
               detail, calculated_days, accumulated_leave, penalty, update_date,
               update_time, updated_by, weekday_label
        FROM leave_records
        WHERE leave_date=:d AND source_sheet_id=:sid
        ORDER BY source_row, id
        FOR UPDATE
    """), {"d": target, "sid": LEAVE_SHEET_ID}).mappings().all()
    changed = []
    counters: dict[str, int] = {key: 0 for key in keys}
    for raw in rows:
        key = _progressive_key(raw["leave_reason"])
        if key not in keys:
            continue
        counters[key] += 1
        ordinal = counters[key]
        item = _reason_item(conn, raw["leave_reason"])
        base = float(item.get("penalty") or 0)
        detail = _strip_progressive_prefix(raw.get("detail", ""))
        prefix = f"Người Thứ {ordinal} {item['name'].lower()}"
        new_detail = f"{prefix} | {detail}" if detail else prefix
        new_penalty = base + max(0, ordinal - 2) * 100000.0
        if new_detail == str(raw.get("detail") or "") and abs(new_penalty - float(raw.get("penalty") or 0)) < 1e-9:
            continue
        record = _row_record(dict(raw))
        record["detail"] = new_detail
        record["penalty"] = new_penalty
        _update_record(conn, record, int(raw["source_row"]))
        changed.append((int(raw["source_row"]), record))
    return changed


def _reindex_after_delete(conn, deleted_rows: list[int]) -> None:
    deleted = sorted({int(value) for value in deleted_rows})
    if not deleted:
        return
    rows = conn.execute(text("""
        SELECT record_uid, source_row
        FROM leave_records
        WHERE source_sheet_id=:sid AND source_row>:minimum
        ORDER BY source_row, id
        FOR UPDATE
    """), {"sid": LEAVE_SHEET_ID, "minimum": min(deleted)}).mappings().all()
    moves = []
    for row in rows:
        old = int(row["source_row"])
        shift = sum(1 for removed in deleted if removed < old)
        if shift:
            moves.append((str(row["record_uid"]), old - shift))
    for index, (uid, _final) in enumerate(moves, start=1):
        conn.execute(text("UPDATE leave_records SET source_row=:row WHERE record_uid=:uid"), {"row": -(1_000_000_000 + index), "uid": uid})
    for uid, final in moves:
        conn.execute(text("""
            UPDATE leave_records
            SET source_row=:row,
                payload=jsonb_set(
                    jsonb_set(COALESCE(payload,'{}'::jsonb), '{__source_row}', to_jsonb(CAST(:row AS INTEGER)), true),
                    '{__record_uid}', to_jsonb(CAST(:uid AS TEXT)), true
                ), updated_at=NOW()
            WHERE record_uid=:uid
        """), {"row": final, "uid": uid})


def _restore_sheet_updates(ws, backups: dict[int, list[Any]]) -> None:
    for row_number, values in sorted(backups.items()):
        try:
            ws.update(range_name=f"A{row_number}:M{row_number}", values=[values], value_input_option="USER_ENTERED")
        except Exception:
            pass


@app.get("/v2/health")
def health():
    with _engine_instance().connect() as conn:
        conn.execute(text("SELECT 1"))
    return {"ok": True, "service": "vera-web-v2-api", "version": "3.8-payroll-export-settings"}


@app.get("/v2/me")
def me(ident: Identity = Depends(current_identity)):
    with _engine_instance().connect() as conn:
        permission_payload = _permission_payload(conn)
        permissions = {
            feature: _feature_allowed(conn, ident, feature, permission_payload)
            for feature in WEB_V2_FEATURES
        }
        registration_locked = _registration_role_locked(conn, ident.role)
    return {
        **ident.model_dump(),
        "permissions": permissions,
        "registration_locked": registration_locked,
    }


@app.get("/v2/push/config")
def push_config(ident: Identity = Depends(current_identity)):
    with _engine_instance().connect() as conn:
        public_key = _vault_secret(conn, "vera_v2_vapid_public_key")
    return {"enabled": bool(public_key), "public_key": public_key}


@app.post("/v2/push/subscriptions")
def register_push_subscription(
    body: PushSubscriptionCreate,
    ident: Identity = Depends(current_identity),
    user_agent: str | None = Header(default=None),
):
    _require_password_changed(ident)
    endpoint, p256dh, auth_secret = _push_subscription_values(body)
    with _engine_instance().begin() as conn:
        exists = conn.execute(text("""
            SELECT 1 FROM vera_v2_push_subscription WHERE endpoint=:endpoint
        """), {"endpoint": endpoint}).scalar_one_or_none()
        if not exists:
            active_count = conn.execute(text("""
                SELECT COUNT(*) FROM vera_v2_push_subscription
                WHERE auth_user_id=CAST(:auth_user_id AS uuid) AND is_active=true
            """), {"auth_user_id": ident.auth_user_id}).scalar() or 0
            if int(active_count) >= 10:
                raise HTTPException(400, "Mỗi tài khoản chỉ được đăng ký tối đa 10 thiết bị nhận thông báo.")
        conn.execute(text("""
            INSERT INTO vera_v2_push_subscription(
                auth_user_id, employee_username, endpoint, p256dh, auth_secret,
                user_agent, is_active, failure_count, created_at, updated_at
            ) VALUES (
                CAST(:auth_user_id AS uuid), :employee_username, :endpoint, :p256dh, :auth_secret,
                :user_agent, true, 0, NOW(), NOW()
            )
            ON CONFLICT (endpoint) DO UPDATE
            SET auth_user_id=EXCLUDED.auth_user_id,
                employee_username=EXCLUDED.employee_username,
                p256dh=EXCLUDED.p256dh,
                auth_secret=EXCLUDED.auth_secret,
                user_agent=EXCLUDED.user_agent,
                is_active=true,
                failure_count=0,
                last_error=NULL,
                updated_at=NOW()
        """), {
            "auth_user_id": ident.auth_user_id,
            "employee_username": ident.employee_username,
            "endpoint": endpoint,
            "p256dh": p256dh,
            "auth_secret": auth_secret,
            "user_agent": str(user_agent or "")[:1000],
        })
    return {"ok": True, "subscribed": True}


@app.delete("/v2/push/subscriptions")
def unregister_push_subscription(body: PushSubscriptionDelete, ident: Identity = Depends(current_identity)):
    _require_password_changed(ident)
    with _engine_instance().begin() as conn:
        result = conn.execute(text("""
            DELETE FROM vera_v2_push_subscription
            WHERE auth_user_id=CAST(:auth_user_id AS uuid)
              AND endpoint=:endpoint
        """), {"auth_user_id": ident.auth_user_id, "endpoint": body.endpoint.strip()})
    return {"ok": True, "subscribed": False, "removed": int(result.rowcount or 0)}


@app.post("/v2/push/dispatch")
def dispatch_push(
    body: PushDispatch,
    x_vera_push_webhook: str | None = Header(default=None),
):
    with _engine_instance().connect() as conn:
        expected = _vault_secret(conn, "vera_v2_push_webhook_secret")
    supplied = str(x_vera_push_webhook or "")
    if not expected or not supplied or not hmac.compare_digest(expected, supplied):
        raise HTTPException(403, "Webhook Web Push không hợp lệ.")
    return {"ok": True, **_dispatch_paid_watch_pushes(body.dates)}


@app.post("/v2/push/admin-daily-dispatch")
def dispatch_admin_daily_push(x_vera_push_webhook: str | None = Header(default=None)):
    with _engine_instance().connect() as conn:
        expected = _vault_secret(conn, "vera_v2_push_webhook_secret")
    supplied = str(x_vera_push_webhook or "")
    if not expected or not supplied or not hmac.compare_digest(expected, supplied):
        raise HTTPException(403, "Webhook Web Push không hợp lệ.")
    return {"ok": True, **_dispatch_admin_daily_pushes()}


@app.get("/v2/admin/changes")
def admin_changes(days: int = Query(default=7, ge=1, le=31), ident: Identity = Depends(current_identity)):
    if ident.role != "admin":
        raise HTTPException(403, "Chỉ Admin được xem nhật ký thay đổi hệ thống.")
    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "audit_admin_view")
        rows = conn.execute(text("""
            SELECT id, dataset_key, event_type, detail, created_at
            FROM vera_sync_event WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
              AND dataset_key='leave_records'
              AND event_type IN ('insert','update','delete')
            ORDER BY created_at DESC, id DESC LIMIT 1000
        """), {"days": days}).mappings().all()
    return {"changes": [{**dict(row), "created_at": row["created_at"].isoformat()} for row in rows], "count": len(rows), "days": days}


@app.get("/v2/employees")
def employees(ident: Identity = Depends(current_identity)):
    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "leave")
        if ident.role in {"admin", "quanly", "letan"}:
            rows = conn.execute(text("""
                SELECT username, COALESCE(full_name,'') full_name, COALESCE(role,'') role
                FROM employees
                WHERE COALESCE(login_locked,false)=false
                  AND COALESCE(payload->>'__deleted','false') <> 'true'
                  AND COALESCE(payload->>'Trạng thái làm việc', payload->>'employment_status', 'Đang làm việc') = 'Đang làm việc'
                  AND lower(btrim(COALESCE(role,''))) IN ('leader','nhanvien')
                ORDER BY username
            """)).mappings().all()
        else:
            rows = conn.execute(text("""
                SELECT username, COALESCE(full_name,'') full_name, COALESCE(role,'') role
                FROM employees
                WHERE COALESCE(login_locked,false)=false
                  AND COALESCE(payload->>'__deleted','false') <> 'true'
                  AND lower(btrim(username))=lower(btrim(:username))
                  AND COALESCE(payload->>'Trạng thái làm việc', payload->>'employment_status', 'Đang làm việc') = 'Đang làm việc'
                LIMIT 1
            """), {"username": ident.employee_username}).mappings().all()
    return {"employees": [dict(r) for r in rows]}


@app.get("/v2/leave/reasons")
def reasons(date_value: date = Query(alias="date"), ident: Identity = Depends(current_identity)):
    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "leave")
        can_view_penalty = _feature_allowed(conn, ident, "employee_penalty_view")
        output = []
        for row in _policy_rows(conn):
            name = str(_field(row, "Lý do nghỉ", default="") or "").strip()
            if not name:
                continue
            item = _reason_item(conn, name)
            allowed = _role_tokens(item["allowed_roles"])
            if allowed and ident.role not in allowed:
                continue
            if not _day_allowed(item["allowed_days"], date_value):
                continue
            output.append({
                "name": item["name"], "days": item["days"],
                "penalty": item["penalty"] if can_view_penalty else None,
                "requires_manual_penalty": item["requires_manual_penalty"],
            })
    return {"reasons": output}


@app.get("/v2/leave/watch-dates")
def leave_watch_dates(ident: Identity = Depends(current_identity)):
    with _engine_instance().begin() as conn:
        _require_feature(conn, ident, "leave")
        rows = _refresh_leave_watches(conn, ident)
    return {
        "watch_dates": rows,
        "unread_count": sum(1 for row in rows if row["has_unread"]),
    }


@app.post("/v2/leave/watch-dates")
def set_leave_watch(body: LeaveWatchUpdate, ident: Identity = Depends(current_identity)):
    with _engine_instance().begin() as conn:
        _require_feature(conn, ident, "leave")
        if not body.watching:
            conn.execute(text("""
                DELETE FROM vera_v2_leave_watch
                WHERE auth_user_id=CAST(:auth_user_id AS uuid)
                  AND watched_date=:watched_date
            """), {
                "auth_user_id": ident.auth_user_id,
                "watched_date": body.watched_date,
            })
            return {"ok": True, "watching": False, "date": body.watched_date.isoformat()}

        existing = conn.execute(text("""
            SELECT 1 FROM vera_v2_leave_watch
            WHERE auth_user_id=CAST(:auth_user_id AS uuid)
              AND watched_date=:watched_date
        """), {
            "auth_user_id": ident.auth_user_id,
            "watched_date": body.watched_date,
        }).scalar_one_or_none()
        if not existing:
            total = conn.execute(text("""
                SELECT COUNT(*) FROM vera_v2_leave_watch
                WHERE auth_user_id=CAST(:auth_user_id AS uuid)
            """), {"auth_user_id": ident.auth_user_id}).scalar() or 0
            if int(total) >= 100:
                raise HTTPException(400, "Mỗi tài khoản chỉ được quan tâm tối đa 100 ngày.")

        current = _paid_interest_counts(conn, [body.watched_date])[body.watched_date]
        conn.execute(text("""
            INSERT INTO vera_v2_leave_watch(
                auth_user_id, employee_username, watched_date,
                last_seen_paid_count, current_paid_count, has_unread,
                created_at, updated_at
            ) VALUES (
                CAST(:auth_user_id AS uuid), :employee_username, :watched_date,
                :paid_count, :paid_count, false, NOW(), NOW()
            )
            ON CONFLICT (auth_user_id, watched_date) DO UPDATE
            SET employee_username=EXCLUDED.employee_username,
                updated_at=NOW()
        """), {
            "auth_user_id": ident.auth_user_id,
            "employee_username": ident.employee_username,
            "watched_date": body.watched_date,
            "paid_count": current,
        })
    return {
        "ok": True,
        "watching": True,
        "date": body.watched_date.isoformat(),
        "current_paid_count": current,
    }


@app.post("/v2/leave/watch-dates/acknowledge")
def acknowledge_leave_watches(body: LeaveWatchAcknowledge, ident: Identity = Depends(current_identity)):
    unique_dates = list(dict.fromkeys(body.watched_dates))
    with _engine_instance().begin() as conn:
        _require_feature(conn, ident, "leave")
        updated = 0
        for target in unique_dates:
            result = conn.execute(text("""
                UPDATE vera_v2_leave_watch
                SET last_seen_paid_count=current_paid_count,
                    has_unread=false,
                    updated_at=NOW()
                WHERE auth_user_id=CAST(:auth_user_id AS uuid)
                  AND watched_date=:watched_date
            """), {
                "auth_user_id": ident.auth_user_id,
                "watched_date": target,
            })
            updated += int(result.rowcount or 0)
    return {"ok": True, "acknowledged": updated}


@app.get("/v2/leave/records")
def leave_records(
    date_value: date | None = Query(default=None, alias="date"),
    start_date: date | None = Query(default=None, alias="start"),
    end_date: date | None = Query(default=None, alias="end"),
    ident: Identity = Depends(current_identity),
):
    if date_value is not None:
        start_date = end_date = date_value
    if start_date is None or end_date is None:
        raise HTTPException(400, "Phải chọn ngày hoặc khoảng thời gian cần xem.")
    if end_date < start_date:
        raise HTTPException(400, "Khoảng thời gian không hợp lệ.")
    if (end_date - start_date).days > 365:
        raise HTTPException(400, "Khoảng danh sách tối đa là 366 ngày.")

    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "leave")
        can_view_penalty = _feature_allowed(conn, ident, "employee_penalty_view")
        rows = conn.execute(text("""
            SELECT record_uid, leave_date, weekday_label, employee_name, leave_reason,
                   detail, penalty, updated_by, updated_at
            FROM leave_records
            WHERE leave_date BETWEEN :start_date AND :end_date
            ORDER BY leave_date, employee_name, record_uid
        """), {"start_date": start_date, "end_date": end_date}).mappings().all()
    records = []
    for row in rows:
        item = dict(row)
        if not can_view_penalty:
            item.pop("penalty", None)
        records.append(item)
    return {"records": records}


LEAVE_IMPORT_MAX_BYTES = 5 * 1024 * 1024
LEAVE_IMPORT_MAX_ROWS = 10_000


def _leave_import_value(row: dict[str, Any], *names: str, default: Any = "") -> Any:
    for name in names:
        key = _norm(name)
        if key in row:
            return row[key]
    return default


def _leave_import_date(value: Any, row_number: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date):
                return parsed
        except Exception:
            pass
    raw = str(value or "").strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError as exc:
        raise HTTPException(400, f"Dòng {row_number}: cột Ngày không hợp lệ ({raw or 'trống'}).") from exc


def _leave_import_number(
    value: Any,
    *,
    row_number: int,
    field_name: str,
    default: float | None = None,
    money: bool = False,
    maximum: float = 1_000_000_000_000,
) -> float | None:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, bool):
        raise HTTPException(400, f"Dòng {row_number}: {field_name} phải là số.")
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        raw = str(value).strip()
        token = re.sub(r"[^0-9,.-]", "", raw)
        if money:
            if re.fullmatch(r"-?\d{1,3}([.,]\d{3})+", token):
                token = token.replace(".", "").replace(",", "")
            elif "." in token and "," in token:
                token = token.replace(".", "").replace(",", "")
            elif token.count(".") + token.count(",") == 1:
                separator = "." if "." in token else ","
                left, right = token.split(separator)
                token = left + right if len(right) == 3 else f"{left}.{right}"
            else:
                token = token.replace(".", "").replace(",", "")
        else:
            token = token.replace(",", ".")
        try:
            number = float(token)
        except ValueError as exc:
            raise HTTPException(400, f"Dòng {row_number}: {field_name} không phải số hợp lệ ({raw}).") from exc
    if not math.isfinite(number) or number < 0 or number > maximum:
        raise HTTPException(400, f"Dòng {row_number}: {field_name} nằm ngoài phạm vi cho phép.")
    return number


def _leave_import_text(value: Any, *, limit: int = 3000) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()[:limit]


def _leave_import_fingerprint(record: dict[str, Any]) -> str:
    canonical = [
        record["leave_date"].isoformat(), _norm(record["employee_name"]),
        _norm(record["leave_reason"]), _norm(record.get("leave_type")),
        str(record.get("detail") or "").strip(), record.get("calculated_days"),
        record.get("accumulated_leave"), record.get("penalty"),
        str(record.get("update_date") or ""), str(record.get("update_time") or ""),
        _norm(record.get("updated_by")),
    ]
    payload = json.dumps(canonical, ensure_ascii=False, separators=(",", ":"), default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _leave_import_identity(record: dict[str, Any]) -> tuple[str, str, str]:
    return (
        record["leave_date"].isoformat(),
        _norm(record["employee_name"]),
        _norm(record["leave_reason"]),
    )


def _parse_leave_import(content: bytes) -> tuple[list[dict[str, Any]], int]:
    workbook = None
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
        worksheet = workbook["Lịch nghỉ"] if "Lịch nghỉ" in workbook.sheetnames else workbook.active
        rows = worksheet.iter_rows()
        try:
            header_cells = next(rows)
        except StopIteration as exc:
            raise HTTPException(400, "File Excel không có dữ liệu.") from exc
        if any(cell.data_type == "f" for cell in header_cells):
            raise HTTPException(400, "Dòng tiêu đề Excel không được chứa công thức.")
        headers = [_norm(cell.value) for cell in header_cells]
        nonempty_headers = [header for header in headers if header]
        if len(nonempty_headers) != len(set(nonempty_headers)):
            raise HTTPException(400, "File Excel có tên cột bị trùng.")
        header_keys = set(nonempty_headers)
        required_groups = {
            "Ngày": {_norm("Ngày"), _norm("Ngày nghỉ"), "leave_date"},
            "Tên nhân viên": {_norm("Tên nhân viên"), _norm("Nhân viên"), _norm("Tên Hệ thống"), "employee_name"},
            "Lý do nghỉ": {_norm("Lý do nghỉ"), _norm("Lý do"), "leave_reason"},
        }
        missing = [label for label, aliases in required_groups.items() if not aliases.intersection(header_keys)]
        if missing:
            raise HTTPException(400, "File import thiếu cột bắt buộc: " + ", ".join(missing) + ".")

        output: list[dict[str, Any]] = []
        fingerprints_by_uid: dict[str, str] = {}
        fingerprints_by_identity: dict[tuple[str, str, str], str] = {}
        duplicate_rows = 0
        for row_number, cells in enumerate(rows, start=2):
            if not any(cell.value not in (None, "") for cell in cells):
                continue
            if len(output) + duplicate_rows >= LEAVE_IMPORT_MAX_ROWS:
                raise HTTPException(400, f"File Excel chỉ được tối đa {LEAVE_IMPORT_MAX_ROWS:,} dòng dữ liệu.")
            formula_columns = [
                str(header_cells[index].value or index + 1)
                for index, cell in enumerate(cells)
                if cell.data_type == "f"
            ]
            if formula_columns:
                raise HTTPException(400, f"Dòng {row_number}: không nhận ô công thức Excel ({', '.join(formula_columns)}).")
            row = {
                headers[index]: cell.value
                for index, cell in enumerate(cells)
                if index < len(headers) and headers[index]
            }
            leave_date_value = _leave_import_date(
                _leave_import_value(row, "Ngày", "Ngày nghỉ", "leave_date"),
                row_number,
            )
            employee_name = _leave_import_text(
                _leave_import_value(row, "Tên nhân viên", "Nhân viên", "Tên Hệ thống", "employee_name"),
                limit=200,
            )
            leave_reason = _leave_import_text(
                _leave_import_value(row, "Lý do nghỉ", "Lý do", "leave_reason"),
                limit=300,
            )
            if not employee_name:
                raise HTTPException(400, f"Dòng {row_number}: Tên nhân viên đang trống.")
            if not leave_reason:
                raise HTTPException(400, f"Dòng {row_number}: Lý do nghỉ đang trống.")
            calculated_raw = _leave_import_value(row, "Số ngày tính", "Số ngày tính phép", "calculated_days", default=None)
            accumulated_raw = _leave_import_value(row, "Số ngày phép cộng dồn", "Phép cộng dồn", "accumulated_leave", default=None)
            penalty_raw = _leave_import_value(row, "Phạt", "Phạt vi phạm", "Tiền phạt", "penalty", default=None)
            record = {
                "leave_date": leave_date_value,
                "employee_name": employee_name,
                "leave_reason": leave_reason,
                "leave_type": _leave_import_text(_leave_import_value(row, "Loại nghỉ", "Loại", "leave_type"), limit=100),
                "detail": _leave_import_text(_leave_import_value(row, "Chi tiết", "Ghi chú", "detail")),
                "calculated_days": _leave_import_number(calculated_raw, row_number=row_number, field_name="Số ngày tính", default=None, maximum=366),
                "accumulated_leave": _leave_import_number(accumulated_raw, row_number=row_number, field_name="Số ngày phép cộng dồn", default=None, maximum=10_000),
                "penalty": _leave_import_number(penalty_raw, row_number=row_number, field_name="Phạt", default=None, money=True),
                "update_date": _leave_import_text(_leave_import_value(row, "Ngày ghi", "Ngày cập nhật", "update_date"), limit=30),
                "update_time": _leave_import_text(_leave_import_value(row, "Giờ ghi", "Giờ cập nhật", "update_time"), limit=30),
                "updated_by": _leave_import_text(_leave_import_value(row, "Người ghi", "Người cập nhật", "updated_by"), limit=200),
                "weekday_label": _weekday_label(leave_date_value),
                "_row": row_number,
            }
            fingerprint = _leave_import_fingerprint(record)
            identity = _leave_import_identity(record)
            supplied_uid = _leave_import_text(
                _leave_import_value(row, "Mã bản ghi", "record_uid", "__record_uid"),
                limit=201,
            )
            if supplied_uid and (len(supplied_uid) > 200 or not re.fullmatch(r"[A-Za-z0-9._:-]+", supplied_uid)):
                raise HTTPException(400, f"Dòng {row_number}: Mã bản ghi không hợp lệ.")
            identity_hash = hashlib.sha256("|".join(identity).encode("utf-8")).hexdigest()
            record_uid = supplied_uid or f"lr-import-{identity_hash[:32]}"
            if record_uid in fingerprints_by_uid:
                if fingerprints_by_uid[record_uid] != fingerprint:
                    raise HTTPException(400, f"Dòng {row_number}: Mã bản ghi bị trùng nhưng nội dung khác nhau.")
                duplicate_rows += 1
                continue
            if identity in fingerprints_by_identity:
                if fingerprints_by_identity[identity] != fingerprint:
                    raise HTTPException(
                        400,
                        f"Dòng {row_number}: trùng Ngày + Tên nhân viên + Lý do nghỉ nhưng nội dung khác nhau.",
                    )
                duplicate_rows += 1
                continue
            fingerprints_by_uid[record_uid] = fingerprint
            fingerprints_by_identity[identity] = fingerprint
            record["record_uid"] = record_uid
            output.append(record)
        if not output:
            raise HTTPException(400, "File Excel không có lịch nghỉ hợp lệ.")
        return output, duplicate_rows
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được file Excel lịch nghỉ: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


@app.post("/v2/leave/import.xlsx")
async def import_leave_excel(request: Request, ident: Identity = Depends(current_identity)):
    length = int(request.headers.get("content-length") or 0)
    if length > LEAVE_IMPORT_MAX_BYTES:
        raise HTTPException(413, "File Excel vượt quá 5 MB.")
    content = await request.body()
    if not content:
        raise HTTPException(400, "Chưa chọn file Excel.")
    if len(content) > LEAVE_IMPORT_MAX_BYTES:
        raise HTTPException(413, "File Excel vượt quá 5 MB.")
    if not content.startswith(b"PK"):
        raise HTTPException(400, "File không đúng định dạng Excel .xlsx.")
    imported, duplicate_rows = _parse_leave_import(content)

    engine = _engine_instance()
    conn = engine.connect()
    tx = conn.begin()
    worksheet = None
    sheet_range = ""
    try:
        conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:leave_primary'))"))
        if ident.role != "admin":
            raise HTTPException(403, "Chỉ tài khoản admin được Import dữ liệu lịch nghỉ cũ.")
        _require_feature(conn, ident, "leave")
        now = datetime.now(VN_TZ)
        for record in imported:
            policy = {}
            if not record["leave_type"] or record["calculated_days"] is None or record["penalty"] is None:
                try:
                    policy = _reason_item(conn, record["leave_reason"])
                except HTTPException:
                    policy = {}
            record["leave_type"] = record["leave_type"] or str(policy.get("leave_type") or "")
            record["calculated_days"] = float(
                policy.get("days") if record["calculated_days"] is None and policy else record["calculated_days"] or 0
            )
            record["accumulated_leave"] = float(record["accumulated_leave"] or 0)
            record["penalty"] = float(
                policy.get("penalty") if record["penalty"] is None and policy else record["penalty"] or 0
            )
            record["update_date"] = record["update_date"] or now.strftime("%d/%m/%Y")
            record["update_time"] = record["update_time"] or now.strftime("%H:%M:%S")
            record["updated_by"] = record["updated_by"] or ident.employee_username

        existing_uids: set[str] = set()
        uid_query = text("SELECT record_uid FROM leave_records WHERE record_uid IN :uids").bindparams(
            bindparam("uids", expanding=True)
        )
        for offset in range(0, len(imported), 500):
            chunk = [record["record_uid"] for record in imported[offset:offset + 500]]
            existing_uids.update(str(value) for value in conn.execute(uid_query, {"uids": chunk}).scalars())
        existing_identities: set[tuple[str, str, str]] = set()
        date_query = text("""
            SELECT leave_date, employee_name, leave_reason
            FROM leave_records
            WHERE leave_date IN :dates
        """).bindparams(bindparam("dates", expanding=True))
        imported_dates = sorted({record["leave_date"] for record in imported})
        for offset in range(0, len(imported_dates), 300):
            chunk = imported_dates[offset:offset + 300]
            rows = conn.execute(date_query, {"dates": chunk}).mappings().all()
            existing_identities.update(_leave_import_identity(dict(row)) for row in rows)
        new_records = [
            record for record in imported
            if record["record_uid"] not in existing_uids
            and _leave_import_identity(record) not in existing_identities
        ]
        skipped = duplicate_rows + len(imported) - len(new_records)
        if not new_records:
            tx.rollback()
            return {
                "ok": True,
                "imported": 0,
                "skipped": skipped,
                "message": f"Không có dữ liệu mới. Đã bỏ qua {skipped} dòng đã tồn tại hoặc bị trùng.",
            }

        worksheet = _google_client().open_by_key(LEAVE_SHEET_ID).get_worksheet(0)
        all_values = worksheet.get_all_values()
        headers = all_values[0][:13] if all_values else []
        if len(headers) < 13:
            raise HTTPException(503, "MainData chưa có đủ 13 cột A:M để nhận dữ liệu import.")
        sheet_last_row = 1
        for index, values in enumerate(all_values[1:], start=2):
            if any(str(value or "").strip() for value in values[:13]):
                sheet_last_row = index
        database_last_row = conn.execute(text("""
            SELECT COALESCE(MAX(source_row), 1)
            FROM leave_records
            WHERE source_sheet_id=:source_sheet_id AND source_row IS NOT NULL
        """), {"source_sheet_id": LEAVE_SHEET_ID}).scalar() or 1
        start_row = max(sheet_last_row + 1, int(database_last_row) + 1, 2)
        sheet_values = []
        for index, record in enumerate(new_records):
            source_row = start_row + index
            record["source_row"] = source_row
            sheet_values.append(_sheet_values_for_record(headers, record, source_row))
        insert_rows = [
            {
                **record,
                "sid": LEAVE_SHEET_ID,
                "srow": record["source_row"],
                "payload": json_text(_record_payload(record, record["source_row"])),
            }
            for record in new_records
        ]
        conn.execute(text("""
            INSERT INTO leave_records(
                source_sheet_id, source_row, leave_date, employee_name, leave_reason,
                leave_type, detail, calculated_days, accumulated_leave, penalty,
                update_date, update_time, updated_by, weekday_label, payload, record_uid,
                created_at, updated_at
            ) VALUES (
                :sid, :srow, :leave_date, :employee_name, :leave_reason,
                :leave_type, :detail, :calculated_days, :accumulated_leave, :penalty,
                :update_date, :update_time, :updated_by, :weekday_label, CAST(:payload AS jsonb), :record_uid,
                NOW(), NOW()
            )
        """), insert_rows)
        try:
            with conn.begin_nested():
                conn.execute(text("""
                    INSERT INTO vera_sync_event(dataset_key,event_type,detail,created_at)
                    VALUES ('leave_primary','web_v2_leave_import',:detail,NOW())
                """), {"detail": f"count={len(new_records)}; actor={ident.employee_username}"})
        except Exception:
            pass
        end_row = start_row + len(new_records) - 1
        sheet_range = f"A{start_row}:M{end_row}"
        worksheet.update(range_name=sheet_range, values=sheet_values, value_input_option="USER_ENTERED")
        try:
            tx.commit()
        except Exception:
            try:
                worksheet.batch_clear([sheet_range])
            except Exception:
                pass
            raise
        dates = [record["leave_date"] for record in new_records]
        return {
            "ok": True,
            "imported": len(new_records),
            "skipped": skipped,
            "start": min(dates).isoformat(),
            "end": max(dates).isoformat(),
            "message": (
                f"Đã Import {len(new_records)} lịch nghỉ cũ THÀNH CÔNG"
                + (f"; bỏ qua {skipped} dòng đã tồn tại hoặc bị trùng." if skipped else ".")
            ),
        }
    except HTTPException:
        if tx.is_active:
            tx.rollback()
        if worksheet is not None and sheet_range:
            try:
                worksheet.batch_clear([sheet_range])
            except Exception:
                pass
        raise
    except Exception as exc:
        if tx.is_active:
            tx.rollback()
        if worksheet is not None and sheet_range:
            try:
                worksheet.batch_clear([sheet_range])
            except Exception:
                pass
        raise HTTPException(500, f"Không Import được lịch nghỉ an toàn: {type(exc).__name__}: {exc}") from exc
    finally:
        conn.close()


@app.get("/v2/leave/export.xlsx")
def export_leave_excel(
    start_date: date = Query(alias="start"),
    end_date: date = Query(alias="end"),
    employee: str = Query(default="", max_length=200),
    ident: Identity = Depends(current_identity),
):
    if ident.role != "admin":
        raise HTTPException(403, "Chỉ tài khoản admin được xuất danh sách Excel.")
    if end_date < start_date:
        raise HTTPException(400, "Khoảng thời gian không hợp lệ.")
    if (end_date - start_date).days > 365:
        raise HTTPException(400, "Khoảng xuất Excel tối đa là 366 ngày.")

    needle = _norm(employee)
    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "leave")
        rows = conn.execute(text("""
            SELECT leave_date, employee_name, leave_reason, leave_type, detail,
                   calculated_days, accumulated_leave, penalty,
                   update_date, update_time, updated_by, record_uid
            FROM leave_records
            WHERE leave_date BETWEEN :start_date AND :end_date
            ORDER BY leave_date, employee_name, record_uid
        """), {
            "start_date": start_date,
            "end_date": end_date,
        }).mappings().all()
    if needle:
        rows = [row for row in rows if _employee_name_matches(row["employee_name"], needle)]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lịch nghỉ"
    headers = [
        "Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ", "Loại nghỉ", "Chi tiết",
        "Số ngày tính", "Số ngày phép cộng dồn", "Phạt", "Ngày ghi", "Giờ ghi",
        "Người ghi", "Mã bản ghi",
    ]
    sheet.append(headers)
    for row in rows:
        leave_date_value = row["leave_date"]
        sheet.append([
            leave_date_value,
            _weekday_short_label(leave_date_value),
            row.get("employee_name") or "",
            row.get("leave_reason") or "",
            row.get("leave_type") or "",
            row.get("detail") or "",
            float(row.get("calculated_days") or 0),
            float(row.get("accumulated_leave") or 0),
            float(row.get("penalty") or 0),
            row.get("update_date") or "",
            row.get("update_time") or "",
            row.get("updated_by") or "",
            row.get("record_uid") or "",
        ])

    header_fill = PatternFill("solid", fgColor="214639")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [13, 12, 24, 34, 16, 42, 14, 23, 16, 14, 12, 20, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, 1).number_format = "dd/mm/yyyy"
        sheet.cell(row_number, 9).number_format = '#,##0" đ"'
        for column in range(1, len(headers) + 1):
            sheet.cell(row_number, column).alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = quote(
        f"vera-lich-nghi-{start_date.strftime('%d%m%Y')}-den-{end_date.strftime('%d%m%Y')}.xlsx"
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.get("/v2/leave/daily-stats")
def leave_daily_stats(
    start_date: date = Query(alias="start"),
    end_date: date = Query(alias="end"),
    employee: str = Query(default="", max_length=200),
    ident: Identity = Depends(current_identity),
):
    if end_date < start_date:
        raise HTTPException(400, "Khoảng thời gian không hợp lệ.")
    if (end_date - start_date).days > 365:
        raise HTTPException(400, "Khoảng thống kê tối đa là 366 ngày.")

    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "leave")
        can_view_penalty = _feature_allowed(conn, ident, "employee_penalty_view")
        quota = _daily_quota_config(conn)
        rows = conn.execute(text("""
            SELECT l.leave_date, l.employee_name, l.leave_reason, l.leave_type,
                   COALESCE(l.penalty, 0) AS penalty
            FROM leave_records l
            WHERE l.leave_date BETWEEN :start_date AND :end_date
              AND EXISTS (
                SELECT 1
                FROM employees e
                WHERE lower(btrim(e.username)) = lower(btrim(l.employee_name))
                  AND lower(COALESCE(e.role, '')) NOT IN ('admin','letan','locker','tapvu')
              )
            ORDER BY l.leave_date, l.employee_name, l.record_uid
        """), {"start_date": start_date, "end_date": end_date}).mappings().all()

    employee_needle = _norm(employee)
    if employee_needle:
        rows = [row for row in rows if _employee_name_matches(row["employee_name"], employee_needle)]

    buckets: dict[date, dict[str, Any]] = {}
    for row in rows:
        day = row["leave_date"]
        bucket = buckets.setdefault(day, {
            "paid": 0,
            "generated": 0,
            "unpaid": 0,
            "total_penalty": 0.0,
        })
        group = _stats_group(row.get("leave_type", ""), row.get("leave_reason", ""))
        if group == "co_phep":
            bucket["paid"] += 1
        elif group == "phat_sinh":
            bucket["generated"] += 1
        elif group == "khong_phep":
            bucket["unpaid"] += 1
        bucket["total_penalty"] += float(row.get("penalty") or 0)

    output = []
    for day in sorted(buckets):
        bucket = buckets[day]
        day_quota = quota["days"][day.weekday()]
        paid_limit = int(day_quota["paid_limit"])
        generated_limit = int(day_quota["generated_limit"])
        item = {
            "date": day.isoformat(),
            "weekday_label": _weekday_short_label(day),
            "total_leave": bucket["paid"] + bucket["generated"] + bucket["unpaid"],
            "paid": bucket["paid"],
            "generated": bucket["generated"],
            "unpaid": bucket["unpaid"],
            "paid_limit": paid_limit,
            "generated_limit": generated_limit,
            "paid_full": bucket["paid"] >= paid_limit,
            "generated_full": (
                bucket["generated"] > 0 if generated_limit == 0
                else bucket["generated"] >= generated_limit
            ),
        }
        if can_view_penalty:
            item["total_penalty"] = bucket["total_penalty"]
        output.append(item)
    return {"days": output}


@app.get("/v2/leave/summary")
def leave_summary(date_value: date = Query(alias="date"), ident: Identity = Depends(current_identity)):
    with _engine_instance().connect() as conn:
        _require_feature(conn, ident, "leave")
        active = conn.execute(text("""
            SELECT count(*)
            FROM employees
            WHERE COALESCE(login_locked,false)=false
              AND COALESCE(payload->>'__deleted','false') <> 'true'
              AND COALESCE(payload->>'Trạng thái làm việc', payload->>'employment_status', 'Đang làm việc') = 'Đang làm việc'
              AND lower(COALESCE(role,'')) NOT IN ('admin','letan','locker','tapvu')
        """)).scalar() or 0
        rows = conn.execute(text("""
            SELECT employee_name, leave_reason, leave_type, calculated_days
            FROM leave_records WHERE leave_date=:d
        """), {"d": date_value}).mappings().all()
    return summarize_leave_day(rows, int(active))


@app.post("/v2/leave/records")
def create_leave(body: LeaveCreate, ident: Identity = Depends(current_identity)):
    engine = _engine_instance()
    conn = engine.connect()
    tx = conn.begin()
    ws = None
    sheet_row = None
    try:
        # Same advisory lock key used by Phase 4, so Web V2 and Streamlit serialize writes.
        conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:leave_primary'))"))
        _require_feature(conn, ident, "leave_create")
        if _registration_role_locked(conn, ident.role):
            raise HTTPException(
                403,
                "Quyền đăng ký nghỉ của vai trò này đang bị Admin tạm khóa.",
            )
        record, warnings = _validate_and_prepare(conn, body, ident)

        ws = _google_client().open_by_key(LEAVE_SHEET_ID).get_worksheet(0)
        sheet_row, row_values = _sheet_row_for_record(ws, record)
        record["source_row"] = sheet_row
        _insert_record(conn, record, sheet_row)

        # Mirror before committing PG.  If Google fails, transaction rolls back.
        ws.update(range_name=f"A{sheet_row}:M{sheet_row}", values=[row_values], value_input_option="USER_ENTERED")
        try:
            tx.commit()
        except Exception:
            # Commit failure after a successful mirror: best-effort compensation of the new Sheet row.
            try:
                ws.delete_rows(sheet_row)
            except Exception:
                pass
            raise
        return {
            "ok": True, "record_uid": record["record_uid"], "record": record,
            "warnings": warnings,
            "message": "Đã ghi lịch nghỉ THÀNH CÔNG",
        }
    except HTTPException:
        if tx.is_active:
            tx.rollback()
        raise
    except Exception as exc:
        if tx.is_active:
            tx.rollback()
        raise HTTPException(500, f"Không ghi được lịch nghỉ an toàn: {type(exc).__name__}: {exc}") from exc
    finally:
        conn.close()


@app.patch("/v2/leave/records/{record_uid}")
def update_leave(record_uid: str, body: LeaveUpdate, ident: Identity = Depends(current_identity)):
    engine = _engine_instance()
    conn = engine.connect()
    tx = conn.begin()
    ws = None
    backups: dict[int, list[Any]] = {}
    wrote_sheet = False
    try:
        conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:leave_primary'))"))
        old = conn.execute(text("""
            SELECT record_uid, source_sheet_id, source_row, leave_date, employee_name,
                   leave_reason, leave_type, detail, calculated_days, accumulated_leave,
                   penalty, update_date, update_time, updated_by, weekday_label
            FROM leave_records WHERE record_uid=:uid FOR UPDATE
        """), {"uid": str(record_uid or "").strip()}).mappings().first()
        if not old:
            raise HTTPException(404, "Không tìm thấy lịch nghỉ cần sửa.")
        old = dict(old)
        if str(old.get("source_sheet_id") or "") != LEAVE_SHEET_ID or int(old.get("source_row") or 0) < 2:
            raise HTTPException(409, "Bản ghi không có vị trí MainData hợp lệ; từ chối sửa để tránh lệch dữ liệu.")

        item, future_conversion = _validate_edit_permission(conn, old, body.leave_reason, ident)
        manual_penalty = body.manual_penalty
        if item.get("requires_manual_penalty") and manual_penalty is None:
            raise HTTPException(400, "Lý do này bắt buộc nhập Mức phạt vi phạm khi sửa.")
        old_key = _progressive_key(old["leave_reason"])
        new_key = _progressive_key(item["name"])
        existing_ordinal = _existing_progressive_ordinal(old.get("detail", "")) if old_key and old_key == new_key else None
        request_body = LeaveCreate(
            leave_date=old["leave_date"],
            employee_name=old["employee_name"],
            leave_reason=item["name"],
            detail=_strip_progressive_prefix(old.get("detail", "")),
            manual_penalty=manual_penalty,
        )
        record, warnings = _validate_and_prepare(
            conn,
            request_body,
            ident,
            exclude_record_uid=old["record_uid"],
            skip_registration_timing=future_conversion,
            record_uid=old["record_uid"],
            existing_ordinal=existing_ordinal,
        )
        source_row = int(old["source_row"])
        _update_record(conn, record, source_row)
        rebalanced = _rebalance_progressive_rows(conn, old["leave_date"], {old_key, new_key})

        ws = _google_client().open_by_key(LEAVE_SHEET_ID).get_worksheet(0)
        all_values = ws.get_all_values()
        headers = all_values[0][:13] if all_values else []
        if not headers:
            raise RuntimeError("MainData chưa có header A:M")
        sheet_updates: dict[int, dict] = {source_row: record}
        sheet_updates.update({row_number: changed for row_number, changed in rebalanced})
        for row_number in sheet_updates:
            backups[row_number] = list(all_values[row_number - 1][:13]) if row_number <= len(all_values) else []
        for row_number, changed in sorted(sheet_updates.items()):
            values = _sheet_values_for_record(headers, changed, row_number)
            ws.update(range_name=f"A{row_number}:M{row_number}", values=[values], value_input_option="USER_ENTERED")
            wrote_sheet = True
        tx.commit()
        return {
            "ok": True,
            "record_uid": old["record_uid"],
            "warnings": warnings,
            "message": "Đã sửa lịch nghỉ và đồng bộ PostgreSQL/MainData.",
        }
    except HTTPException:
        if tx.is_active:
            tx.rollback()
        if wrote_sheet and ws is not None:
            _restore_sheet_updates(ws, backups)
        raise
    except Exception as exc:
        if tx.is_active:
            tx.rollback()
        if wrote_sheet and ws is not None:
            _restore_sheet_updates(ws, backups)
        raise HTTPException(500, f"Không sửa được lịch nghỉ an toàn: {type(exc).__name__}: {exc}") from exc
    finally:
        conn.close()


def _delete_leave_uids(record_uids: list[str], ident: Identity):
    uids = list(dict.fromkeys(str(uid or "").strip() for uid in record_uids if str(uid or "").strip()))
    if not uids:
        raise HTTPException(400, "Chưa chọn lịch nghỉ cần xóa.")
    engine = _engine_instance()
    conn = engine.connect()
    tx = conn.begin()
    source_rows: list[int] = []
    rebalanced: list[tuple[int, dict]] = []
    deleted_count = 0
    try:
        conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:phase4:leave_primary'))"))
        rows = []
        for uid in uids:
            row = conn.execute(text("""
                SELECT record_uid, source_sheet_id, source_row, leave_date, employee_name,
                       leave_reason, leave_type, detail, calculated_days, accumulated_leave,
                       penalty, update_date, update_time, updated_by, weekday_label
                FROM leave_records WHERE record_uid=:uid FOR UPDATE
            """), {"uid": uid}).mappings().first()
            if not row:
                raise HTTPException(404, "Một lịch nghỉ đã chọn không còn tồn tại; vui lòng làm mới dữ liệu.")
            item = dict(row)
            _validate_delete_permission(conn, item, ident)
            rows.append(item)
        source_rows = sorted({
            int(row.get("source_row") or 0)
            for row in rows
            if str(row.get("source_sheet_id") or "") == LEAVE_SHEET_ID
            and int(row.get("source_row") or 0) >= 2
        })

        affected: dict[date, set[str]] = {}
        for row in rows:
            key = _progressive_key(row["leave_reason"])
            if key:
                affected.setdefault(row["leave_date"], set()).add(key)
            result = conn.execute(text("DELETE FROM leave_records WHERE record_uid=:uid"), {"uid": row["record_uid"]})
            if int(result.rowcount or 0) != 1:
                raise RuntimeError("Xóa PostgreSQL không đúng một bản ghi.")
        _reindex_after_delete(conn, source_rows)
        for target, keys in affected.items():
            rebalanced.extend(_rebalance_progressive_rows(conn, target, keys))
        deleted_count = len(rows)
        # PostgreSQL/record_uid is canonical. A historical Auto Check row may
        # legitimately have no MainData source_row when its optional mirror
        # failed. Commit the deletion first instead of blocking or resurrecting it.
        tx.commit()
    except HTTPException:
        if tx.is_active:
            tx.rollback()
        conn.close()
        raise
    except Exception as exc:
        if tx.is_active:
            tx.rollback()
        conn.close()
        raise HTTPException(500, f"Không xóa được lịch nghỉ an toàn: {type(exc).__name__}: {exc}") from exc

    mirror_pending = False
    if source_rows or rebalanced:
        try:
            ws = _google_client().open_by_key(LEAVE_SHEET_ID).get_worksheet(0)
            all_values = ws.get_all_values()
            headers = all_values[0][:13] if all_values else []
            if not headers:
                raise RuntimeError("MainData chưa có header A:M")
            for source_row in sorted(source_rows, reverse=True):
                if source_row <= len(all_values):
                    ws.delete_rows(source_row)
            after_delete_values = ws.get_all_values() if rebalanced else []
            for row_number, changed in rebalanced:
                if row_number > len(after_delete_values):
                    continue
                ws.update(
                    range_name=f"A{row_number}:M{row_number}",
                    values=[_sheet_values_for_record(headers, changed, row_number)],
                    value_input_option="USER_ENTERED",
                )
        except Exception:
            mirror_pending = True
    conn.close()
    message = f"Đã xóa {deleted_count} bản ghi THÀNH CÔNG."
    if mirror_pending:
        message += " PostgreSQL đã cập nhật; MainData đang chờ đồng bộ lại."
    return {
        "ok": True,
        "deleted": deleted_count,
        "mirror_pending": mirror_pending,
        "message": message,
    }


@app.delete("/v2/leave/records")
def delete_leave(body: LeaveDelete, ident: Identity = Depends(current_identity)):
    return _delete_leave_uids(body.record_uids, ident)


# Employee routes are kept in a separate module so this migration remains
# reviewable while sharing the same authenticated FastAPI application.
from vera_web_v2_staff import install_staff_routes

install_staff_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    feature_allowed=_feature_allowed,
    norm=_norm,
    google_client=_google_client,
    leave_sheet_id=LEAVE_SHEET_ID,
    progressive_key=_progressive_key,
    reindex_after_delete=_reindex_after_delete,
    rebalance_progressive_rows=_rebalance_progressive_rows,
    sheet_values_for_record=_sheet_values_for_record,
    restore_sheet_updates=_restore_sheet_updates,
    identity_type=Identity,
    vn_tz=VN_TZ,
)

# Official policy routes use the same auth/permission boundary. PostgreSQL is
# canonical and every successful write mirrors the legacy LoaiNghi worksheet.
from vera_web_v2_rules import install_rules_routes

install_rules_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    feature_allowed=_feature_allowed,
    google_client=_google_client,
    leave_sheet_id=LEAVE_SHEET_ID,
    identity_type=Identity,
    vn_tz=VN_TZ,
)

# Annual-leave / long-leave requests reuse the existing Phase-14 PostgreSQL
# dataset and the current NghiDaiHan approval workflow.
from vera_web_v2_long_leave import install_long_leave_routes

install_long_leave_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    feature_allowed=_feature_allowed,
    norm=_norm,
    google_client=_google_client,
    leave_sheet_id=LEAVE_SHEET_ID,
    identity_type=Identity,
    vn_tz=VN_TZ,
)

from vera_web_v2_permissions import install_permission_routes

install_permission_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    google_client=_google_client,
    identity_type=Identity,
    vn_tz=VN_TZ,
    permissions_changed=_clear_permission_cache,
)

from vera_web_v2_profile import install_profile_routes

install_profile_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    identity_type=Identity,
)

from vera_web_v2_snapshot import install_snapshot_routes

install_snapshot_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    identity_type=Identity,
)

from vera_web_v2_payroll import install_payroll_routes

install_payroll_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    norm=_norm,
    identity_type=Identity,
    google_client=_google_client,
)

from vera_web_v2_people import install_people_routes

install_people_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    identity_type=Identity,
    vn_tz=VN_TZ,
)

from vera_web_v2_storage import install_storage_routes

install_storage_routes(
    app,
    engine_instance=_engine_instance,
    current_identity=current_identity,
    require_feature=_require_feature,
    delete_leave_uids=_delete_leave_uids,
    identity_type=Identity,
)
