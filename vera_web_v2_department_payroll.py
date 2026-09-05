"""Editable monthly payroll for Quản lý, Locker, Lễ tân and Tạp vụ."""
from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from email.utils import formataddr
from html import escape
from io import BytesIO
import os
import smtplib
from typing import Any, Callable, Literal
import uuid

from fastapi import Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import text

import vera_web_v2_payroll as payroll
import vera_web_v2_snapshot as attendance
import vera_web_v2_work_schedule as work_schedule


RELEASE = "department-payroll-combined-history-schedule-excel-2026-09-05-v5"
DEPARTMENTS = {
    "quanly": "Quản lý",
    "locker": "Locker",
    "letan": "Lễ tân",
    "tapvu": "Tạp vụ",
}
CALCULATION_MODES = {
    "quanly": "hourly",
    "locker": "hourly",
    "letan": "hourly",
    "tapvu": "monthly",
}
VN_TZ = timezone(timedelta(hours=7))

DEFAULT_CONFIG = {
    "quanly": {
        "calculation_mode": "hourly",
        "rate_ca1": 0,
        "rate_ca2_before_22": 0,
        "rate_ca2_after_22": 0,
        "standard_day_hours": 8,
        "standard_month_days": 26,
        "full_day_hours": 12,
        "full_day_allowance": 0,
        "default_base_salary": 0,
        "default_attendance_bonus": 0,
        "default_responsibility": 0,
        "default_seniority": 0,
        "default_combo_sales": 0,
    },
    "locker": {
        "calculation_mode": "hourly",
        "rate_ca1": 27000,
        "rate_ca2_before_22": 27000,
        "rate_ca2_after_22": 30000,
        "standard_day_hours": 8,
        "standard_month_days": 26,
        "full_day_hours": 12,
        "full_day_allowance": 30000,
        "default_base_salary": 0,
        "default_attendance_bonus": 0,
        "default_responsibility": 0,
        "default_seniority": 0,
        "default_combo_sales": 0,
    },
    "letan": {
        "calculation_mode": "hourly",
        "rate_ca1": 27000,
        "rate_ca2_before_22": 30000,
        "rate_ca2_after_22": 33000,
        "standard_day_hours": 8,
        "standard_month_days": 26,
        "full_day_hours": 12,
        "full_day_allowance": 30000,
        "default_base_salary": 0,
        "default_attendance_bonus": 500000,
        "default_responsibility": 0,
        "default_seniority": 0,
        "default_combo_sales": 0,
    },
    "tapvu": {
        "calculation_mode": "monthly",
        "rate_ca1": 0,
        "rate_ca2_before_22": 0,
        "rate_ca2_after_22": 0,
        "standard_day_hours": 8,
        "standard_month_days": 26,
        "full_day_hours": 12,
        "full_day_allowance": 0,
        "default_base_salary": 0,
        "default_attendance_bonus": 0,
        "default_responsibility": 0,
        "default_seniority": 0,
        "default_combo_sales": 0,
    },
}

DEFAULT_EMAIL_TEMPLATE = {
    "subject": "Bảng lương {bo_phan} tháng {thang} - {ten_nhan_vien}",
    "body": (
        "Chào {ten_nhan_vien},\n\n"
        "VERA SPA gửi bảng lương {bo_phan} tháng {thang}.\n\n"
        "{bang_chi_tiet}\n\n"
        "Tổng lương: {tong_luong}\n"
        "Tổng phạt: {tong_phat}\n"
        "Số tiền thực nhận: {thuc_nhan}\n\n"
        "Vui lòng kiểm tra và phản hồi nếu có sai sót.\n\nTrân trọng,\nVERA SPA"
    ),
}

ROW_MONEY_FIELDS = (
    "base_salary", "salary", "full_allowance", "attendance_bonus", "responsibility",
    "seniority", "combo_sales", "other_income_1", "other_income_2", "total_salary",
    "violation_penalty", "late_penalty", "advance", "net_salary",
)

EXCEL_COLUMNS = (
    ("tt", "TT"), ("employee_name", "Họ tên"), ("department_label", "Bộ phận"),
    ("base_salary", "Lương cơ bản"), ("work_days", "Số ngày đi làm"),
    ("hours_ca1", "Giờ Ca 1"), ("hours_ca2_before_22", "Giờ Ca 2 trước 22h"),
    ("hours_ca2_after_22", "Giờ Ca 2 sau 22h"), ("salary", "Tiền lương"),
    ("full_allowance", "Phụ cấp Full"), ("attendance_bonus", "Tiền chuyên cần"),
    ("responsibility", "Tiền trách nhiệm"), ("seniority", "Phụ cấp thâm niên"),
    ("combo_sales", "Bán combo"), ("other_income_1", "Khoản cộng khác 1"),
    ("other_income_2", "Khoản cộng khác 2"), ("total_salary", "Tổng lương"),
    ("violation_penalty", "Tiền phạt vi phạm"), ("late_penalty", "Tiền phạt đi trễ"),
    ("advance", "Tiền đã ứng"), ("net_salary", "Số tiền thực nhận"),
)


class PenaltyRule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()), max_length=80)
    name: str = Field(min_length=1, max_length=300)
    amount: float = Field(default=0, ge=0, le=1_000_000_000)
    note: str = Field(default="", max_length=1000)
    enabled: bool = True


class DepartmentSettingsUpdate(BaseModel):
    config: dict[str, Any]
    penalty_rules: list[PenaltyRule] = Field(default_factory=list, max_length=300)
    email_template: dict[str, str]


class DepartmentDraft(BaseModel):
    department: Literal["quanly", "locker", "letan", "tapvu"]
    month: str = Field(pattern=r"^\d{4}-\d{2}$")
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=300)


class DepartmentEmail(DepartmentDraft):
    employees: list[str] = Field(default_factory=list, max_length=300)


class CombinedPayrollDraft(BaseModel):
    month: str = Field(pattern=r"^\d{4}-\d{2}$")
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=1200)
    history_id: str = Field(default="", max_length=100)


class EmployeeSalaryConfigUpdate(BaseModel):
    rows: list[dict[str, Any]] = Field(default_factory=list, max_length=300)


def _number(value: Any) -> int:
    return max(0, payroll._number(value))


def _month_range(month: str) -> tuple[date, date, str]:
    try:
        year, month_number = (int(part) for part in month.split("-", 1))
        if year < 2020 or year > 2100 or month_number < 1 or month_number > 12:
            raise ValueError
    except Exception as exc:
        raise HTTPException(400, "Tháng lương không hợp lệ.") from exc
    return (
        date(year, month_number, 1),
        date(year, month_number, calendar.monthrange(year, month_number)[1]),
        f"{month_number:02d}/{year}",
    )


def _setting_key(department: str, suffix: str) -> str:
    return f"department_{department}_{suffix}"


def _clean_config(department: str, raw: Any) -> dict[str, Any]:
    defaults = DEFAULT_CONFIG[department]
    source = raw if isinstance(raw, dict) else {}
    result = {"calculation_mode": CALCULATION_MODES[department]}
    for key, value in defaults.items():
        if key == "calculation_mode":
            continue
        result[key] = _number(source.get(key, value))
    result["standard_day_hours"] = max(1, result["standard_day_hours"])
    result["standard_month_days"] = max(1, result["standard_month_days"])
    result["full_day_hours"] = max(1, result["full_day_hours"])
    if department == "tapvu":
        result["standard_month_days"] = 26
    return result


def _settings(conn, department: str) -> dict[str, Any]:
    config = _clean_config(
        department,
        payroll._setting(conn, _setting_key(department, "config"), DEFAULT_CONFIG[department]),
    )
    rules = payroll._setting(conn, _setting_key(department, "penalty_rules"), [])
    template = payroll._setting(conn, _setting_key(department, "email_template"), DEFAULT_EMAIL_TEMPLATE)
    if not isinstance(rules, list):
        rules = []
    if not isinstance(template, dict):
        template = DEFAULT_EMAIL_TEMPLATE
    return {
        "department": department,
        "department_label": DEPARTMENTS[department],
        "config": config,
        "penalty_rules": [dict(item) for item in rules if isinstance(item, dict)],
        "email_template": {
            "subject": str((template or {}).get("subject") or DEFAULT_EMAIL_TEMPLATE["subject"]),
            "body": str((template or {}).get("body") or DEFAULT_EMAIL_TEMPLATE["body"]),
        },
    }


def _salary_employee_catalog(conn) -> list[dict[str, Any]]:
    employees = conn.execute(text("""
        SELECT username,COALESCE(full_name,'') AS full_name,lower(COALESCE(role,'')) AS role
        FROM employees
        WHERE lower(COALESCE(role,'')) IN ('quanly','letan','locker','tapvu')
          AND COALESCE(payload->>'__deleted','false') <> 'true'
          AND lower(COALESCE(payload->>'Trạng thái làm việc',payload->>'employment_status','đang làm việc'))='đang làm việc'
        ORDER BY CASE lower(COALESCE(role,''))
          WHEN 'quanly' THEN 0 WHEN 'letan' THEN 1 WHEN 'locker' THEN 2 ELSE 3 END,
          COALESCE(stt,2147483647),username
    """)).mappings().all()
    return [{
        "employee_username": str(employee.get("username") or "").strip(),
        "employee_name": str(employee.get("full_name") or employee.get("username") or "").strip(),
        "department": str(employee.get("role") or "").strip().lower(),
        "department_label": DEPARTMENTS[str(employee.get("role") or "").strip().lower()],
    } for employee in employees]


def _employee_config_rows(conn) -> list[dict[str, Any]]:
    stored = payroll._setting(conn, "department_employee_salary_configs", {})
    stored = stored if isinstance(stored, dict) else {}
    catalog = _salary_employee_catalog(conn)
    rows = []
    for employee in catalog:
        username = employee["employee_username"]
        role = employee["department"]
        saved = stored.get(username)
        if not isinstance(saved, dict):
            saved = stored.get(username.casefold())
        if not isinstance(saved, dict):
            continue
        config = _clean_config(role, {**DEFAULT_CONFIG[role], **saved})
        rows.append({
            "employee_username": username,
            "employee_name": employee["employee_name"],
            "department": role,
            "department_label": DEPARTMENTS[role],
            **config,
        })
    return rows


def _employee_config_map(conn) -> dict[str, dict[str, Any]]:
    return {str(row["employee_username"]).casefold(): row for row in _employee_config_rows(conn)}


def _salary_config_tables(conn) -> dict[str, list[dict[str, Any]]]:
    rows = _employee_config_rows(conn)
    return {
        "operations": [row for row in rows if row["department"] in {"quanly", "letan", "locker"}],
        "tapvu": [row for row in rows if row["department"] == "tapvu"],
    }


def _parse_clock(value: Any, work_day: date) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    for candidate in (raw, f"{work_day.strftime('%d/%m/%Y')} {raw}"):
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%H:%M:%S", "%H:%M"):
            try:
                parsed = datetime.strptime(candidate, fmt)
                if parsed.year == 1900:
                    parsed = parsed.replace(year=work_day.year, month=work_day.month, day=work_day.day)
                return parsed
            except ValueError:
                continue
    return None


def _interval_minutes(item: dict[str, Any], work_day: date) -> tuple[int, datetime | None, datetime | None]:
    start = _parse_clock(item.get("check_in"), work_day)
    end = _parse_clock(item.get("check_out"), work_day)
    if start and end:
        if end < start:
            end += timedelta(days=1)
        return max(0, int((end - start).total_seconds() // 60)), start, end
    return max(0, payroll._number(item.get("total_minutes"))), start, end


def _attendance_totals(records: list[dict[str, Any]], employee: str, norm: Callable[[Any], str], cfg: dict[str, Any]) -> dict[str, Any]:
    totals = {"minutes_ca1": 0, "minutes_ca2_before_22": 0, "minutes_ca2_after_22": 0, "full_days": 0, "incomplete_days": 0}
    for item in records:
        if norm(item.get("employee_name")) != norm(employee):
            continue
        try:
            work_day = datetime.strptime(str(item.get("date") or ""), "%d/%m/%Y").date()
        except ValueError:
            continue
        minutes, start, end = _interval_minutes(item, work_day)
        if minutes <= 0:
            totals["incomplete_days"] += 1
            continue
        if minutes >= cfg["full_day_hours"] * 60:
            totals["full_days"] += 1
        shift = norm(item.get("shift"))
        if "ca 2" not in shift:
            totals["minutes_ca1"] += minutes
            continue
        if not start or not end:
            totals["minutes_ca2_before_22"] += minutes
            continue
        cutoff = datetime.combine(work_day, datetime.strptime("22:00", "%H:%M").time())
        before = max(0, int((min(end, cutoff) - start).total_seconds() // 60)) if start < cutoff else 0
        before = min(minutes, before)
        totals["minutes_ca2_before_22"] += before
        totals["minutes_ca2_after_22"] += max(0, minutes - before)
    return totals


def _recalculate(row: dict[str, Any], cfg: dict[str, Any]) -> dict[str, Any]:
    result = dict(row)
    for field in ROW_MONEY_FIELDS:
        result[field] = _number(result.get(field))
    for field in ("hours_ca1", "hours_ca2_before_22", "hours_ca2_after_22", "work_days"):
        try:
            result[field] = round(max(0.0, float(result.get(field) or 0)), 2)
        except (TypeError, ValueError):
            result[field] = 0.0
    if cfg["calculation_mode"] == "monthly":
        result["salary"] = int(round(result["base_salary"] * result["work_days"] / cfg["standard_month_days"]))
    else:
        result["salary"] = int(round(
            result["hours_ca1"] * cfg["rate_ca1"]
            + result["hours_ca2_before_22"] * cfg["rate_ca2_before_22"]
            + result["hours_ca2_after_22"] * cfg["rate_ca2_after_22"]
        ))
    result["total_salary"] = sum(result[field] for field in (
        "salary", "full_allowance", "attendance_bonus", "responsibility", "seniority",
        "combo_sales", "other_income_1", "other_income_2",
    ))
    result["net_salary"] = result["total_salary"] - sum(
        result[field] for field in ("violation_penalty", "late_penalty", "advance")
    )
    return result


def _employees(conn, department: str) -> list[dict[str, Any]]:
    return [dict(row) for row in conn.execute(text("""
        SELECT username,COALESCE(full_name,'') AS full_name,COALESCE(email,'') AS email
        FROM employees
        WHERE lower(COALESCE(role,''))=:department
          AND COALESCE(payload->>'__deleted','false') <> 'true'
          AND lower(COALESCE(payload->>'Trạng thái làm việc',payload->>'employment_status','đang làm việc'))='đang làm việc'
        ORDER BY COALESCE(stt,2147483647),username
    """), {"department": department}).mappings().all()]


def _penalty_maps(conn, start: date, end: date, norm: Callable[[Any], str]) -> tuple[dict[str, int], dict[str, int]]:
    rows = conn.execute(text("""
        SELECT employee_name,leave_reason,COALESCE(penalty,0) AS penalty
        FROM leave_records WHERE leave_date BETWEEN :start AND :end AND COALESCE(penalty,0)>0
    """), {"start": start, "end": end}).mappings().all()
    late: dict[str, int] = {}
    other: dict[str, int] = {}
    for item in rows:
        key = norm(item.get("employee_name"))
        reason = norm(item.get("leave_reason"))
        target = late if ("di tre" in reason or "ra ngoai vao muon" in reason or "vao lai tre" in reason) else other
        target[key] = target.get(key, 0) + _number(item.get("penalty"))
    return other, late


def _calculation(conn, department: str, month: str, norm: Callable[[Any], str]) -> dict[str, Any]:
    start, end, label = _month_range(month)
    settings = _settings(conn, department)
    cfg = settings["config"]
    employee_configs = _employee_config_map(conn)
    employees = _employees(conn, department)
    records = attendance._records(conn, start, end)
    violation_map, late_map = _penalty_maps(conn, start, end, norm)
    rows = []
    for index, employee in enumerate(employees, start=1):
        username = str(employee.get("username") or "").strip()
        employee_cfg = employee_configs.get(username.casefold(), cfg)
        totals = _attendance_totals(records, username, norm, employee_cfg)
        hours_ca1 = round(totals["minutes_ca1"] / 60, 2)
        hours_before = round(totals["minutes_ca2_before_22"] / 60, 2)
        hours_after = round(totals["minutes_ca2_after_22"] / 60, 2)
        total_hours = hours_ca1 + hours_before + hours_after
        row = {
            "tt": index, "employee_username": username,
            "employee_name": str(employee.get("full_name") or username),
            "email": str(employee.get("email") or ""), "department": department,
            "department_label": DEPARTMENTS[department], "base_salary": employee_cfg["default_base_salary"],
            "hours_ca1": hours_ca1, "hours_ca2_before_22": hours_before,
            "hours_ca2_after_22": hours_after,
            "work_days": round(total_hours / employee_cfg["standard_day_hours"], 2),
            "full_days": totals["full_days"],
            "full_allowance": totals["full_days"] * employee_cfg["full_day_allowance"],
            "attendance_bonus": employee_cfg["default_attendance_bonus"],
            "responsibility": employee_cfg["default_responsibility"], "seniority": employee_cfg["default_seniority"],
            "combo_sales": employee_cfg["default_combo_sales"], "other_income_1": 0, "other_income_2": 0,
            "violation_penalty": violation_map.get(norm(username), 0),
            "late_penalty": late_map.get(norm(username), 0), "advance": 0,
            "incomplete_days": totals["incomplete_days"],
        }
        rows.append(_recalculate(row, employee_cfg))
    return {**settings, "month": month, "month_label": label, "start": start.isoformat(), "end": end.isoformat(), "rows": rows}


def _minutes_between(start_time: Any, end_time: Any, work_day: date) -> tuple[int, datetime | None, datetime | None]:
    start = _parse_clock(start_time, work_day)
    end = _parse_clock(end_time, work_day)
    if not start or not end or start == end:
        return 0, start, end
    if end < start:
        end += timedelta(days=1)
    return max(0, int((end - start).total_seconds() // 60)), start, end


def _add_planned_interval(totals: dict[str, int], start_time: Any, end_time: Any, work_day: date, bucket: str = "") -> int:
    minutes, start, end = _minutes_between(start_time, end_time, work_day)
    if minutes <= 0 or not start or not end:
        return 0
    if bucket == "ca1" or (not bucket and start.hour < 12):
        totals["minutes_ca1"] += minutes
        return minutes
    cutoff = datetime.combine(work_day, datetime.strptime("22:00", "%H:%M").time())
    before = max(0, int((min(end, cutoff) - start).total_seconds() // 60)) if start < cutoff else 0
    before = min(minutes, before)
    totals["minutes_ca2_before_22"] += before
    totals["minutes_ca2_after_22"] += max(0, minutes - before)
    return minutes


def _schedule_totals(
    records: list[dict[str, Any]],
    employee: str,
    department: str,
    definitions: dict[str, Any],
    cfg: dict[str, Any],
    norm: Callable[[Any], str],
) -> dict[str, Any]:
    totals = {
        "minutes_ca1": 0, "minutes_ca2_before_22": 0, "minutes_ca2_after_22": 0,
        "work_days": 0, "full_days": 0,
    }
    for item in records:
        if norm(item.get("employee_username")) != norm(employee) or item.get("shift_code") == "Nghỉ":
            continue
        work_day = item.get("work_date")
        if isinstance(work_day, datetime):
            work_day = work_day.date()
        if not isinstance(work_day, date):
            continue
        totals["work_days"] += 1
        shift_code = str(item.get("shift_code") or "")
        regular_minutes = 0
        if department == "quanly":
            regular_minutes = _add_planned_interval(totals, item.get("start_time"), item.get("end_time"), work_day)
        else:
            spec = (definitions.get(department) or {}).get(shift_code) or {}
            bucket = "ca1" if "ca 1" in norm(shift_code) else "ca2"
            regular_minutes = _add_planned_interval(totals, spec.get("start"), spec.get("end"), work_day, bucket)

        overtime_minutes = 0
        overtime_shift = str(item.get("overtime_shift") or "")
        if overtime_shift in {"TC Ca 1", "TC Ca 2"}:
            target_shift = overtime_shift.replace("TC ", "", 1)
            spec = (definitions.get(department) or {}).get(target_shift) or {}
            if spec:
                overtime_minutes = _add_planned_interval(
                    totals, spec.get("start"), spec.get("end"), work_day,
                    "ca1" if target_shift == "Ca 1" else "ca2",
                )
            else:
                fallback_start, fallback_end = (("09:00", "17:00") if target_shift == "Ca 1" else ("17:00", "01:00"))
                overtime_minutes = _add_planned_interval(
                    totals, fallback_start, fallback_end, work_day,
                    "ca1" if target_shift == "Ca 1" else "ca2",
                )
        elif overtime_shift == "Từ giờ tới giờ" or item.get("overtime_start_time") or item.get("overtime_end_time"):
            overtime_minutes = _add_planned_interval(
                totals, item.get("overtime_start_time"), item.get("overtime_end_time"), work_day,
            )
        if regular_minutes + overtime_minutes >= cfg["full_day_hours"] * 60:
            totals["full_days"] += 1
    return totals


def _schedule_calculation(conn, department: str, month: str, norm: Callable[[Any], str]) -> dict[str, Any]:
    start, end, label = _month_range(month)
    work_schedule._ensure_schema(conn)
    settings = _settings(conn, department)
    cfg = settings["config"]
    employee_configs = _employee_config_map(conn)
    employees = _employees(conn, department)
    records = [dict(item) for item in conn.execute(text("""
        SELECT work_date,employee_username,shift_code,overtime_shift,start_time,end_time,
               overtime_start_time,overtime_end_time
        FROM vera_work_schedule
        WHERE work_date BETWEEN :start AND :end AND department=:department
        ORDER BY work_date,employee_username
    """), {"start": start, "end": end, "department": department}).mappings().all()]
    definition_rows = conn.execute(text("""
        SELECT department,shift_code,start_time,end_time
        FROM vera_work_shift_definition
        WHERE department IN ('locker','letan','tapvu')
    """)).mappings().all()
    definitions: dict[str, dict[str, dict[str, str]]] = {key: {} for key in DEPARTMENTS}
    for item in definition_rows:
        definitions[str(item["department"])][str(item["shift_code"])] = {
            "start": str(item.get("start_time") or "")[:5],
            "end": str(item.get("end_time") or "")[:5],
        }
    violation_map, late_map = _penalty_maps(conn, start, end, norm)
    rows = []
    for index, employee in enumerate(employees, start=1):
        username = str(employee.get("username") or "").strip()
        employee_cfg = employee_configs.get(username.casefold(), cfg)
        totals = _schedule_totals(records, username, department, definitions, employee_cfg, norm)
        row = {
            "tt": index, "employee_username": username,
            "employee_name": str(employee.get("full_name") or username),
            "email": str(employee.get("email") or ""), "department": department,
            "department_label": DEPARTMENTS[department],
            "base_salary": employee_cfg["default_base_salary"],
            "hours_ca1": round(totals["minutes_ca1"] / 60, 2),
            "hours_ca2_before_22": round(totals["minutes_ca2_before_22"] / 60, 2),
            "hours_ca2_after_22": round(totals["minutes_ca2_after_22"] / 60, 2),
            "work_days": totals["work_days"], "full_days": totals["full_days"],
            "full_allowance": totals["full_days"] * employee_cfg["full_day_allowance"],
            "attendance_bonus": employee_cfg["default_attendance_bonus"],
            "responsibility": employee_cfg["default_responsibility"],
            "seniority": employee_cfg["default_seniority"], "combo_sales": employee_cfg["default_combo_sales"],
            "other_income_1": 0, "other_income_2": 0,
            "violation_penalty": violation_map.get(norm(username), 0),
            "late_penalty": late_map.get(norm(username), 0), "advance": 0,
            "incomplete_days": 0, "calculation_source": "schedule",
            "calculation_config": employee_cfg,
        }
        rows.append(_recalculate(row, employee_cfg))
    return {**settings, "month": month, "month_label": label, "start": start.isoformat(), "end": end.isoformat(), "rows": rows}


def _combined_calculation(conn, month: str, norm: Callable[[Any], str], source: str) -> dict[str, Any]:
    if source not in {"attendance", "schedule"}:
        raise HTTPException(400, "Nguồn tính lương không hợp lệ.")
    calculator = _schedule_calculation if source == "schedule" else _calculation
    rows: list[dict[str, Any]] = []
    settings: dict[str, Any] = {}
    for department in DEPARTMENTS:
        result = calculator(conn, department, month, norm)
        settings[department] = result
        for row in result["rows"]:
            cfg = _employee_config_map(conn).get(str(row["employee_username"]).casefold(), result["config"])
            row["calculation_config"] = cfg
            row["calculation_source"] = source
            rows.append(row)
    saved_rows = payroll._setting(conn, f"department_payroll_combined_draft_{month}", [])
    saved_advances = {
        norm(item.get("employee_username")): _number(item.get("advance"))
        for item in saved_rows if isinstance(item, dict)
    } if isinstance(saved_rows, list) else {}
    for index, row in enumerate(rows, start=1):
        row["tt"] = index
        row["advance"] = saved_advances.get(norm(row.get("employee_username")), _number(row.get("advance")))
        row.update(_recalculate(row, row["calculation_config"]))
    _, _, label = _month_range(month)
    return {
        "ok": True, "month": month, "month_label": label, "source": source,
        "source_label": "Lịch làm việc" if source == "schedule" else "Chấm công",
        "departments": settings, "rows": rows,
    }


def _combined_employee_catalog(conn) -> dict[str, dict[str, Any]]:
    rows = conn.execute(text("""
        SELECT username,COALESCE(full_name,'') AS full_name,COALESCE(email,'') AS email,
               lower(COALESCE(role,'')) AS role
        FROM employees
        WHERE lower(COALESCE(role,'')) IN ('quanly','letan','locker','tapvu')
    """)).mappings().all()
    return {str(item["username"]).strip().casefold(): dict(item) for item in rows}


def _clean_combined_rows(conn, rows: list[dict[str, Any]], norm: Callable[[Any], str]) -> list[dict[str, Any]]:
    catalog = _combined_employee_catalog(conn)
    employee_configs = _employee_config_map(conn)
    output = []
    seen = set()
    for supplied in rows:
        key = str(supplied.get("employee_username") or "").strip().casefold()
        employee = catalog.get(key)
        if not key or key in seen or not employee:
            raise HTTPException(400, "Bảng lương có nhân viên trống, trùng hoặc không thuộc Lương hành chánh.")
        seen.add(key)
        department = str(employee.get("role") or "").lower()
        supplied_cfg = supplied.get("calculation_config")
        cfg = _clean_config(department, supplied_cfg) if isinstance(supplied_cfg, dict) else employee_configs.get(key, _settings(conn, department)["config"])
        row = dict(supplied)
        row.update({
            "tt": len(output) + 1, "employee_username": employee["username"],
            "employee_name": employee.get("full_name") or employee["username"],
            "email": employee.get("email") or "", "department": department,
            "department_label": DEPARTMENTS[department], "calculation_config": cfg,
        })
        output.append(_recalculate(row, cfg))
    if not output:
        raise HTTPException(400, "Bảng Lương hành chánh chưa có nhân viên.")
    return output


def _clean_rows(conn, department: str, rows: list[dict[str, Any]], cfg: dict[str, Any], norm) -> list[dict[str, Any]]:
    catalog = {norm(item["username"]): item for item in _employees(conn, department)}
    employee_configs = _employee_config_map(conn)
    output = []
    seen = set()
    for supplied in rows:
        key = norm(supplied.get("employee_username"))
        employee = catalog.get(key)
        if not key or key in seen or not employee:
            raise HTTPException(400, "Bảng lương có nhân viên trống, trùng hoặc không thuộc đúng bộ phận.")
        seen.add(key)
        row = dict(supplied)
        row.update({
            "tt": len(output) + 1, "employee_username": employee["username"],
            "employee_name": employee.get("full_name") or employee["username"],
            "email": employee.get("email") or "", "department": department,
            "department_label": DEPARTMENTS[department],
        })
        output.append(_recalculate(row, employee_configs.get(str(employee["username"]).casefold(), cfg)))
    if not output:
        raise HTTPException(400, "Bảng lương chưa có nhân viên.")
    return output


def _workbook(rows: list[dict[str, Any]], department: str, label: str) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = f"Lương {DEPARTMENTS[department]}"
    ws.append([f"BẢNG LƯƠNG {DEPARTMENTS[department].upper()} THÁNG {label}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXCEL_COLUMNS))
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F513F")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([title for _, title in EXCEL_COLUMNS])
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="A99D97")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(key, "") for key, _ in EXCEL_COLUMNS])
    ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:{ws.cell(2, len(EXCEL_COLUMNS)).column_letter}{ws.max_row}"
    for index, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(2, index).column_letter].width = 24 if key == "employee_name" else 16
        if key in ROW_MONEY_FIELDS:
            for cell in ws.iter_cols(min_col=index, max_col=index, min_row=3):
                for item in cell: item.number_format = '#,##0"đ"'
    stream = BytesIO(); wb.save(stream); return stream.getvalue()


def _combined_workbook(rows: list[dict[str, Any]], label: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Lương hành chánh"
    worksheet.append([f"BẢNG LƯƠNG HÀNH CHÁNH THÁNG {label}"])
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXCEL_COLUMNS))
    worksheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor="1F513F")
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet.append([title for _, title in EXCEL_COLUMNS])
    for cell in worksheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="A99D97")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        worksheet.append([row.get(key, "") for key, _ in EXCEL_COLUMNS])
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:{worksheet.cell(2, len(EXCEL_COLUMNS)).column_letter}{worksheet.max_row}"
    for index, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
        worksheet.column_dimensions[worksheet.cell(2, index).column_letter].width = 26 if key == "employee_name" else 17
        if key in ROW_MONEY_FIELDS:
            for column in worksheet.iter_cols(min_col=index, max_col=index, min_row=3):
                for cell in column:
                    cell.number_format = '#,##0"đ"'
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _money(value: Any) -> str:
    return f"{payroll._number(value):,}".replace(",", ".") + "đ"


def _detail_text(row: dict[str, Any]) -> str:
    labels = dict(EXCEL_COLUMNS)
    keys = ("work_days", "salary", "full_allowance", "attendance_bonus", "responsibility", "seniority", "combo_sales", "other_income_1", "other_income_2", "violation_penalty", "late_penalty", "advance")
    return "\n".join(f"{labels[key]}: {_money(row.get(key)) if key != 'work_days' else row.get(key, 0)}" for key in keys)


def _render_template(template: dict[str, str], row: dict[str, Any], month_label: str) -> tuple[str, str, str]:
    values = {
        "ten_nhan_vien": str(row.get("employee_name") or row.get("employee_username") or ""),
        "bo_phan": str(row.get("department_label") or ""), "thang": month_label,
        "tong_luong": _money(row.get("total_salary")),
        "tong_phat": _money(_number(row.get("violation_penalty")) + _number(row.get("late_penalty"))),
        "thuc_nhan": _money(row.get("net_salary")), "bang_chi_tiet": _detail_text(row),
    }
    subject = str(template.get("subject") or DEFAULT_EMAIL_TEMPLATE["subject"])
    body = str(template.get("body") or DEFAULT_EMAIL_TEMPLATE["body"])
    for key, value in values.items():
        subject = subject.replace("{" + key + "}", value)
        body = body.replace("{" + key + "}", value)
    html = "<br>".join(escape(line) for line in body.splitlines())
    return subject, body, f"<!doctype html><html lang='vi'><body style='font:14px Arial;line-height:1.5;color:#22352d;max-width:720px;padding:24px'>{html}</body></html>"


def install_department_payroll_routes(app, *, engine_instance, current_identity, require_feature, identity_type, norm) -> None:
    if getattr(app.state, "department_payroll_installed", False):
        return

    def valid_department(department: str) -> str:
        value = str(department or "").lower().strip()
        if value not in DEPARTMENTS:
            raise HTTPException(404, "Bộ phận không hỗ trợ bảng lương này.")
        return value

    @app.get("/v2/department-payroll/settings")
    def get_settings(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            return {
                "ok": True, "release": RELEASE,
                "departments": {key: _settings(conn, key) for key in DEPARTMENTS},
                "salary_config_tables": _salary_config_tables(conn),
                "salary_employee_catalog": _salary_employee_catalog(conn),
                "email_layout": payroll.PAYROLL_EMAIL_TEMPLATE_RELEASE,
            }

    @app.put("/v2/department-payroll/settings/employees")
    def save_employee_settings(body: EmployeeSalaryConfigUpdate, ident: identity_type = Depends(current_identity)):
        if str(ident.role or "").lower() != "admin":
            raise HTTPException(403, "Chỉ Admin được sửa cấu hình lương theo nhân viên.")
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_config_edit")
            catalog = {row["employee_username"]: row for row in _salary_employee_catalog(conn)}
            supplied = {}
            for row in body.rows:
                username = str(row.get("employee_username") or "").strip()
                employee = catalog.get(username)
                if not employee or username in supplied:
                    raise HTTPException(400, "Cấu hình có nhân viên trống, trùng hoặc không thuộc đúng bộ phận.")
                supplied[username] = _clean_config(employee["department"], row)
            payroll._put_setting(conn, "department_employee_salary_configs", supplied, ident.employee_username)
            tables = _salary_config_tables(conn)
        return {"ok": True, "salary_config_tables": tables, "message": "Đã lưu cấu hình lương theo từng nhân viên."}

    @app.put("/v2/department-payroll/settings/{department}")
    def save_settings(department: str, body: DepartmentSettingsUpdate, ident: identity_type = Depends(current_identity)):
        department = valid_department(department)
        if str(ident.role or "").lower() != "admin":
            raise HTTPException(403, "Chỉ Admin được sửa quy định phạt, công thức lương và mẫu email.")
        config = _clean_config(department, body.config)
        template = {
            "subject": str(body.email_template.get("subject") or "")[:500],
            "body": str(body.email_template.get("body") or "")[:10000],
        }
        if not template["subject"] or not template["body"]:
            raise HTTPException(400, "Mẫu email phải có tiêu đề và nội dung.")
        rules = [item.model_dump() if hasattr(item, "model_dump") else item.dict() for item in body.penalty_rules]
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_config_edit")
            payroll._put_setting(conn, _setting_key(department, "config"), config, ident.employee_username)
            payroll._put_setting(conn, _setting_key(department, "penalty_rules"), rules, ident.employee_username)
            payroll._put_setting(conn, _setting_key(department, "email_template"), template, ident.employee_username)
            saved = _settings(conn, department)
        return {"ok": True, "message": f"Đã lưu cấu hình {DEPARTMENTS[department]}.", **saved}

    @app.get("/v2/department-payroll/calculate")
    def calculate(department: str = Query(...), month: str = Query(...), ident: identity_type = Depends(current_identity)):
        department = valid_department(department)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            return _calculation(conn, department, month, norm)

    @app.get("/v2/department-payroll/combined/calculate")
    def calculate_combined(
        month: str = Query(...),
        source: str = Query("attendance"),
        ident: identity_type = Depends(current_identity),
    ):
        _month_range(month)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            return _combined_calculation(conn, month, norm, source.strip().lower())

    @app.get("/v2/department-payroll/combined/draft")
    def get_combined_draft(month: str = Query(...), ident: identity_type = Depends(current_identity)):
        _month_range(month)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            rows = payroll._setting(conn, f"department_payroll_combined_draft_{month}", [])
        return {"ok": True, "month": month, "rows": rows if isinstance(rows, list) else []}

    @app.put("/v2/department-payroll/combined/draft")
    def save_combined_draft(body: CombinedPayrollDraft, ident: identity_type = Depends(current_identity)):
        _month_range(body.month)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            rows = _clean_combined_rows(conn, body.rows, norm)
            payroll._put_setting(conn, f"department_payroll_combined_draft_{body.month}", rows, ident.employee_username)
        return {"ok": True, "rows": rows, "message": "Đã lưu nháp bảng Lương hành chánh."}

    @app.get("/v2/department-payroll/combined/history")
    def get_combined_history(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            history = payroll._setting(conn, "department_payroll_combined_history", [])
        if not isinstance(history, list):
            history = []
        items = []
        for item in history:
            if not isinstance(item, dict):
                continue
            history_rows = item.get("rows") if isinstance(item.get("rows"), list) else []
            items.append({
                **{key: item.get(key) for key in ("id", "month", "month_label", "saved_at", "saved_by", "source_label")},
                "employee_count": len(history_rows),
                "total_net": sum(_number(row.get("net_salary")) for row in history_rows if isinstance(row, dict)),
            })
        return {"ok": True, "items": sorted(items, key=lambda item: str(item.get("saved_at") or ""), reverse=True)}

    @app.post("/v2/department-payroll/combined/history/{history_id}/open")
    def open_combined_history(history_id: str, ident: identity_type = Depends(current_identity)):
        wanted = str(history_id or "").strip()
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            history = payroll._setting(conn, "department_payroll_combined_history", [])
            item = next((entry for entry in history if isinstance(entry, dict) and str(entry.get("id")) == wanted), None) if isinstance(history, list) else None
            if not item:
                raise HTTPException(404, "Không tìm thấy lịch sử bảng Lương hành chánh.")
            rows = item.get("rows") if isinstance(item.get("rows"), list) else []
            payroll._put_setting(conn, f"department_payroll_combined_draft_{item['month']}", rows, ident.employee_username)
        return {
            "ok": True, "history_id": wanted, "month": item["month"], "rows": rows,
            "message": f"Đã mở bảng lương tháng {item.get('month_label')} để sửa. Bấm Hoàn thành bảng lương để cập nhật lịch sử.",
        }

    @app.post("/v2/department-payroll/combined/complete")
    def complete_combined(body: CombinedPayrollDraft, ident: identity_type = Depends(current_identity)):
        _, _, label = _month_range(body.month)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            rows = _clean_combined_rows(conn, body.rows, norm)
            history = payroll._setting(conn, "department_payroll_combined_history", [])
            history = history if isinstance(history, list) else []
            if body.history_id:
                existing = next((item for item in history if isinstance(item, dict) and str(item.get("id")) == body.history_id), None)
                if not existing:
                    raise HTTPException(404, "Không tìm thấy lịch sử bảng Lương hành chánh đang sửa.")
                if existing.get("month") != body.month:
                    raise HTTPException(400, "Không được chuyển lịch sử bảng lương sang tháng khác.")
            else:
                existing = next((item for item in history if isinstance(item, dict) and item.get("month") == body.month), None)
            history_id = str(existing.get("id")) if existing else str(uuid.uuid4())
            source_label = next((str(row.get("calculation_source") or "") for row in rows if row.get("calculation_source")), "")
            completed = {
                "id": history_id, "month": body.month, "month_label": label,
                "saved_at": datetime.now(VN_TZ).isoformat(), "saved_by": ident.employee_username,
                "source_label": "Lịch làm việc" if source_label == "schedule" else "Chấm công",
                "rows": rows,
            }
            history = [item for item in history if not (isinstance(item, dict) and (str(item.get("id")) == history_id or item.get("month") == body.month))]
            history.append(completed)
            payroll._put_setting(conn, "department_payroll_combined_history", history[-120:], ident.employee_username)
            payroll._put_setting(conn, f"department_payroll_combined_draft_{body.month}", rows, ident.employee_username)
        return {
            "ok": True, "history_id": history_id, "rows": rows,
            "message": f"Đã hoàn thành và lưu bảng Lương hành chánh tháng {label} vào lịch sử.",
        }

    @app.post("/v2/department-payroll/combined/export.xlsx")
    def export_combined_xlsx(body: CombinedPayrollDraft, ident: identity_type = Depends(current_identity)):
        _, _, label = _month_range(body.month)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_export")
            rows = _clean_combined_rows(conn, body.rows, norm)
        content = _combined_workbook(rows, label)
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Luong_hanh_chanh_{body.month}.xlsx"},
        )

    @app.get("/v2/department-payroll/draft")
    def get_draft(department: str = Query(...), month: str = Query(...), ident: identity_type = Depends(current_identity)):
        department = valid_department(department); _month_range(month)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            return {"rows": payroll._setting(conn, _setting_key(department, f"draft_{month}"), []), "department": department, "month": month}

    @app.put("/v2/department-payroll/draft")
    def save_draft(body: DepartmentDraft, ident: identity_type = Depends(current_identity)):
        department = valid_department(body.department); _month_range(body.month)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            cfg = _settings(conn, department)["config"]
            rows = _clean_rows(conn, department, body.rows, cfg, norm)
            payroll._put_setting(conn, _setting_key(department, f"draft_{body.month}"), rows, ident.employee_username)
        return {"ok": True, "rows": rows, "message": f"Đã lưu bảng lương nháp {DEPARTMENTS[department]}."}

    @app.post("/v2/department-payroll/save")
    def save_official(body: DepartmentDraft, ident: identity_type = Depends(current_identity)):
        department = valid_department(body.department); _, _, label = _month_range(body.month)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            cfg = _settings(conn, department)["config"]
            rows = _clean_rows(conn, department, body.rows, cfg, norm)
            history = payroll._setting(conn, _setting_key(department, "history"), [])
            if not isinstance(history, list):
                history = []
            history = [item for item in history if not (isinstance(item, dict) and item.get("month") == body.month)]
            history.append({"id": str(uuid.uuid4()), "month": body.month, "month_label": label, "saved_at": datetime.now(VN_TZ).isoformat(), "saved_by": ident.employee_username, "rows": rows})
            payroll._put_setting(conn, _setting_key(department, "history"), history[-120:], ident.employee_username)
            payroll._put_setting(conn, _setting_key(department, f"draft_{body.month}"), rows, ident.employee_username)
        return {"ok": True, "rows": rows, "message": f"Đã lưu chính thức bảng lương {DEPARTMENTS[department]} tháng {label}."}

    @app.post("/v2/department-payroll/export.xlsx")
    def export_xlsx(body: DepartmentDraft, ident: identity_type = Depends(current_identity)):
        department = valid_department(body.department); _, _, label = _month_range(body.month)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_export")
            rows = _clean_rows(conn, department, body.rows, _settings(conn, department)["config"], norm)
        content = _workbook(rows, department, label)
        return StreamingResponse(BytesIO(content), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Bang_luong_{department}_{body.month}.xlsx"})

    @app.post("/v2/department-payroll/email")
    def email_payroll(body: DepartmentEmail, ident: identity_type = Depends(current_identity)):
        department = valid_department(body.department); start, end, label = _month_range(body.month)
        sender = os.getenv("SMTP_SENDER_EMAIL", "veraspabienhoa@gmail.com").strip()
        password = os.getenv("SMTP_APP_PASSWORD", "")
        if not password:
            raise HTTPException(503, "Máy chủ chưa cấu hình mật khẩu gửi email bảng lương.")
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_email")
            settings = _settings(conn, department)
            rows = _clean_rows(conn, department, body.rows, settings["config"], norm)
            violation_rows = conn.execute(text("""
                SELECT employee_name,leave_date,leave_reason,detail,COALESCE(penalty,0) penalty
                FROM leave_records
                WHERE leave_date BETWEEN :start AND :end AND COALESCE(penalty,0)>0
                ORDER BY leave_date,COALESCE(source_row,0),id
            """), {"start": start, "end": end}).mappings().all()
        violations_by_employee: dict[str, list[dict[str, Any]]] = {}
        for item in violation_rows:
            violations_by_employee.setdefault(norm(item.get("employee_name")), []).append(dict(item))
        selected = {norm(value) for value in body.employees if norm(value)}
        rows = [row for row in rows if not selected or norm(row.get("employee_username")) in selected]
        sent, failed = [], []
        try:
            smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20); smtp.login(sender, password)
        except Exception as exc:
            raise HTTPException(502, f"Không kết nối được máy chủ gửi email: {str(exc)[:200]}") from exc
        try:
            for row in rows:
                recipient = str(row.get("email") or "").strip()
                if "@" not in recipient:
                    failed.append({"employee": row.get("employee_name"), "error": "Chưa có email hợp lệ"}); continue
                email_row = {
                    "Tiền Lương": row.get("salary"),
                    "Tiền Hỗ Trợ Hoàn Lại": sum(_number(row.get(key)) for key in (
                        "full_allowance", "attendance_bonus", "responsibility", "seniority",
                        "combo_sales", "other_income_1", "other_income_2",
                    )),
                    "Hoàn trả tiền tích lũy": 0, "Tích lũy": 0, "Chi Phí Sinh Hoạt": 0,
                    "Tiền phạt trong tháng": _number(row.get("violation_penalty")) + _number(row.get("late_penalty")),
                    "Vi phạm kỳ trước": 0, "Tiền ứng lương": row.get("advance"),
                    "Tiền hỗ trợ Locker": 0, "Số tiền thực nhận": row.get("net_salary"),
                }
                name = str(row.get("employee_name") or row.get("employee_username") or "")
                employee_violations = (
                    violations_by_employee.get(norm(row.get("employee_username")))
                    or violations_by_employee.get(norm(name), [])
                )
                message = EmailMessage(); message["Subject"] = payroll._payroll_email_subject(row.get("employee_username"), start, end)
                message["From"] = formataddr(("VERA SPA", sender)); message["To"] = recipient
                message.set_content(payroll._payroll_email_text(name, start, end, email_row, employee_violations))
                message.add_alternative(payroll._payroll_email_html(name, start, end, email_row, employee_violations), subtype="html")
                message.add_attachment(_workbook([row], department, label), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"Bang_luong_{row['employee_username']}_{body.month}.xlsx")
                try:
                    smtp.send_message(message); sent.append(row.get("employee_name"))
                except Exception as exc:
                    failed.append({"employee": row.get("employee_name"), "error": str(exc)[:200]})
        finally:
            try: smtp.quit()
            except Exception: pass
        return {"ok": not failed, "sent": sent, "failed": failed, "message": f"Đã gửi {len(sent)} email; lỗi {len(failed)}."}

    @app.get("/v2/department-payroll/health")
    def health():
        return {"ok": True, "release": RELEASE, "departments": DEPARTMENTS, "source": "Web V2 attendance + work schedule", "employee_config_tables": True, "email_layout": payroll.PAYROLL_EMAIL_TEMPLATE_RELEASE}

    app.state.department_payroll_installed = True
