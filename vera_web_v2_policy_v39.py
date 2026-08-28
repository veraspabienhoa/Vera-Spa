"""VERA Web V2 V3.9 business-policy additions.

- Employees who start/resume work on or after day 16 are capped at 3 paid-leave
  days for that calendar month.
- Re-activation is stamped in PostgreSQL so the cap also applies to staff who
  were temporarily away during days 1-15.
- Payroll TimeSoft input always reads physical Sheet2.
- A small authenticated detector lets the browser select month/period from the
  actual dates contained in Sheet2 before payroll calculation.
"""
from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from io import BytesIO
import numbers
from typing import Any, Callable

import pandas as pd
from fastapi import Body, Depends, HTTPException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import text

import vera_web_v2_api_shared as _shared
import vera_web_v2_payroll as _payroll
import vera_web_v2_staff as _staff

RELEASE = "3.9-policy-payroll-period"
LATE_MONTH_PAID_LIMIT = 3.0
ACTIVE_STATUS = "Đang làm việc"


def _remove_route(app, path: str, method: str):
    method = method.upper()
    for route in list(app.router.routes):
        if getattr(route, "path", None) == path and method in (getattr(route, "methods", set()) or set()):
            app.router.routes.remove(route)
            return getattr(route, "endpoint", None)
    return None


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except ValueError:
            pass
    return None


def _effective_start_or_resume(row: dict[str, Any]) -> date | None:
    payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
    resume = _parse_date(payload.get("return_to_work_date") or payload.get("Ngày làm lại gần nhất"))
    return resume or _parse_date(row.get("employment_start_date"))


def _read_source_sheet2(content: bytes) -> pd.DataFrame:
    """Read the physical second worksheet only, regardless of workbook active tab."""
    if not content:
        raise HTTPException(400, "File Excel đang trống.")
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(413, "File Excel vượt quá 15 MB.")
    if not content.startswith(b"PK"):
        raise HTTPException(400, "File không đúng định dạng Excel .xlsx. Vui lòng xuất lại từ TimeSoft.")

    workbook = None
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if len(workbook.worksheets) < 2:
            raise HTTPException(400, "File TimeSoft phải có Sheet2. Bảng lương chỉ đọc dữ liệu từ Sheet2.")
        worksheet = workbook.worksheets[1]
        if worksheet.max_row < 4 or worksheet.max_column < len(_payroll.TIMESOFT_PAYROLL_HEADERS):
            raise HTTPException(400, "Sheet2 TimeSoft phải có header dòng 3 và đủ 11 cột A:K.")
        header_row = next(worksheet.iter_rows(
            min_row=3, max_row=3, min_col=1,
            max_col=len(_payroll.TIMESOFT_PAYROLL_HEADERS), values_only=True,
        ), ())
        actual_headers = [str(value or "").strip() for value in header_row]
        if actual_headers != _payroll.TIMESOFT_PAYROLL_HEADERS:
            mismatches = [
                f"{get_column_letter(index + 1)}: '{actual}' ≠ '{expected}'"
                for index, (actual, expected) in enumerate(zip(actual_headers, _payroll.TIMESOFT_PAYROLL_HEADERS))
                if actual != expected
            ]
            raise HTTPException(
                400,
                f"Sheet2 ({worksheet.title}) không đúng header TimeSoft chuẩn ở dòng 3. "
                + " | ".join(mismatches[:6]),
            )

        selected_rows: list[tuple[Any, Any, Any, Any]] = []
        for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=9, values_only=True):
            padded = tuple(row) + (None,) * max(0, 9 - len(row))
            time_value, item_value, amount_value, employee_value = padded[1], padded[5], padded[6], padded[8]
            if any(value not in (None, "") for value in (time_value, item_value, amount_value, employee_value)):
                selected_rows.append((time_value, item_value, amount_value, employee_value))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được Sheet2 của file TimeSoft: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()

    output = pd.DataFrame(selected_rows, columns=["time", "item", "amount", "employee"])
    numeric_input = output["time"].apply(
        lambda value: isinstance(value, numbers.Number) and not isinstance(value, bool)
    )
    numeric_time = pd.to_numeric(output["time"], errors="coerce")
    output["time"] = pd.to_datetime(output["time"], dayfirst=True, errors="coerce", format="mixed")
    numeric_mask = numeric_input & numeric_time.between(1, 100_000, inclusive="both")
    if numeric_mask.any():
        output.loc[numeric_mask, "time"] = (
            pd.Timestamp("1899-12-30") + pd.to_timedelta(numeric_time[numeric_mask], unit="D")
        )
    output["amount"] = output["amount"].apply(_payroll._number)
    output["item"] = output["item"].fillna("").astype(str).str.strip()
    output["employee"] = output["employee"].fillna("").astype(str).str.strip()
    return output


def _period_from_source(source: pd.DataFrame) -> tuple[str, int, date, date]:
    dated = source.dropna(subset=["time"]).copy()
    if dated.empty:
        raise HTTPException(400, "Sheet2 không có ngày hợp lệ ở cột B để tự nhận Kỳ lương.")
    item_norm = dated["item"].astype(str).str.strip().str.casefold()
    tips = dated[item_norm.str.match(_payroll.TIP_ITEM_PATTERN, na=False)].copy()
    sample = tips if not tips.empty else dated
    groups: Counter[tuple[int, int, int]] = Counter()
    latest: dict[tuple[int, int, int], date] = {}
    for value in sample["time"].tolist():
        d = pd.Timestamp(value).date()
        key = (d.year, d.month, 1 if d.day <= 15 else 2)
        groups[key] += 1
        latest[key] = max(latest.get(key, d), d)
    if not groups:
        raise HTTPException(400, "Không xác định được Kỳ lương từ Sheet2.")
    year, month, period_no = max(groups, key=lambda key: (groups[key], latest[key]))
    month_text = f"{year:04d}-{month:02d}"
    start, end, _ = _payroll._period(month_text, period_no)
    return month_text, period_no, start, end


def install_policy_v39(
    app,
    *,
    engine_instance: Callable[[], Any],
    current_identity,
    require_feature,
    vn_tz,
) -> None:
    if getattr(app.state, "policy_v39_installed", False):
        return

    # All payroll upload/calculation paths now consume physical Sheet2.
    _payroll._read_source = _read_source_sheet2

    original_validate = _shared._validate_and_prepare

    def validate_with_late_month_cap(
        conn,
        body,
        ident,
        *,
        exclude_record_uid="",
        skip_registration_timing=False,
        record_uid="",
        existing_ordinal=None,
    ):
        record, warnings = original_validate(
            conn, body, ident,
            exclude_record_uid=exclude_record_uid,
            skip_registration_timing=skip_registration_timing,
            record_uid=record_uid,
            existing_ordinal=existing_ordinal,
        )
        if str(getattr(ident, "role", "") or "").strip().lower() == "admin":
            return record, warnings
        if float(record.get("calculated_days") or 0) <= 0:
            return record, warnings
        try:
            group_now = _shared._policy_group(conn, record.get("leave_reason", ""))
        except Exception:
            group_now = ""
        if group_now != "co_phep":
            return record, warnings

        employee = conn.execute(text("""
            SELECT employment_start_date, payload
            FROM employees
            WHERE lower(btrim(username))=lower(btrim(:username))
            LIMIT 1
        """), {"username": record.get("employee_name", "")}).mappings().first()
        if not employee:
            return record, warnings
        effective = _effective_start_or_resume(dict(employee))
        target = body.leave_date
        if not effective or effective.year != target.year or effective.month != target.month or effective.day < 16:
            return record, warnings

        rows = conn.execute(text("""
            SELECT leave_reason, calculated_days
            FROM leave_records
            WHERE lower(btrim(employee_name))=lower(btrim(:employee))
              AND leave_date >= :start AND leave_date <= :end
              AND (:uid = '' OR record_uid <> :uid)
        """), {
            "employee": record.get("employee_name", ""),
            "start": target.replace(day=1),
            "end": _payroll._period(f"{target.year:04d}-{target.month:02d}", 2)[1],
            "uid": str(exclude_record_uid or ""),
        }).mappings().all()
        used = 0.0
        for row in rows:
            try:
                if _shared._policy_group(conn, row.get("leave_reason", "")) == "co_phep":
                    used += max(0.0, float(row.get("calculated_days") or 0))
            except Exception:
                continue
        requested = max(0.0, float(record.get("calculated_days") or 0))
        if used + requested > LATE_MONTH_PAID_LIMIT + 1e-9:
            raise HTTPException(
                400,
                f"Nhân viên bắt đầu/làm lại từ {effective.strftime('%d/%m/%Y')} (từ ngày 16). "
                f"Tháng {target.month}/{target.year} chỉ được tối đa 3 ngày nghỉ CÓ phép; "
                f"đã dùng {used:g} ngày, đăng ký thêm {requested:g} ngày.",
            )
        return record, warnings

    _shared._validate_and_prepare = validate_with_late_month_cap
    _shared._api._validate_and_prepare = validate_with_late_month_cap

    # Stamp the exact date a temporarily-away employee is switched back active.
    original_staff_update = _remove_route(app, "/v2/staff/{username}", "PATCH")
    if not callable(original_staff_update):
        raise RuntimeError("Không tìm thấy route sửa nhân viên để cài ngày làm lại.")

    @app.patch("/v2/staff/{username}")
    def update_staff_with_resume_date(username: str, body: _staff.StaffUpdate, ident=Depends(current_identity)):
        with engine_instance().connect() as conn:
            before = conn.execute(text("""
                SELECT COALESCE(payload->>'Trạng thái làm việc', payload->>'employment_status', 'Đang làm việc')
                FROM employees WHERE lower(btrim(username))=lower(btrim(:username)) LIMIT 1
            """), {"username": username}).scalar_one_or_none()
        result = original_staff_update(username=username, body=body, ident=ident)
        requested = getattr(body, "employment_status", None)
        if requested == ACTIVE_STATUS and str(before or ACTIVE_STATUS) != ACTIVE_STATUS:
            resume_text = datetime.now(vn_tz).strftime("%d/%m/%Y")
            with engine_instance().begin() as conn:
                conn.execute(text("""
                    UPDATE employees
                    SET payload=jsonb_set(
                        COALESCE(payload,'{}'::jsonb), '{return_to_work_date}',
                        to_jsonb(CAST(:resume_date AS text)), true
                    ), updated_at=NOW()
                    WHERE lower(btrim(username))=lower(btrim(:username))
                """), {"resume_date": resume_text, "username": username})
        return result

    @app.post("/v2/payroll/detect-period")
    def detect_payroll_period(
        payload: bytes = Body(..., media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ident=Depends(current_identity),
    ):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
        source = _read_source_sheet2(payload)
        month, period_no, start, end = _period_from_source(source)
        return {
            "ok": True,
            "month": month,
            "period_no": period_no,
            "start": start.isoformat(),
            "end": end.isoformat(),
            "sheet_index": 2,
            "sheet_name": "Sheet2",
            "message": f"Đã nhận Kỳ {period_no} - Tháng {int(month[-2:])}/{month[:4]} từ Sheet2.",
        }

    @app.get("/v2/policy-v39/health")
    def policy_v39_health():
        return {"ok": True, "release": RELEASE, "payroll_sheet": "Sheet2", "auto_penalty_minutes": 5}

    app.state.policy_v39_installed = True
