"""Vera Spa daily operations Cloud Run Job - V93.2.

Action:
  python vera_daily_ops_job.py violations
      - Chạy lúc 21:00 ICT.
      - Đọc trực tiếp cột Vào trễ (V) của Bảng tour.
      - Vào trễ >= 5 phút => tự chọn đúng Lý do nghỉ theo LoaiNghi.
      - Nếu chỉ có một trong Giờ ra / Giờ vào => ghi lỗi "Ra ngoài chỉ có dữ liệu một lần"
        (hoặc lý do tương đương tìm được trong LoaiNghi).
      - Chống trùng theo Ngày + Nhân viên + Lý do nghỉ.
      - Gửi email cho nhân viên; CC veraspabienhoa@gmail.com + quanly + letan.

Job này chạy độc lập, không import app.py.
"""

from __future__ import annotations

import json
import math
import os
import re
import smtplib
import sys
import tempfile
import unicodedata
from datetime import date, datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import gspread
import pandas as pd
from google.auth.transport.requests import AuthorizedSession
from google.oauth2.service_account import Credentials
from sqlalchemy import text

import auto_penalty_daily_job as daily_mail
import vera_postgres as vpg


VN_TZ = timezone(timedelta(hours=7))

SHEET_MAT_KHAU_ID = os.getenv(
    "SHEET_MAT_KHAU_ID",
    "1DGXy3kPyMPwtz-3CnG8i6BiQbXFDApasoXVFzSmUe24",
)
SHEET_DU_PHONG_ID = os.getenv(
    "SHEET_DU_PHONG_ID",
    "1Kz0aw-JatptAN9G7YSwZ6rJO09urOPaD-rS-18eZSY0",
)

# TourVera hiện hành; có thể thay đổi an toàn qua biến môi trường khi cần.
BANG_TOUR_FILE_ID = (
    os.getenv("VERA_TOUR_FILE_ID", "15nDSicFhEHstxQjGrETuSK8Z7q6cSQyS")
    or "15nDSicFhEHstxQjGrETuSK8Z7q6cSQyS"
).strip()

BANG_TOUR_AUDIT_WORKSHEET = os.getenv(
    "BANG_TOUR_AUDIT_WORKSHEET",
    "DoiSoatRaNgoai",
)
AUTO_LATE_MINUTES = max(
    5,
    int(os.getenv("AUTO_LATE_MINUTES", "5") or 5),
)

BANG_TOUR_SINGLE_SIDE_REASON = (
    os.getenv("BANG_TOUR_SINGLE_SIDE_REASON", "Ra ngoài chỉ có dữ liệu một lần")
    or "Ra ngoài chỉ có dữ liệu một lần"
).strip()

SMTP_SENDER_EMAIL = "veraspabienhoa@gmail.com"
SMTP_APP_PASSWORD = (os.getenv("SMTP_APP_PASSWORD", "") or "").strip()
AUTO_CC_EMAIL = "veraspabienhoa@gmail.com"

AUDIT_HEADERS = [
    "Ngày",
    "Tên nhân viên",
    "Loại vi phạm",
    "Số phút",
    "Mức phạt",
    "Ghi phiếu phạt",
    "Email",
    "Chi tiết",
    "Cập nhật lúc",
    "Người cập nhật",
]


PRIMARY_LEAVE_HEADERS = [
    "Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ", "Loại nghỉ",
    "Chi tiết", "Số ngày tính", "Số ngày phép cộng dồn", "Phạt vi phạm",
    "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật",
]
LEAVE_HEADER_RANGE = "A1:L1"
LEAVE_DATA_RANGE = "A:L"



def vn_now() -> datetime:
    return datetime.now(VN_TZ)


def weekday_vi(d: date) -> str:
    return {
        0: "Thứ Hai", 1: "Thứ Ba", 2: "Thứ Tư", 3: "Thứ Năm",
        4: "Thứ Sáu", 5: "Thứ Bảy", 6: "Chủ Nhật",
    }.get(d.weekday(), "") if isinstance(d, date) else ""


def ensure_leave_header(ws) -> None:
    current = ws.get(LEAVE_HEADER_RANGE)
    row = list(current[0]) if current else []
    row += [""] * max(0, 12 - len(row))
    normalized = [str(x).strip() for x in row[:12]]
    expected = [str(x).strip() for x in PRIMARY_LEAVE_HEADERS]
    if normalized == expected:
        return
    if not any(normalized):
        ws.update(
            LEAVE_HEADER_RANGE,
            [PRIMARY_LEAVE_HEADERS],
            value_input_option="USER_ENTERED",
        )
        return

    legacy = ws.get("A1:M1")
    legacy_row = list(legacy[0]) if legacy else []
    legacy_row += [""] * max(0, 13 - len(legacy_row))
    legacy_expected = [
        "Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ", "", "Loại nghỉ",
        "Chi tiết", "Số ngày tính", "Số ngày phép cộng dồn", "Phạt vi phạm",
        "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật",
    ]
    if [str(x).strip() for x in legacy_row[:13]] == legacy_expected:
        raise RuntimeError("Sheet1 vẫn đang ở schema A:M cũ. Hãy chạy migration A:M -> A:L trước khi chạy Job 21:00.")
    raise RuntimeError(
        "Header Sheet1 không đúng schema A:L. Không tự sửa để tránh lệch dữ liệu; "
        "hãy chạy migrate_leave_sheet_to_AL.py trước."
    )


def normalize_text(value) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(
        ch for ch in text if unicodedata.category(ch) != "Mn"
    )
    text = text.replace("đ", "d").replace("Đ", "D")
    return " ".join(text.strip().split()).casefold()


def clean_employee_name(value) -> str:
    """Cẩm Nhung * -> Cẩm Nhung."""
    text = str(value or "").strip()
    text = re.sub(r"\s*\*+\s*$", "", text).strip()
    return text


def parse_money(value) -> float:
    if value is None:
        return 0.0
    text = (
        str(value)
        .strip()
        .replace("đ", "")
        .replace("Đ", "")
        .replace("VND", "")
        .replace("VNĐ", "")
        .replace(" ", "")
    )
    if not text or text.casefold() in {"nan", "none", "-"}:
        return 0.0
    text = text.replace(".", "").replace(",", "")
    try:
        return float(text)
    except Exception:
        return 0.0


def _credentials(scopes):
    env_json = (os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "") or "").strip()
    if env_json:
        return Credentials.from_service_account_info(
            json.loads(env_json),
            scopes=scopes,
        )
    import google.auth

    creds, _ = google.auth.default(scopes=scopes)
    return creds


def gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    return gspread.authorize(_credentials(scopes))


def drive_session():
    return AuthorizedSession(
        _credentials(["https://www.googleapis.com/auth/drive.readonly"])
    )


def download_drive_file(file_id: str, path: str):
    session = drive_session()
    url = (
        f"https://www.googleapis.com/drive/v3/files/{file_id}"
        "?alt=media&supportsAllDrives=true"
    )
    r = session.get(url, timeout=120, stream=True)
    if r.status_code != 200:
        raise RuntimeError(
            f"Drive download HTTP {r.status_code}: {str(r.text)[:400]}"
        )
    with open(path, "wb") as f:
        for chunk in r.iter_content(1024 * 1024):
            if chunk:
                f.write(chunk)
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        raise RuntimeError("Drive returned an empty file.")


def parse_datetime(
    value,
    fallback_date: date | None = None,
) -> datetime | None:
    if value is None or value == "":
        return None

    if isinstance(value, pd.Timestamp):
        dt = value.to_pydatetime()
    elif isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    else:
        parsed = pd.to_datetime(
            str(value).strip(),
            dayfirst=True,
            errors="coerce",
        )
        if pd.isna(parsed):
            return None
        dt = (
            parsed.to_pydatetime()
            if isinstance(parsed, pd.Timestamp)
            else parsed
        )

    if dt.tzinfo is not None:
        dt = dt.astimezone(VN_TZ).replace(tzinfo=None)

    if dt.year in {1899, 1900, 1970} and fallback_date:
        dt = datetime.combine(fallback_date, dt.time())

    return dt


def parse_late_minutes(value) -> float | None:
    """
    Đọc cột Vào trễ:
    - số phút: 18 -> 18 phút;
    - timedelta / pandas Timedelta;
    - time;
    - chuỗi HH:MM[:SS];
    - chuỗi "18 phút".
    """
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, timedelta):
        return max(0.0, value.total_seconds() / 60.0)

    if isinstance(value, pd.Timedelta):
        return max(0.0, value.total_seconds() / 60.0)

    # Excel/openpyxl có thể trả datetime.time.
    if hasattr(value, "hour") and hasattr(value, "minute") and not isinstance(
        value, (datetime, date)
    ):
        try:
            return max(
                0.0,
                float(value.hour * 60 + value.minute + value.second / 60.0),
            )
        except Exception:
            pass

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        n = float(value)
        if math.isnan(n):
            return None

        # Nếu Excel lưu duration dạng fraction of day (ví dụ 18 phút = 0.0125),
        # đổi sang phút. Nếu là số nguyên/lớn hơn 1, hiểu là số phút trực tiếp.
        if 0 < abs(n) < 1:
            n = n * 24 * 60
        return max(0.0, n)

    s = str(value).strip()
    if not s or s.casefold() in {"nan", "none", "nat", "<na>"}:
        return None

    # "18 phút", "18 phut", "18"
    m = re.fullmatch(
        r"\s*(-?\d+(?:[.,]\d+)?)\s*(?:phút|phut|min|minutes?)?\s*",
        s,
        flags=re.IGNORECASE,
    )
    if m:
        try:
            return max(
                0.0,
                float(m.group(1).replace(",", ".")),
            )
        except Exception:
            pass

    # HH:MM[:SS] là duration.
    m = re.search(
        r"(-?\d{1,3}):(\d{2})(?::(\d{2}))?",
        s,
    )
    if m:
        try:
            h = int(m.group(1))
            mi = int(m.group(2))
            sec = int(m.group(3) or 0)
            sign = -1 if h < 0 else 1
            total = sign * (
                abs(h) * 60 + mi + sec / 60.0
            )
            return max(0.0, total)
        except Exception:
            pass

    return None


def open_bang_tour(data_only=False):
    from openpyxl import load_workbook

    fd, path = tempfile.mkstemp(suffix=".xlsm")
    os.close(fd)
    download_drive_file(BANG_TOUR_FILE_ID, path)

    wb = load_workbook(
        path,
        keep_vba=True,
        data_only=data_only,
    )

    if "Input" not in wb.sheetnames:
        try:
            os.remove(path)
        except Exception:
            pass
        raise RuntimeError(
            "Không tìm thấy sheet Input trong Bảng tour."
        )

    return wb, wb["Input"], path


def leave_catalog(client) -> dict:
    """
    LoaiNghi:
      B = Lý do nghỉ
      C = Loại nghỉ
      F = Phạt vi phạm (theo cấu trúc hiện tại).
    """
    ss = client.open_by_key(SHEET_DU_PHONG_ID)
    try:
        ws = ss.worksheet("LoaiNghi")
    except Exception:
        return {}

    vals = ws.get_all_values()
    catalog = {}

    for row in vals:
        if len(row) < 2:
            continue

        name = str(row[1] or "").strip()
        if (
            not name
            or normalize_text(name)
            in {"loai nghi", "ly do nghi"}
        ):
            continue

        leave_type = (
            str(row[2] or "").strip()
            if len(row) > 2
            else ""
        )
        penalty = parse_money(
            row[5] if len(row) > 5 else 0
        )

        catalog[normalize_text(name)] = {
            "name": name,
            "type": leave_type,
            "penalty": penalty,
        }

    return catalog


def _catalog_reason(catalog: dict, preferred: list[str]) -> str:
    """
    Tìm tên lý do đúng theo sheet LoaiNghi.
    Ưu tiên exact normalize, sau đó contains.
    """
    for wanted in preferred:
        key = normalize_text(wanted)
        item = catalog.get(key)
        if item:
            return str(item.get("name", wanted)).strip()

    for wanted in preferred:
        key = normalize_text(wanted)
        for item_key, item in catalog.items():
            if key and (
                key in item_key
                or item_key in key
            ):
                return str(item.get("name", wanted)).strip()

    return preferred[0]


def pick_late_reason(
    catalog: dict,
    late_minutes: int,
) -> str:
    """
    Quy tắc 21:00 đồng nhất với Auto Bảng tour:
      5-29  -> dưới 30
      30-59 -> dưới 60
      60-120 -> dưới 120
      >120 -> trên 120
    """
    m = int(math.ceil(float(late_minutes)))

    if m < 30:
        return _catalog_reason(
            catalog,
            ["Ra ngoài vào muộn dưới 30 phút"],
        )

    if m < 60:
        return _catalog_reason(
            catalog,
            ["Ra ngoài vào muộn dưới 60 phút"],
        )

    if m <= 120:
        return _catalog_reason(
            catalog,
            ["Ra ngoài vào muộn dưới 120 phút"],
        )

    return _catalog_reason(
        catalog,
        ["Ra ngoài vào muộn trên 120 phút"],
    )


def pick_single_side_reason(catalog: dict) -> str:
    preferred = [
        "Ra ngoài chỉ có dữ liệu một lần",
        "Ra ngoài thiếu giờ ra/vào",
    ]

    for wanted in preferred:
        item = catalog.get(normalize_text(wanted))
        if item:
            return str(item.get("name", wanted)).strip()

    token_pairs = [
        ("ra ngoai", "mot lan"),
        ("ra ngoai", "1 lan"),
        ("ra ngoai", "thieu"),
        ("ra ngoai", "khong du"),
        ("ra ngoai", "khong co gio"),
    ]

    for item in catalog.values():
        name = str(item.get("name", "")).strip()
        key = normalize_text(name)
        if any(
            all(token in key for token in pair)
            for pair in token_pairs
        ):
            return name

    return BANG_TOUR_SINGLE_SIDE_REASON


def penalty_for(
    catalog: dict,
    reason: str,
) -> float:
    return float(
        (
            catalog.get(normalize_text(reason))
            or {}
        ).get("penalty", 0)
        or 0
    )


def get_or_create_ws(
    spreadsheet,
    title,
    rows=3000,
    cols=12,
):
    try:
        return spreadsheet.worksheet(title)
    except Exception:
        return spreadsheet.add_worksheet(
            title=title,
            rows=rows,
            cols=cols,
        )


def audit_ws(client):
    ss = client.open_by_key(SHEET_MAT_KHAU_ID)
    ws = get_or_create_ws(
        ss,
        BANG_TOUR_AUDIT_WORKSHEET,
        rows=3000,
        cols=12,
    )

    current = ws.row_values(1)
    if current[: len(AUDIT_HEADERS)] != AUDIT_HEADERS:
        ws.update(
            "A1:J1",
            [AUDIT_HEADERS],
            value_input_option="USER_ENTERED",
        )
    return ws


def sent_keys(client) -> set:
    ws = audit_ws(client)
    vals = ws.get_all_values()
    if len(vals) <= 1:
        return set()

    idx = {h: i for i, h in enumerate(vals[0])}
    result = set()

    for row in vals[1:]:
        def cell(name):
            i = idx.get(name, -1)
            return (
                row[i]
                if i >= 0 and i < len(row)
                else ""
            )

        if normalize_text(cell("Email")) not in {
            "1",
            "true",
            "yes",
            "da gui",
        }:
            continue

        result.add(
            (
                cell("Ngày"),
                normalize_text(
                    clean_employee_name(
                        cell("Tên nhân viên")
                    )
                ),
                normalize_text(cell("Loại vi phạm")),
            )
        )

    return result


def employee_directory():
    """
    TO = email nhân viên.
    CC = veraspabienhoa@gmail.com + quanly + letan.
    Không tự CC toàn bộ admin.
    Hồ sơ nhân viên chỉ đọc từ PostgreSQL; Sheet1 không còn là nguồn.
    """
    if not vpg.is_enabled():
        raise RuntimeError("PostgreSQL chưa được cấu hình cho danh bạ nhân viên.")
    with vpg.get_engine().connect() as conn:
        vals = conn.execute(text("""
            SELECT username, role, email
            FROM employees
            WHERE btrim(COALESCE(username, '')) <> ''
              AND COALESCE(payload->>'__deleted', 'false') <> 'true'
            ORDER BY COALESCE(stt, 2147483647), username
        """)).mappings().all()

    emails = {}
    canonical_names = {}
    cc = [AUTO_CC_EMAIL]

    for row in vals:
        name = str(row.get("username") or "").strip()
        role = str(row.get("role") or "").strip().lower()
        email = str(row.get("email") or "").strip()

        key = normalize_text(name)
        if name:
            canonical_names[key] = name

        if name and "@" in email:
            emails[key] = email

        if (
            role in {"letan", "quanly"}
            and "@" in email
        ):
            cc.append(email)

    dedup = []
    seen = set()
    for email in cc:
        key = email.casefold()
        if email and key not in seen:
            dedup.append(email)
            seen.add(key)

    return emails, canonical_names, dedup


def existing_violation_keys(client) -> set:
    ws = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
    ensure_leave_header(ws)
    vals = ws.get(LEAVE_DATA_RANGE)

    out = set()
    for row in vals[1:]:
        row = list(row[:12]) + [""] * max(0, 12 - len(row))
        if not any(str(v).strip() for v in row):
            continue
        out.add(
            (
                str(row[0]).strip(),
                normalize_text(clean_employee_name(row[2])),
                normalize_text(row[3]),
            )
        )

    return out

def append_violation(
    client,
    target_date: date,
    employee: str,
    reason: str,
    detail: str,
    penalty: float,
) -> tuple[bool, str]:
    """Ghi Job 21:00 vào Sheet1 A:L; E là Loại nghỉ từ LoaiNghi."""
    ws = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)

    date_text = target_date.strftime("%d/%m/%Y")
    employee = clean_employee_name(employee)

    key = (
        date_text,
        normalize_text(employee),
        normalize_text(reason),
    )
    if key in existing_violation_keys(client):
        return False, "Đã tồn tại"

    catalog = leave_catalog(client)
    leave_type = str(
        (catalog.get(normalize_text(reason)) or {}).get("type", "")
        or ""
    ).strip()

    # Bảo đảm header A:L mới.
    ensure_leave_header(ws)

    now = vn_now()
    ws.append_row(
        [
            date_text,                         # A Ngày
            weekday_vi(target_date),           # B Thứ ngày
            employee,                          # C Tên nhân viên
            reason,                            # D Lý do nghỉ
            leave_type,                        # E Loại nghỉ
            detail,                            # F Chi tiết
            0,                                 # G Số ngày tính
            0,                                 # H Số ngày phép cộng dồn
            float(penalty or 0),               # I Phạt vi phạm
            now.strftime("%d/%m/%Y"),          # J Ngày cập nhật
            now.strftime("%H:%M:%S"),          # K Giờ cập nhật
            "AUTO UPDATE 21:00 - BẢNG TOUR",   # L Người cập nhật
        ],
        value_input_option="USER_ENTERED",
    )
    return True, "Đã ghi"


def send_email(
    to_email: str,
    cc: list[str],
    employee: str,
    target_date: date,
    rows: list[dict],
) -> tuple[bool, str]:
    if not SMTP_SENDER_EMAIL or not SMTP_APP_PASSWORD:
        return (
            False,
            "SMTP_SENDER_EMAIL/SMTP_APP_PASSWORD chưa cấu hình.",
        )

    if not to_email or "@" not in to_email:
        return (
            False,
            "Nhân viên chưa có email hợp lệ.",
        )

    cc = [
        e
        for e in cc
        if e
        and "@" in e
        and e.casefold() != to_email.casefold()
    ]

    detail_html = "".join(
        "<tr>"
        f"<td style='padding:6px;border:1px solid #ddd'>{r['reason']}</td>"
        f"<td style='padding:6px;border:1px solid #ddd'>{int(r.get('minutes', 0) or 0)}</td>"
        f"<td style='padding:6px;border:1px solid #ddd'>{float(r.get('penalty', 0) or 0):,.0f} VNĐ</td>"
        f"<td style='padding:6px;border:1px solid #ddd'>{r.get('detail', '')}</td>"
        "</tr>"
        for r in rows
    )

    html = f"""
    <html>
    <body style='font-family:Arial,sans-serif;color:#222'>
      <p>Chào <b>{employee}</b>,</p>

      <p>
        Hệ thống Vera Spa ghi nhận vi phạm Bảng tour ngày
        <b>{target_date.strftime('%d/%m/%Y')}</b>:
      </p>

      <table style='border-collapse:collapse;width:100%'>
        <thead>
          <tr>
            <th style='padding:6px;border:1px solid #ddd'>Vi phạm</th>
            <th style='padding:6px;border:1px solid #ddd'>Số phút</th>
            <th style='padding:6px;border:1px solid #ddd'>Mức phạt</th>
            <th style='padding:6px;border:1px solid #ddd'>Chi tiết</th>
          </tr>
        </thead>
        <tbody>{detail_html}</tbody>
      </table>

      <p>
        Nếu dữ liệu chưa chính xác, vui lòng phản hồi với Lễ tân/Quản lý.
      </p>

      <p>
        Trân trọng,<br>
        <b>VERA SPA</b><br>
        <b>Địa chỉ:</b> 193 Trương Định, Tam Hiệp, Đồng Nai<br>
        <b>Điện Thoại:</b> 0833229939
      </p>
    </body>
    </html>
    """

    msg = MIMEMultipart()
    msg["From"] = (
        f"Vera Spa <{SMTP_SENDER_EMAIL}>"
    )
    msg["To"] = to_email

    if cc:
        msg["Cc"] = ", ".join(cc)

    msg["Subject"] = (
        f"[VERA SPA] Auto Update Bảng tour 21:00 "
        f"{target_date.strftime('%d/%m/%Y')} - {employee}"
    )
    msg.attach(MIMEText(html, "html"))

    server = smtplib.SMTP(
        "smtp.gmail.com",
        587,
        timeout=30,
    )

    try:
        server.starttls()
        server.login(
            SMTP_SENDER_EMAIL,
            SMTP_APP_PASSWORD,
        )
        server.send_message(msg)
    finally:
        try:
            server.quit()
        except Exception:
            pass

    return True, "Đã gửi"


def collect_violations(
    target_date: date,
    catalog: dict,
    canonical_names: dict,
) -> list[dict]:
    """
    Input columns:
      B = NV
      S = Giờ ra
      U = Giờ vào
      V = Vào trễ

    Quy tắc chính:
      Vào trễ >= 5 phút -> Auto Update theo ngưỡng.

    Quy tắc phụ:
      Nếu chỉ có đúng một mốc Giờ ra/Giờ vào trong ngày
      -> "Ra ngoài chỉ có dữ liệu một lần".
    """
    wb, ws, path = open_bang_tour(
        data_only=False
    )
    results = []

    try:
        for row in range(21, 501):
            raw_employee = str(
                ws[f"B{row}"].value or ""
            ).strip()

            if not raw_employee:
                continue

            cleaned = clean_employee_name(
                raw_employee
            )
            employee = canonical_names.get(
                normalize_text(cleaned),
                cleaned,
            )

            out_dt = parse_datetime(
                ws[f"S{row}"].value,
                fallback_date=target_date,
            )
            in_dt = parse_datetime(
                ws[f"U{row}"].value,
                fallback_date=(
                    out_dt.date()
                    if out_dt
                    else target_date
                ),
            )

            out_matches = (
                out_dt is not None
                and out_dt.date() == target_date
            )
            in_matches = (
                in_dt is not None
                and in_dt.date() == target_date
            )

            # Không xử lý dòng lịch sử không thuộc hôm nay.
            if not out_matches and not in_matches:
                continue

            late_value = ws[f"V{row}"].value
            late = parse_late_minutes(
                late_value
            )

            # Ưu tiên tuyệt đối cột Vào trễ.
            if (
                late is not None
                and float(late) >= AUTO_LATE_MINUTES
            ):
                minutes = int(
                    math.ceil(float(late))
                )
                reason = pick_late_reason(
                    catalog,
                    minutes,
                )

                detail_bits = [
                    (
                        "Auto Update 21:00 · Bảng tour"
                        f" · Vào trễ {minutes} phút"
                    )
                ]

                if out_dt is not None:
                    detail_bits.append(
                        "Giờ ra "
                        + out_dt.strftime("%H:%M:%S")
                    )

                if in_dt is not None:
                    detail_bits.append(
                        "Giờ vào "
                        + in_dt.strftime("%H:%M:%S")
                    )

                results.append(
                    {
                        "employee": employee,
                        "reason": reason,
                        "minutes": minutes,
                        "penalty": penalty_for(
                            catalog,
                            reason,
                        ),
                        "detail": " · ".join(
                            detail_bits
                        ),
                    }
                )
                continue

            # Nếu không có vi phạm Vào trễ >= 5,
            # vẫn kiểm tra trường hợp chỉ có 1 mốc ra/vào.
            if out_matches != in_matches:
                reason = pick_single_side_reason(
                    catalog
                )
                missing = (
                    "Giờ vào"
                    if out_matches
                    else "Giờ ra"
                )

                detail = (
                    "Auto Update 21:00 · Bảng tour"
                    f" · chỉ có một mốc thời gian"
                    f" · thiếu {missing}"
                )

                results.append(
                    {
                        "employee": employee,
                        "reason": reason,
                        "minutes": 0,
                        "penalty": penalty_for(
                            catalog,
                            reason,
                        ),
                        "detail": detail,
                    }
                )

        return results

    finally:
        try:
            wb.close()
        except Exception:
            pass

        try:
            os.remove(path)
        except Exception:
            pass


def process_violations() -> str:
    client = gspread_client()
    target_date = vn_now().date()
    catalog = leave_catalog(client)

    emails, canonical_names, cc = employee_directory()

    violations = collect_violations(
        target_date,
        catalog,
        canonical_names,
    )

    already_sent = sent_keys(client)
    audit = audit_ws(client)

    grouped = {}
    audit_rows = []
    saved_count = 0

    for item in violations:
        ok, save_msg = append_violation(
            client,
            target_date,
            item["employee"],
            item["reason"],
            item["detail"],
            item["penalty"],
        )

        if ok:
            saved_count += 1

        key = (
            target_date.strftime("%d/%m/%Y"),
            normalize_text(
                clean_employee_name(
                    item["employee"]
                )
            ),
            normalize_text(item["reason"]),
        )

        if key not in already_sent:
            grouped.setdefault(
                item["employee"],
                [],
            ).append(item)

        audit_rows.append(
            {
                "Ngày": target_date.strftime(
                    "%d/%m/%Y"
                ),
                "Tên nhân viên": item["employee"],
                "Loại vi phạm": item["reason"],
                "Số phút": item["minutes"],
                "Mức phạt": item["penalty"],
                "Ghi phiếu phạt": (
                    "1"
                    if ok
                    else "Đã tồn tại"
                ),
                "Email": "",
                "Chi tiết": (
                    item["detail"]
                    + " | "
                    + save_msg
                ),
                "Cập nhật lúc": vn_now().strftime(
                    "%d/%m/%Y %H:%M:%S"
                ),
                "Người cập nhật": (
                    "AUTO UPDATE 21:00 - BẢNG TOUR"
                ),
            }
        )

    # V86.12: email được gửi bởi cơ chế thống nhất trong auto_penalty_daily_job.
    # Cơ chế này quét tất cả Auto Update phạt hôm nay chưa có log SENT,
    # nên không gửi trùng và có thể retry email lỗi.
    try:
        unified_mail = daily_mail.send_pending_auto_penalty_notifications(
            client,
            target_date=target_date,
        )
        sent_count = int(unified_mail.get("sent", 0) or 0)
        failed_count = int(unified_mail.get("failed", 0) or 0)
    except Exception as exc:
        sent_count = 0
        failed_count = 1
        print(f"EMAIL UNIFIED ERROR: {type(exc).__name__}: {exc}")

    zero_penalty = sum(
        1
        for x in violations
        if float(x.get("penalty", 0) or 0)
        <= 0
    )
    return (
        "violations success: "
        f"detected={len(violations)}, "
        f"saved={saved_count}, "
        f"emails={sent_count}, "
        f"zero_penalty={zero_penalty}, "
        f"tour_file={BANG_TOUR_FILE_ID}, "
        f"late_threshold={AUTO_LATE_MINUTES}"
    )


def main():
    action = (
        sys.argv[1]
        if len(sys.argv) > 1
        else os.getenv(
            "DAILY_OPS_ACTION",
            "",
        )
    ).strip().lower()

    if action in {
        "violations",
        "violation",
        "21:00",
        "2100",
    }:
        print(process_violations())
        return

    raise SystemExit(
        "Usage: python vera_daily_ops_job.py violations"
    )


if __name__ == "__main__":
    main()
