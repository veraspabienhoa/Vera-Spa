"""Authenticated Web V2 control/status API for PostgreSQL Auto Check."""
from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any, Callable
from urllib.parse import quote

from fastapi import Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

import vera_auto_check as core


class AutoCheckConfig(BaseModel):
    status: str | None = None
    threshold_minutes: int | None = None


EXPORT_COLUMNS = [
    ("work_date", "Ngày"),
    ("employee_name", "Nhân viên"),
    ("reason", "Lý do"),
    ("source", "Nguồn"),
    ("minutes", "Phút"),
    ("status", "Trạng thái"),
    ("detail", "Chi tiết"),
    ("created_at", "Thời gian ghi nhận"),
]


def _date_range(start: date | None, end: date | None, *, required: bool = False) -> tuple[date | None, date | None]:
    if required and (start is None or end is None):
        raise HTTPException(status_code=400, detail="Vui lòng chọn đủ Từ ngày và Đến ngày.")
    if (start is None) != (end is None):
        raise HTTPException(status_code=400, detail="Vui lòng chọn đủ Từ ngày và Đến ngày.")
    if start is not None and end is not None and end < start:
        raise HTTPException(status_code=400, detail="Đến ngày phải bằng hoặc sau Từ ngày.")
    return start, end


def _workbook(events: list[dict], start: date, end: date) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Auto Check"
    sheet.append([label for _, label in EXPORT_COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F513F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for event in events:
        sheet.append([event.get(key, "") for key, _ in EXPORT_COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 28, 30, 20, 11, 16, 55, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("Thông tin")
    summary.append(["Báo cáo", "AUTO CHECK"])
    summary.append(["Từ ngày", start.isoformat()])
    summary.append(["Đến ngày", end.isoformat()])
    summary.append(["Số dòng", len(events)])
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 24
    summary["A1"].font = Font(bold=True)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def install_auto_check_routes(app, *, engine_instance: Callable[[], Any], current_identity, require_feature, identity_type):
    @app.get("/v2/auto-check")
    def auto_check_dashboard(
        limit: int = 100,
        start: date | None = Query(default=None),
        end: date | None = Query(default=None),
        identity: identity_type = Depends(current_identity),
    ):
        start, end = _date_range(start, end)
        with engine_instance().begin() as conn:
            require_feature(conn, identity, "auto_penalty")
            return core.dashboard(conn, limit, start=start, end=end)

    @app.get("/v2/auto-check/export.xlsx")
    def export_auto_check(
        start: date = Query(...),
        end: date = Query(...),
        identity: identity_type = Depends(current_identity),
    ):
        start, end = _date_range(start, end, required=True)
        with engine_instance().begin() as conn:
            require_feature(conn, identity, "auto_penalty")
            events = core.event_rows(conn, start=start, end=end, limit=None)
        content = _workbook(events, start, end)
        filename = f"VERA_Auto_Check_{start.isoformat()}_{end.isoformat()}.xlsx"
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    @app.put("/v2/auto-check/config")
    def update_auto_check_config(body: AutoCheckConfig, identity: identity_type = Depends(current_identity)):
        updates = body.model_dump(exclude_none=True) if hasattr(body, "model_dump") else body.dict(exclude_none=True)
        if "status" in updates and str(updates["status"]).upper() not in {"RUNNING", "PAUSED"}:
            raise HTTPException(status_code=422, detail="Trạng thái chỉ nhận RUNNING hoặc PAUSED.")
        with engine_instance().begin() as conn:
            require_feature(conn, identity, "auto_penalty_control")
            return {"ok": True, "config": core.save_config(conn, updates, identity.employee_username)}

    @app.post("/v2/auto-check/run")
    def request_auto_check_run(identity: identity_type = Depends(current_identity)):
        with engine_instance().begin() as conn:
            require_feature(conn, identity, "auto_penalty_run")
            cfg = core.load_config(conn)
            if cfg.get("status") == "PAUSED":
                raise HTTPException(status_code=409, detail="Auto Check đang tạm dừng. Hãy mở lại trước khi chạy.")
            cfg = core.save_config(conn, {"manual_run_requested": True}, identity.employee_username)
        return {"ok": True, "queued": True, "message": "Đã xếp hàng. Job gần nhất sẽ chạy Auto Check.", "config": cfg}
