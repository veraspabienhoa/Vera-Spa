from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook

import vera_auto_check as core
from vera_web_v2_auto_check import _date_range, _workbook


class _Result:
    def __init__(self, rows=None):
        self._rows = rows or []

    def mappings(self):
        return self._rows


class _Connection:
    def __init__(self):
        self.calls = []

    def execute(self, statement, params=None):
        sql = str(statement)
        self.calls.append((sql, params or {}))
        if "FROM vera_auto_check_event" in sql:
            return _Result([{
                "work_date": date(2026, 8, 31),
                "employee_name": "Nhân viên A",
                "reason": "Đi trễ không phép",
                "source": "TimeSoft",
                "minutes": Decimal("12"),
                "status": "added",
                "detail": "Trễ 12 phút",
                "created_at": date(2026, 8, 31),
            }])
        return _Result()


def test_event_rows_filters_by_work_date_without_export_limit():
    conn = _Connection()

    rows = core.event_rows(
        conn,
        start=date(2026, 8, 1),
        end=date(2026, 8, 31),
        limit=None,
    )

    query, params = next(call for call in conn.calls if "FROM vera_auto_check_event" in call[0])
    assert "work_date >= :start" in query
    assert "work_date <= :end" in query
    assert " LIMIT " not in query
    assert params == {"start": date(2026, 8, 1), "end": date(2026, 8, 31)}
    assert rows[0]["work_date"] == "2026-08-31"
    assert rows[0]["minutes"] == 12.0


def test_auto_check_workbook_contains_filtered_rows_and_summary():
    content = _workbook([{
        "work_date": "2026-08-31",
        "employee_name": "Nhân viên A",
        "reason": "Đi trễ không phép",
        "source": "TimeSoft",
        "minutes": 12,
        "status": "added",
        "detail": "Trễ 12 phút",
        "created_at": "2026-08-31T08:12:00+07:00",
    }], date(2026, 8, 1), date(2026, 8, 31))

    workbook = load_workbook(BytesIO(content), data_only=True)
    assert workbook.sheetnames == ["Auto Check", "Thông tin"]
    assert workbook["Auto Check"]["A2"].value == "2026-08-31"
    assert workbook["Auto Check"]["B2"].value == "Nhân viên A"
    assert workbook["Thông tin"]["B4"].value == 1


def test_auto_check_date_range_rejects_reversed_dates():
    try:
        _date_range(date(2026, 9, 1), date(2026, 8, 31), required=True)
    except HTTPException as exc:
        assert exc.status_code == 400
        assert "Đến ngày" in exc.detail
    else:
        raise AssertionError("Expected a reversed Auto Check range to be rejected")


def test_single_device_logout_guard_is_removed_from_web_v2():
    root = Path(__file__).resolve().parents[1]
    app_source = (root / "web-v2/src/App.jsx").read_text(encoding="utf-8")
    login_source = (root / "web-v2/src/pages/LoginPage.jsx").read_text(encoding="utf-8")
    api_source = (root / "vera_web_v2_api_v38.py").read_text(encoding="utf-8")
    workflow_source = (root / ".github/workflows/letan-leave-guard.yml").read_text(encoding="utf-8")

    combined = "\n".join((app_source, login_source, api_source))
    assert "vera-device-conflict" not in combined
    assert "claimCurrentDevice" not in combined
    assert "install_single_device_guard" not in combined
    assert "assert 'install_single_device_guard(' not in source" in workflow_source
    assert "signOut({ scope: 'local' })" in app_source
    assert not (root / "web-v2/src/lib/deviceSession.js").exists()
    assert not (root / "vera_web_v2_single_device.py").exists()
