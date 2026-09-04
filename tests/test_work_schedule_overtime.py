from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from vera_web_v2_work_schedule import ScheduleRow, _combo_import_rows, _validate_row


ROOT = Path(__file__).resolve().parents[1]
DEFINITIONS = {
    "locker": {"Ca 1": {}, "Ca 2": {}},
    "letan": {"Ca 1": {}, "Ca 2": {}},
    "tapvu": {"Ca 1": {}, "Ca 2": {}},
}


def row(department, **updates):
    values = {
        "work_date": date(2026, 9, 3),
        "employee_username": "Test",
        "employee_name": "Test",
        "department": department,
        "shift_code": "Giờ làm" if department == "quanly" else "Ca 1",
        "start_time": "09:00" if department == "quanly" else "",
        "end_time": "17:00" if department == "quanly" else "",
    }
    values.update(updates)
    return ScheduleRow(**values)


def test_all_four_departments_accept_shift_overtime():
    for department in ("quanly", "locker", "letan", "tapvu"):
        normalized = _validate_row(row(department, overtime_shift="TC Ca 2"), DEFINITIONS)
        assert normalized[2] == "TC Ca 2"


def test_all_four_departments_accept_custom_overtime_range():
    for department in ("quanly", "locker", "letan", "tapvu"):
        normalized = _validate_row(row(
            department,
            overtime_shift="Từ giờ tới giờ",
            overtime_start_time="20:00",
            overtime_end_time="01:00",
        ), DEFINITIONS)
        assert normalized[2:] == ("Từ giờ tới giờ", "20:00", "01:00")


def test_schedule_page_has_monthly_statistics_and_clickable_total_highlight():
    source = (ROOT / "web-v2/src/pages/WorkSchedulePage.jsx").read_text(encoding="utf-8")
    assert "THỐNG KÊ THÁNG" in source
    assert "employeeMatchesTotal" in source
    assert "Từ giờ tới giờ" in source
    assert "Tổng bộ phận" in source
    assert "row.work_date <= yesterdayIso" in source
    assert "mobile-cell-summary ${mobileShiftClass}" in source
    assert "BÁN COMBO" in source
    assert "combo_customer_phone" in source


def test_schedule_persists_combo_sale_details_for_management_and_reception():
    source = (ROOT / "vera_web_v2_work_schedule.py").read_text(encoding="utf-8")
    for field in (
        "combo_sold", "combo_sale_date", "combo_customer_name",
        "combo_customer_phone", "combo_ticket", "combo_note",
    ):
        assert field in source
    assert 'row.department in {"quanly", "letan"}' in source
    assert 'vera_work_schedule_combo_sale' in source
    assert '@app.get("/v2/work-schedule/combo-sales")' in source
    assert '@app.post("/v2/work-schedule/combo-sales")' in source
    assert '@app.put("/v2/work-schedule/combo-sales/{sale_id}")' in source
    assert '@app.delete("/v2/work-schedule/combo-sales/{sale_id}")' in source
    assert '@app.get("/v2/work-schedule/combo-sales/export.xlsx")' in source
    assert '@app.post("/v2/work-schedule/combo-sales/import.xlsx")' in source
    assert 'COMBO_EDITOR_ROLES = {"admin", "quanly", "letan"}' in source


def test_combo_excel_import_reads_each_employee_sheet():
    workbook = Workbook()
    first = workbook.active
    first.title = "Anh Nguyễn"
    first.append(["ID", "Ngày bán", "Tên hệ thống", "Nhân viên", "Tên khách hàng", "Số điện thoại", "Vé combo", "Ghi chú"])
    first.append(["sale-1", date(2026, 4, 9), "Anh Nguyễn", "Anh Nguyễn", "Khách A", "0901000001", "Combo 10", "Đã cọc"])
    second = workbook.create_sheet("Bình")
    second.append(["ID", "Ngày bán", "Tên hệ thống", "Nhân viên", "Tên khách hàng", "Số điện thoại", "Vé combo", "Ghi chú"])
    second.append(["", "10/04/2026", "Bình", "Bình", "Khách B", "0901000002", "Combo 20", ""])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    rows = _combo_import_rows(stream.getvalue(), "letan")

    assert len(rows) == 2
    assert rows[0]["id"] == "sale-1"
    assert rows[0]["sale_date"] == date(2026, 4, 9)
    assert rows[1]["employee_username"] == "Bình"
    assert rows[1]["sale_date"] == date(2026, 4, 10)
    assert all(row["department"] == "letan" for row in rows)


def test_combo_excel_import_infers_employee_from_export_metadata():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anh Nguyễn"
    sheet.append(["ID", "Ngày bán", "Tên hệ thống", "Nhân viên", "Tên khách hàng", "Số điện thoại", "Vé combo", "Ghi chú", "", "__VERA_EMPLOYEE_USERNAME__"])
    sheet.cell(row=2, column=10, value="anh.nguyen")
    sheet.append(["", "11/04/2026", "", "", "Khách C", 901000003, "Combo 30", ""])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    rows = _combo_import_rows(stream.getvalue(), "letan")

    assert rows[0]["employee_username"] == "anh.nguyen"
    assert rows[0]["customer_phone"] == "901000003"


def test_staff_status_update_survives_google_credentials_failure():
    source = (ROOT / "vera_web_v2_staff.py").read_text(encoding="utf-8")
    assert '"sync_warning": f"{type(sync_exc).__name__}: {sync_exc}"' in source
    assert "Trạng thái đã lưu; chưa dọn lịch nghỉ tương lai" in source


def test_combo_sales_table_is_rendered_after_monthly_statistics():
    source = (ROOT / "web-v2/src/pages/WorkSchedulePage.jsx").read_text(encoding="utf-8")
    assert source.index('className="schedule-scroll monthly-statistics"') < source.index("{!loading && comboEditor}")
    assert "comboEmployees.map" in source
    assert "BẢNG CỦA" in source
    assert "canEditCombo = ['admin', 'quanly', 'letan'].includes(role)" in source
    assert "Import Excel" in source
    assert "Export Excel" in source
    assert "Lưu sửa" in source
