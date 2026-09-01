from io import BytesIO

from openpyxl import load_workbook

from vera_web_v2_leave_source_export import build_leave_source_workbook


def _norm(value):
    return str(value or "").strip().casefold().replace("ạ", "a")


def test_export_has_legacy_sheets_and_excludes_violation_rows_and_penalty_fields():
    payload = build_leave_source_workbook([
        ["Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ", "Phạt"],
        ["01/09/2026", "Thứ Ba", "Cẩm Vân", "Nghỉ CÓ phép", 0],
        ["01/09/2026", "Thứ Ba", "Cẩm Vân", "Nghỉ CÓ phép", 0],
        ["01/09/2026", "Thứ Ba", "Cẩm Vân", "Đi trễ", 50_000],
    ], [
        ["STT", "Lý do nghỉ", "Loại nghỉ"],
        [1, "Nghỉ CÓ phép", "Có phép"],
        [2, "Đi trễ", "Vi phạm"],
    ], norm=_norm)

    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ["MainData", "LoaiNghi"]
    main = workbook["MainData"]
    assert [cell.value for cell in main[1]] == ["Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ"]
    assert main.max_row == 2
    assert main["C2"].value == "Cẩm Vân"
    assert main["D2"].value == "Nghỉ CÓ phép"
    assert "Phạt" not in [cell.value for cell in main[1]]
    assert workbook["LoaiNghi"]["B2"].value == "Nghỉ CÓ phép"
