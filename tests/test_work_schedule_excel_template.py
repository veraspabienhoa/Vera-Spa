from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from vera_web_v2_work_schedule import _schedule_excel_bytes, _schedule_import_rows


def test_schedule_template_has_dropdown_catalog_and_round_trips_blank_cells():
    content = _schedule_excel_bytes(
        start=date(2026, 9, 5),
        end=date(2026, 9, 5),
        department="locker",
        employees=[{"username": "locker-a", "full_name": "Locker A"}],
        rows=[],
        shift_definitions={
            "locker": {
                "Ca 1": {"start": "09:30", "end": "17:30"},
                "Ca 2": {"start": "17:30", "end": "01:30"},
            },
        },
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook["DanhMuc"].sheet_state == "hidden"
    assert workbook["LichLamViec"].max_row == 2
    assert len(workbook["LichLamViec"].data_validations.dataValidation) == 4
    workbook.close()

    rows = _schedule_import_rows(content, "locker")
    assert rows[0]["work_date"] == date(2026, 9, 5)
    assert rows[0]["employee_username"] == "locker-a"
    assert rows[0]["shift_code"] == ""


def test_schedule_frontend_import_waits_for_manual_save():
    source = (Path(__file__).resolve().parents[1] / "web-v2/src/pages/WorkSchedulePage.jsx").read_text(encoding="utf-8")
    assert "Xuất Excel mẫu" in source
    assert "Import Excel" in source
    assert "importedAwaitingManualSaveRef.current = true" in source
    assert "Excel chờ Lưu lịch" in source
