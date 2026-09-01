from datetime import date, datetime
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

import vera_web_v2_tour_leave_sync as sync


def _workbook_bytes():
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = "Input"
    reference = workbook.create_sheet("Nghi")
    headers = [
        "STT", "Tên nhân viên", "Lịch hẹn", "Dịch vụ", "Yêu cầu", "Phòng",
        "Trạng Thái", "Thời lượng", "TG bắt đầu thực hiện", "TG bắt đầu thực hiện YC",
        "Thời gian còn lại", "Chờ thanh toán", "SL tua", "SL yêu cầu", "Tổng SL",
        "Đi làm", "Vào ca", "Break", "Giờ ra", "Thời gian", "Giờ vào", "Ghi chú",
        "Giờ Booking", "TG Xông Hơi",
    ]
    for column, value in enumerate(headers, start=1):
        input_sheet.cell(20, column, value)
    rows = [
        (1, "An *", "", "Di lam", datetime(2026, 8, 31, 9, 0)),
        (2, "Bình", "cũ", "Nghi phep", datetime(2026, 8, 31, 10, 0)),
        (3, "Chi", "cũ", "Nghi phep", datetime(2026, 8, 31, 11, 0)),
        (4, "Dũng", "", "Di lam", datetime(2026, 8, 31, 8, 0)),
    ]
    for row_number, (stt, employee, reason, status, start) in enumerate(rows, start=21):
        input_sheet.cell(row_number, 1, stt)
        input_sheet.cell(row_number, 2, employee)
        input_sheet.cell(row_number, 3, reason)
        input_sheet.cell(row_number, 9, start)
        input_sheet.cell(row_number, 11, f'=IF(B{row_number}="","",10)')
        input_sheet.cell(row_number, 16, status)

    reference["B1"] = "Loại nghỉ"
    for index, value in enumerate([
        "Nghỉ phép", "Nghỉ không phép", "Nghỉ không phép CUỐI TUẦN",
        "Nghỉ phát sinh", "PHEP NAM", "Hỗ trợ Ca 1 đi trễ 2 tiếng",
    ], start=2):
        reference.cell(index, 2, value)
    reference["H1"] = "Đi trễ"
    reference["H2"] = "Đi trễ CP"
    reference["K1"] = "Về sớm"
    reference["K2"] = "Về sớm CP"
    reference["N1"] = "Hỗ trợ"
    reference["N2"] = "Hỗ trợ Ca 1 đi trễ 2 tiếng"
    reference["Q1"] = "Loại nghỉ"
    reference["Q2"] = "Nghỉ phép"

    output = BytesIO()
    workbook.save(output)
    source = BytesIO(output.getvalue())
    target = BytesIO()
    with ZipFile(source, "r") as archive, ZipFile(target, "w", ZIP_DEFLATED) as updated:
        for info in archive.infolist():
            updated.writestr(info, archive.read(info.filename))
        updated.writestr("xl/vbaProject.bin", b"VBA-MUST-STAY-UNCHANGED")
        updated.writestr("xl/drawings/test-drawing.xml", b"DRAWING-MUST-STAY-UNCHANGED")
    return target.getvalue()


CATALOG = [
    ("Nghỉ CÓ phép", "Có phép"),
    ("Hỗ trợ Ca 1 sau 23H đi trễ 2 tiếng", "Hỗ trợ"),
    ("Xin đi tua cuối-qua tua có phép", "Tua cuối"),
]
SOURCE_ROWS = [
    ["31/08/2026", "Thứ 2", "An", "Nghỉ CÓ phép"],
    ["31/08/2026", "Thứ 2", "An", "Xin đi tua cuối-qua tua có phép"],
    ["31/08/2026", "Thứ 2", "Bình", "Hỗ trợ Ca 1 sau 23H đi trễ 2 tiếng"],
    ["31/08/2026", "Thứ 2", "Dũng", "Xin đi tua cuối-qua tua có phép"],
]


def _read_rows(payload):
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    sheet = workbook["Input"]
    output = {
        str(sheet.cell(row, 2).value or "").replace(" *", ""): {
            "reason": sheet.cell(row, 3).value or "",
            "status": sheet.cell(row, 16).value or "",
        }
        for row in range(21, 25)
    }
    workbook.close()
    return output


def test_sync_all_matches_first_reason_and_vba_status_rules():
    editor = sync._TourWorkbook(_workbook_bytes())
    stats = sync._apply_action(
        editor, "sync_all", date(2026, 8, 31), SOURCE_ROWS, CATALOG
    )
    payload = editor.to_bytes()
    rows = _read_rows(payload)

    assert stats == {
        "matched": 3,
        "reason_updated": 3,
        "status_updated": 1,
        "source_total": 4,
        "source_permit": 1,
        "source_no_permit": 0,
        "source_special": 3,
    }
    assert rows["An"] == {"reason": "Nghỉ phép", "status": "Nghi phep"}
    assert rows["Bình"] == {
        "reason": "Hỗ trợ Ca 1 đi trễ 2 tiếng", "status": "Nghi phep"
    }
    assert rows["Chi"] == {"reason": "cũ", "status": "Nghi phep"}
    assert rows["Dũng"] == {
        "reason": "Xin đi tua cuối-qua tua có phép", "status": "Di lam"
    }


def test_clear_leave_status_matches_vba_cleanup_rules():
    editor = sync._TourWorkbook(_workbook_bytes())
    stats = sync._apply_action(
        editor, "clear_leave_status", date(2026, 8, 31), SOURCE_ROWS, CATALOG
    )
    rows = _read_rows(editor.to_bytes())

    assert stats["matched"] == 1
    assert stats["reason_updated"] == 1
    assert stats["status_updated"] == 2
    assert rows["Bình"] == {
        "reason": "Hỗ trợ Ca 1 đi trễ 2 tiếng", "status": "Di lam"
    }
    assert rows["Chi"] == {"reason": "cũ", "status": "Di lam"}


def test_internal_vba_actions_use_exact_tourvera_reference_columns():
    editor = sync._TourWorkbook(_workbook_bytes())
    editor.set_input_text(21, 3, "Đi trễ CP")
    stats = sync._apply_action(editor, "late_to_leave", date(2026, 8, 31))
    assert stats["matched"] == 1
    assert stats["status_updated"] == 1
    assert editor.input_value(21, 16) == "Nghi phep"

    editor.set_input_text(21, 3, "Hỗ trợ Ca 1 đi trễ 2 tiếng")
    stats = sync._apply_action(editor, "support_to_working", date(2026, 8, 31))
    assert stats["matched"] == 1
    assert stats["status_updated"] == 1
    assert editor.input_value(21, 16) == "Di lam"


def test_xlsm_update_preserves_vba_and_drawings_exactly():
    original = _workbook_bytes()
    editor = sync._TourWorkbook(original)
    sync._apply_action(editor, "update_reasons", date(2026, 8, 31), SOURCE_ROWS, CATALOG)
    updated = editor.to_bytes()

    with ZipFile(BytesIO(original), "r") as before, ZipFile(BytesIO(updated), "r") as after:
        assert before.read("xl/vbaProject.bin") == after.read("xl/vbaProject.bin")
        assert before.read("xl/drawings/test-drawing.xml") == after.read("xl/drawings/test-drawing.xml")
        assert after.testzip() is None


def test_date_and_reason_aliases_match_vba():
    assert sync._same_date("2026-08-31", date(2026, 8, 31))
    assert sync._same_date(46265, date(2026, 8, 31))
    assert sync._convert_reason("Nghỉ CUỐI TUẦN KHÔNG phép") == "Nghi khong phep CUOI TUAN"
    assert sync._convert_reason("Về sớm CÓ phép") == "Ve som CP"
    assert sync._should_use_leave_status("Nghỉ phép quay video") is True
    assert sync._should_use_leave_status("Hỗ trợ Ca 1 đi trễ 2 tiếng") is False
