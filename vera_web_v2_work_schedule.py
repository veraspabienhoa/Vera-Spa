"""Daily work-schedule module for VERA Web V2.

Schedules are assigned by individual calendar date, never generated from a recurring
cycle. The UI may load a whole selected month (up to 63 days per request).

Quản lý is scheduled with a start/end time for each date. Locker, Lễ tân and Tạp vụ use
configurable shift definitions stored in PostgreSQL. All four departments share
the same overtime choices: TC Ca 1, TC Ca 2, or an explicit time range.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import re
from typing import Any, Callable, Literal
import unicodedata
from urllib.parse import quote
import uuid

from fastapi import Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import text


DEFAULT_SHIFT_DEFINITIONS = {
    "locker": {
        "Ca 1": {"start": "09:30", "end": "17:30"},
        "Ca 2": {"start": "17:30", "end": "01:30"},
    },
    "letan": {
        "Ca 1": {"start": "09:00", "end": "17:00"},
        "Ca 2": {"start": "16:30", "end": "00:30"},
    },
    "tapvu": {
        "Ca 1": {"start": "09:00", "end": "17:00"},
        "Ca 2": {"start": "16:30", "end": "00:30"},
    },
}

WORK_SCHEDULE_FEATURES = {
    "quanly": "work_schedule_quanly",
    "letan": "work_schedule_letan",
    "locker": "work_schedule_locker",
    "tapvu": "work_schedule_tapvu",
}

_TIME_RE = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")
OVERTIME_SHIFT_CHOICES = {"", "TC Ca 1", "TC Ca 2", "Từ giờ tới giờ"}
COMBO_EDITOR_ROLES = {"admin", "quanly", "letan"}
COMBO_EXCEL_HEADERS = (
    "ID", "Ngày bán", "Tên hệ thống", "Nhân viên",
    "Tên khách hàng", "Số điện thoại", "Vé combo", "Ghi chú",
)
COMBO_EXCEL_USERNAME_MARKER = "__VERA_EMPLOYEE_USERNAME__"


class ScheduleRow(BaseModel):
    work_date: date
    employee_username: str = Field(min_length=1, max_length=200)
    employee_name: str = Field(default="", max_length=300)
    department: Literal["quanly", "locker", "letan", "tapvu"]
    shift_code: str = Field(min_length=1, max_length=100)
    overtime_shift: str = Field(default="", max_length=100)
    start_time: str = Field(default="", max_length=5)
    end_time: str = Field(default="", max_length=5)
    overtime_start_time: str = Field(default="", max_length=5)
    overtime_end_time: str = Field(default="", max_length=5)
    note: str = Field(default="", max_length=500)
    combo_sold: bool = False
    combo_sale_date: date | None = None
    combo_customer_name: str = Field(default="", max_length=300)
    combo_customer_phone: str = Field(default="", max_length=50)
    combo_ticket: str = Field(default="", max_length=300)
    combo_note: str = Field(default="", max_length=500)


class ScheduleSave(BaseModel):
    rows: list[ScheduleRow] = Field(default_factory=list, max_length=1000)


class ShiftDefinitionRow(BaseModel):
    shift_code: str = Field(min_length=1, max_length=100)
    start_time: str = Field(min_length=5, max_length=5)
    end_time: str = Field(min_length=5, max_length=5)


class ShiftDefinitionSave(BaseModel):
    department: Literal["locker", "letan", "tapvu"]
    shifts: list[ShiftDefinitionRow] = Field(min_length=1, max_length=20)


class ComboSaleSave(BaseModel):
    sale_date: date
    employee_username: str = Field(min_length=1, max_length=200)
    employee_name: str = Field(default="", max_length=300)
    department: Literal["quanly", "letan"]
    customer_name: str = Field(min_length=1, max_length=300)
    customer_phone: str = Field(default="", max_length=50)
    combo_ticket: str = Field(min_length=1, max_length=300)
    note: str = Field(default="", max_length=500)


def _time_is_next_day(start_time: str, end_time: str) -> bool:
    return end_time <= start_time


def _ensure_schema(conn) -> None:
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS vera_work_schedule (
            work_date DATE NOT NULL,
            employee_username TEXT NOT NULL,
            employee_name TEXT NOT NULL DEFAULT '',
            department TEXT NOT NULL,
            shift_code TEXT NOT NULL,
            overtime_shift TEXT NOT NULL DEFAULT '',
            start_time TEXT NOT NULL DEFAULT '',
            end_time TEXT NOT NULL DEFAULT '',
            overtime_start_time TEXT NOT NULL DEFAULT '',
            overtime_end_time TEXT NOT NULL DEFAULT '',
            note TEXT NOT NULL DEFAULT '',
            combo_sold BOOLEAN NOT NULL DEFAULT FALSE,
            combo_sale_date DATE,
            combo_customer_name TEXT NOT NULL DEFAULT '',
            combo_customer_phone TEXT NOT NULL DEFAULT '',
            combo_ticket TEXT NOT NULL DEFAULT '',
            combo_note TEXT NOT NULL DEFAULT '',
            updated_by TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY(work_date, employee_username)
        )
    """))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS start_time TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS end_time TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS overtime_start_time TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS overtime_end_time TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS combo_sold BOOLEAN NOT NULL DEFAULT FALSE"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS combo_sale_date DATE"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS combo_customer_name TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS combo_customer_phone TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS combo_ticket TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("ALTER TABLE vera_work_schedule ADD COLUMN IF NOT EXISTS combo_note TEXT NOT NULL DEFAULT ''"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_vera_work_schedule_department_date ON vera_work_schedule(department, work_date)"))

    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS vera_work_shift_definition (
            department TEXT NOT NULL,
            shift_code TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            updated_by TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY(department, shift_code)
        )
    """))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_vera_work_shift_definition_department ON vera_work_shift_definition(department, sort_order, shift_code)"))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS vera_work_schedule_combo_sale (
            id TEXT PRIMARY KEY,
            sale_date DATE NOT NULL,
            employee_username TEXT NOT NULL,
            employee_name TEXT NOT NULL DEFAULT '',
            department TEXT NOT NULL,
            customer_name TEXT NOT NULL,
            customer_phone TEXT NOT NULL DEFAULT '',
            combo_ticket TEXT NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            updated_by TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
    """))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_vera_combo_sale_department_date ON vera_work_schedule_combo_sale(department, sale_date)"))

    for department, defaults in DEFAULT_SHIFT_DEFINITIONS.items():
        existing = conn.execute(text("""
            SELECT COUNT(*) FROM vera_work_shift_definition WHERE department=:department
        """), {"department": department}).scalar_one()
        if int(existing or 0) > 0:
            continue
        for sort_order, (shift_code, spec) in enumerate(defaults.items(), start=1):
            conn.execute(text("""
                INSERT INTO vera_work_shift_definition(
                    department, shift_code, start_time, end_time, sort_order, updated_by
                ) VALUES (
                    :department, :shift_code, :start_time, :end_time, :sort_order, 'system-default'
                )
                ON CONFLICT(department, shift_code) DO NOTHING
            """), {
                "department": department,
                "shift_code": shift_code,
                "start_time": spec["start"],
                "end_time": spec["end"],
                "sort_order": sort_order,
            })


def _role(ident: Any) -> str:
    return str(getattr(ident, "role", "") or "").strip().lower()


def _actor(ident: Any) -> str:
    return str(getattr(ident, "employee_username", "") or getattr(ident, "email", "") or "web_v2")


def _require_combo_editor(ident: Any) -> None:
    if _role(ident) not in COMBO_EDITOR_ROLES:
        raise HTTPException(403, "Chỉ Admin, Lễ tân hoặc Quản lý được cập nhật bảng bán combo.")


def _normalize_excel_header(value: Any) -> str:
    raw = unicodedata.normalize("NFD", str(value or "").strip().lower())
    raw = "".join(char for char in raw if unicodedata.category(char) != "Mn").replace("đ", "d")
    return re.sub(r"[^a-z0-9]+", " ", raw).strip()


def _combo_excel_date(value: Any, *, location: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            pass
    raise HTTPException(400, f"{location}: Ngày bán không hợp lệ; dùng DD/MM/YYYY hoặc YYYY-MM-DD.")


def _combo_excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _combo_import_rows(payload: bytes, department: str) -> list[dict[str, Any]]:
    if not payload:
        raise HTTPException(400, "File Excel đang trống.")
    if len(payload) > 10 * 1024 * 1024:
        raise HTTPException(413, "File Excel bán combo vượt quá 10 MB.")
    if not payload.startswith(b"PK"):
        raise HTTPException(400, "File bán combo phải là Excel .xlsx hợp lệ.")
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, "Không đọc được file Excel bán combo.") from exc

    aliases = {
        "id": "id",
        "ngay ban": "sale_date",
        "ten he thong": "employee_username",
        "nhan vien": "employee_name",
        "ten khach hang": "customer_name",
        "so dien thoai": "customer_phone",
        "ve combo": "combo_ticket",
        "ghi chu": "note",
    }
    required = {"sale_date", "employee_username", "customer_name", "combo_ticket"}
    result: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    try:
        for worksheet in workbook.worksheets:
            headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
            marker = _normalize_excel_header(worksheet.cell(row=1, column=10).value)
            sheet_username = (
                _combo_excel_text(worksheet.cell(row=2, column=10).value)
                if marker == _normalize_excel_header(COMBO_EXCEL_USERNAME_MARKER)
                else ""
            )
            indexes = {
                aliases[normalized]: index
                for index, header in enumerate(headers)
                if (normalized := _normalize_excel_header(header)) in aliases
            }
            if not any(str(value or "").strip() for value in headers):
                continue
            missing = required - set(indexes)
            if missing:
                raise HTTPException(
                    400,
                    f"Sheet {worksheet.title}: thiếu cột Ngày bán, Tên hệ thống, Tên khách hàng hoặc Vé combo.",
                )
            for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                def cell(field: str) -> Any:
                    index = indexes.get(field, -1)
                    return values[index] if 0 <= index < len(values) else ""

                raw_date = cell("sale_date")
                raw_username = _combo_excel_text(cell("employee_username"))
                employee_name = _combo_excel_text(cell("employee_name"))
                customer_name = _combo_excel_text(cell("customer_name"))
                customer_phone = _combo_excel_text(cell("customer_phone"))
                combo_ticket = _combo_excel_text(cell("combo_ticket"))
                note = _combo_excel_text(cell("note"))
                if not any((raw_date, raw_username, employee_name, customer_name, customer_phone, combo_ticket, note)):
                    continue
                username = raw_username or sheet_username or worksheet.title.strip()
                location = f"Sheet {worksheet.title}, dòng {row_number}"
                if not username or not customer_name or not combo_ticket:
                    raise HTTPException(400, f"{location}: cần đủ Tên hệ thống, Tên khách hàng và Vé combo.")
                sale_id = _combo_excel_text(cell("id"))
                if len(sale_id) > 100:
                    raise HTTPException(400, f"{location}: ID quá dài.")
                if sale_id and sale_id in seen_ids:
                    raise HTTPException(400, f"{location}: ID bị trùng trong file.")
                if sale_id:
                    seen_ids.add(sale_id)
                item = {
                    "id": sale_id,
                    "sale_date": _combo_excel_date(raw_date, location=location),
                    "employee_username": username,
                    "employee_name": employee_name,
                    "department": department,
                    "customer_name": customer_name,
                    "customer_phone": customer_phone,
                    "combo_ticket": combo_ticket,
                    "note": note,
                }
                limits = {
                    "employee_username": 200, "employee_name": 300, "customer_name": 300,
                    "customer_phone": 50, "combo_ticket": 300, "note": 500,
                }
                for field, limit in limits.items():
                    if len(item[field]) > limit:
                        raise HTTPException(400, f"{location}: {field} vượt quá {limit} ký tự.")
                result.append(item)
                if len(result) > 2000:
                    raise HTTPException(413, "Mỗi lần chỉ Import tối đa 2.000 dòng bán combo.")
    finally:
        workbook.close()
    if not result:
        raise HTTPException(400, "File Excel không có dòng bán combo hợp lệ.")
    return result


def _safe_combo_sheet_title(value: Any, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", " ", str(value or "Bán combo")).strip()[:31] or "Bán combo"
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        tail = f" {suffix}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _feature_for_department(department: str) -> str:
    try:
        return WORK_SCHEDULE_FEATURES[department]
    except KeyError as exc:
        raise HTTPException(400, "Bộ phận chỉ hỗ trợ Quản lý, Lễ tân, Locker hoặc Tạp vụ.") from exc


def _allowed_department(conn, ident, department: str, feature_allowed) -> bool:
    return bool(feature_allowed(conn, ident, _feature_for_department(department)))


def _employee_catalog(conn, department: str) -> list[dict[str, Any]]:
    """Schedule viewers do not need the broad staff_list permission."""
    rows = conn.execute(text("""
        SELECT username,
               COALESCE(NULLIF(full_name,''), username) AS full_name,
               lower(COALESCE(role,'')) AS role,
               COALESCE(payload->>'Trạng thái làm việc','') AS employment_status
        FROM employees
        WHERE lower(COALESCE(role,''))=:department
          AND COALESCE(payload->>'__deleted','false') <> 'true'
        ORDER BY lower(COALESCE(NULLIF(full_name,''), username)), lower(username)
    """), {"department": department}).mappings().all()
    return [dict(row) for row in rows]


def _combo_employee(conn, department: str, username: str) -> dict[str, Any]:
    wanted = str(username or "").strip().casefold()
    employee = next(
        (item for item in _employee_catalog(conn, department) if str(item.get("username") or "").strip().casefold() == wanted),
        None,
    )
    if not employee:
        raise HTTPException(400, f"Nhân viên '{username}' không thuộc bộ phận {department}.")
    return employee


def _combo_employee_name(employee: dict[str, Any], fallback: str) -> str:
    return str(
        employee.get("system_name") or employee.get("full_name") or fallback
    ).strip()


def _load_shift_definitions(conn) -> dict[str, Any]:
    output: dict[str, Any] = {
        "quanly": {"mode": "daily_time_range"},
        "locker": {},
        "letan": {},
        "tapvu": {},
    }
    rows = conn.execute(text("""
        SELECT department, shift_code, start_time, end_time, sort_order
        FROM vera_work_shift_definition
        WHERE department IN ('locker','letan','tapvu')
        ORDER BY department, sort_order, lower(shift_code)
    """)).mappings().all()
    for row in rows:
        department = str(row.get("department") or "")
        shift_code = str(row.get("shift_code") or "")
        start_time = str(row.get("start_time") or "")[:5]
        end_time = str(row.get("end_time") or "")[:5]
        if not department or not shift_code:
            continue
        output.setdefault(department, {})[shift_code] = {
            "start": start_time,
            "end": end_time,
            "end_next_day": _time_is_next_day(start_time, end_time),
        }
    return output


def _validate_shift_definition(row: ShiftDefinitionRow) -> tuple[str, str, str]:
    shift_code = str(row.shift_code or "").strip()
    start_time = str(row.start_time or "").strip()
    end_time = str(row.end_time or "").strip()
    if not shift_code or shift_code.casefold() in {"nghỉ", "nghi", "giờ làm", "gio lam"}:
        raise HTTPException(400, "Tên ca không hợp lệ.")
    if not (_TIME_RE.fullmatch(start_time) and _TIME_RE.fullmatch(end_time)):
        raise HTTPException(400, f"Ca {shift_code}: giờ bắt đầu/kết thúc phải theo HH:MM.")
    if start_time == end_time:
        raise HTTPException(400, f"Ca {shift_code}: giờ bắt đầu và kết thúc không được trùng nhau.")
    return shift_code, start_time, end_time


def _validate_row(row: ScheduleRow, shift_definitions: dict[str, Any]) -> tuple[str, str, str, str, str]:
    """Validate and normalize schedule-specific fields."""
    shift_code = str(row.shift_code or "").strip()
    overtime_shift = str(row.overtime_shift or "").strip()
    start_time = str(row.start_time or "").strip()
    end_time = str(row.end_time or "").strip()
    overtime_start_time = str(row.overtime_start_time or "").strip()
    overtime_end_time = str(row.overtime_end_time or "").strip()

    if row.department == "quanly":
        if shift_code not in {"Giờ làm", "Nghỉ"}:
            raise HTTPException(400, "Lịch Quản lý chỉ dùng Giờ làm hoặc Nghỉ.")
        if shift_code == "Nghỉ":
            if overtime_shift or overtime_start_time or overtime_end_time:
                raise HTTPException(400, "Ngày nghỉ không thể đồng thời có tăng ca.")
            return "", "", "", "", ""
        if not (_TIME_RE.fullmatch(start_time) and _TIME_RE.fullmatch(end_time)):
            raise HTTPException(
                400,
                f"Quản lý {row.employee_name or row.employee_username} ngày {row.work_date:%d/%m/%Y} cần đủ giờ bắt đầu và giờ kết thúc dạng HH:MM.",
            )
    elif shift_code == "Nghỉ":
        if overtime_shift or overtime_start_time or overtime_end_time:
            raise HTTPException(400, "Ngày nghỉ không thể đồng thời có tăng ca.")
        return "", "", "", "", ""
    elif shift_code not in (shift_definitions.get(row.department) or {}):
        raise HTTPException(400, f"Ca '{shift_code}' không còn trong cấu hình {row.department}.")

    if overtime_shift not in OVERTIME_SHIFT_CHOICES:
        raise HTTPException(400, "Tăng ca chỉ dùng TC Ca 1, TC Ca 2 hoặc Từ giờ tới giờ.")
    has_overtime_times = bool(overtime_start_time or overtime_end_time)
    custom_overtime = overtime_shift == "Từ giờ tới giờ" or (not overtime_shift and has_overtime_times)
    if overtime_shift in {"TC Ca 1", "TC Ca 2"}:
        if overtime_start_time or overtime_end_time:
            raise HTTPException(400, "TC Ca 1/TC Ca 2 không nhập thêm khoảng giờ tùy chỉnh.")
        return start_time if row.department == "quanly" else "", end_time if row.department == "quanly" else "", overtime_shift, "", ""
    if custom_overtime and not (_TIME_RE.fullmatch(overtime_start_time) and _TIME_RE.fullmatch(overtime_end_time)):
        raise HTTPException(
            400,
            f"{row.employee_name or row.employee_username} ngày {row.work_date:%d/%m/%Y}: tăng ca cần đủ giờ bắt đầu và kết thúc dạng HH:MM.",
        )
    if custom_overtime and overtime_start_time == overtime_end_time:
        raise HTTPException(400, "Giờ bắt đầu và kết thúc tăng ca không được trùng nhau.")
    return (
        start_time if row.department == "quanly" else "",
        end_time if row.department == "quanly" else "",
        "Từ giờ tới giờ" if custom_overtime else "",
        overtime_start_time if custom_overtime else "",
        overtime_end_time if custom_overtime else "",
    )


def install_work_schedule_routes(
    app,
    *,
    engine_instance: Callable[[], Any],
    current_identity,
    feature_allowed,
) -> None:
    if getattr(app.state, "work_schedule_installed", False):
        return

    @app.get("/v2/work-schedule")
    def get_work_schedule(
        start: date = Query(...),
        end: date = Query(...),
        department: str = Query(""),
        ident=Depends(current_identity),
    ):
        if end < start:
            raise HTTPException(400, "Ngày kết thúc phải từ ngày bắt đầu trở đi.")
        if (end - start).days > 62:
            raise HTTPException(400, "Chỉ xem tối đa 63 ngày mỗi lần.")
        dep = department.strip().lower()
        if dep and dep not in WORK_SCHEDULE_FEATURES:
            raise HTTPException(400, "Bộ phận chỉ hỗ trợ Quản lý, Locker, Lễ tân hoặc Tạp vụ.")

        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            allowed_departments = [
                item for item in ("quanly", "letan", "tapvu", "locker")
                if _allowed_department(conn, ident, item, feature_allowed)
            ]
            if dep and dep not in allowed_departments:
                raise HTTPException(403, "Bạn không có quyền xem lịch làm việc của bộ phận này.")
            if not dep and not allowed_departments:
                raise HTTPException(403, "Bạn không có quyền xem Lịch làm việc.")

            sql = """
                SELECT work_date, employee_username, employee_name, department,
                       shift_code, overtime_shift, start_time, end_time,
                       overtime_start_time, overtime_end_time, note,
                       combo_sold, combo_sale_date, combo_customer_name,
                       combo_customer_phone, combo_ticket, combo_note,
                       updated_by, updated_at
                FROM vera_work_schedule
                WHERE work_date BETWEEN :start AND :end
            """
            params: dict[str, Any] = {"start": start, "end": end}
            if dep:
                sql += " AND department=:department"
                params["department"] = dep
            else:
                sql += " AND department = ANY(:departments)"
                params["departments"] = allowed_departments
            sql += " ORDER BY department, employee_name, employee_username, work_date"
            rows = conn.execute(text(sql), params).mappings().all()
            employees = _employee_catalog(conn, dep) if dep else []
            shift_definitions = _load_shift_definitions(conn)

        return {
            "ok": True,
            "start": start.isoformat(),
            "end": end.isoformat(),
            "shift_definitions": shift_definitions,
            "rows": [dict(row) for row in rows],
            "employees": employees,
            "can_edit": _role(ident) in {"admin", "quanly"},
            "allowed_departments": allowed_departments,
            "assignment_mode": "daily",
            "display_mode": "selected_month_all_days",
            "overtime_mode": {"locker": "shared", "letan": "shared", "tapvu": "shared", "quanly": "shared"},
            "overtime_choices": ["TC Ca 1", "TC Ca 2", "Từ giờ tới giờ"],
        }

    @app.put("/v2/work-schedule/shifts")
    def save_shift_definitions(body: ShiftDefinitionSave, ident=Depends(current_identity)):
        if _role(ident) not in {"admin", "quanly"}:
            raise HTTPException(403, "Chỉ Admin hoặc Quản lý được tạo/sửa ca làm việc.")
        actor = _actor(ident)
        normalized: list[tuple[str, str, str]] = []
        seen: set[str] = set()
        for item in body.shifts:
            shift_code, start_time, end_time = _validate_shift_definition(item)
            key = shift_code.casefold()
            if key in seen:
                raise HTTPException(400, f"Tên ca bị trùng: {shift_code}.")
            seen.add(key)
            normalized.append((shift_code, start_time, end_time))

        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            if not _allowed_department(conn, ident, body.department, feature_allowed):
                raise HTTPException(403, "Bạn không có quyền cấu hình ca của bộ phận này.")
            conn.execute(text("DELETE FROM vera_work_shift_definition WHERE department=:department"), {"department": body.department})
            for sort_order, (shift_code, start_time, end_time) in enumerate(normalized, start=1):
                conn.execute(text("""
                    INSERT INTO vera_work_shift_definition(
                        department, shift_code, start_time, end_time, sort_order,
                        updated_by, created_at, updated_at
                    ) VALUES (
                        :department, :shift_code, :start_time, :end_time, :sort_order,
                        :updated_by, NOW(), NOW()
                    )
                """), {
                    "department": body.department,
                    "shift_code": shift_code,
                    "start_time": start_time,
                    "end_time": end_time,
                    "sort_order": sort_order,
                    "updated_by": actor,
                })

        return {"ok": True, "saved": len(normalized), "message": "Đã cập nhật cấu hình ca làm việc."}

    @app.get("/v2/work-schedule/combo-sales")
    def get_combo_sales(
        start: date = Query(...), end: date = Query(...), department: str = Query(...),
        ident=Depends(current_identity),
    ):
        dep = department.strip().lower()
        if dep not in {"quanly", "letan"}:
            raise HTTPException(400, "Bảng bán combo chỉ áp dụng cho Quản lý và Lễ tân.")
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            if not _allowed_department(conn, ident, dep, feature_allowed):
                raise HTTPException(403, "Bạn không có quyền xem bảng bán combo của bộ phận này.")
            rows = conn.execute(text("""
                SELECT id, sale_date, employee_username, employee_name, department,
                       customer_name, customer_phone, combo_ticket, note, updated_by, updated_at
                FROM vera_work_schedule_combo_sale
                WHERE department=:department AND sale_date BETWEEN :start AND :end
                ORDER BY sale_date DESC, lower(employee_name), created_at DESC
            """), {"department": dep, "start": start, "end": end}).mappings().all()
        return {
            "ok": True,
            "rows": [dict(row) for row in rows],
            "can_edit": _role(ident) in COMBO_EDITOR_ROLES,
            "layout": "one_table_per_employee",
        }

    @app.post("/v2/work-schedule/combo-sales")
    def save_combo_sale(body: ComboSaleSave, ident=Depends(current_identity)):
        _require_combo_editor(ident)
        sale_id = str(uuid.uuid4())
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            if not _allowed_department(conn, ident, body.department, feature_allowed):
                raise HTTPException(403, "Bạn không có quyền cập nhật bảng bán combo của bộ phận này.")
            employee = _combo_employee(conn, body.department, body.employee_username)
            conn.execute(text("""
                INSERT INTO vera_work_schedule_combo_sale(
                    id, sale_date, employee_username, employee_name, department,
                    customer_name, customer_phone, combo_ticket, note, updated_by
                ) VALUES (
                    :id, :sale_date, :employee_username, :employee_name, :department,
                    :customer_name, :customer_phone, :combo_ticket, :note, :updated_by
                )
            """), {
                "id": sale_id, "sale_date": body.sale_date,
                "employee_username": body.employee_username.strip(),
                "employee_name": _combo_employee_name(employee, body.employee_username),
                "department": body.department, "customer_name": body.customer_name.strip(),
                "customer_phone": body.customer_phone.strip(), "combo_ticket": body.combo_ticket.strip(),
                "note": body.note.strip(), "updated_by": _actor(ident),
            })
        return {"ok": True, "id": sale_id, "message": "Đã thêm lượt bán combo."}

    @app.put("/v2/work-schedule/combo-sales/{sale_id}")
    def update_combo_sale(sale_id: str, body: ComboSaleSave, ident=Depends(current_identity)):
        _require_combo_editor(ident)
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            existing = conn.execute(text("""
                SELECT id, department FROM vera_work_schedule_combo_sale WHERE id=:id
            """), {"id": sale_id}).mappings().first()
            if not existing:
                raise HTTPException(404, "Không tìm thấy dữ liệu bán combo cần sửa.")
            existing_department = str(existing.get("department") or "")
            if not _allowed_department(conn, ident, existing_department, feature_allowed):
                raise HTTPException(403, "Bạn không có quyền sửa dữ liệu này.")
            if body.department != existing_department:
                raise HTTPException(400, "Không được chuyển dữ liệu bán combo sang bộ phận khác.")
            employee = _combo_employee(conn, body.department, body.employee_username)
            conn.execute(text("""
                UPDATE vera_work_schedule_combo_sale
                SET sale_date=:sale_date,
                    employee_username=:employee_username,
                    employee_name=:employee_name,
                    customer_name=:customer_name,
                    customer_phone=:customer_phone,
                    combo_ticket=:combo_ticket,
                    note=:note,
                    updated_by=:updated_by,
                    updated_at=NOW()
                WHERE id=:id
            """), {
                "id": sale_id, "sale_date": body.sale_date,
                "employee_username": body.employee_username.strip(),
                "employee_name": _combo_employee_name(employee, body.employee_username),
                "customer_name": body.customer_name.strip(),
                "customer_phone": body.customer_phone.strip(),
                "combo_ticket": body.combo_ticket.strip(),
                "note": body.note.strip(),
                "updated_by": _actor(ident),
            })
        return {"ok": True, "id": sale_id, "message": "Đã cập nhật lượt bán combo."}

    @app.delete("/v2/work-schedule/combo-sales/{sale_id}")
    def delete_combo_sale(sale_id: str, ident=Depends(current_identity)):
        _require_combo_editor(ident)
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            existing = conn.execute(text("SELECT department FROM vera_work_schedule_combo_sale WHERE id=:id"), {"id": sale_id}).mappings().first()
            if existing and not _allowed_department(conn, ident, str(existing["department"]), feature_allowed):
                raise HTTPException(403, "Bạn không có quyền xóa dữ liệu này.")
            result = conn.execute(text("DELETE FROM vera_work_schedule_combo_sale WHERE id=:id"), {"id": sale_id})
        return {"ok": True, "deleted": int(result.rowcount or 0)}

    @app.get("/v2/work-schedule/combo-sales/export.xlsx")
    def export_combo_sales(
        start: date = Query(...), end: date = Query(...), department: str = Query(...),
        ident=Depends(current_identity),
    ):
        _require_combo_editor(ident)
        dep = department.strip().lower()
        if dep not in {"quanly", "letan"}:
            raise HTTPException(400, "Bảng bán combo chỉ áp dụng cho Quản lý và Lễ tân.")
        if end < start or (end - start).days > 366:
            raise HTTPException(400, "Khoảng ngày Export bán combo không hợp lệ.")
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            if not _allowed_department(conn, ident, dep, feature_allowed):
                raise HTTPException(403, "Bạn không có quyền Export bảng bán combo của bộ phận này.")
            employees = _employee_catalog(conn, dep)
            rows = conn.execute(text("""
                SELECT id, sale_date, employee_username, employee_name,
                       customer_name, customer_phone, combo_ticket, note
                FROM vera_work_schedule_combo_sale
                WHERE department=:department AND sale_date BETWEEN :start AND :end
                ORDER BY lower(employee_name), sale_date, created_at
            """), {"department": dep, "start": start, "end": end}).mappings().all()

        groups: dict[str, dict[str, Any]] = {}
        for employee in employees:
            username = str(employee.get("username") or "").strip()
            groups[username.casefold()] = {
                "username": username,
                "name": _combo_employee_name(employee, username),
                "rows": [],
            }
        for row in rows:
            username = str(row.get("employee_username") or "").strip()
            group = groups.setdefault(username.casefold(), {
                "username": username,
                "name": str(row.get("employee_name") or username).strip(),
                "rows": [],
            })
            group["rows"].append(dict(row))

        workbook = Workbook()
        workbook.remove(workbook.active)
        used_titles: set[str] = set()
        for group in groups.values():
            worksheet = workbook.create_sheet(_safe_combo_sheet_title(group["name"] or group["username"], used_titles))
            worksheet.append(COMBO_EXCEL_HEADERS)
            worksheet.cell(row=1, column=10, value=COMBO_EXCEL_USERNAME_MARKER)
            worksheet.cell(row=2, column=10, value=group["username"])
            worksheet.column_dimensions["J"].hidden = True
            for row in group["rows"]:
                worksheet.append([
                    row.get("id"), row.get("sale_date"), group["username"],
                    row.get("employee_name") or group["name"], row.get("customer_name"),
                    row.get("customer_phone"), row.get("combo_ticket"), row.get("note"),
                ])
                for cell in worksheet[worksheet.max_row]:
                    if cell.column != 2 and cell.value is not None:
                        cell.data_type = "s"
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = f"A1:H{max(1, worksheet.max_row)}"
            worksheet.column_dimensions["A"].hidden = True
            worksheet.column_dimensions["B"].width = 14
            worksheet.column_dimensions["C"].width = 24
            worksheet.column_dimensions["D"].width = 24
            worksheet.column_dimensions["E"].width = 30
            worksheet.column_dimensions["F"].width = 18
            worksheet.column_dimensions["G"].width = 25
            worksheet.column_dimensions["H"].width = 36
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F513F")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cell in worksheet["B"][1:]:
                cell.number_format = "DD/MM/YYYY"
        if not workbook.worksheets:
            worksheet = workbook.create_sheet("Bán combo")
            worksheet.append(COMBO_EXCEL_HEADERS)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        stream.seek(0)
        filename = f"VERA_Ban_Combo_{dep}_{start:%Y-%m-%d}_{end:%Y-%m-%d}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    @app.post("/v2/work-schedule/combo-sales/import.xlsx")
    async def import_combo_sales(
        request: Request, department: str = Query(...), ident=Depends(current_identity),
    ):
        _require_combo_editor(ident)
        dep = department.strip().lower()
        if dep not in {"quanly", "letan"}:
            raise HTTPException(400, "Bảng bán combo chỉ áp dụng cho Quản lý và Lễ tân.")
        imported = _combo_import_rows(await request.body(), dep)
        inserted = updated = 0
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            if not _allowed_department(conn, ident, dep, feature_allowed):
                raise HTTPException(403, "Bạn không có quyền Import bảng bán combo của bộ phận này.")
            catalog = {
                str(item.get("username") or "").strip().casefold(): item
                for item in _employee_catalog(conn, dep)
            }
            ids = [str(item["id"]) for item in imported if item.get("id")]
            existing = {}
            if ids:
                existing = {
                    str(row["id"]): str(row["department"])
                    for row in conn.execute(text("""
                        SELECT id, department FROM vera_work_schedule_combo_sale WHERE id = ANY(:ids)
                    """), {"ids": ids}).mappings().all()
                }
            for item in imported:
                key = str(item["employee_username"]).casefold()
                employee = catalog.get(key)
                if not employee:
                    raise HTTPException(400, f"Tên hệ thống '{item['employee_username']}' không thuộc bộ phận {dep}.")
                sale_id = str(item.get("id") or uuid.uuid4())
                if sale_id in existing and existing[sale_id] != dep:
                    raise HTTPException(400, f"ID {sale_id} đang thuộc bộ phận khác.")
                was_existing = sale_id in existing
                conn.execute(text("""
                    INSERT INTO vera_work_schedule_combo_sale(
                        id, sale_date, employee_username, employee_name, department,
                        customer_name, customer_phone, combo_ticket, note, updated_by,
                        created_at, updated_at
                    ) VALUES (
                        :id, :sale_date, :employee_username, :employee_name, :department,
                        :customer_name, :customer_phone, :combo_ticket, :note, :updated_by,
                        NOW(), NOW()
                    )
                    ON CONFLICT(id) DO UPDATE SET
                        sale_date=EXCLUDED.sale_date,
                        employee_username=EXCLUDED.employee_username,
                        employee_name=EXCLUDED.employee_name,
                        customer_name=EXCLUDED.customer_name,
                        customer_phone=EXCLUDED.customer_phone,
                        combo_ticket=EXCLUDED.combo_ticket,
                        note=EXCLUDED.note,
                        updated_by=EXCLUDED.updated_by,
                        updated_at=NOW()
                """), {
                    **item,
                    "id": sale_id,
                    "employee_name": _combo_employee_name(employee, item["employee_username"]),
                    "updated_by": _actor(ident),
                })
                if was_existing:
                    updated += 1
                else:
                    inserted += 1
        return {
            "ok": True,
            "inserted": inserted,
            "updated": updated,
            "message": f"Đã Import {inserted} dòng mới và cập nhật {updated} dòng bán combo.",
        }

    @app.put("/v2/work-schedule")
    def save_work_schedule(body: ScheduleSave, ident=Depends(current_identity)):
        if _role(ident) not in {"admin", "quanly"}:
            raise HTTPException(403, "Chỉ Admin hoặc Quản lý được sắp xếp lịch làm việc.")
        actor = _actor(ident)

        unique_keys: set[tuple[date, str]] = set()
        normalized_rows: list[tuple[ScheduleRow, str, str, str, str, str]] = []
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            shift_definitions = _load_shift_definitions(conn)
            for row in body.rows:
                key = (row.work_date, row.employee_username.strip())
                if key in unique_keys:
                    raise HTTPException(400, f"Trùng lịch {row.employee_username} ngày {row.work_date:%d/%m/%Y}.")
                unique_keys.add(key)
                if not _allowed_department(conn, ident, row.department, feature_allowed):
                    raise HTTPException(403, f"Bạn không có quyền sửa lịch {row.department}.")
                start_time, end_time, overtime_shift, overtime_start_time, overtime_end_time = _validate_row(row, shift_definitions)
                normalized_rows.append((row, start_time, end_time, overtime_shift, overtime_start_time, overtime_end_time))

            for row, start_time, end_time, overtime_shift, overtime_start_time, overtime_end_time in normalized_rows:
                conn.execute(text("""
                    INSERT INTO vera_work_schedule(
                        work_date, employee_username, employee_name, department,
                        shift_code, overtime_shift, start_time, end_time,
                        overtime_start_time, overtime_end_time, note,
                        combo_sold, combo_sale_date, combo_customer_name,
                        combo_customer_phone, combo_ticket, combo_note,
                        updated_by, created_at, updated_at
                    ) VALUES (
                        :work_date, :employee_username, :employee_name, :department,
                        :shift_code, :overtime_shift, :start_time, :end_time,
                        :overtime_start_time, :overtime_end_time, :note,
                        :combo_sold, :combo_sale_date, :combo_customer_name,
                        :combo_customer_phone, :combo_ticket, :combo_note,
                        :updated_by, NOW(), NOW()
                    )
                    ON CONFLICT(work_date, employee_username) DO UPDATE SET
                        employee_name=EXCLUDED.employee_name,
                        department=EXCLUDED.department,
                        shift_code=EXCLUDED.shift_code,
                        overtime_shift=EXCLUDED.overtime_shift,
                        start_time=EXCLUDED.start_time,
                        end_time=EXCLUDED.end_time,
                        overtime_start_time=EXCLUDED.overtime_start_time,
                        overtime_end_time=EXCLUDED.overtime_end_time,
                        note=EXCLUDED.note,
                        combo_sold=EXCLUDED.combo_sold,
                        combo_sale_date=EXCLUDED.combo_sale_date,
                        combo_customer_name=EXCLUDED.combo_customer_name,
                        combo_customer_phone=EXCLUDED.combo_customer_phone,
                        combo_ticket=EXCLUDED.combo_ticket,
                        combo_note=EXCLUDED.combo_note,
                        updated_by=EXCLUDED.updated_by,
                        updated_at=NOW()
                """), {
                    "work_date": row.work_date,
                    "employee_username": row.employee_username.strip(),
                    "employee_name": row.employee_name.strip(),
                    "department": row.department,
                    "shift_code": str(row.shift_code or "").strip(),
                    "overtime_shift": overtime_shift,
                    "start_time": start_time,
                    "end_time": end_time,
                    "overtime_start_time": overtime_start_time,
                    "overtime_end_time": overtime_end_time,
                    "note": row.note.strip(),
                    "combo_sold": bool(row.combo_sold) if row.department in {"quanly", "letan"} else False,
                    "combo_sale_date": row.combo_sale_date if row.combo_sold and row.department in {"quanly", "letan"} else None,
                    "combo_customer_name": row.combo_customer_name.strip() if row.combo_sold and row.department in {"quanly", "letan"} else "",
                    "combo_customer_phone": row.combo_customer_phone.strip() if row.combo_sold and row.department in {"quanly", "letan"} else "",
                    "combo_ticket": row.combo_ticket.strip() if row.combo_sold and row.department in {"quanly", "letan"} else "",
                    "combo_note": row.combo_note.strip() if row.combo_sold and row.department in {"quanly", "letan"} else "",
                    "updated_by": actor,
                })

        return {"ok": True, "saved": len(normalized_rows), "message": "Đã lưu lịch làm việc theo từng ngày."}

    @app.delete("/v2/work-schedule")
    def delete_work_schedule(
        work_date: date = Query(...),
        employee_username: str = Query(..., min_length=1),
        ident=Depends(current_identity),
    ):
        if _role(ident) not in {"admin", "quanly"}:
            raise HTTPException(403, "Chỉ Admin hoặc Quản lý được xóa lịch làm việc.")
        with engine_instance().begin() as conn:
            _ensure_schema(conn)
            existing = conn.execute(text("""
                SELECT department FROM vera_work_schedule
                WHERE work_date=:work_date AND employee_username=:employee_username
            """), {"work_date": work_date, "employee_username": employee_username.strip()}).mappings().first()
            if existing and not _allowed_department(conn, ident, str(existing.get("department") or ""), feature_allowed):
                raise HTTPException(403, "Bạn không có quyền xóa lịch làm việc của bộ phận này.")
            result = conn.execute(text("""
                DELETE FROM vera_work_schedule
                WHERE work_date=:work_date AND employee_username=:employee_username
            """), {"work_date": work_date, "employee_username": employee_username.strip()})
        return {"ok": True, "deleted": int(result.rowcount or 0)}

    app.state.work_schedule_installed = True
