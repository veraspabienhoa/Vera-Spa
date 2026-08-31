"""Attendance-only routes. Revenue data is intentionally excluded.

Web V2 now restores the legacy mid-shift-break view. Planned break duration and
FaceID clustering come from the same PostgreSQL shift settings used by the old
system; actual break intervals are derived from TimeSoft's sequential punch
fields when those punches are present.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
import json
import re
from typing import Any, Callable
import unicodedata
from urllib.parse import quote

from fastapi import Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text


def _day(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    return raw.split(" ", 1)[0]


def _json_value(value: Any, default):
    if isinstance(value, (dict, list)):
        return value
    if value in (None, ""):
        return default
    try:
        parsed = json.loads(str(value))
        return parsed if isinstance(parsed, type(default)) else default
    except Exception:
        return default


def _shift_break_settings(conn) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = conn.execute(text("""
        SELECT setting_key, value_json
        FROM vera_app_setting
        WHERE category='shift' AND setting_key IN ('shift_definitions','shift_break_config')
    """)).mappings().all()
    values = {str(row.get("setting_key") or ""): row.get("value_json") for row in rows}
    definitions = _json_value(values.get("shift_definitions"), [])
    break_config = _json_value(values.get("shift_break_config"), {})
    return [dict(item) for item in definitions if isinstance(item, dict)], dict(break_config)


def _bool(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip().casefold() in {"1", "true", "yes", "on", "có", "co", "bật", "bat"}
    return bool(value)


def _number(value: Any, default=0) -> int:
    try:
        return int(float(value or 0))
    except Exception:
        return int(default)


def _norm(value: Any) -> str:
    raw = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return " ".join("".join(ch for ch in raw if unicodedata.category(ch) != "Mn").replace("đ", "d").split())


def _support_late_allowances(conn, start: date, end: date) -> dict[tuple[date, str], tuple[int, str]]:
    rows = conn.execute(text("""
        SELECT leave_date, employee_name, leave_reason
        FROM leave_records
        WHERE leave_date BETWEEN :start_date AND :end_date
    """), {"start_date": start, "end_date": end}).mappings().all()
    output: dict[tuple[date, str], tuple[int, str]] = {}
    for row in rows:
        reason = str(row.get("leave_reason") or "").strip()
        normalized = _norm(reason)
        if "ho tro" not in normalized:
            continue
        match = re.search(r"di tre\s+(\d+(?:[.,]\d+)?)\s*(?:tieng|gio)", normalized)
        if not match:
            continue
        allowance = int(round(float(match.group(1).replace(",", ".")) * 60))
        key = (row["leave_date"], _norm(row.get("employee_name")))
        if key not in output or allowance > output[key][0]:
            output[key] = (allowance, reason)
    return output


def _apply_support_shift_start(item: dict[str, Any], support_allowances: dict[tuple[date, str], tuple[int, str]]) -> dict[str, Any]:
    try:
        work_date = datetime.strptime(str(item.get("date") or ""), "%d/%m/%Y").date()
    except ValueError:
        return item
    support = support_allowances.get((work_date, _norm(item.get("employee_name"))))
    if not support:
        return item
    allowance, reason = support
    shift_start = _parse_punch(item.get("shift_start"), item["date"])
    if shift_start is None:
        return item
    effective_start = shift_start + timedelta(minutes=allowance)
    check_in = _parse_punch(item.get("check_in"), item["date"])
    late_minutes = item.get("late_minutes", 0)
    if check_in is not None:
        late_minutes = max(0, int((check_in - effective_start).total_seconds() // 60))
    return {
        **item,
        "shift_start": effective_start.strftime("%H:%M"),
        "late_minutes": late_minutes,
        "arrival_status": "Đi trễ" if late_minutes > 0 else "Đúng giờ",
        "support_reason": reason,
        "support_late_allowance_minutes": allowance,
    }


def _parse_punch(value: Any, work_day: str) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    candidates = [raw]
    if work_day and ":" in raw and "/" not in raw and "-" not in raw.split(" ", 1)[0]:
        candidates.insert(0, f"{work_day} {raw}")
    formats = [
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M", "%H:%M:%S", "%H:%M",
    ]
    for candidate in candidates:
        for fmt in formats:
            try:
                parsed = datetime.strptime(candidate, fmt)
                if parsed.year == 1900 and work_day:
                    base = datetime.strptime(work_day, "%d/%m/%Y")
                    parsed = parsed.replace(year=base.year, month=base.month, day=base.day)
                return parsed
            except ValueError:
                continue
    return None


def _shift_config(item: dict[str, Any], definitions: list[dict[str, Any]], break_config: dict[str, Any]) -> dict[str, Any]:
    shift_name = str(item.get("WorkTimeName") or "").strip()
    active = []
    for definition in definitions:
        if str(definition.get("Tên ca") or "").strip().casefold() != shift_name.casefold():
            continue
        if str(definition.get("Trạng thái") or "Đang dùng").strip().casefold() == "đã xóa".casefold():
            continue
        active.append(definition)
    definition = active[-1] if active else {}
    department = str(definition.get("Bộ phận") or "Nhân viên + Leader").strip() or "Nhân viên + Leader"
    department_cfg = break_config.get(department) if isinstance(break_config.get(department), dict) else {}
    enabled = _bool(definition.get("Áp dụng nghỉ giữa ca", department_cfg.get("enabled", False)))
    duration = _number(
        definition.get("Duration nghỉ giữa ca (phút)", department_cfg.get("duration_minutes", 0)),
        department_cfg.get("duration_minutes", 0),
    )
    cluster = max(1, _number(definition.get("Khoảng gom FaceID (phút)", 10), 10))
    return {
        "break_enabled": enabled,
        "break_planned_minutes": max(0, duration) if enabled else 0,
        "faceid_cluster_minutes": cluster,
        "break_department": department,
    }


def _punch_sequence(item: dict[str, Any], work_day: str, cluster_minutes: int) -> list[datetime]:
    raw_values = []
    first = item.get("MachineTimeCheckInStr") or item.get("LocalTimeCheckInStr")
    if first:
        raw_values.append(first)
    for index in range(2, 11):
        value = item.get(f"MachineTimeCheckIn{index}Str") or item.get(f"LocalTimeCheckIn{index}Str")
        if value:
            raw_values.append(value)
    last = item.get("MachineTimeCheckOutStr") or item.get("LocalTimeCheckOutStr")
    if last:
        raw_values.append(last)

    parsed = sorted({dt for value in raw_values if (dt := _parse_punch(value, work_day)) is not None})
    if not parsed:
        return []

    # Same principle as the legacy FaceID grouping: repeated scans within the
    # configured cluster window are one attendance event, not a new break edge.
    clustered = [parsed[0]]
    for current in parsed[1:]:
        delta = (current - clustered[-1]).total_seconds() / 60
        if delta <= max(1, cluster_minutes):
            continue
        clustered.append(current)
    return clustered


def _break_payload(item: dict[str, Any], cfg: dict[str, Any]) -> dict[str, Any]:
    work_day = _day(item.get("WorkDateStr"))
    punches = _punch_sequence(item, work_day, int(cfg.get("faceid_cluster_minutes") or 10))
    middle = punches[1:-1] if len(punches) >= 3 else []
    intervals: list[tuple[datetime, datetime]] = []
    for index in range(0, len(middle) - 1, 2):
        start, end = middle[index], middle[index + 1]
        if end >= start:
            intervals.append((start, end))
    actual_minutes = int(sum((end - start).total_seconds() for start, end in intervals) // 60)
    details = [f"{start.strftime('%H:%M')} → {end.strftime('%H:%M')}" for start, end in intervals]
    if middle and len(middle) % 2:
        details.append(f"{middle[-1].strftime('%H:%M')} → chưa chấm vào lại")

    planned = int(cfg.get("break_planned_minutes") or 0)
    enabled = bool(cfg.get("break_enabled"))
    if not enabled:
        status = "Không áp dụng"
    elif intervals:
        over = actual_minutes - planned
        status = f"Quá {over} phút" if planned > 0 and over > 0 else "Trong giới hạn"
    elif middle:
        status = "Chưa đủ cặp chấm công"
    else:
        status = "Chưa ghi nhận FaceID nghỉ"

    return {
        **cfg,
        "break_actual_minutes": actual_minutes,
        "break_count": len(intervals),
        "break_detail": "; ".join(details),
        "break_status": status,
        "punch_times": [value.strftime("%H:%M") for value in punches],
    }


def _record(item: dict[str, Any], definitions: list[dict[str, Any]], break_config: dict[str, Any]) -> dict[str, Any]:
    cfg = _shift_config(item, definitions, break_config)
    break_data = _break_payload(item, cfg)
    return {
        "date": _day(item.get("WorkDateStr")),
        "employee_code": str(item.get("employeeInfo.EmployeeCode") or item.get("EnrollNumber") or ""),
        "employee_name": str(item.get("employeeInfo.Name") or item.get("EmployeeName") or "").strip(),
        "shift": str(item.get("WorkTimeName") or "").strip(),
        "shift_start": str(item.get("StartWorkTime") or "").strip(),
        "shift_end": str(item.get("EndWorkTime") or "").strip(),
        "check_in": str(item.get("MachineTimeCheckInStr") or item.get("LocalTimeCheckInStr") or "").strip(),
        "check_out": str(item.get("MachineTimeCheckOutStr") or item.get("LocalTimeCheckOutStr") or "").strip(),
        "arrival_status": str(item.get("GoWorkTypeName") or "").strip(),
        "departure_status": str(item.get("LastCheckInTypeName") or "").strip(),
        "late_minutes": _number(item.get("TotalMinuteInGoLate")),
        "early_minutes": _number(item.get("TotalMinuteBackHomeEarly")),
        "total_minutes": _number(item.get("TotalMinuteInDay")),
        "punch_count": _number(item.get("TotalCheckInOneDay")),
        **break_data,
    }


def _records(conn, start: date, end: date) -> list[dict[str, Any]]:
    definitions, break_config = _shift_break_settings(conn)
    support_allowances = _support_late_allowances(conn, start, end)
    rows = conn.execute(text("""
        SELECT dataset_key, payload
        FROM vera_dataset_cache
        WHERE dataset_key='timesoft_employee_checkin_today'
           OR dataset_key LIKE 'timesoft_employee_checkin_20%'
        ORDER BY CASE WHEN dataset_key='timesoft_employee_checkin_today' THEN 0 ELSE 1 END, dataset_key DESC
    """)).mappings().all()
    seen: set[tuple[str, str, str, str]] = set()
    output = []
    for dataset in rows:
        for raw in dataset.get("payload") or []:
            if not isinstance(raw, dict):
                continue
            item = _apply_support_shift_start(_record(raw, definitions, break_config), support_allowances)
            try:
                item_date = datetime.strptime(item["date"], "%d/%m/%Y").date()
            except ValueError:
                continue
            if not start <= item_date <= end:
                continue
            key = (item["date"], item["employee_code"], item["check_in"], item["check_out"])
            if key in seen:
                continue
            seen.add(key)
            output.append(item)
    return sorted(output, key=lambda item: (datetime.strptime(item["date"], "%d/%m/%Y"), item["employee_name"].casefold()))


def install_snapshot_routes(app, *, engine_instance: Callable[[], Any], current_identity, require_feature, identity_type):
    def dates(start: date, end: date) -> tuple[date, date]:
        if end < start:
            raise HTTPException(400, "Đến ngày phải bằng hoặc sau Từ ngày.")
        if end - start > timedelta(days=62):
            raise HTTPException(400, "Chấm công chỉ cho xem tối đa 63 ngày mỗi lần.")
        return start, end

    @app.get("/v2/snapshot")
    def snapshot(start: date = Query(...), end: date = Query(...), ident: identity_type = Depends(current_identity)):
        dates(start, end)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "snapshot_today")
            records = _records(conn, start, end)
        return {"records": records, "count": len(records), "start": start.isoformat(), "end": end.isoformat(), "data_scope": "attendance_only"}

    @app.get("/v2/snapshot/export.xlsx")
    def snapshot_export(start: date = Query(...), end: date = Query(...), ident: identity_type = Depends(current_identity)):
        dates(start, end)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "snapshot_export")
            records = _records(conn, start, end)
        columns = [
            ("date", "Ngày"), ("employee_code", "Mã nhân viên"), ("employee_name", "Nhân viên"),
            ("shift", "Ca"), ("shift_start", "Bắt đầu ca"), ("shift_end", "Kết thúc ca"),
            ("check_in", "Giờ vào"), ("check_out", "Giờ ra"),
            ("break_planned_minutes", "Nghỉ giữa ca quy định (phút)"),
            ("break_detail", "FaceID nghỉ giữa ca"), ("break_actual_minutes", "Nghỉ giữa ca thực tế (phút)"),
            ("break_status", "Trạng thái nghỉ giữa ca"), ("punch_times", "Các lần chấm"),
            ("arrival_status", "Trạng thái vào"), ("departure_status", "Trạng thái ra"),
            ("late_minutes", "Phút trễ"), ("early_minutes", "Phút về sớm"),
            ("total_minutes", "Tổng phút"), ("punch_count", "Số lần chấm"),
        ]
        wb = Workbook(); ws = wb.active; ws.title = "Chấm công"
        ws.append([label for _, label in columns])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F513F")
        for item in records:
            ws.append([" · ".join(item.get(key, [])) if key == "punch_times" else item.get(key, "") for key, _ in columns])
        stream = BytesIO(); wb.save(stream); stream.seek(0)
        filename = f"VERA_ChamCong_{start.isoformat()}_{end.isoformat()}.xlsx"
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})
