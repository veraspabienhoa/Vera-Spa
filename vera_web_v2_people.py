"""Birthday notifications and the read-only TourVera board for Web V2."""
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
import json
import math
import re
import threading
import time
from typing import Any, Callable
import unicodedata

import requests
from fastapi import Depends, HTTPException, Query
from openpyxl import load_workbook
from sqlalchemy import text

from vera_tour_break_counter import next_break_event_state, tour_business_date
import vera_web_v2_snapshot as attendance_snapshot


BANG_TOUR_FILE_ID = "151d1ueCwH2KXX-HPQF1uj340uWSCS2dW"
TOUR_CACHE_SECONDS = 60
_tour_cache: dict[str, Any] = {"loaded_at": 0.0, "columns": [], "records": [], "rooms": {}, "source_updated_at": ""}
_tour_lock = threading.Lock()


def invalidate_tour_cache() -> None:
    """Force the next Bảng tua request to read the just-updated XLSM."""
    with _tour_lock:
        _tour_cache.update({
            "loaded_at": 0.0,
            "columns": [],
            "records": [],
            "rooms": {},
            "source_updated_at": "",
        })


def _birthday(value: Any) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return str(value).strip() if not isinstance(value, (int, float, bool)) else value


def _room_snapshot(sheet) -> dict[str, Any]:
    occupancy: dict[str, bool] = {}
    for values in sheet.iter_rows(min_row=3, max_col=7, values_only=True):
        room_value = _clean_cell(values[2] if len(values) > 2 else "")
        room = str(room_value).strip()
        if not room or room.upper() in {"#N/A", "N/A"}:
            continue
        status = _token(values[5] if len(values) > 5 else "")
        occupied = status in {"dang thuc hien", "dang cho"}
        occupancy[room] = occupancy.get(room, False) or occupied
    available = [room for room, occupied in occupancy.items() if not occupied]
    occupied = [room for room, is_occupied in occupancy.items() if is_occupied]
    return {
        "all": list(occupancy),
        "available": available,
        "occupied": occupied,
        "total_count": len(occupancy),
        "available_count": len(available),
        "occupied_count": len(occupied),
        "source_sheet": "Room",
    }


def _room_key(value: Any) -> str:
    token = _token(value)
    return re.sub(r"^phong\s*", "", token).strip()


def _available_rooms_for_tour(rooms: dict[str, Any], prepared: dict[str, Any]) -> list[str]:
    """Exclude rooms already displayed in the Web Tour PHÒNG column."""
    columns = list(prepared.get("columns") or [])
    room_column = _find_column_any(columns, ("Phòng", "Phong"))
    used_on_tour = {
        _room_key(record.get(room_column))
        for record in (prepared.get("records") or [])
        if room_column and _room_key(record.get(room_column))
    }
    return [
        str(room)
        for room in (rooms.get("available") or [])
        if _room_key(room) not in used_on_tour
    ]


def _download_tour() -> tuple[list[str], list[dict[str, Any]], str]:
    errors = []
    for url in (
        f"https://drive.usercontent.google.com/download?id={BANG_TOUR_FILE_ID}&export=download&confirm=t",
        f"https://drive.google.com/uc?export=download&id={BANG_TOUR_FILE_ID}&confirm=t",
    ):
        try:
            response = requests.get(url, timeout=25, allow_redirects=True)
            response.raise_for_status()
            if "text/html" in str(response.headers.get("Content-Type", "")).lower():
                raise RuntimeError("Google Drive trả về trang HTML thay vì file XLSM")
            workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
            if "Input" not in workbook.sheetnames:
                raise RuntimeError("File Bảng tua không có sheet Input")
            sheet = workbook["Input"]
            raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=20, max_row=20, max_col=24))]
            columns: list[str] = []
            used: dict[str, int] = {}
            for index, value in enumerate(raw_headers, start=1):
                label = str(value or "").strip() or f"Cột {index}"
                used[label] = used.get(label, 0) + 1
                columns.append(label if used[label] == 1 else f"{label} ({used[label]})")
            records = []
            for values in sheet.iter_rows(min_row=21, max_col=24, values_only=True):
                if not any(value not in (None, "") for value in values):
                    continue
                records.append({columns[index]: _clean_cell(value) for index, value in enumerate(values)})
                if len(records) >= 500:
                    break
            rooms = _room_snapshot(workbook["Room"]) if "Room" in workbook.sheetnames else {
                "all": [], "available": [], "occupied": [],
                "total_count": 0, "available_count": 0, "occupied_count": 0,
                "source_sheet": "Room",
            }
            # Preserve the existing three-value return contract used by the
            # attendance module while making the Room snapshot available to
            # the Tour API through the shared cache.
            _tour_cache["rooms"] = rooms
            workbook.close()
            return columns, records, str(response.headers.get("Last-Modified") or "")
        except Exception as exc:
            errors.append(str(exc))
    raise RuntimeError("Không tải được Bảng tua: " + " | ".join(errors[-2:]))


def _token(value: Any) -> str:
    raw = unicodedata.normalize("NFD", str(value or "").strip().lower())
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn").replace("đ", "d")
    return " ".join(raw.replace("_", " ").replace("-", " ").split())


def _find_column(columns: list[str], wanted: str) -> str:
    wanted_key = _token(wanted)
    exact = [column for column in columns if _token(column) == wanted_key]
    if exact:
        return exact[0]
    contains = [column for column in columns if wanted_key in _token(column)]
    return contains[0] if contains else ""


def _find_column_any(columns: list[str], candidates: tuple[str, ...]) -> str:
    normalized = {column: _token(column) for column in columns}
    for candidate in candidates:
        candidate_key = _token(candidate)
        exact = next((column for column, key in normalized.items() if key == candidate_key), "")
        if exact:
            return exact
    for candidate in candidates:
        candidate_key = _token(candidate)
        contains = next((column for column, key in normalized.items() if candidate_key in key), "")
        if contains:
            return contains
    return ""


def _tour_number(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return float(raw.replace(",", ".") if "," in raw and "." not in raw else raw.replace(",", ""))
    except ValueError:
        return None


def _duration_minutes(value: Any) -> float | None:
    number = _tour_number(value)
    if number is not None:
        return abs(number) * 1440 if 0 < abs(number) < 1 else abs(number)
    raw = str(value or "").strip()
    match = re.fullmatch(r"(\d{1,3}):(\d{1,2})(?::(\d{1,2}))?", raw)
    if match:
        return int(match.group(1)) * 60 + int(match.group(2)) + int(match.group(3) or 0) / 60
    match = re.search(r"[-+]?\d+(?:[.,]\d+)?", raw)
    return abs(float(match.group(0).replace(",", "."))) if match else None


def _start_datetime(value: Any, now: datetime, duration_minutes: float | None) -> datetime | None:
    if value in (None, ""):
        return None
    parsed: datetime | None = None
    time_only = False
    number = _tour_number(value)
    if isinstance(value, (int, float)) and number is not None:
        if 0 <= number < 1:
            seconds = int(round(number * 86400)) % 86400
            parsed = now.replace(hour=seconds // 3600, minute=(seconds % 3600) // 60, second=seconds % 60, microsecond=0)
            time_only = True
        elif number >= 1:
            parsed = datetime(1899, 12, 30, tzinfo=now.tzinfo) + timedelta(days=number)
    else:
        raw = str(value).strip()
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                parsed = datetime.strptime(raw, fmt).replace(tzinfo=now.tzinfo)
                break
            except ValueError:
                pass
        if parsed is None:
            for fmt in ("%H:%M:%S", "%H:%M"):
                try:
                    clock = datetime.strptime(raw, fmt)
                    parsed = now.replace(hour=clock.hour, minute=clock.minute, second=clock.second, microsecond=0)
                    time_only = True
                    break
                except ValueError:
                    pass
    if parsed is None:
        return None
    if time_only and parsed > now:
        previous = parsed - timedelta(days=1)
        elapsed = (now - previous).total_seconds() / 60
        if 0 <= elapsed <= max(float(duration_minutes or 0) + 240, 480):
            return previous
    return parsed


def _display_value(value: Any) -> Any:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    raw = str(value).strip()
    if raw.lower() in {"none", "nan", "nat", "<na>"}:
        return ""
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", raw):
        return str(int(round(float(raw))))
    return raw


def _prepare_tour(columns: list[str], source_records: list[dict[str, Any]], now: datetime) -> dict[str, Any]:
    request_column = _find_column_any(columns, ("Yêu cầu", "Yeu cau"))
    duration_column = _find_column_any(columns, ("Thời lượng", "Thoi luong", "Thời lượng (phút)", "Thời lượng phút"))
    start_column = _find_column_any(columns, ("TG bắt đầu thực hiện", "TG bat dau thuc hien", "BĐ thực hiện", "BD thuc hien", "Bắt đầu thực hiện"))
    start_yc_column = _find_column_any(columns, ("TG bắt đầu thực hiện YC", "TG bat dau thuc hien YC", "BĐ thực hiện YC", "BD thực hiện YC", "BĐ YC", "BD YC", "Bắt đầu YC"))
    source_remaining_column = _find_column_any(columns, ("Thời gian còn lại", "TG còn lại"))
    remaining_column = "TG CÒN LẠI"
    output_columns = [column for column in columns if column not in {source_remaining_column, remaining_column}]
    output_columns.append(remaining_column)
    name_column = _find_column(columns, "Tên nhân viên")
    status_column = _find_column(columns, "Trạng thái")
    room_column = _find_column_any(columns, ("Phòng", "Phong"))
    work_column = _find_column(columns, "Đi làm")
    shift_column = _find_column(columns, "Vào ca") or next((column for column in columns if _token(column) == "ca"), "")

    moved = [column for column in (status_column, room_column, remaining_column, request_column) if column]
    ordered_columns = [column for column in output_columns if column not in moved]
    if name_column and name_column in ordered_columns:
        position = ordered_columns.index(name_column) + 1
        for column in reversed(moved):
            ordered_columns.insert(position, column)
    else:
        ordered_columns = moved + ordered_columns

    prepared: list[dict[str, Any]] = []
    counters = {"doing": 0, "waiting": 0, "finishing": 0, "idle": 0, "break": 0, "working": 0, "leave": 0}
    countdown_error = "" if duration_column else "Không tìm thấy cột Thời lượng."
    for source in source_records:
        request_mode = _token(source.get(request_column, "")) if request_column else ""
        duration = _duration_minutes(source.get(duration_column)) if duration_column else None
        active_start_column = start_yc_column if request_mode == "yc" else (start_column if request_mode == "" else "")
        started = _start_datetime(source.get(active_start_column), now, duration) if active_start_column else None
        remaining: int | None = None
        if duration is not None and started is not None:
            elapsed = (now - started).total_seconds() / 60
            if elapsed >= 0:
                remaining = int(math.ceil(duration - elapsed))

        status_token = _token(source.get(status_column, "")) if status_column else ""
        work_token = _token(source.get(work_column, "")) if work_column else ""
        shift_token = _token(source.get(shift_column, "")) if shift_column else ""
        expired = remaining is not None and remaining <= -15
        status_number = _tour_number(source.get(status_column)) if status_column else None
        finishing = bool(
            (status_number is not None and status_number <= 30)
            or (status_token == "dang thuc hien" and remaining is not None and -15 < remaining <= 30)
        )
        idle = bool(
            not expired
            and not finishing
            and status_token not in {"dang thuc hien", "dang cho"}
            and work_token == "di lam"
            and shift_token in {"ca 1", "ca 2"}
            and remaining is None
        )

        counters["doing"] += int(status_token == "dang thuc hien")
        counters["waiting"] += int(status_token == "dang cho")
        counters["finishing"] += int(finishing)
        counters["idle"] += int(idle)
        counters["working"] += int(work_token == "di lam")
        counters["leave"] += int(work_token == "nghi phep")

        row_style = "work" if work_token == "di lam" else ("leave" if work_token == "nghi phep" else "default")
        if remaining is not None and work_token != "nghi phep":
            if remaining >= 15:
                row_style = "green"
            elif 0 <= remaining < 15:
                row_style = "yellow"
            elif -15 < remaining < 0:
                row_style = "red"
        if idle:
            row_style = "idle"

        values = {column: _display_value(source.get(column, "")) for column in ordered_columns}
        values[remaining_column] = "" if expired or remaining is None else remaining
        if status_column:
            values[status_column] = {"dang cho": "Đang chờ", "dang thuc hien": "Đang thực hiện"}.get(status_token, values[status_column])
        time_column = _find_column(columns, "Thời gian")
        if time_column and time_column != remaining_column:
            time_number = _tour_number(source.get(time_column))
            if time_number is not None and time_number < -180:
                values[time_column] = ""
        tour_groups = []
        if finishing or idle:
            tour_groups.append("available")
        if finishing:
            tour_groups.append("finishing")
        if work_token == "di lam":
            tour_groups.append("working")
        if work_token == "nghi phep":
            tour_groups.append("leave")
        if status_token == "dang thuc hien":
            tour_groups.append("doing")
        if status_token == "dang cho":
            tour_groups.append("waiting")
        prepared.append({**values, "_row_style": row_style, "_tour_groups": tour_groups})

    available = counters["finishing"] + counters["idle"]
    return {
        "columns": ordered_columns,
        "records": prepared,
        "employee_count": len(prepared),
        "available": available,
        "stats": [{"label": "Có thể lên tua", "value": available, "detail": "Sắp xong + Đang rảnh"}],
        "break_count": counters["break"],
        "working_count": counters["working"],
        "leave_count": counters["leave"],
        "doing_count": counters["doing"],
        "waiting_count": counters["waiting"],
        "finishing_count": counters["finishing"],
        "countdown_error": countdown_error,
        "countdown_at": now.isoformat(),
    }


def _tour_shift_bucket(record: dict[str, Any], columns: list[str]) -> str:
    column = _find_column_any(columns, ("Vào ca", "Giờ vào ca", "Thời gian vào ca", "Ca"))
    raw = str(record.get(column, "") or "").strip()
    token = _token(raw).replace(" ", "")
    if token in {"ca1", "10", "10h", "10h00"}:
        return "ca1"
    if token in {"ca2", "12", "12h", "12h00", "14", "14h", "14h00"}:
        return "ca2"
    match = re.search(r"(\d{1,2})\s*[:hH]", raw)
    if match:
        return "ca1" if int(match.group(1)) < 12 else "ca2"
    return ""


def _tour_metric_snapshots(prepared: dict[str, Any]) -> dict[str, dict[str, int]]:
    columns = list(prepared.get("columns") or [])
    records = list(prepared.get("records") or [])
    total_column = _find_column_any(columns, ("Tổng SL", "Tổng số lượng"))
    output: dict[str, dict[str, int]] = {}
    for bucket in ("all", "ca1", "ca2"):
        scoped = records if bucket == "all" else [
            row for row in records if _tour_shift_bucket(row, columns) == bucket
        ]
        total_quantity = sum(int(round(_tour_number(row.get(total_column)) or 0)) for row in scoped) if total_column else 0
        waiting = sum(1 for row in scoped if "waiting" in (row.get("_tour_groups") or []))
        breaks = sum(1 for row in scoped if "break" in (row.get("_tour_groups") or []))
        output[bucket] = {
            "total_quantity": total_quantity,
            "waiting_count": waiting,
            "customer_count": total_quantity + waiting,
            "break_count": breaks,
        }
    return output


def _attendance_shift_bucket(record: dict[str, Any]) -> str:
    token = _token(record.get("shift"))
    return "ca1" if token == "ca 1" else "ca2" if token == "ca 2" else ""


def _tracked_break_metrics(
    conn, attendance_records: list[dict[str, Any]], now: datetime, actor: str,
) -> tuple[dict[str, dict[str, int]], str, set[str]]:
    """Persist Chấm công break events in a workday rolling over at 10:00 VN."""
    business_date = tour_business_date(now)
    setting_key = business_date.isoformat()
    observed_events: dict[str, str] = {}
    active_employees: dict[str, str] = {}
    active_names: set[str] = set()

