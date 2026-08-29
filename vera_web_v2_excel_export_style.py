"""Global Excel export styling for every Web V2 .xlsx response.

Every workbook returned by the API is normalized before it leaves Cloud Run:
- colored/bold header row
- frozen pane immediately below the header
- AutoFilter across the used range
- AutoFit column widths with sane caps

Keeping this as response middleware prevents individual export routes from
silently drifting into different formatting conventions.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from starlette.responses import Response


RELEASE = "excel-export-style-2026-08-29"
HEADER_FILL = "1F513F"
HEADER_FONT = "FFFFFF"


def _header_row(ws) -> int:
    """Find the first table-like row; fall back to row 1 for ordinary exports."""
    for row_index in range(1, min(ws.max_row, 20) + 1):
        nonempty = sum(1 for cell in ws[row_index] if str(cell.value or "").strip())
        if nonempty >= 2:
            return row_index
    return 1


def _display_length(value: Any) -> int:
    if value is None:
        return 0
    text = str(value)
    return max((len(part) for part in text.splitlines()), default=0)


def style_workbook_bytes(payload: bytes) -> bytes:
    if not payload or not payload.startswith(b"PK"):
        return payload
    stream = BytesIO(payload)
    try:
        workbook = load_workbook(stream)
    except Exception:
        return payload

    for ws in workbook.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        header_row = _header_row(ws)
        for cell in ws[header_row]:
            if cell.column > ws.max_column:
                continue
            cell.font = Font(bold=True, color=HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = f"A{header_row + 1}"
        if ws.max_column >= 1 and ws.max_row >= header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        for column_index in range(1, ws.max_column + 1):
            width = 0
            for row_index in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_index, column=column_index)
                width = max(width, _display_length(cell.value))
            ws.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), 60)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def install_excel_export_style(app) -> None:
    if getattr(app.state, "excel_export_style_installed", False):
        return

    @app.middleware("http")
    async def excel_export_style_middleware(request, call_next):
        response = await call_next(request)
        path = str(request.url.path or "").lower()
        content_type = str(response.headers.get("content-type") or "").lower()
        disposition = str(response.headers.get("content-disposition") or "").lower()
        is_excel = (
            path.endswith(".xlsx")
            or "spreadsheetml" in content_type
            or ".xlsx" in disposition
        )
        if not is_excel or response.status_code < 200 or response.status_code >= 300:
            return response

        body = b""
        async for chunk in response.body_iterator:
            body += bytes(chunk)
        styled = style_workbook_bytes(body)
        headers = dict(response.headers)
        headers["content-length"] = str(len(styled))
        return Response(
            content=styled,
            status_code=response.status_code,
            headers=headers,
            media_type=response.media_type,
            background=response.background,
        )

    @app.get("/v2/excel-export-style/health")
    def excel_export_style_health():
        return {
            "ok": True,
            "release": RELEASE,
            "header_color": HEADER_FILL,
            "freeze_header": True,
            "autofilter": True,
            "autofit": True,
        }

    app.state.excel_export_style_installed = True
    app.state.excel_export_style_release = RELEASE
