"""Complete Web V2 payroll workflow: calculate, configure, save and email."""
from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from email.utils import formataddr
import hashlib
from html import escape
from io import BytesIO
import json
import numbers
import os
import re
import smtplib
from typing import Any, Callable
from urllib.parse import quote
import uuid

import pandas as pd
from fastapi import Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import text


PUBLIC_FIELDS = [
    "Mã bản lưu", "Từ ngày", "Đến ngày", "Ngày lưu", "Giờ lưu", "Tên Hệ thống", "Họ và tên",
    "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Hoàn trả tiền tích lũy", "Tích lũy", "Chi Phí Sinh Hoạt",
    "Tiền phạt trong tháng", "Tiền ứng lương", "Tiền hỗ trợ Locker", "Vi phạm kỳ trước", "Số tiền thực nhận",
]
ADMIN_FIELDS = PUBLIC_FIELDS + ["Email", "Số tài khoản ngân hàng", "Tên ngân hàng", "Người lưu", "Nguồn dữ liệu"]
MONEY_FIELDS = {
    "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Hoàn trả tiền tích lũy", "Tích lũy", "Chi Phí Sinh Hoạt",
    "Tiền phạt trong tháng", "Tiền ứng lương", "Tiền hỗ trợ Locker", "Vi phạm kỳ trước", "Số tiền thực nhận",
}
DRAFT_FIELDS = [
    "TT", "Tên Hệ thống", "Họ và tên", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Hoàn trả tiền tích lũy",
    "Tích lũy", "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương",
    "Tiền hỗ trợ Locker", "Số tiền thực nhận", "Email", "Số tài khoản ngân hàng", "Tên ngân hàng", "Số dòng Tip",
]
EDITABLE_FIELDS = {
    "Tiền Hỗ Trợ Hoàn Lại", "Hoàn trả tiền tích lũy", "Tích lũy", "Chi Phí Sinh Hoạt",
    "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương", "Tiền hỗ trợ Locker",
}
DEFAULT_CONFIG = {
    "default_living_expense": 150000,
    "default_locker_support": 80000,
    "leader_responsibility_allowance": 0,
}
LEGACY_SPREADSHEET_ID = os.getenv(
    "VERA_CREDENTIAL_SHEET_ID", "1DGXy3kPyMPwtz-3CnG8i6BiQbXFDApasoXVFzSmUe24"
)
LEGACY_PAYROLL_WORKSHEET = "BangLuong"
LEGACY_OBLIGATION_WORKSHEET = "NoViPham"
LEGACY_OBLIGATION_HEADERS = [
    "STT", "Tên nhân viên", "Số tiền", "Nội dung", "Loại",
    "Kỳ phát sinh từ", "Kỳ phát sinh đến", "Bắt đầu trừ từ",
    "Trạng thái", "Mã nguồn", "Ngày cập nhật", "Giờ cập nhật",
    "Người cập nhật", "Kỳ đã khấu trừ",
]
TIMESOFT_PAYROLL_HEADERS = [
    "STT", "Thời gian", "Mã hóa đơn", "Mã KH2", "Tên KH",
    "Sản phẩm/ Dịch vụ/ PT", "Tổng tiền", "Giảm giá",
    "NV tư vấn", "Ghi chú", "Nhân viên tư vấn",
]
PAYROLL_SOURCE_WORKSHEET = "Báo cáo doanh thu hóa đơn"
PAYROLL_SOURCE_READER_RELEASE = "payroll-timesoft-sheet-reader-2026-09-01.3"
PAYROLL_SOURCE_READER_MODE = "openpyxl-normal-workbook"
PAYROLL_EMAIL_TEMPLATE_RELEASE = "payroll-email-layout-2026-09-01.1"
TIP_ITEM_PATTERN = r"^tip(?:\b|[_\-\s])"


class PayrollConfigUpdate(BaseModel):
    default_living_expense: float = Field(ge=0, le=1_000_000_000)
    default_locker_support: float = Field(ge=0, le=1_000_000_000)
    leader_responsibility_allowance: float = Field(ge=0, le=1_000_000_000)


class PayrollSave(BaseModel):
    start: date
    end: date
    source_name: str = Field(default="Excel upload", max_length=300)
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=500)


class PayrollEmail(BaseModel):
    start: date
    end: date
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=100)


class PayrollExport(BaseModel):
    start: date
    end: date
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=500)


class ObligationCreate(BaseModel):
    employee_name: str = Field(min_length=1, max_length=200)
    amount: float = Field(gt=0, le=1_000_000_000)
    content: str = Field(default="Chưa hoàn thành nghĩa vụ Vi phạm", max_length=1000)
    due_from: date


class AccumulationRefundCreate(BaseModel):
    employee_name: str = Field(min_length=1, max_length=200)
    amount: float = Field(gt=0, le=1_000_000_000)
    start: date
    end: date
    note: str = Field(default="Hoàn trả tiền tích lũy khi nghỉ việc", max_length=1000)


def _number(value: Any) -> int:
    if isinstance(value, numbers.Number):
        return int(round(float(value)))
    raw = str(value or "").strip()
    if not raw:
        return 0
    if re.fullmatch(r"-?\d+", raw):
        return int(raw)
    decimal_match = re.fullmatch(r"(-?\d+)([.,])(\d+)", raw)
    if decimal_match:
        whole, _separator, fraction = decimal_match.groups()
        # VERA/Google thường dùng dấu chấm phân cách hàng nghìn: 150.000.
        if len(fraction) == 3 and len(whole.lstrip("-")) <= 3:
            return int(whole + fraction)
        return int(round(float(f"{whole}.{fraction}")))
    negative = raw.startswith("-")
    digits = re.sub(r"[^0-9]", "", raw)
    return (-1 if negative else 1) * int(digits or 0)


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def _email_money(value: Any) -> str:
    return f"{_number(value):,} VNĐ"


def _email_date(value: Any) -> str:
    parsed = _parse_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else str(value or "").strip()


def _payroll_email_subject(employee_name: str, start: date, end: date) -> str:
    return (
        f"Bảng lương {str(employee_name or '').strip()} - "
        f"{start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}"
    )


def _payroll_email_summary_rows(row: dict[str, Any]) -> list[tuple[str, int]]:
    rows = [
        ("Tiền Lương", _number(row.get("Tiền Lương"))),
        ("Hỗ trợ/Hoàn tiền", _number(row.get("Tiền Hỗ Trợ Hoàn Lại"))),
    ]
    accumulation_refund = _number(row.get("Hoàn trả tiền tích lũy"))
    if accumulation_refund:
        rows.append(("Hoàn trả tiền tích lũy", accumulation_refund))
    rows.extend([
        ("Tích lũy", _number(row.get("Tích lũy"))),
        ("Phí sinh hoạt", _number(row.get("Chi Phí Sinh Hoạt"))),
        ("Vi phạm trong kỳ", _number(row.get("Tiền phạt trong tháng"))),
        ("Vi phạm kỳ trước", _number(row.get("Vi phạm kỳ trước"))),
        ("Tiền ứng lương", _number(row.get("Tiền ứng lương"))),
        ("Tiền hỗ trợ Locker", _number(row.get("Tiền hỗ trợ Locker"))),
        ("Số tiền thực nhận", _number(row.get("Số tiền thực nhận"))),
    ])
    return rows


def _payroll_email_text(
    name: str,
    start: date,
    end: date,
    row: dict[str, Any],
    violations: list[dict[str, Any]],
) -> str:
    lines = [
        f"Chào {name},",
        "",
        f"VERA SPA gửi bảng lương kỳ từ {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}.",
        "",
    ]
    lines.extend(f"{label}: {_email_money(amount)}" for label, amount in _payroll_email_summary_rows(row))
    lines.extend(["", f"Số tiền thực nhận: {_email_money(row.get('Số tiền thực nhận'))}"])
    if violations:
        lines.extend(["", "Chi tiết vi phạm trong kỳ:"])
        for item in violations:
            lines.append(
                " - " + " | ".join([
                    _email_date(item.get("leave_date")),
                    str(item.get("leave_reason") or ""),
                    str(item.get("detail") or ""),
                    _email_money(item.get("penalty")),
                ])
            )
        lines.extend(["", f"Tổng vi phạm: {_email_money(row.get('Tiền phạt trong tháng'))}"])
    lines.extend(["", "Vui lòng kiểm tra và phản hồi nếu có sai sót.", "", "Trân trọng,", "VERA SPA"])
    return "\n".join(lines)


def _payroll_email_html(
    name: str,
    start: date,
    end: date,
    row: dict[str, Any],
    violations: list[dict[str, Any]],
) -> str:
    summary_html = "".join(
        "<tr>"
        f"<td style=\"padding:8px 7px;border:1px solid #dddddd;\">{escape(label)}</td>"
        f"<td style=\"padding:8px 7px;border:1px solid #dddddd;text-align:right;white-space:nowrap;\">{escape(_email_money(amount))}</td>"
        "</tr>"
        for label, amount in _payroll_email_summary_rows(row)
    )
    violation_section = ""
    if violations:
        violation_rows = "".join(
            "<tr>"
            f"<td style=\"padding:7px 6px;border:1px solid #dddddd;white-space:nowrap;\">{escape(_email_date(item.get('leave_date')))}</td>"
            f"<td style=\"padding:7px 6px;border:1px solid #dddddd;\">{escape(str(item.get('leave_reason') or ''))}</td>"
            f"<td style=\"padding:7px 6px;border:1px solid #dddddd;\">{escape(str(item.get('detail') or ''))}</td>"
            f"<td style=\"padding:7px 6px;border:1px solid #dddddd;text-align:right;white-space:nowrap;\">{escape(_email_money(item.get('penalty')))}</td>"
            "</tr>"
            for item in violations
        )
        violation_section = f"""
          <p style="margin:26px 0 12px;font-weight:700;">Chi tiết vi phạm trong kỳ:</p>
          <table role="presentation" cellspacing="0" cellpadding="0" style="border-collapse:collapse;width:100%;max-width:620px;font-size:14px;">
            <thead><tr style="background:#a99d97;color:#111111;">
              <th style="padding:8px 6px;border:1px solid #dddddd;width:20%;">Ngày</th>
              <th style="padding:8px 6px;border:1px solid #dddddd;width:44%;">Lý do</th>
              <th style="padding:8px 6px;border:1px solid #dddddd;width:14%;">Chi tiết</th>
              <th style="padding:8px 6px;border:1px solid #dddddd;width:22%;">Phạt</th>
            </tr></thead>
            <tbody>{violation_rows}</tbody>
          </table>
          <p style="margin:14px 0 0;font-weight:700;">Tổng vi phạm: {escape(_email_money(row.get('Tiền phạt trong tháng')))}</p>
        """
    return f"""<!doctype html>
<html lang="vi"><body style="margin:0;padding:0;background:#ffffff;color:#222222;font-family:Arial,Helvetica,sans-serif;font-size:14px;line-height:1.45;">
  <div style="max-width:680px;padding:24px;">
    <p style="margin:0 0 20px;">Chào <strong>{escape(name)}</strong>,</p>
    <p style="margin:0 0 14px;">VERA SPA gửi bảng lương kỳ từ <strong>{start.strftime('%d/%m/%Y')}</strong> đến <strong>{end.strftime('%d/%m/%Y')}</strong>.</p>
    <table role="presentation" cellspacing="0" cellpadding="0" style="border-collapse:collapse;width:100%;max-width:520px;font-size:14px;">
      <thead><tr style="background:#a99d97;color:#111111;">
        <th style="padding:8px 7px;border:1px solid #dddddd;width:56%;">Khoản mục</th>
        <th style="padding:8px 7px;border:1px solid #dddddd;width:44%;">Số tiền</th>
      </tr></thead>
      <tbody>{summary_html}</tbody>
    </table>
    <p style="margin:14px 0 0;font-weight:700;">Số tiền thực nhận: {escape(_email_money(row.get('Số tiền thực nhận')))}</p>
    {violation_section}
    <p style="margin:18px 0 0;">Vui lòng kiểm tra và phản hồi nếu có sai sót.</p>
    <p style="margin:18px 0 0;">Trân trọng,<br><strong>VERA SPA</strong></p>
  </div>
</body></html>"""


def _period(month: str, period_no: int) -> tuple[date, date, str]:
    try:
        year, month_number = (int(item) for item in month.split("-", 1))
        if not 2020 <= year <= 2100 or not 1 <= month_number <= 12 or period_no not in {1, 2}:
            raise ValueError
    except Exception as exc:
        raise HTTPException(400, "Kỳ lương không hợp lệ.") from exc
    start = date(year, month_number, 1 if period_no == 1 else 16)
    end = date(year, month_number, 15 if period_no == 1 else calendar.monthrange(year, month_number)[1])
    return start, end, f"Kỳ {period_no} - Tháng {month_number}/{year}"


def _period_label(start: date, end: date) -> str:
    if start.year != end.year or start.month != end.month:
        raise HTTPException(400, "Chỉ được lưu một kỳ lương trong cùng tháng.")
    expected_end = 15 if start.day == 1 else calendar.monthrange(start.year, start.month)[1]
    period_no = 1 if start.day == 1 and end.day == 15 else 2 if start.day == 16 and end.day == expected_end else 0
    if not period_no:
        raise HTTPException(400, "Kỳ lương phải là ngày 01–15 hoặc ngày 16–cuối tháng.")
    return f"Kỳ {period_no} - Tháng {start.month}/{start.year}"


def _setting(conn, key: str, default: Any) -> Any:
    value = conn.execute(text("""
        SELECT value_json FROM vera_app_setting WHERE category='payroll' AND setting_key=:key LIMIT 1
    """), {"key": key}).scalar_one_or_none()
    return value if value is not None else default


def _put_setting(conn, key: str, value: Any, actor: str) -> None:
    conn.execute(text("""
        INSERT INTO vera_app_setting(category,setting_key,value_json,source,updated_by,revision,created_at,updated_at)
        VALUES ('payroll',:key,CAST(:value AS jsonb),'web_v2',:actor,1,NOW(),NOW())
        ON CONFLICT(category,setting_key) DO UPDATE SET value_json=EXCLUDED.value_json,source='web_v2',
          updated_by=EXCLUDED.updated_by,revision=vera_app_setting.revision+1,updated_at=NOW()
    """), {"key": key, "value": json.dumps(value, ensure_ascii=False), "actor": actor})


def _config(conn) -> dict[str, int]:
    raw = _setting(conn, "config", DEFAULT_CONFIG)
    if not isinstance(raw, dict):
        raw = {}
    return {key: max(0, _number(raw.get(key, value))) for key, value in DEFAULT_CONFIG.items()}


def _payload(conn) -> tuple[list[dict[str, Any]], str, str]:
    cached = conn.execute(text("SELECT payload,checksum,updated_at::text FROM vera_dataset_cache WHERE dataset_key='payroll_history' LIMIT 1")).first()
    records = [dict(item) for item in ((cached[0] if cached else []) or []) if isinstance(item, dict)]
    normalized = conn.execute(text("SELECT payload FROM payroll_history_rows ORDER BY saved_at DESC,id DESC")).scalars().all()
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for item in [dict(value) for value in normalized if isinstance(value, dict)] + records:
        key = (str(item.get("Mã bản lưu") or ""), str(item.get("Tên Hệ thống") or ""))
        by_key.setdefault(key, item)
    return list(by_key.values()), str(cached[1] or "") if cached else "", str(cached[2] or "") if cached else ""


def _sheet_records_from_values(values: list[list[Any]], fallback_headers: list[str]) -> list[dict[str, Any]]:
    if len(values) < 2:
        return []
    headers = [str(value or "").strip() for value in values[0]]
    if not any(headers):
        headers = list(fallback_headers)
    records: list[dict[str, Any]] = []
    for source_row, values_row in enumerate(values[1:], start=2):
        padded = list(values_row) + [""] * max(0, len(headers) - len(values_row))
        item = {
            header: padded[index]
            for index, header in enumerate(headers)
            if header
        }
        if any(str(value or "").strip() for value in item.values()):
            item["__legacy_sheet_row"] = source_row
            records.append(item)
    return records


def _legacy_sheet_datasets(spreadsheet) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read both legacy payroll datasets in one Google Sheets values request."""
    response = spreadsheet.values_batch_get([
        f"'{LEGACY_PAYROLL_WORKSHEET}'!A:AZ",
        f"'{LEGACY_OBLIGATION_WORKSHEET}'!A:N",
    ], params={"majorDimension": "ROWS", "valueRenderOption": "FORMATTED_VALUE"})
    ranges = response.get("valueRanges", []) if isinstance(response, dict) else []
    if len(ranges) != 2:
        raise RuntimeError("Google Sheets không trả đủ BangLuong và NoViPham.")
    payroll_values = ranges[0].get("values", []) if isinstance(ranges[0], dict) else []
    obligation_values = ranges[1].get("values", []) if isinstance(ranges[1], dict) else []
    return (
        _sheet_records_from_values(payroll_values, PUBLIC_FIELDS),
        _sheet_records_from_values(obligation_values, LEGACY_OBLIGATION_HEADERS),
    )


def _write_dataset_cache(conn, dataset_key: str, records: list[dict[str, Any]], source_version: str) -> str:
    serialized = json.dumps(records, ensure_ascii=False, separators=(",", ":"), default=str)
    checksum = hashlib.sha256(serialized.encode("utf-8")).hexdigest()
    conn.execute(text("""
        INSERT INTO vera_dataset_cache(dataset_key,payload,row_count,checksum,source_version,updated_at,expires_at)
        VALUES (:dataset_key,CAST(:payload AS jsonb),:count,:checksum,:source_version,NOW(),NOW()+INTERVAL '3650 days')
        ON CONFLICT(dataset_key) DO UPDATE SET payload=EXCLUDED.payload,row_count=EXCLUDED.row_count,
          checksum=EXCLUDED.checksum,source_version=EXCLUDED.source_version,updated_at=NOW(),expires_at=EXCLUDED.expires_at
    """), {
        "dataset_key": dataset_key, "payload": serialized, "count": len(records),
        "checksum": checksum, "source_version": source_version,
    })
    return checksum


def _visible(records, ident, norm):
    if str(ident.role or "").lower() in {"admin", "quanly", "letan"}:
        return records
    keys = {norm(ident.employee_username), norm(ident.full_name)} - {""}
    return [item for item in records if norm(item.get("Tên Hệ thống")) in keys or norm(item.get("Họ và tên")) in keys]


def _filter_rows(records, batch: str, search: str, norm):
    if batch:
        records = [item for item in records if str(item.get("Mã bản lưu") or "") == batch]
    needle = norm(search)
    if needle:
        records = [item for item in records if needle in {norm(item.get("Tên Hệ thống")), norm(item.get("Họ và tên"))}]
    return records


def _net(row: dict[str, Any]) -> dict[str, Any]:
    result = dict(row)
    for field in MONEY_FIELDS:
        result[field] = _number(result.get(field))
    if result["Tiền Lương"] == 0:
        result["Chi Phí Sinh Hoạt"] = 0
        result["Tiền hỗ trợ Locker"] = 0
    result["Số tiền thực nhận"] = (
        result["Tiền Lương"] + result["Tiền Hỗ Trợ Hoàn Lại"] + result["Hoàn trả tiền tích lũy"]
        - result["Tích lũy"] - result["Chi Phí Sinh Hoạt"] - result["Tiền phạt trong tháng"]
        - result["Vi phạm kỳ trước"] - result["Tiền ứng lương"] - result["Tiền hỗ trợ Locker"]
    )
    return result


def _draft_key(start: date, end: date) -> str:
    return f"draft:{start.isoformat()}:{end.isoformat()}"


def _employee_catalog(conn, norm) -> dict[str, dict[str, Any]]:
    return {
        norm(item["username"]): dict(item)
        for item in conn.execute(text("""
            SELECT username,COALESCE(full_name,'') full_name,COALESCE(email,'') email,
                   COALESCE(bank_account,'') bank_account,COALESCE(bank_name,'') bank_name,
                   COALESCE(payload->>'Trạng thái làm việc',payload->>'employment_status','Đang làm việc') employment_status
            FROM employees
            WHERE lower(COALESCE(role,'')) IN ('nhanvien','leader')
              AND COALESCE(payload->>'__deleted', 'false') <> 'true'
        """)).mappings().all()
    }


def _clean_draft_rows(conn, supplied_rows: list[dict[str, Any]], norm) -> list[dict[str, Any]]:
    employees = _employee_catalog(conn, norm)
    clean_rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for supplied in supplied_rows:
        key = norm(supplied.get("Tên Hệ thống"))
        if not key or key in seen or key not in employees:
            raise HTTPException(400, "Bảng lương có nhân viên trống, trùng tên hoặc không tồn tại.")
        seen.add(key)
        employee = employees[key]
        row = _net({field: supplied.get(field, "") for field in DRAFT_FIELDS})
        row.update({
            "TT": len(clean_rows) + 1,
            "Tên Hệ thống": employee["username"],
            "Họ và tên": employee["full_name"],
            "Email": employee["email"],
            "Số tài khoản ngân hàng": employee["bank_account"],
            "Tên ngân hàng": employee["bank_name"],
        })
        clean_rows.append(row)
    if not clean_rows:
        raise HTTPException(400, "Bảng lương nháp chưa có nhân viên.")
    return clean_rows


def _read_source(content: bytes) -> pd.DataFrame:
    if not content:
        raise HTTPException(400, "File Excel đang trống.")
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(413, "File Excel vượt quá 15 MB.")
    if not content.startswith(b"PK"):
        raise HTTPException(400, "File không đúng định dạng Excel .xlsx. Vui lòng xuất lại từ TimeSoft.")
    workbook = None
    try:
        # Không dùng read_only cho file TimeSoft. Một số file có XML dimension
        # sai A1:K4 dù dữ liệu thật kéo dài hàng nghìn dòng. Normal worksheet
        # quét các cell XML thực tế và không cắt ở dòng dữ liệu đầu tiên 90'.
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        sheet_name = PAYROLL_SOURCE_WORKSHEET
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(400, f"File TimeSoft không có sheet '{sheet_name}'.")
        worksheet = workbook[sheet_name]

        header_row = next(worksheet.iter_rows(
            min_row=3, max_row=3, min_col=1, max_col=len(TIMESOFT_PAYROLL_HEADERS), values_only=True
        ), ())
        actual_headers = [str(value or "").strip() for value in header_row]
        if actual_headers != TIMESOFT_PAYROLL_HEADERS:
            mismatches = [
                f"{get_column_letter(index + 1)}: '{actual}' ≠ '{expected}'"
                for index, (actual, expected) in enumerate(zip(actual_headers, TIMESOFT_PAYROLL_HEADERS))
                if actual != expected
            ]
            raise HTTPException(400, "File TimeSoft không đúng header chuẩn ở dòng 3. " + " | ".join(mismatches[:6]))
        selected_rows: list[tuple[Any, Any, Any, Any]] = []

        for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=9, values_only=True):
            padded = tuple(row) + (None,) * max(0, 9 - len(row))
            time_value, item_value, amount_value, employee_value = padded[1], padded[5], padded[6], padded[8]
            if any(value not in (None, "") for value in (time_value, item_value, amount_value, employee_value)):
                selected_rows.append((time_value, item_value, amount_value, employee_value))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được sheet 'Báo cáo doanh thu hóa đơn': {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()

    output = pd.DataFrame(selected_rows, columns=["time", "item", "amount", "employee"])
    if output.empty:
        raise HTTPException(400, f"Sheet '{PAYROLL_SOURCE_WORKSHEET}' không có dữ liệu từ dòng 4 trở xuống.")
    numeric_input = output["time"].apply(
        lambda value: isinstance(value, numbers.Number) and not isinstance(value, bool)
    )
    numeric_time = pd.to_numeric(output["time"], errors="coerce")
    output["time"] = pd.to_datetime(output["time"], dayfirst=True, errors="coerce", format="mixed")
    numeric_mask = numeric_input & numeric_time.between(1, 100_000, inclusive="both")
    if numeric_mask.any():
        output.loc[numeric_mask, "time"] = (
            pd.Timestamp("1899-12-30") + pd.to_timedelta(numeric_time[numeric_mask], unit="D")
        )
    output["amount"] = output["amount"].apply(_number)
    output["item"] = output["item"].fillna("").astype(str).str.strip()
    output["employee"] = output["employee"].fillna("").astype(str).str.strip()
    return output


def _tip_rows(source: pd.DataFrame, start: date, end: date, norm) -> tuple[pd.DataFrame, dict[str, Any]]:
    dated = source.dropna(subset=["time"]).copy()
    if dated.empty:
        preview = ", ".join(str(value) for value in source["time"].head(5).tolist())
        raise HTTPException(400, f"Không đọc được ngày ở cột B của file TimeSoft. Giá trị đầu tiên: {preview}")
    filtered = dated[(dated["time"].dt.date >= start) & (dated["time"].dt.date <= end)].copy()
    if filtered.empty:
        source_start = dated["time"].min().strftime("%d/%m/%Y")
        source_end = dated["time"].max().strftime("%d/%m/%Y")
        raise HTTPException(
            400,
            f"File TimeSoft có dữ liệu {source_start}–{source_end}, không thuộc kỳ "
            f"{start.strftime('%d/%m/%Y')}–{end.strftime('%d/%m/%Y')}.",
        )
    item_norm = filtered["item"].astype(str).str.strip().str.casefold()
    tips = filtered[item_norm.str.match(TIP_ITEM_PATTERN, na=False)].copy()
    if tips.empty:
        preview = ", ".join(
            filtered["item"].loc[lambda values: values.astype(str).str.strip().ne("")]
            .astype(str).drop_duplicates().head(10).tolist()
        )
        raise HTTPException(
            400,
            "File TimeSoft không có dòng Tip hợp lệ ở cột F trong kỳ đã chọn. "
            f"Giá trị đầu tiên: {preview}",
        )
    tips["key"] = tips["employee"].apply(norm)
    if not tips["key"].astype(str).str.strip().ne("").any():
        raise HTTPException(400, "Các dòng Tip chưa có tên nhân viên ở cột I (NV tư vấn).")
    salary_total = int(tips["amount"].sum())
    if salary_total <= 0:
        raise HTTPException(400, "Tổng tiền các dòng Tip đang bằng 0; hệ thống dừng để tránh tạo bảng lương sai.")
    return tips, {
        "source_rows": int(len(source)),
        "dated_rows": int(len(dated)),
        "period_rows": int(len(filtered)),
        "tip_rows": int(len(tips)),
        "salary_total": salary_total,
    }


def _tichluy_map(conn, employees: list[dict[str, Any]], start: date, end: date, norm) -> dict[str, int]:
    payload = conn.execute(text("SELECT payload FROM vera_dataset_cache WHERE dataset_key='tichluy' LIMIT 1")).scalar_one_or_none()
    source = {norm(item.get("Tên nhân viên")): item for item in (payload or []) if isinstance(item, dict)}
    result: dict[str, int] = {}
    for employee in employees:
        key = norm(employee["username"])
        item = source.get(key)
        if not item:
            result[key] = 0
            continue
        target = _number(item.get("Mục tiêu tích lũy")) or 5_000_000
        accumulated = _number(item.get("Đã tích lũy"))
        # D/E là nguồn xác nhận hoàn thành. Một số dòng legacy còn giữ F=Còn lại=5.000.000
        # dù Đã tích lũy đã đủ mục tiêu; không được tiếp tục khấu trừ các dòng này.
        completed = target > 0 and accumulated >= target
        if completed:
            remaining = 0
        else:
            remaining_raw = str(item.get("Còn lại") or "").strip()
            remaining = (
                _number(remaining_raw)
                if remaining_raw and remaining_raw not in {"-", "–", "—"}
                else max(0, target - accumulated)
            )
            remaining = max(0, min(remaining, max(0, target - accumulated)))
        history = {}
        try:
            history = json.loads(str(item.get("Chi tiết các kỳ") or "{}"))
        except Exception:
            pass
        period_key = f"{start.isoformat()}|{end.isoformat()}"
        if period_key in history:
            result[key] = max(0, min(500_000, _number(history[period_key])))
            continue
        start_work = _parse_date(employee.get("employment_start_date")) or _parse_date(item.get("Ngày bắt đầu làm"))
        if start_work and start <= start_work <= end and (end - start_work).days + 1 < 10:
            result[key] = 0
        else:
            result[key] = max(0, min(500_000, remaining))
    return result


def _obligations(conn) -> list[dict[str, Any]]:
    custom = _setting(conn, "violation_obligations", [])
    return [dict(item) for item in (custom or []) if isinstance(item, dict)]


def _accumulation_refunds(conn) -> list[dict[str, Any]]:
    custom = _setting(conn, "accumulation_refunds", [])
    return [dict(item) for item in (custom or []) if isinstance(item, dict)]


def _accumulation_refund_map(conn, start: date, end: date, norm) -> dict[str, int]:
    result: dict[str, int] = {}
    for item in _accumulation_refunds(conn):
        if _parse_date(item.get("start")) != start or _parse_date(item.get("end")) != end:
            continue
        key = norm(item.get("employee_name"))
        if key:
            result[key] = result.get(key, 0) + max(0, _number(item.get("amount")))
    return result


def _legacy_obligations(conn) -> list[dict[str, Any]]:
    payload = conn.execute(text(
        "SELECT payload FROM vera_dataset_cache WHERE dataset_key='violation_debt' LIMIT 1"
    )).scalar_one_or_none()
    return [dict(item) for item in (payload or []) if isinstance(item, dict)]


def _open_obligation(value: Any, norm) -> bool:
    return norm(value or "Chưa hoàn thành") in {"", "chua hoan thanh"}


def _obligation_group(records: list[dict[str, Any]], obligation_type: str, norm) -> dict[str, Any]:
    details: list[dict[str, Any]] = []
    for item in records:
        if norm(item.get("Loại")) != norm(obligation_type) or not _open_obligation(item.get("Trạng thái"), norm):
            continue
        employee_name = str(item.get("Tên nhân viên") or "").strip()
        amount = max(0, _number(item.get("Số tiền")))
        if not employee_name or amount <= 0:
            continue
        details.append({
            "employee_name": employee_name,
            "amount": amount,
            "period_start": str(item.get("Kỳ phát sinh từ") or "").strip(),
            "period_end": str(item.get("Kỳ phát sinh đến") or "").strip(),
            "due_from": str(item.get("Bắt đầu trừ từ") or "").strip(),
            "content": str(item.get("Nội dung") or "Chưa hoàn thành nghĩa vụ Vi phạm").strip(),
            "type": str(item.get("Loại") or obligation_type).strip(),
            "status": str(item.get("Trạng thái") or "Chưa hoàn thành").strip(),
        })
    details.sort(key=lambda item: (_parse_date(item["period_start"]) or date.max, norm(item["employee_name"])))
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in details:
        grouped.setdefault(norm(item["employee_name"]), []).append(item)
    summary = []
    for items in grouped.values():
        latest = max(items, key=lambda item: _parse_date(item["period_start"]) or date.min)
        due_dates = [_parse_date(item["due_from"]) for item in items]
        due_dates = [value for value in due_dates if value]
        summary.append({
            "employee_name": items[0]["employee_name"],
            "total": sum(item["amount"] for item in items),
            "period_count": len(items),
            "latest_period": f'{latest["period_start"]} - {latest["period_end"]}',
            "due_from": min(due_dates).strftime("%d/%m/%Y") if due_dates else latest["due_from"],
        })
    summary.sort(key=lambda item: (-item["total"], norm(item["employee_name"])))
    return {"type": obligation_type, "summary": summary, "details": details}


def _obligation_map(conn, start: date, norm) -> dict[str, int]:
    result: dict[str, int] = {}
    sources = list(_obligations(conn))
    legacy = conn.execute(text("SELECT payload FROM vera_dataset_cache WHERE dataset_key='violation_debt' LIMIT 1")).scalar_one_or_none()
    sources.extend(dict(item) for item in (legacy or []) if isinstance(item, dict))
    for item in sources:
        status = norm(item.get("status") or item.get("Trạng thái") or "Chưa hoàn thành")
        if status not in {"", "chua hoan thanh"}:
            continue
        due = _parse_date(item.get("due_from") or item.get("Bắt đầu trừ từ"))
        if due and due > start:
            continue
        key = norm(item.get("employee_name") or item.get("Tên nhân viên"))
        if key:
            result[key] = result.get(key, 0) + max(0, _number(item.get("amount") or item.get("Số tiền")))
    return result


def _workbook(records: list[dict[str, Any]], fields: list[str], title: str = "Bảng lương") -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = title[:31]; ws.append(fields)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F513F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for item in records:
        ws.append([_number(item.get(key)) if key in MONEY_FIELDS else item.get(key, "") for key in fields])
    for column_index, field in enumerate(fields, start=1):
        letter = get_column_letter(column_index)
        cells = ws[letter]
        if field in MONEY_FIELDS:
            for cell in cells[1:]:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            width = 18
        else:
            max_length = max([len(str(field))] + [len(str(cell.value or "")) for cell in cells[1:]])
            width = min(max(max_length + 2, 12), 34)
        ws.column_dimensions[letter].width = width
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
        for row_index in range(2, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 21
    stream = BytesIO()
    wb.save(stream)
    wb.close()
    return stream.getvalue()


def _new_payroll_sheet_title(start: date) -> str:
    period_no = 1 if start.day == 1 else 2
    return f"Bảng lương bản mới K{period_no} Tháng {start.month}"[:31]


def _new_payroll_workbook(records: list[dict[str, Any]], fields: list[str], start: date, end: date) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _new_payroll_sheet_title(start)
    ws.sheet_view.showGridLines = False
    last_column = get_column_letter(len(fields))

    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = "BẢNG LƯƠNG NHÂN VIÊN"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F513F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = "KỲ LƯƠNG"
    ws["A2"].font = Font(bold=True, color="1F513F")
    ws["A2"].fill = PatternFill("solid", fgColor="E5EFEA")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    if len(fields) > 1:
        ws.merge_cells(f"B2:{last_column}2")
    ws["B2"] = f"Từ ngày {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}"
    ws["B2"].font = Font(bold=True, color="1F513F")
    ws["B2"].fill = PatternFill("solid", fgColor="E5EFEA")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="E5EFEA")
    ws.row_dimensions[2].height = 24

    header_row = 4
    for column_index, field in enumerate(fields, start=1):
        cell = ws.cell(header_row, column_index, field)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F513F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    for item in records:
        ws.append([_number(item.get(key)) if key in MONEY_FIELDS else item.get(key, "") for key in fields])

    for column_index, field in enumerate(fields, start=1):
        letter = get_column_letter(column_index)
        data_cells = [ws.cell(row_index, column_index) for row_index in range(header_row + 1, ws.max_row + 1)]
        if field in MONEY_FIELDS:
            for cell in data_cells:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            width = 18
        elif field == "TT":
            for cell in data_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            width = 7
        else:
            max_length = max([len(str(field))] + [len(str(cell.value or "")) for cell in data_cells])
            width = min(max(max_length + 2, 12), 34)
        ws.column_dimensions[letter].width = width

    if ws.max_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{last_column}{ws.max_row}"
        for row_index in range(header_row + 1, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 21
    ws.freeze_panes = f"A{header_row + 1}"
    stream = BytesIO()
    wb.save(stream)
    wb.close()
    return stream.getvalue()


def _saved_draft(conn, start: date, end: date) -> dict[str, Any] | None:
    stored = conn.execute(text("""
        SELECT value_json,updated_by,updated_at
        FROM vera_app_setting
        WHERE category='payroll' AND setting_key=:key
        LIMIT 1
    """), {"key": _draft_key(start, end)}).mappings().first()
    if not stored or not isinstance(stored["value_json"], dict):
        return None
    value = dict(stored["value_json"])
    rows = [
        _net({field: item.get(field, "") for field in DRAFT_FIELDS})
        for item in (value.get("rows") or [])
        if isinstance(item, dict)
    ]
    if not rows:
        return None
    return {
        "period_label": _period_label(start, end),
        "start": start.isoformat(),
        "end": end.isoformat(),
        "fields": DRAFT_FIELDS,
        "editable_fields": sorted(EDITABLE_FIELDS),
        "rows": rows,
        "source_name": str(value.get("source_name") or "Bảng lương nháp"),
        "saved_at": stored["updated_at"].isoformat() if stored["updated_at"] else "",
        "saved_by": str(stored["updated_by"] or ""),
    }


def _read_draft_workbook(content: bytes) -> tuple[list[dict[str, Any]], set[str], set[tuple[date, date]]]:
    if not content:
        raise HTTPException(400, "File Excel đang trống.")
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(413, "File Excel vượt quá 15 MB.")
    if not content.startswith(b"PK"):
        raise HTTPException(400, "File không đúng định dạng Excel .xlsx.")
    workbook = None
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        source_rows = list(worksheet.iter_rows(values_only=True))
        header_index = next((
            index for index, values in enumerate(source_rows[:12])
            if "Tên Hệ thống" in [str(value or "").strip() for value in values]
        ), None)
        if header_index is None:
            raise HTTPException(400, "File Import thiếu cột 'Tên Hệ thống'. Hãy dùng file Export to Excel từ Bảng lương.")
        headers = [str(value or "").strip() for value in source_rows[header_index]]
        nonempty_headers = [header for header in headers if header]
        if len(nonempty_headers) != len(set(nonempty_headers)):
            raise HTTPException(400, "File Import có tiêu đề cột bị trùng.")
        rows: list[dict[str, Any]] = []
        period_labels: set[str] = set()
        period_ranges: set[tuple[date, date]] = set()
        for values in source_rows[:header_index]:
            text_value = " ".join(str(value or "").strip() for value in values if str(value or "").strip())
            match = re.search(r"Từ ngày\s+(\d{2}/\d{2}/\d{4})\s+đến\s+(\d{2}/\d{2}/\d{4})", text_value, re.IGNORECASE)
            if match:
                parsed_start, parsed_end = _parse_date(match.group(1)), _parse_date(match.group(2))
                if parsed_start and parsed_end:
                    period_ranges.add((parsed_start, parsed_end))
        for row_number, values in enumerate(source_rows[header_index + 1:], start=header_index + 2):
            if row_number > header_index + 501:
                raise HTTPException(400, "File Import vượt quá 500 dòng bảng lương.")
            padded = list(values) + [None] * max(0, len(headers) - len(values))
            item = {header: padded[index] for index, header in enumerate(headers) if header}
            if not any(str(value or "").strip() for value in item.values()):
                continue
            employee_name = str(item.get("Tên Hệ thống") or "").strip()
            if not employee_name:
                raise HTTPException(400, f"Dòng {row_number} chưa có Tên Hệ thống.")
            item["Tên Hệ thống"] = employee_name
            label = str(item.get("Mã bản lưu") or "").strip()
            if label:
                period_labels.add(label)
            rows.append(item)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được file Bảng lương: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()
    if not rows:
        raise HTTPException(400, "File Import không có dòng bảng lương.")
    return rows, period_labels, period_ranges


def install_payroll_routes(app, *, engine_instance: Callable[[], Any], current_identity, require_feature, norm, identity_type, google_client):
    @app.get("/v2/payroll/history")
    def payroll_history(batch: str = Query(default=""), search: str = Query(default=""), ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_history")
            records, checksum, updated_at = _payload(conn)
        visible_records = _visible(records, ident, norm)
        batches = list(dict.fromkeys(str(item.get("Mã bản lưu") or "") for item in visible_records if str(item.get("Mã bản lưu") or "")))
        employees = sorted({str(item.get("Tên Hệ thống") or "") for item in visible_records if str(item.get("Tên Hệ thống") or "")}, key=norm)
        records = _filter_rows(visible_records, batch, search, norm)
        fields = ADMIN_FIELDS if str(ident.role).lower() == "admin" else PUBLIC_FIELDS
        clean = [{key: (_number(item.get(key)) if key in MONEY_FIELDS else str(item.get(key) or "")) for key in fields} for item in records]
        return {"records": clean, "batches": batches, "employees": employees, "fields": fields, "count": len(clean), "checksum": checksum, "updated_at": updated_at}

    @app.get("/v2/payroll/history/export.xlsx")
    def payroll_export(batch: str = Query(default=""), search: str = Query(default=""), ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_export")
            records, _checksum, _updated = _payload(conn)
        records = _filter_rows(_visible(records, ident, norm), batch, search, norm)
        fields = ADMIN_FIELDS if str(ident.role).lower() == "admin" else PUBLIC_FIELDS
        filename = "VERA_BangLuong_BanCu.xlsx" if not batch else f"VERA_BangLuong_BanCu_{batch.replace(' ', '_').replace('/', '-')}.xlsx"
        return StreamingResponse(BytesIO(_workbook(records, fields, "Bảng lương bản cũ")), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})

    @app.get("/v2/payroll/draft")
    def payroll_draft(month: str = Query(...), period_no: int = Query(..., ge=1, le=2), ident: identity_type = Depends(current_identity)):
        start, end, label = _period(month, period_no)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            draft = _saved_draft(conn, start, end)
        return {"draft": draft, "period_label": label}

    @app.put("/v2/payroll/draft")
    def save_payroll_draft(body: PayrollSave, ident: identity_type = Depends(current_identity)):
        label = _period_label(body.start, body.end)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:v2:payroll-draft:' || :label))"), {"label": label})
            clean_rows = _clean_draft_rows(conn, body.rows, norm)
            _put_setting(conn, _draft_key(body.start, body.end), {
                "start": body.start.isoformat(),
                "end": body.end.isoformat(),
                "source_name": body.source_name,
                "rows": clean_rows,
            }, ident.employee_username)
            draft = _saved_draft(conn, body.start, body.end)
        return {"ok": True, "draft": draft, "message": f"Đã lưu bảng lương nháp {label} cho {len(clean_rows)} nhân viên."}

    @app.delete("/v2/payroll/draft")
    def delete_payroll_draft(month: str = Query(...), period_no: int = Query(..., ge=1, le=2), ident: identity_type = Depends(current_identity)):
        start, end, label = _period(month, period_no)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_save")
            deleted = conn.execute(text("""
                DELETE FROM vera_app_setting
                WHERE category='payroll' AND setting_key=:key
            """), {"key": _draft_key(start, end)}).rowcount
        message = f"Đã xóa bảng lương nháp {label}." if deleted else f"Đã bỏ bảng lương nháp {label} khỏi màn hình; máy chủ chưa có bản nháp đã lưu."
        return {"ok": True, "deleted": bool(deleted), "message": message}

    @app.post("/v2/payroll/draft/import.xlsx")
    async def import_payroll_draft(month: str = Query(...), period_no: int = Query(..., ge=1, le=2), payload: bytes = Body(..., media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ident: identity_type = Depends(current_identity)):
        start, end, label = _period(month, period_no)
        supplied_rows, labels, period_ranges = _read_draft_workbook(payload)
        if labels and labels != {label}:
            raise HTTPException(400, f"File Excel thuộc kỳ {', '.join(sorted(labels))}; màn hình đang chọn {label}.")
        if period_ranges and period_ranges != {(start, end)}:
            descriptions = ", ".join(
                f"{period_start.strftime('%d/%m/%Y')}–{period_end.strftime('%d/%m/%Y')}"
                for period_start, period_end in sorted(period_ranges)
            )
            raise HTTPException(400, f"File Excel thuộc kỳ {descriptions}; màn hình đang chọn {label}.")
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            clean_rows = _clean_draft_rows(conn, supplied_rows, norm)
        return {
            "period_label": label,
            "start": start.isoformat(),
            "end": end.isoformat(),
            "fields": DRAFT_FIELDS,
            "editable_fields": sorted(EDITABLE_FIELDS),
            "rows": clean_rows,
            "source_name": "Excel Import",
            "imported": len(clean_rows),
            "message": f"Đã Import {len(clean_rows)} dòng vào bảng lương nháp {label}.",
        }

    @app.post("/v2/payroll/draft/export.xlsx")
    def payroll_draft_export(body: PayrollExport, ident: identity_type = Depends(current_identity)):
        label = _period_label(body.start, body.end)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_export")
        records = []
        for supplied in _visible(body.rows, ident, norm):
            row = _net({field: supplied.get(field, "") for field in DRAFT_FIELDS})
            row.update({
                "Từ ngày": body.start.strftime("%d/%m/%Y"),
                "Đến ngày": body.end.strftime("%d/%m/%Y"),
            })
            records.append(row)
        if not records:
            raise HTTPException(400, "Không có dữ liệu bảng lương mới để xuất.")
        fields = ["Từ ngày", "Đến ngày"] + [
            field for field in DRAFT_FIELDS if field not in {"Email", "Số dòng Tip"}
        ]
        if str(ident.role or "").lower() != "admin":
            fields = [field for field in fields if field not in {"Số tài khoản ngân hàng", "Tên ngân hàng"}]
        filename = f"VERA_BangLuong_BanMoi_{label.replace(' ', '_').replace('/', '-')}.xlsx"
        return StreamingResponse(
            BytesIO(_new_payroll_workbook(records, fields, body.start, body.end)),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    @app.post("/v2/payroll/history/sync-legacy")
    def sync_legacy_payroll(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_history_edit")
        try:
            spreadsheet = google_client().open_by_key(LEGACY_SPREADSHEET_ID)
            payroll_records, obligation_records = _legacy_sheet_datasets(spreadsheet)
        except Exception as exc:
            raise HTTPException(502, f"Không đọc được dữ liệu bảng lương cũ từ Google Sheets: {str(exc)[:300]}") from exc

        for item in payroll_records:
            item.pop("__legacy_sheet_row", None)
            item.setdefault("Hoàn trả tiền tích lũy", 0)
        for item in obligation_records:
            item.pop("__legacy_sheet_row", None)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_history_edit")
            _write_dataset_cache(conn, "payroll_history", payroll_records, "legacy_google_sheet")
            _write_dataset_cache(conn, "violation_debt", obligation_records, "legacy_google_sheet")
        return {
            "ok": True,
            "payroll_count": len(payroll_records),
            "obligation_count": len(obligation_records),
            "message": f"Đã tải {len(payroll_records)} dòng lịch sử lương và {len(obligation_records)} dòng nghĩa vụ từ hệ thống cũ.",
        }

    @app.get("/v2/payroll/config")
    def payroll_config(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll")
            return {"config": _config(conn)}

    @app.put("/v2/payroll/config")
    def save_payroll_config(body: PayrollConfigUpdate, ident: identity_type = Depends(current_identity)):
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_config_edit")
            value = {key: _number(item) for key, item in body.model_dump().items()}
            _put_setting(conn, "config", value, ident.employee_username)
        return {"ok": True, "config": value, "message": "Đã lưu cài đặt khấu trừ mặc định và tiền trách nhiệm Leader."}

    @app.get("/v2/payroll/accumulation-refunds")
    def accumulation_refunds(ident: identity_type = Depends(current_identity)):
        if str(ident.role or "").lower() != "admin":
            raise HTTPException(403, "Chỉ Admin được cài đặt hoàn trả tiền tích lũy khi nghỉ việc.")
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_config_edit")
            employees = [
                item for item in _employee_catalog(conn, norm).values()
                if norm(item.get("employment_status")) != "dang lam viec"
            ]
            refunds = _accumulation_refunds(conn)
        employees.sort(key=lambda item: norm(item["username"]))
        refunds.sort(key=lambda item: (_parse_date(item.get("start")) or date.max, norm(item.get("employee_name"))))
        return {
            "refunds": refunds,
            "employees": [
                {
                    "employee_name": item["username"],
                    "full_name": item["full_name"],
                    "employment_status": item["employment_status"],
                }
                for item in employees
            ],
        }

    @app.post("/v2/payroll/accumulation-refunds")
    def create_accumulation_refund(body: AccumulationRefundCreate, ident: identity_type = Depends(current_identity)):
        if str(ident.role or "").lower() != "admin":
            raise HTTPException(403, "Chỉ Admin được cài đặt hoàn trả tiền tích lũy khi nghỉ việc.")
        label = _period_label(body.start, body.end)
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_config_edit")
            employees = _employee_catalog(conn, norm)
            key = norm(body.employee_name)
            employee = employees.get(key)
            if not employee:
                raise HTTPException(400, "Tên nhân viên không tồn tại trong hồ sơ Nhân viên/Leader.")
            if norm(employee.get("employment_status")) == "dang lam viec":
                raise HTTPException(400, "Chỉ được hoàn trả tích lũy cho nhân viên Tạm thời nghỉ việc hoặc Đã nghỉ việc.")
            rows = _accumulation_refunds(conn)
            existing = next((
                item for item in rows
                if norm(item.get("employee_name")) == key
                and _parse_date(item.get("start")) == body.start
                and _parse_date(item.get("end")) == body.end
            ), None)
            item = {
                "id": str(existing.get("id") if existing else uuid.uuid4()),
                "employee_name": employee["username"],
                "amount": _number(body.amount),
                "start": body.start.isoformat(),
                "end": body.end.isoformat(),
                "period_label": label,
                "note": body.note.strip() or "Hoàn trả tiền tích lũy khi nghỉ việc",
                "updated_by": ident.employee_username,
            }
            keep = [
                value for value in rows
                if not (
                    norm(value.get("employee_name")) == key
                    and _parse_date(value.get("start")) == body.start
                    and _parse_date(value.get("end")) == body.end
                )
            ]
            keep.append(item)
            _put_setting(conn, "accumulation_refunds", keep, ident.employee_username)
        return {"ok": True, "refund": item, "message": f"Đã cài hoàn trả tích lũy cho {employee['username']} trong {label}."}

    @app.delete("/v2/payroll/accumulation-refunds/{refund_id}")
    def delete_accumulation_refund(refund_id: str, ident: identity_type = Depends(current_identity)):
        if str(ident.role or "").lower() != "admin":
            raise HTTPException(403, "Chỉ Admin được xóa cài đặt hoàn trả tiền tích lũy.")
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_config_edit")
            rows = _accumulation_refunds(conn)
            keep = [item for item in rows if str(item.get("id")) != refund_id]
            if len(keep) == len(rows):
                raise HTTPException(404, "Không tìm thấy cài đặt hoàn trả tiền tích lũy.")
            _put_setting(conn, "accumulation_refunds", keep, ident.employee_username)
        return {"ok": True, "message": "Đã xóa cài đặt hoàn trả tiền tích lũy."}

    @app.post("/v2/payroll/calculate")
    async def calculate_payroll(month: str = Query(...), period_no: int = Query(..., ge=1, le=2), payload: bytes = Body(..., media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ident: identity_type = Depends(current_identity)):
        start, end, label = _period(month, period_no)
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_calculate")
            source = _read_source(payload)
            cfg = _config(conn)
            employees = [dict(row) for row in conn.execute(text("""
                SELECT username,COALESCE(full_name,'') full_name,lower(COALESCE(role,'')) role,
                       COALESCE(email,'') email,COALESCE(bank_account,'') bank_account,COALESCE(bank_name,'') bank_name,
                       COALESCE(employment_start_date,'') employment_start_date
                FROM employees
                WHERE lower(COALESCE(role,'')) IN ('nhanvien','leader')
                  AND COALESCE(payload->>'__deleted', 'false') <> 'true'
                ORDER BY COALESCE(stt,2147483647),username
            """)).mappings().all()]
            penalties = conn.execute(text("""
                SELECT employee_name,SUM(COALESCE(penalty,0)) amount FROM leave_records
                WHERE leave_date BETWEEN :start AND :end GROUP BY employee_name
            """), {"start": start, "end": end}).mappings().all()
            penalty_map = {norm(item["employee_name"]): _number(item["amount"]) for item in penalties}
            tichluy = _tichluy_map(conn, employees, start, end, norm)
            obligations = _obligation_map(conn, start, norm)
            accumulation_refunds = _accumulation_refund_map(conn, start, end, norm)
        tips, source_summary = _tip_rows(source, start, end, norm)
        known = {norm(item["username"]) for item in employees}
        matched_tips = tips[tips["key"].isin(known)].copy()
        unmatched = sorted({
            value for value in tips.loc[~tips["key"].isin(known), "employee"].tolist() if value
        }, key=norm)
        if matched_tips.empty:
            preview = ", ".join(unmatched[:10])
            raise HTTPException(
                400,
                "Đã đọc được dòng Tip nhưng không tên nào khớp Tên Hệ thống của Nhân viên/Leader. "
                f"Tên TimeSoft đầu tiên: {preview}",
            )
        matched_salary_total = int(matched_tips["amount"].sum())
        if matched_salary_total <= 0:
            raise HTTPException(400, "Tổng Tiền Lương sau khi khớp tên nhân viên đang bằng 0; hệ thống dừng tính.")
        source_summary.update({
            "matched_tip_rows": int(len(matched_tips)),
            "matched_salary_total": matched_salary_total,
            "employee_rows": int(len(employees)),
            "unmatched_tip_names": int(len(unmatched)),
        })
        salaries = matched_tips.groupby("key")["amount"].sum().to_dict()
        counts = matched_tips.groupby("key").size().to_dict()
        rows = []
        for index, employee in enumerate(employees, start=1):
            key = norm(employee["username"])
            accumulation_refund = accumulation_refunds.get(key, 0)
            row = {
                "TT": index, "Tên Hệ thống": employee["username"], "Họ và tên": employee["full_name"],
                "Tiền Lương": _number(salaries.get(key)),
                "Tiền Hỗ Trợ Hoàn Lại": cfg["leader_responsibility_allowance"] if employee["role"] == "leader" and period_no == 2 else 0,
                "Hoàn trả tiền tích lũy": accumulation_refund,
                "Tích lũy": 0 if accumulation_refund else tichluy.get(key, 0),
                "Chi Phí Sinh Hoạt": cfg["default_living_expense"], "Tiền phạt trong tháng": penalty_map.get(key, 0),
                "Vi phạm kỳ trước": obligations.get(key, 0), "Tiền ứng lương": 0,
                "Tiền hỗ trợ Locker": cfg["default_locker_support"], "Số tiền thực nhận": 0,
                "Email": employee["email"], "Số tài khoản ngân hàng": employee["bank_account"],
                "Tên ngân hàng": employee["bank_name"], "Số dòng Tip": int(counts.get(key, 0)),
            }
            rows.append(_net(row))
        return {
            "period_label": label, "start": start.isoformat(), "end": end.isoformat(),
            "fields": DRAFT_FIELDS, "editable_fields": sorted(EDITABLE_FIELDS),
            "rows": rows, "unmatched": unmatched, "config": cfg, "source_summary": source_summary,
        }

    @app.post("/v2/payroll/save")
    def save_payroll(body: PayrollSave, ident: identity_type = Depends(current_identity)):
        label = _period_label(body.start, body.end)
        now = datetime.now(timezone(timedelta(hours=7)))
        clean_rows = []
        engine = engine_instance(); conn = engine.connect(); tx = conn.begin()
        try:
            require_feature(conn, ident, "payroll_save")
            conn.execute(text("SELECT pg_advisory_xact_lock(hashtext('vera:v2:payroll:' || :label))"), {"label": label})
            prepared_rows = _clean_draft_rows(conn, body.rows, norm)
            if sum(_number(row.get("Tiền Lương")) for row in prepared_rows) <= 0:
                raise HTTPException(
                    400,
                    "Tổng Tiền Lương đang bằng 0. Hệ thống không lưu bảng lương chính thức; "
                    "hãy kiểm tra lại file TimeSoft và số dòng Tip.",
                )
            for row in prepared_rows:
                row.update({
                    "Mã bản lưu": label,
                    "Từ ngày": body.start.strftime("%d/%m/%Y"), "Đến ngày": body.end.strftime("%d/%m/%Y"),
                    "Ngày lưu": now.strftime("%d/%m/%Y"), "Giờ lưu": now.strftime("%H:%M:%S"),
                    "Người lưu": ident.employee_username, "Nguồn dữ liệu": body.source_name,
                })
                clean_rows.append(row)
            existing, _checksum, _updated = _payload(conn)
            merged = [item for item in existing if str(item.get("Mã bản lưu") or "") != label] + clean_rows
            serialized = json.dumps(merged, ensure_ascii=False, separators=(",", ":"), default=str)
            checksum = hashlib.sha256(serialized.encode("utf-8")).hexdigest()
            conn.execute(text("DELETE FROM payroll_history_rows WHERE batch_id=:label"), {"label": label})
            for row in clean_rows:
                conn.execute(text("""
                    INSERT INTO payroll_history_rows(batch_id,employee_name,period_start,period_end,payload,saved_at)
                    VALUES (:label,:employee,:start,:end,CAST(:payload AS jsonb),NOW())
                """), {"label": label, "employee": row["Tên Hệ thống"], "start": body.start, "end": body.end, "payload": json.dumps(row, ensure_ascii=False)})
            conn.execute(text("""
                INSERT INTO vera_dataset_cache(dataset_key,payload,row_count,checksum,source_version,updated_at,expires_at)
                VALUES ('payroll_history',CAST(:payload AS jsonb),:count,:checksum,'web_v2',NOW(),NOW()+INTERVAL '3650 days')
                ON CONFLICT(dataset_key) DO UPDATE SET payload=EXCLUDED.payload,row_count=EXCLUDED.row_count,
                  checksum=EXCLUDED.checksum,source_version='web_v2',updated_at=NOW(),expires_at=EXCLUDED.expires_at
            """), {"payload": serialized, "count": len(merged), "checksum": checksum})
            tx.commit()
        except Exception:
            if tx.is_active:
                tx.rollback()
            raise
        finally:
            conn.close()
        return {"ok": True, "batch": label, "saved": len(clean_rows), "message": f"Đã lưu {label} cho {len(clean_rows)} nhân viên."}

    @app.get("/v2/payroll/obligations")
    def obligations(ident: identity_type = Depends(current_identity)):
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_penalty_obligation")
            rows = _obligations(conn)
            legacy_rows = _legacy_obligations(conn)
        groups = [
            _obligation_group(legacy_rows, "Âm thực nhận", norm),
            _obligation_group(legacy_rows, "Tạm hoãn vi phạm", norm),
        ]
        return {
            "obligations": rows, "count": len(rows), "legacy_count": len(legacy_rows),
            "groups": groups,
        }

    @app.post("/v2/payroll/obligations")
    def create_obligation(body: ObligationCreate, ident: identity_type = Depends(current_identity)):
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_penalty_obligation")
            canonical_name = conn.execute(text("""
                SELECT username FROM employees
                WHERE lower(btrim(username))=lower(btrim(:username))
                  AND lower(COALESCE(role,'')) IN ('nhanvien','leader') LIMIT 1
            """), {"username": body.employee_name.strip()}).scalar_one_or_none()
            if not canonical_name:
                raise HTTPException(400, "Tên nhân viên không khớp chính xác với hồ sơ Nhân viên/Leader.")
            rows = _obligations(conn)
            item = {"id": str(uuid.uuid4()), "employee_name": canonical_name, "amount": _number(body.amount), "content": body.content.strip(), "due_from": body.due_from.isoformat(), "status": "Chưa hoàn thành", "updated_by": ident.employee_username}
            rows.append(item)
            _put_setting(conn, "violation_obligations", rows, ident.employee_username)
        return {"ok": True, "obligation": item, "message": "Đã thêm Nghĩa vụ vi phạm."}

    @app.delete("/v2/payroll/obligations/{obligation_id}")
    def delete_obligation(obligation_id: str, ident: identity_type = Depends(current_identity)):
        with engine_instance().begin() as conn:
            require_feature(conn, ident, "payroll_penalty_obligation")
            rows = _obligations(conn)
            keep = [item for item in rows if str(item.get("id")) != obligation_id]
            if len(keep) == len(rows):
                raise HTTPException(404, "Không tìm thấy Nghĩa vụ vi phạm.")
            _put_setting(conn, "violation_obligations", keep, ident.employee_username)
        return {"ok": True, "message": "Đã xóa Nghĩa vụ vi phạm."}

    @app.post("/v2/payroll/email")
    def email_payroll(body: PayrollEmail, ident: identity_type = Depends(current_identity)):
        label = _period_label(body.start, body.end)
        sender = os.getenv("SMTP_SENDER_EMAIL", "veraspabienhoa@gmail.com").strip()
        password = os.getenv("SMTP_APP_PASSWORD", "")
        if not password:
            raise HTTPException(503, "Máy chủ chưa cấu hình mật khẩu gửi email bảng lương.")
        with engine_instance().connect() as conn:
            require_feature(conn, ident, "payroll_email")
            employee_rows = conn.execute(text("""
                SELECT username,COALESCE(full_name,'') full_name,COALESCE(email,'') email,
                       COALESCE(bank_account,'') bank_account,COALESCE(bank_name,'') bank_name
                FROM employees WHERE lower(COALESCE(role,'')) IN ('nhanvien','leader')
            """)).mappings().all()
            violation_rows = conn.execute(text("""
                SELECT employee_name,leave_date,leave_reason,detail,COALESCE(penalty,0) penalty
                FROM leave_records
                WHERE leave_date BETWEEN :start AND :end
                  AND COALESCE(penalty,0) > 0
                ORDER BY leave_date,COALESCE(source_row,0),id
            """), {"start": body.start, "end": body.end}).mappings().all()
        employees = {norm(item["username"]): dict(item) for item in employee_rows}
        violations_by_employee: dict[str, list[dict[str, Any]]] = {}
        for item in violation_rows:
            violations_by_employee.setdefault(norm(item["employee_name"]), []).append(dict(item))
        sent, failed = [], []
        try:
            smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20)
            smtp.login(sender, password)
        except Exception as exc:
            raise HTTPException(502, f"Không kết nối được máy chủ gửi email: {str(exc)[:200]}") from exc
        try:
            for supplied in body.rows:
                row = _net(supplied)
                employee = employees.get(norm(row.get("Tên Hệ thống")))
                if not employee:
                    failed.append({"employee": str(row.get("Tên Hệ thống") or ""), "error": "Không khớp hồ sơ nhân viên"})
                    continue
                row.update({"Tên Hệ thống": employee["username"], "Họ và tên": employee["full_name"], "Email": employee["email"], "Số tài khoản ngân hàng": employee["bank_account"], "Tên ngân hàng": employee["bank_name"]})
                recipient = str(employee["email"] or "").strip()
                name = str(employee["full_name"] or employee["username"])
                if not recipient or "@" not in recipient:
                    failed.append({"employee": name, "error": "Chưa có email hợp lệ"})
                    continue
                message = EmailMessage()
                message["Subject"] = _payroll_email_subject(employee["username"], body.start, body.end)
                message["From"] = formataddr(("VERA SPA", sender))
                message["To"] = recipient
                employee_violations = violations_by_employee.get(norm(employee["username"]), [])
                message.set_content(_payroll_email_text(
                    name, body.start, body.end, row, employee_violations,
                ))
                message.add_alternative(_payroll_email_html(
                    name, body.start, body.end, row, employee_violations,
                ), subtype="html")
                attachment = _workbook([row], [field for field in DRAFT_FIELDS if field != "Email"], "Bảng lương")
                message.add_attachment(attachment, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"Bang_luong_{row.get('Tên Hệ thống','nhan_vien')}.xlsx")
                try:
                    smtp.send_message(message)
                    sent.append(name)
                except Exception as exc:
                    failed.append({"employee": name, "error": str(exc)[:200]})
        finally:
            try:
                smtp.quit()
            except Exception:
                pass
        return {"ok": not failed, "sent": sent, "failed": failed, "message": f"Đã gửi {len(sent)} email; lỗi {len(failed)}."}
