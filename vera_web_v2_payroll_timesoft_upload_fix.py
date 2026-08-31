"""Repair TimeSoft payroll uploads whose XLSX worksheet dimension is stale.

Some TimeSoft exports contain thousands of rows but advertise the worksheet
range as only ``A1:K4``.  ``openpyxl`` read-only mode trusts that stale range,
so the canonical payroll parser used to read only the first data row (often a
90' service line) and incorrectly report that there were no Tip rows.

This startup patch keeps the canonical payroll calculation intact while
replacing only the XLSX source reader.  It resets read-only worksheet
dimensions before streaming the real rows and slightly hardens the Tip prefix
pattern used by the existing calculator.
"""
from __future__ import annotations

from io import BytesIO
import numbers
from typing import Any

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import vera_web_v2_payroll as _payroll


RELEASE = "payroll-timesoft-upload-dimension-fix-2026-09-01.1"
TIP_ITEM_PATTERN = r"^tip(?:$|[_\-\s])"


def _read_source_resilient(content: bytes) -> pd.DataFrame:
    if not content:
        raise HTTPException(400, "File Excel đang trống.")
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(413, "File Excel vượt quá 15 MB.")
    if not content.startswith(b"PK"):
        raise HTTPException(400, "File không đúng định dạng Excel .xlsx. Vui lòng xuất lại từ TimeSoft.")

    workbook = None
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet_name = "Báo cáo doanh thu hóa đơn"
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(400, f"File TimeSoft không có sheet '{sheet_name}'.")
        worksheet = workbook[sheet_name]

        # TimeSoft may write a stale <dimension ref="A1:K4"> even when the
        # sheet actually contains thousands of rows.  ReadOnlyWorksheet uses
        # that metadata as its iteration boundary unless reset_dimensions()
        # is called first.
        reset_dimensions = getattr(worksheet, "reset_dimensions", None)
        if callable(reset_dimensions):
            reset_dimensions()

        expected_headers = list(_payroll.TIMESOFT_PAYROLL_HEADERS)
        header_row = next(
            worksheet.iter_rows(
                min_row=3,
                max_row=3,
                min_col=1,
                max_col=len(expected_headers),
                values_only=True,
            ),
            (),
        )
        actual_headers = [str(value or "").strip() for value in header_row]
        if actual_headers != expected_headers:
            mismatches = [
                f"{get_column_letter(index + 1)}: '{actual}' ≠ '{expected}'"
                for index, (actual, expected) in enumerate(zip(actual_headers, expected_headers))
                if actual != expected
            ]
            if len(actual_headers) < len(expected_headers):
                mismatches.append(
                    f"Thiếu cột: cần {len(expected_headers)} cột A:K, đọc được {len(actual_headers)} cột."
                )
            raise HTTPException(
                400,
                "File TimeSoft không đúng header chuẩn ở dòng 3. " + " | ".join(mismatches[:6]),
            )

        selected_rows: list[tuple[Any, Any, Any, Any]] = []
        for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=9, values_only=True):
            padded = tuple(row) + (None,) * max(0, 9 - len(row))
            time_value, item_value, amount_value, employee_value = (
                padded[1], padded[5], padded[6], padded[8]
            )
            if any(
                value not in (None, "")
                for value in (time_value, item_value, amount_value, employee_value)
            ):
                selected_rows.append((time_value, item_value, amount_value, employee_value))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được sheet 'Báo cáo doanh thu hóa đơn': {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()

    output = pd.DataFrame(selected_rows, columns=["time", "item", "amount", "employee"])
    if output.empty:
        raise HTTPException(400, "File TimeSoft không có dữ liệu từ dòng 4 trở xuống.")

    numeric_input = output["time"].apply(
        lambda value: isinstance(value, numbers.Number) and not isinstance(value, bool)
    )
    numeric_time = pd.to_numeric(output["time"], errors="coerce")
    output["time"] = pd.to_datetime(
        output["time"], dayfirst=True, errors="coerce", format="mixed"
    )
    numeric_mask = numeric_input & numeric_time.between(1, 100_000, inclusive="both")
    if numeric_mask.any():
        output.loc[numeric_mask, "time"] = (
            pd.Timestamp("1899-12-30")
            + pd.to_timedelta(numeric_time[numeric_mask], unit="D")
        )

    output["amount"] = output["amount"].apply(_payroll._number)
    output["item"] = output["item"].fillna("").astype(str).str.strip()
    output["employee"] = output["employee"].fillna("").astype(str).str.strip()
    return output


def install_payroll_timesoft_upload_fix() -> None:
    """Patch the canonical payroll module once for every Web V2 calculator."""
    if getattr(_payroll, "_timesoft_upload_dimension_fix_installed", False):
        return
    _payroll._read_source = _read_source_resilient
    _payroll.TIP_ITEM_PATTERN = TIP_ITEM_PATTERN
    _payroll._timesoft_upload_dimension_fix_installed = True
    _payroll._timesoft_upload_dimension_fix_release = RELEASE
