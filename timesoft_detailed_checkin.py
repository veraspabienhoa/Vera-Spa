"""Augment TimeSoft SearchElastic attendance with raw FaceID history.

TimeSoft's SearchElastic report keeps one aggregate row per employee/day and can
collapse intermediate FaceID scans into only first/last values. VERA needs the
raw events to reconstruct mid-shift break out/return pairs, so this patch reads
the site's own ExportCheckinLogElastic workbook and appends standardized raw
rows to the dataframe returned by ``timesoft_sync_job.fetch_checkin``.

The TimeSoft workbook can publish a stale worksheet dimension. It must be read
with a normal openpyxl worksheet (``read_only=False``), otherwise only the first
few rows can be visible even when the XLSX contains the complete XML data.
"""
from __future__ import annotations

from datetime import date, datetime, time as dt_time
from io import BytesIO
import json
import re
import time
import unicodedata
from urllib.parse import urljoin
import zipfile

import pandas as pd
from openpyxl import load_workbook


RELEASE = "timesoft-detailed-checkin-2026-09-02-v1"
EXPORT_PATH = "/Report/ReportEmployeeCheckin/ExportCheckinLogElastic"
EXPORT_TIMEOUT_SECONDS = 120
DETAIL_RETRIES = 2


def _norm(value) -> str:
    raw = unicodedata.normalize("NFD", str(value or "").strip().lower())
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    raw = raw.replace("đ", "d")
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return " ".join(raw.split())


def _display(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, dt_time):
        return value.strftime("%H:%M:%S")
    return " ".join(str(value).strip().split())


def _event_text(value, target_date: date) -> str:
    text = _display(value)
    if re.fullmatch(r"\d{1,2}:\d{2}(?::\d{2})?", text):
        return f"{target_date.strftime('%d/%m/%Y')} {text}"
    return text


def _header_map(values) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        key = _norm(value)
        if key and key not in result:
            result[key] = index
    return result


def _first_index(headers: dict[str, int], aliases) -> int | None:
    for alias in aliases:
        key = _norm(alias)
        if key in headers:
            return headers[key]
    return None


def _find_detail_sheet(workbook):
    candidates = []
    for ws in workbook.worksheets:
        max_scan = min(max(int(ws.max_row or 0), 1), 20)
        for row_number in range(1, max_scan + 1):
            values = [cell.value for cell in ws[row_number]]
            headers = _header_map(values)
            name_idx = _first_index(headers, ("Tên nhân viên", "Ten nhan vien", "EmployeeName"))
            time_idx = _first_index(headers, ("Thời gian", "Thời gian checkin", "Gio checkin", "Checkin time"))
            if name_idx is None or time_idx is None:
                continue
            normalized_sheet = _norm(ws.title)
            score = 0
            if "lich su" in normalized_sheet or "chi tiet" in normalized_sheet:
                score += 100
            if _first_index(headers, ("Mã nhân viên", "Ma nhan vien", "EmployeeCode")) is not None:
                score += 50
            if _first_index(headers, ("Thời gian checkout", "Checkout time")) is not None:
                score -= 100
            candidates.append((score, ws, row_number, headers))
    if not candidates:
        raise RuntimeError("TimeSoft check-in log XLSX không có bảng chi tiết FaceID nhận diện được.")
    candidates.sort(key=lambda item: item[0], reverse=True)
    _, ws, header_row, headers = candidates[0]
    return ws, header_row, headers


def _parse_workbook(content: bytes, target_date: date) -> pd.DataFrame:
    bio = BytesIO(content)
    if not zipfile.is_zipfile(bio):
        raise RuntimeError("TimeSoft ExportCheckinLogElastic không trả file XLSX hợp lệ.")
    bio.seek(0)
    workbook = load_workbook(bio, read_only=False, data_only=True)
    try:
        ws, header_row, headers = _find_detail_sheet(workbook)
        name_idx = _first_index(headers, ("Tên nhân viên", "Ten nhan vien", "EmployeeName"))
        time_idx = _first_index(headers, ("Thời gian", "Thời gian checkin", "Gio checkin", "Checkin time"))
        date_idx = _first_index(headers, ("Ngày", "Ngay", "WorkDate"))
        employee_code_idx = _first_index(headers, ("Mã nhân viên", "Ma nhan vien", "EmployeeCode"))
        enroll_idx = _first_index(headers, ("Mã chấm công", "Ma cham cong", "EnrollNumber"))
        phone_idx = _first_index(headers, ("Số điện thoại", "So dien thoai", "Mobile"))
        stt_idx = _first_index(headers, ("STT",))

        rows = []
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            employee = _display(values[name_idx]) if name_idx is not None and name_idx < len(values) else ""
            event = _event_text(values[time_idx], target_date) if time_idx is not None and time_idx < len(values) else ""
            if not employee or not event:
                continue
            work_date = _display(values[date_idx]) if date_idx is not None and date_idx < len(values) else ""
            if not work_date:
                match = re.match(r"\s*(\d{1,2}/\d{1,2}/\d{4})\b", event)
                work_date = match.group(1) if match else target_date.strftime("%d/%m/%Y")
            row = {
                "EmployeeName": employee,
                "WorkDateStr": work_date,
                "MachineTimeStr": event,
                "TimeSoftRawSource": "ExportCheckinLogElastic",
            }
            if employee_code_idx is not None and employee_code_idx < len(values):
                row["EmployeeCode"] = _display(values[employee_code_idx])
            if enroll_idx is not None and enroll_idx < len(values):
                row["EnrollNumber"] = _display(values[enroll_idx])
            if phone_idx is not None and phone_idx < len(values):
                row["Mobile"] = _display(values[phone_idx])
            if stt_idx is not None and stt_idx < len(values):
                row["TimeSoftRawSTT"] = _display(values[stt_idx])
            rows.append(row)
        return pd.DataFrame(rows)
    finally:
        workbook.close()


def fetch_detailed_checkin(ts, session, target_date: date) -> pd.DataFrame:
    url = urljoin(ts.BASE_URL + "/", EXPORT_PATH.lstrip("/"))
    referer = urljoin(ts.BASE_URL + "/", ts.REPORT_CHECKIN_PAGE.lstrip("/"))
    object_search = {
        "CreateDateRange": ts._date_range_text(target_date, target_date),
        "isSortByEmp": False,
    }
    response = session.get(
        url,
        params={
            "objectSearchStr": json.dumps(
                object_search,
                ensure_ascii=False,
                separators=(",", ":"),
            )
        },
        headers={
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "Referer": referer,
        },
        timeout=EXPORT_TIMEOUT_SECONDS,
        allow_redirects=True,
    )
    if response.status_code in {401, 403} or "/user/login" in str(response.url or "").lower():
        raise RuntimeError("Phiên TimeSoft hết hạn khi tải lịch sử FaceID chi tiết.")
    response.raise_for_status()
    return _parse_workbook(response.content, target_date)


def install(ts) -> None:
    """Patch ``ts.fetch_checkin`` so all VERA snapshot/live readers get raw FaceID."""
    if getattr(ts, "_detailed_checkin_patch_release", "") == RELEASE:
        return

    original_fetch_checkin = ts.fetch_checkin

    def fetch_checkin_with_detail(session, target_date):
        summary_df, meta = original_fetch_checkin(session, target_date)
        summary_rows = int(len(summary_df)) if isinstance(summary_df, pd.DataFrame) else 0
        raw_df = pd.DataFrame()
        last_error = None
        for attempt in range(1, DETAIL_RETRIES + 1):
            try:
                raw_df = fetch_detailed_checkin(ts, session, target_date)
                if not (summary_rows > 0 and raw_df.empty):
                    break
                last_error = RuntimeError(
                    "TimeSoft có dữ liệu SearchElastic nhưng ExportCheckinLogElastic trả 0 FaceID."
                )
            except Exception as exc:
                last_error = exc
            if attempt < DETAIL_RETRIES:
                time.sleep(1.0 * attempt)

        if summary_rows > 0 and (not isinstance(raw_df, pd.DataFrame) or raw_df.empty):
            raise RuntimeError(
                "Không lấy được lịch sử FaceID chi tiết TimeSoft; từ chối ghi snapshot tổng hợp thiếu mốc giữa ca: "
                f"{type(last_error).__name__ if last_error else 'RuntimeError'}: {last_error or 'raw log empty'}"
            )

        if not isinstance(summary_df, pd.DataFrame) or summary_df.empty:
            combined = raw_df.copy() if isinstance(raw_df, pd.DataFrame) else pd.DataFrame()
        elif not isinstance(raw_df, pd.DataFrame) or raw_df.empty:
            combined = summary_df.copy()
        else:
            combined = pd.concat([summary_df, raw_df], ignore_index=True, sort=False)

        out_meta = dict(meta or {})
        out_meta.update({
            "SummaryRows": summary_rows,
            "RawLogRows": int(len(raw_df)) if isinstance(raw_df, pd.DataFrame) else 0,
            "CombinedRows": int(len(combined)),
            "DetailedLogReady": True,
            "DetailedLogRelease": RELEASE,
        })
        return combined, out_meta

    ts.fetch_detailed_checkin = lambda session, target_date: fetch_detailed_checkin(ts, session, target_date)
    ts.fetch_checkin = fetch_checkin_with_detail
    ts._detailed_checkin_patch_release = RELEASE
